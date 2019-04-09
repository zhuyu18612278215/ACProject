#coding=utf-8
from django.utils.translation import ugettext as _
from django.utils.translation import ugettext_lazy as lazy_
from django.shortcuts import render
from django.contrib import auth
from django.contrib.auth import login,authenticate
from account.models import NewUser,Account_Group,Auto_Update_Rule
from django.contrib.auth.decorators import login_required
from device.models import Device,Probe_config,Probe_audit_basic_status,Probe_audit_dev_status,Probe_audit_place_status,Probe_event,Probe_group
import models
import json
from django.http import JsonResponse
from django.forms.models import model_to_dict
from django.http import HttpResponse,HttpResponseRedirect
from models import Device_ap,AP_event,Group_wlan,Device_wlan,User_policy_config,Guest_policy,ApBlackList,Gpon,Group_gpon,Setting_gpon,Customer,Customer_name,Customer_black_list,Customer_history,Customer_white_list,Customer_black_white_switch,Timing_Policy
from django.utils import timezone
import datetime
from django.core.urlresolvers import reverse
from django.shortcuts import redirect
import time
import pytz
import urllib
import urllib2
import cookielib
import hashlib
import re
# Create your views here.
import sys
sys.path.append("DataQuery")
import DataQuery
import os
import os.path
from system.models import Setting,Page_limit,Oem_limit
from django.core.files import File
import Public_function
import logging
from django.utils.timezone import is_naive, make_aware, utc, is_aware
from django.db.models import Q
from default_settions import user_policy_config,guest_policy_config
import random
import MySQLdb
import MySQLdb.cursors
from django.db import connection
from django.conf import settings as settings_py
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from django.core.servers.basehttp import FileWrapper
import collections
reload(sys)
sys.setdefaultencoding('utf-8')
import logging
# Create your views here.
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
fh = logging.FileHandler("./statics/log/aplog.log")
fh.setLevel(logging.DEBUG)
formater = logging.Formatter('[%(levelname)s][%(asctime)s]--## %(message)s')
fh.setFormatter(formater)
logger.addHandler(fh)

from ap.tasks import task_for_ap_del_wlan,task_for_modify_ap_group_ajax,task_for_add_wlan_ajax,task_for_modify_wlan_ajax,task_for_global_config_ajax,task_for_user_policy_config_ajax,default_config_issued
from nonoperate.tasks import default_config_issued as nonoperate_default_config_issued


def get_plimit():
    if Page_limit.objects.filter(pk = 1).exists():
        plimit = model_to_dict(Page_limit.objects.get(pk = 1))
    else:
        plimit = Public_function.page_limit_dict
    return plimit

def get_oemlimit():
    if Oem_limit.objects.filter(pk = 1).exists():
        oemlimit = model_to_dict(Oem_limit.objects.get(pk = 1))
    else:
        oemlimit = Public_function.oem_limit_dict
    return oemlimit

def upgrade_file_list():
    django_settings = Setting.objects.get(SIGN = 1)
    try:
        # url = "http://115.28.241.216:81/version/byz/"
        url = django_settings.VERSION_SERVER_URL+"version/newap/"
        print url
        website = urllib2.urlopen(url)

        html = website.read()

        # links = re.findall('href="(.*)">witfios', html)
        links = re.findall('href="(.*)">[a-z]', html)
        # print links
        a = []
        for i in links:
            if i[-4:] == '.bin':
                if 'kernel' in i or 'rootfs' in i:
                    pass
                else:
                    a.append(i)
        # print a
        return a
    except Exception as e:
        print e,'urlerror'
        return False

def upgrade_file(dev_type,fl):
    r = dev_type_change(dev_type)
    result = ''
    print r,fl,'uf'
    if r == False or fl == False:
        pass
    else:
        for i in fl:
            if i == r :
                result = i
        if result != "":
            return result

    return False

def dev_type_change(dev_type):
    try:
        f = open('./statics/upgrade/newap_upgrade_dict.info','r')
        try:
            data = json.load(f,'utf-8')
            return data[dev_type]
        except Exception as e:
            print 'dev version file:',e
            return False
        finally:
            f.close()
    except Exception as e:
        print 'dev version file:',e
        return False

def compare_ver(file1,ver1):
    print 'version:',file1,ver1
    if file1 == False:
        return False
    return file1 != str(ver1)

def ap_event_save(pe):
    ap_event = AP_event()
    ap_event.event_time = pe["event_time"]
    ap_event.event = pe['event']
    ap_event.msg = pe['msg']
    ap_event.admin_username = pe['admin_username']
    ap_event.ap_mac = pe['ap_mac']
    try:
        ap_event.save()
        return True
    except Exception as e:
        print e
        return False


@login_required
def ap_access(request):
    errors = ''
    # 接受到的mac
    r_mac = request.GET.get('mac')
    # 数据信息
    data = {'mac':r_mac,'name':'','model':'X-300','sn':'00000001','lastip':'10.200.10.1:20000','privateip':'192.168.1.101','currstanum':'0','version':'2.0001','last_heart_time':'2017-04-07 15:30:00','upload':'0','download':'0','radios_type':'0','own_model':'WitFi-DAP510E'}
    #
    # data = DataQuery.DQGetAccessDev("ap", r_mac)
    if Device.objects.filter(Q(mac = r_mac)&Q(support_mode = '1')).exists():
        errors = _(u"设备已存在！")
    if len(errors) == 0:
        if Device.objects.filter(Q(mac = r_mac)&~Q(support_mode = '1')).exists() :
            Device.objects.get(Q(mac = r_mac)&~Q(support_mode = '1')).delete()
            # try:
            #     Probe_config.objects.get(mac = r_mac).delete()
            # except Exception as e:
            #     print e
            # try:
            #     Probe_audit_basic_status.objects.get(mac = r_mac).delete()
            # except Exception as e:
            #     print e
            # try:
            #     Probe_audit_dev_status.objects.get(mac = r_mac).delete()
            # except Exception as e:
            #     print e
            # try:
            #     Probe_audit_place_status.objects.get(mac = r_mac).delete()
            # except Exception as e:
            #     print e
        ap = Device()
        ap.mac = data['mac']
        if data['name'] == '':
            ap.name = data['model'] + '_' + data['mac'][-6:]
        else:
            ap.name = data['name']
        ap.model = data['model']
        ap.sn = data['sn']
        ap.lastip = data['lastip']
        ap.privateip = data['privateip']
        ap.version = data['version']
        ap.last_heart_time = data['last_heart_time']
        ap.upload = data['upload']
        ap.download = data['download']
        ap.up_pkts = '0'#data['up_pkts']
        ap.down_pkts = '0'#data['down_pkts']
        ap.own_model = data['own_model']
        ap.support_mode = '1'

        if ap is not None:
            try:
                ap.save()

                p = Device.objects.get(mac = r_mac)
                d = Device_ap()
                d.device = p
                d.apusernum = 0#data['apusernum']
                d.guestsnum = 0#data['guestsnum']
                d.radios_type = '0'#data['radios_type']
                d.radios_2_channel = 'auto'#data['radios_2_channel']
                d.radios_2_ht = '20'#data['radios_2_ht']
                d.radios_2_power = 'auto'#data['radios_2_power']
                d.radios_2_com = 'auto'#data['radios_2_com']
                d.radios_5_channel = 'auto'#data['radios_5_channel']
                d.radios_5_ht = '20'#data['radios_5_ht']
                d.radios_5_power = 'auto'#data['radios_5_power']
                d.radios_2_currstanum = 0#data['radios_2_currstanum']
                d.radios_2_guestsnum = 0#data['radios_2_guestsnum']
                d.radios_5_currstanum = 0#data['radios_5_currstanum']
                d.radios_5_guestsnum = 0#data['radios_5_guestsnum']
                d.save()

                DataQuery.DQSetAccessed("ap", r_mac)
                errors = _(u'准入成功！')
                event = {
                    'event_time' : timezone.now(),
                    'event' : 'AP_ADMITED_BY_ADMIN',
                    'msg' : 'Device['+ap.mac+'] was admited by Admin['+request.user.username+']',
                    'admin_username' : request.user.username,
                    'ap_mac' : ap.mac,
                }

                accessed_device2authpuppy(r_mac)

                try:
                    ap_event_save(event)
                except Exception as e:
                    print e
            except Exception as e:
                errors = _(u'准入失败！')

    request.session['error_message'] = errors
    return redirect('/ap-list/')

@login_required
def ap_list(request):
    global ap_ugfl
    DataQuery.ap_ugfl = upgrade_file_list()
    errors = ''

    try:
        errors = request.session.get('error_message')
        # print errors
        del request.session['error_message']
    except Exception as e:
        pass
    errors_json = json.dumps(errors)
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'ap-list.html',{'errors_json':errors_json,'errors':errors,'admin_error_json':admin_error_json,'admin_error':admin_error,'plimit':plimit,'oemlimit':oemlimit})

def ap_wait_access_ajax(request):
    wait_access_list = ''
    if request.user.administrator_permission > 4 or request.user.administrator_permission == 0 :
        # p_all = Device.objects.filter(support_mode = '1')
        wait_access_list = DataQuery.DQGetAccess("ap")
        #

        # ap1 = {'mac':'112233445566','model':'X-300','privateip':'192.168.1.101','lastip':'10.200.10.1:20000','version':'2.00000001','last_heart_time':'2017-04-07 15:30:00'}
        # wait_access_list = [ap1]

    ret = {'data':wait_access_list}
    return JsonResponse(ret,safe = False)

def ap_already_access_ajax(request):
    # print 2131231,request.GET
    draw = int(request.GET.get('draw'))
    start = int(request.GET.get('start'))
    length = int(request.GET.get('length'))
    order_id = request.GET.get('order[0][column]')
    order = request.GET.get('columns['+order_id+'][data]')

    device_type_receive = request.GET.get('device_type')
    dev_online_time = timezone.now() - datetime.timedelta(minutes = 15)
    if device_type_receive == 'online':
        time_condition = Q(last_heart_time__gte = dev_online_time)
    elif device_type_receive == 'offline':
        time_condition = Q(last_heart_time__lt = dev_online_time)
    else:
        time_condition = Q(last_heart_time__gte = dev_online_time)|~Q(last_heart_time__gte = dev_online_time)

    if request.GET.get('order[0][dir]') == 'asc':
        order_type = ''
    elif request.GET.get('order[0][dir]') == 'desc':
        order_type = '-'

    if order == 'state':
        order = 'last_heart_time'

    # print '#######',order_type,'#######',order
    search_value_old = request.GET.get('search[value]')
    search_value = search_value_old
    # search_value = ''
    # for s in search_value_old:
    #     if s>= u"\u4e00" and s<= u"\u9fa5":
    #         pass
    #     else:
    #         search_value = search_value + s

    if request.user.administrator_permission >= 4 or request.user.administrator_permission == 0 :
        recordsTotal = Device.objects.filter(support_mode = '1').filter(time_condition).count()
        if search_value == '':
            recordsFiltered = recordsTotal
            already_access_list = Device.objects.filter(support_mode = '1').filter(time_condition).order_by(order_type+order)[start:start+length]
        else:
            if re.match(r'[\u4e00 -\u9fa5]+',search_value) == None :
                if search_value in [u'退服',u'离线',u'超时',u'在线']:
                    sv = {
                        u'在线':'online',
                        u'超时':'time out',
                        u'离线':'offline',
                        u'退服':'retired',
                    }
                    already_access_list = Device.objects.filter(support_mode = '1').filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = Device.objects.filter(support_mode = '1').filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value)).count()
                else:
                    already_access_list = Device.objects.filter(support_mode = '1').filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = Device.objects.filter(support_mode = '1').filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value)).count()
            else:
                already_access_list = Device.objects.filter(support_mode = '1').filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-'))))[start:start+length]
                recordsFiltered = Device.objects.filter(support_mode = '1').filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-')))).count()
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        recordsTotal = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1')).filter(time_condition).count()
        if search_value == '':
            recordsFiltered = recordsTotal
            already_access_list = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1')).filter(time_condition).order_by(order_type+order)[start:start+length]
        else:
            if re.match(r'[\u4e00 -\u9fa5]+',search_value) == None :
                if search_value in [u'退服',u'离线',u'超时',u'在线']:
                    sv = {
                        u'在线':'online',
                        u'超时':'time out',
                        u'离线':'offline',
                        u'退服':'retired',
                    }
                    already_access_list = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value)).count()
                else:
                    already_access_list = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1')).filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1')).filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value)).count()
            else:
                already_access_list = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-'))))[start:start+length]
                recordsFiltered = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-')))).count()
    DataQuery.DQUpdateAccessed("ap", already_access_list)
    timenow = timezone.now()
    ret = {'draw':draw,'recordsTotal':recordsTotal,'recordsFiltered':recordsFiltered,'data':[]}
    for al in already_access_list:
        a = int((timenow - al.last_heart_time).total_seconds())

        al.last_heart_time = timezone.localtime(al.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")

        al.upgrade_button = ''
        rss = al.reboot_sign.split(",")
        if al.last_heart_time == "1970-1-1 08:00:00":
            al.state = _(u"离线")
        elif a >= 86400:
            al.state = _(u"退服")
        elif a >= 3600:
            al.state = _(u"离线")
        elif a >= 900 :
            al.state = _(u"超时")
        elif a < 900:
            if rss[0] != '0' and rss[0] != '':
                if rss[0] == '1':
                    al.state = _(u"重启")
                elif rss[0] == '2':
                    al.state = _(u"升级")
            else:
                al.state = _(u"在线")
        try:
            print 'update conf:',al.own_model,al.version,dev_type_change(al.own_model)
            if compare_ver(''.join(re.findall("[0-9\.]",''.join(re.findall("(\d\.\d+\.\d+\.\w\d+)",dev_type_change(al.own_model)[:-4])))),al.version) == True:
                al.upgrade_button = True
                al.upgrade_version = dev_type_change(al.own_model)
            print 'update:',al.upgrade_button,al.upgrade_version
        except Exception as e:
            print 'update exc:',e
            al.upgrade_button = False
            al.upgrade_version = ''

        admin_power_control = ""
        if request.user.administrator_permission >= 5:
            if al.account_group_name == 'admin' or al.account_group_name == '':
                admin_power_control = ""
            else:
                admin_power_control = "ban"

        aall = model_to_dict(al)
        aall['admin_power_control'] = admin_power_control
        cpuInfo = DataQuery.DQGetAPCpu("ap",al.mac) or ""
        aall['cpu'] = cpuInfo

        try:
            gpn = Probe_group.objects.get(pk = aall['group_id']).group_name
        except Exception as e:
            print e
            gpn = "DefaultGroup"
        if aall['account_group_name'] == "":
            acpn = "admin"
        else:
            acpn = aall['account_group_name']
        aall['group_id'] = acpn + "/" + gpn
        ret['data'].append(aall)
    return JsonResponse(ret,safe = False)

@login_required
def ap_reboot(request):
    errors = ''
    r_mac = request.POST.get('mac')
    data = {'action':'reboot','mac':r_mac}
    #执行动作
    try:
        #在此调用函数
        DataQuery.DQProcess(data)
        errors = _(u'执行成功！')
        event = {
            'event_time' : timezone.now(),
            'event' : 'AP_RESTARTED_BY_ADMIN',
            'msg' : 'Device['+r_mac+'] was restarted by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'ap_mac' : r_mac,
        }
        try:
            nu = Device.objects.get(mac = r_mac)
            nu.reboot_sign = '1'+','+str(int(time.mktime(datetime.datetime.now().timetuple())))
            nu.save()
        except Exception as e:
            print e
        try:
            # DataQuery.LOGGING_INIT()
            # logging.info('Device['+r_mac+'] was restarted by Admin['+request.user.username+']')
            # logging.info('GGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGG')
            # print 'Device['+r_mac+'] was restarted by Admin['+request.user.username+']'
            ap_event_save(event)
        except Exception as e:
            print e
    except Exception as e:
        print e
        errors = _(u'执行失败！')
    return JsonResponse(errors,safe=False)

@login_required
def ap_update(request):
    django_settings = Setting.objects.get(SIGN = 1)
    errors = ''
    r_mac = request.POST.get('mac')
    #data = {'action':'upgrade','mac':r_mac, 'param':'http://115.28.241.216/img/witfios3.01.03.r191710.bin'}
    u = upgrade_file(Device.objects.get(mac = r_mac).own_model,DataQuery.ap_ugfl)
    print u,'url'
    if u == False:
        errors = _(u'没有此型号对应的版本文件')
    else:
        data = {'action':'upgrade','mac':r_mac,'param':django_settings.VERSION_SERVER_URL+"version/newap/"+u}
        #执行动作
        try:
            #在此调用函数
            DataQuery.DQProcess(data)
            errors = _(u'执行成功！')
            if u != False:
                event = {
                    'event_time' : timezone.now(),
                    'event' : 'AP_UPGRADE_BY_ADMIN_TO_VERSION',
                    'msg' : 'Device['+r_mac+'] was upgrade to version['+u+'] by Admin['+request.user.username+']',
                    'admin_username' : request.user.username,
                    'ap_mac' : r_mac,
                }
                try:
                    nu = Device.objects.get(mac = r_mac)
                    nu.reboot_sign = '2'+','+str(int(time.mktime(datetime.datetime.now().timetuple())))
                    nu.save()
                except Exception as e:
                    print e
                try:
                    ap_event_save(event)
                except Exception as e:
                    print e

        except Exception as e:
            print e
            errors = _(u'执行失败！')
    return JsonResponse(errors,safe=False)

@login_required
def ap_kickmac(request):
    errors = ''
    apmac = request.POST.get('apmac')
    r_mac = request.POST.get('mac')
    data = {'action':'kickmac', 'mac':apmac, 'param':r_mac}
    print "UUUUUUUUUUUUUUUUUUUUUU", apmac, r_mac
    #执行动作
    try:
        #在此调用函数
        Customer.objects.get(mac = r_mac).delete()
        errors = _(u'执行成功！')
        event = {
            'event_time' : timezone.now(),
            'event' : 'USER_FORGOTEN_BY_ADMIN',
            'msg' : 'Device['+r_mac+'] was forgoten by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'user_mac' : r_mac,
        }
        try:
            DataQuery.DQProcess(data)

            ap_event_save(event)
        except Exception as e:
            print e
    except Exception as e:
        print e
        errors = _(u'执行失败！')

    return JsonResponse(errors,safe = False)

@login_required
def ap_del(request):
    errors = ''
    r_mac = request.POST.get('mac')
    data = {'action':'del','mac':r_mac}
    #执行动作
    try:
        #在此调用函数
        Device.objects.get(mac = r_mac).delete()
        try:
            Probe_config.objects.filter(mac = r_mac).delete()
            models.Gpon.objects.filter(ap_mac = r_mac).delete()

        except Exception as e:
            print e
        errors = _(u'执行成功！')
        event = {
            'event_time' : timezone.now(),
            'event' : 'AP_FORGOTEN_BY_ADMIN',
            'msg' : 'Device['+r_mac+'] was forgoten by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'ap_mac' : r_mac,
        }
        try:
            DataQuery.DQDelAccessed("ap", r_mac)
            data = {'action':'reset','mac':r_mac}
            DataQuery.DQProcess(data)

            del_device2authpuppy(r_mac)

            ap_event_save(event)
        except Exception as e:
            print e
    except Exception as e:
        print e
        errors = _(u'执行失败！')
    return JsonResponse(errors,safe = False)

def ap_customers_get_by_mac(apmac):
    mac = apmac
    online_time = int(time.time()) - 900

    print "1212121212121212121212121212", mac
    customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap = mac)&Q(portal_enable = '0')).count()

    return customer_list

def ap_guests_get_by_mac(apmac):
    mac = apmac
    online_time = int(time.time()) - 900

    print "1212121212121212121212121212", mac
    customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap = mac)&Q(portal_enable = '1')).count()

    return customer_list

@login_required
def ap_detail(request):
    r_mac = request.GET.get('mac')
    r = Device.objects.get(mac = r_mac)
    name = r.name
    vpn = r.vpn
    vpnip = r.vpnip

    upload = r.upload
    download = r.download
    up_pkts = r.up_pkts
    down_pkts = r.down_pkts
    apusernum = 0
    guestsnum = 0
    if hasattr(r,'device_ap'):
        # apusernum = r.device_ap.apusernum
        # guestsnum = r.device_ap.guestsnum
        guestsnum = ap_guests_get_by_mac(r_mac)
        apusernum = ap_customers_get_by_mac(r_mac)
        radios_type = r.device_ap.radios_type
        radios_2_channel = r.device_ap.radios_2_channel
        radios_2_ht = r.device_ap.radios_2_ht
        radios_2_power = r.device_ap.radios_2_power
        radios_2_com = r.device_ap.radios_2_com
        radios_5_channel = r.device_ap.radios_5_channel
        radios_5_ht = r.device_ap.radios_5_ht
        radios_5_power = r.device_ap.radios_5_power
        radios_2_currstanum = r.device_ap.radios_2_currstanum
        radios_2_guestsnum = r.device_ap.radios_2_guestsnum
        radios_5_currstanum = r.device_ap.radios_5_currstanum
        radios_5_guestsnum = r.device_ap.radios_5_guestsnum
    else:
        guestsnum = ap_guests_get_by_mac(r_mac)
        apusernum = ap_customers_get_by_mac(r_mac)
        try:
            radios_type = Public_function.device_type_info_dict[r.own_model]['radio_num']
        except Exception as e:
            print e
            radios_type = '0'
        radios_2_channel = 'auto'
        radios_2_ht = '20'
        radios_2_power = 'auto'
        radios_2_com = 'auto'
        radios_5_channel = 'auto'
        radios_5_ht = '20'
        radios_5_power = 'auto'
        radios_2_currstanum = 0
        radios_2_guestsnum = 0
        radios_5_currstanum = 0
        radios_5_guestsnum = 0

    try:
        d = Probe_config.objects.get(mac = r_mac)
        # print d,1111111111
        ac_address = d.ac_address
        log_address = d.log_address
        ip_model = d.ip_model
        ip_address = d.ip_address
        subnet_mask = d.subnet_mask
        gateway = d.gateway
        preferred_dns = d.preferred_dns
        alternative_dns = d.alternative_dns
        preferred_ntp = d.preferred_ntp
        alternative_ntp = d.alternative_ntp
    except Exception as e:
        print e
        ac_address = ''
        log_address = ''
        ip_model = ''
        ip_address = ''
        subnet_mask = ''
        gateway = ''
        preferred_dns = ''
        alternative_dns = ''
        preferred_ntp = ''
        alternative_ntp = ''

    admin_power_control = ""
    if request.user.administrator_permission >= 5:
        if r.account_group_name == 'admin' or r.account_group_name == '':
            admin_power_control = ""
        else:
            admin_power_control = "ban"
    data = {'mac':r_mac,'sn':'','model':'','version':'','ip':'','running_time':'','name':name,'last_time':'','cpu':'','memory':'','flash':'','apusernum':apusernum,'guestsnum':guestsnum,'upload':upload,'download':download,'up_pkts':up_pkts,'down_pkts':down_pkts,'radios_type':radios_type,'radios_2_channel':radios_2_channel,'radios_2_ht':radios_2_ht,'radios_2_power':radios_2_power,'radios_2_com':radios_2_com,'radios_5_channel':radios_5_channel,'radios_5_ht':radios_5_ht,'radios_5_power':radios_5_power,'radios_2_currstanum':radios_2_currstanum,'radios_2_guestsnum':radios_2_guestsnum,'radios_5_currstanum':radios_5_currstanum,'radios_5_guestsnum':radios_5_guestsnum,'ac_address':ac_address,'log_address':log_address,'ip_model':ip_model,'ip_address':ip_address,'subnet_mask':subnet_mask,'gateway':gateway,'preferred_dns':preferred_dns,'alternative_dns':alternative_dns,'vpn':vpn,'vpnip':vpnip,'preferred_ntp':preferred_ntp,'alternative_ntp':alternative_ntp,}
    DataQuery.DQUpdateAccessedDev('ap', r_mac, data)

    data["admin_power_control"] = admin_power_control
    data["locateState"] = r.locateState

    # data_str = set_wlan_str(r_mac)
    # DataQuery.DQSetUpdateConfig('wireless',r_mac,data_str)
    return JsonResponse(data)

@login_required
def ap_config(request):
    r_mac = request.GET.get('mac')
    r = Device.objects.get(mac = r_mac)

    try:
        radios_type = r.device_ap.radios_type
        radios_2_channel = r.device_ap.radios_2_channel
        radios_2_ht = r.device_ap.radios_2_ht
        radios_2_power = r.device_ap.radios_2_power
        radios_2_com = r.device_ap.radios_2_com
        radios_5_channel = r.device_ap.radios_5_channel
        radios_5_ht = r.device_ap.radios_5_ht
        radios_5_power = r.device_ap.radios_5_power
        radios_5_com = r.device_ap.radios_5_com
    except Exception as e:
        print e
        try:
            radios_type = Public_function.device_type_info_dict[r.own_model]['radio_num']
        except Exception as e:
            print e
            radios_type = '0'
        radios_2_channel = 'auto'
        radios_2_ht = '20'
        radios_2_power = 'auto'
        radios_2_com = 'auto'
        radios_5_channel = 'auto'
        radios_5_ht = '20'
        radios_5_power = 'auto'
        radios_5_com = 'auto'

    if hasattr(r,'ap_user_policy_config_extend'):
        max_user = r.ap_user_policy_config_extend.max_user
    else:
        if hasattr(Account_Group.objects.get(groupname = r.account_group_name),'user_policy_config'):
            if radios_type == '0' :
                max_user = Account_Group.objects.get(groupname = r.account_group_name).user_policy_config.dual_max_user
            else:
                max_user = Account_Group.objects.get(groupname = r.account_group_name).user_policy_config.single_max_user
        else:
            if radios_type == '0' :
                max_user = user_policy_config['dual_max_user']
            else:
                max_user = user_policy_config['single_max_user']

    try:
        r_2g_maxp = Public_function.device_type_info_dict[r.own_model]['radio_2_txp']
    except Exception as e:
        r_2g_maxp = '20'
    try:
        r_5g_maxp = Public_function.device_type_info_dict[r.own_model]['radio_5_txp']
    except Exception as e:
        r_5g_maxp = '20'

    data = {'radios_type':radios_type,'radios_2_channel':radios_2_channel,'radios_2_ht':radios_2_ht,'radios_2_power':radios_2_power,'radios_2_com':radios_2_com,'radios_5_channel':radios_5_channel,'radios_5_ht':radios_5_ht,'radios_5_power':radios_5_power,'radios_5_com':radios_5_com,'r_2g_maxp':r_2g_maxp,'r_5g_maxp':r_5g_maxp,'max_user':max_user}

    return JsonResponse(data)

@login_required
def ap_radios_config(request):
    errors = ''
    m_model = request.GET.get('modify_model')
    if request.method == "POST":
        mac = request.POST.get('mac')
        if m_model == '6':
            radios_2_channel = request.POST.get('2g_channel')
            radios_2_ht = request.POST.get('2g_ht')
            radios_2_power = request.POST.get('2G_power')
            radios_2_com = request.POST.get('radios_2_com')
            radios_5_channel = request.POST.get('5g_channel')
            radios_5_ht = request.POST.get('5g_ht')
            radios_5_power = request.POST.get('5G_power')
            radios_2_customize = request.POST.get('2g_customize')
            radios_5_customize = request.POST.get('5g_customize')
            radios_5_com = request.POST.get('radios_5_com')
            try:
                dev = Device.objects.get(mac = mac)
                if hasattr(dev,'device_ap'):
                    rad = dev.device_ap
                else:
                    rad = Device_ap()
                    rad.device_id = dev.pk

                try:
                    rad.radios_type = Public_function.device_type_info_dict[dev.own_model]['radio_num']
                except Exception as e:
                    print e
                    rad.radios_type = '0'
                rad.radios_2_channel = radios_2_channel
                rad.radios_2_ht = radios_2_ht
                if radios_2_power == "customize":
                    rad.radios_2_power = radios_2_customize
                else:
                    rad.radios_2_power = radios_2_power
                if radios_5_power == "customize":
                    rad.radios_5_power = radios_5_customize
                else:
                    rad.radios_5_power = radios_5_power
                rad.radios_2_com = radios_2_com
                rad.radios_5_channel = radios_5_channel
                rad.radios_5_ht = radios_5_ht
                rad.radios_5_com = radios_5_com

                rad.save()

                create_wireless_json(dev)
                # data_str = set_wlan_str(mac)
                # DataQuery.DQSetUpdateConfig('wireless',mac,data_str)
                # ntpstr = preferred_ntp + ';' + alternative_ntp
                # data = {'action':'ntp','mac':mac,'param':ntpstr}
                # DataQuery.DQProcess(data)
                errors = _(u'修改成功！')
            except Exception as e:
                print e,'3213213212'
                errors = _(u'修改失败！')

    return JsonResponse(errors,safe = False)

@login_required
def ap_vpn(request):
    model = request.POST.get('model')
    mac = request.POST.get('mac')
    errors = ''
    if model == 'on':
        serverip = '115.28.241.216'
        timeval = str(int(time.time()))
        m = hashlib.md5()
        m.update(timeval+mac.upper())
        token = m.hexdigest()
        urls = "http://"+serverip+"/vpnserver/vpn_address_asign.php?t="+timeval+"&token="+token+"&mac="+mac.upper()
        result = json.loads(urllib2.urlopen(urls).read())
        try:
            if 'rc' in result['meta'] and result['meta']['rc'] == 'ok' and 'ip' in result['data'] and 'mac' in result['data'] and result['data']['mac'] == mac.upper():

                ip = result['data']['ip']
                #add
                #
                errors = result['meta']['rc']
                ret = {'result':'on','ip':ip}

                event = {
                    'event_time' : timezone.now(),
                    'event' : 'AP_START_REMOTE_CONNECT_BY_ADMIN',
                    'msg' : 'Device['+mac+'] was open remote connect by Admin['+request.user.username+']',
                    'admin_username' : request.user.username,
                    'ap_mac' : mac,
                }
                try:
                    ap_event_save(event)
                except Exception as e:
                    print e
            elif 'rc' in result['meta'] and result['meta']['rc'] == 'error':
                errors = result['meta']['rc']
                ret = {'result':'off','ip':''}
        except Exception as e:
            print e
            ret = {'result':'off','ip':''}
            ip = ''

    elif model == 'off':
        ret = {'result':'off','ip':''}
        ip = ''
        event = {
            'event_time' : timezone.now(),
            'event' : 'AP_STOP_REMOTE_CONNECT_BY_ADMIN',
            'msg' : 'Device['+mac+'] was close remote connect by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'ap_mac' : mac,
        }
        try:
            ap_event_save(event)
        except Exception as e:
            print e

    try:
        v = models.Device.objects.get(mac = mac)
        v.vpn = ret['result']
        v.vpnip = ip
        v.save()

        #send message to devs
        if model == 'on':
            ip = result['data']['ip']
            param = "115.28.241.216;1194;" + ip
            data = {'action':'vpn','mac':mac,'param':param}
        else:
            data = {'action':'vpn','mac':mac,'param':'stop'}

        DataQuery.DQProcess(data)
    except Exception as e:
        print e
    return JsonResponse(ret)


@login_required
def ap_eventajax(request):
    event_type = request.GET.get('type')
    event_time_type = request.GET.get('time')
    seconds = int(event_time_type) * 3600
    t = timezone.now() - datetime.timedelta(hours = int(event_time_type))
    p_list = []
    event = {'data':[]}

    if request.user.administrator_permission >= 4 or request.user.administrator_permission == 0 :
        probe_event = AP_event.objects.filter(event_time__gte = t).order_by('-event_time')

        if event_type == "dev":
            # event = {'data':[{'event_time':'2017/05/04 12:05:35','msg':'Device[A4E6B1300005] was onlined.','action':_(u'存档'),'probe_mac':'A4E6B1300005','event':u'','state':state_check('A4E6B1300005')},{'event_time':'2017/05/04 12:04:44','msg':'Device[8c8401163ae0] was offlined.','action':_(u"存档"),'probe_mac':'8c8401163ae0','event':u'PROBE_WAS_OFFLINED','state':state_check('8c8401163ae0')}]}
            p_list = [a.mac for a in Device.objects.filter(support_mode = '1')]
            event = DataQuery.DQGetEventFromRedis(seconds, p_list)
            for evd in event['data']:
                evd['probe_mac'] = evd['probe_mac'].lower()
                evd['state'] = state_check(evd['probe_mac'].lower())

        if event_type == "admin":
            if probe_event.count() != 0:
                for i in probe_event:
                    state = state_check(i.ap_mac)
                    event['data'].append({'event_time':timezone.localtime(i.event_time).strftime("%Y-%m-%d %H:%M:%S"),'msg':i.msg,'probe_mac':i.ap_mac,'admin_username':i.admin_username,'event':i.event,'state':state})
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        p_all = Device.objects.filter(Q(support_mode = '1')&Q(account_group_name = request.user.groupname))
        u_all = [a.username for a in NewUser.objects.filter(groupname = request.user.groupname)]
        p_list = [a.mac for a in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1'))]
        ##########xuyaoxiugai##################
        probe_event = AP_event.objects.filter(event_time__gte = t).order_by('-event_time')

        if event_type == "dev":
            # event = {'data':[{'event_time':'2017/05/04 12:05:35','msg':'Device[A4E6B1300005] was onlined.','action':_(u'存档'),'probe_mac':'A4E6B1300005','event':u'','state':state_check('A4E6B1300005')},{'event_time':'2017/05/04 12:04:44','msg':'Device[8c8401163ae0] was offlined.','action':_(u"存档"),'probe_mac':'8c8401163ae0','event':u'PROBE_WAS_OFFLINED','state':state_check('8c8401163ae0')}]}

            event = DataQuery.DQGetEventFromRedis(seconds, p_list)
            for evd in event['data']:
                evd['probe_mac'] = evd['probe_mac'].lower()
                evd['state'] = state_check(evd['probe_mac'].lower())

        if event_type == "admin":
            if probe_event.count() != 0:
                for i in probe_event:
                    if i.admin_username in u_all:
                        state = state_check(i.ap_mac)
                        event['data'].append({'event_time':timezone.localtime(i.event_time).strftime("%Y-%m-%d %H:%M:%S"),'msg':i.msg,'probe_mac':i.ap_mac,'admin_username':i.admin_username,'event':i.event,'state':state})

    return JsonResponse(event,safe = False)

def state_check(mac):
    state = ''
    try:
        al = Device.objects.get(mac = mac)

        # a = int((timezone.now()-al.last_heart_time).total_seconds())
        # # print type(a)
        # if a >= 86400:
        #     al.state = _(u"退服")
        # elif a >= 3600:
        #     al.state = _(u"离线")
        # elif a >= 900 :
        #     al.state = _(u"超时")
        # elif a < 900:
        #     al.state = _(u"在线")

        a = int((timezone.now()-al.last_heart_time).total_seconds())
        al.last_heart_time = timezone.localtime(al.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")
        rss = al.reboot_sign.split(",")
        if al.last_heart_time == "1970-1-1 08:00:00":
            al.state = _(u"离线")
        elif a >= 86400:
            al.state = _(u"退服")
        elif a >= 3600:
            al.state = _(u"离线")
        elif a >= 900 :
            al.state = _(u"超时")
        elif a < 900:
            if rss[0] != '0' and rss[0] != '':
                if rss[0] == '1':
                    al.state = _(u"重启")
                elif rss[0] == '2':
                    al.state = _(u"升级")
            else:
                al.state = _(u"在线")
        state = al.state
    except Exception as e:
        print e
        state = _(u'已移除')

    return state

def create_wireless_json(i):
    if i.model == "KL0001" or i.model == "TPY3101-HR":
        pass
    else:
        m = change_mac(i.mac)
        data_str = set_wlan_str(m)
        DataQuery.DQSetUpdateConfig('wireless',m,data_str)

def set_wlan_str(mac):
    dev = Device.objects.get(mac = mac)
    try:
        radios_type = Public_function.device_type_info_dict[dev.own_model]['radio_num']
    except Exception as e:
        print e
        radios_type = '0'

    try:
        if hasattr(dev,'device_ap'):
            if dev.device_ap.radios_2_channel == 'auto':
                radios_2_channel = 0
            else:
                radios_2_channel = int(dev.device_ap.radios_2_channel)

            if dev.device_ap.radios_2_ht == 'auto':
                radios_2_ht = 'auto'
            else:
                radios_2_ht = 'HT' + dev.device_ap.radios_2_ht

            if dev.device_ap.radios_2_power == 'auto':
                radios_2_power = 0
            elif dev.device_ap.radios_2_power == 'high':
                radios_2_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_2_txp'])
            elif dev.device_ap.radios_2_power == 'medium':
                radios_2_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_2_txp']) - 5
            elif dev.device_ap.radios_2_power == 'low':
                radios_2_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_2_txp']) - 10
            else:
                radios_2_power = int(dev.device_ap.radios_2_power)

            if dev.device_ap.radios_2_com == 'auto':
                radios_2_com = '11ng'
            else:
                radios_2_com = dev.device_ap.radios_2_com



            if dev.device_ap.radios_5_channel == 'auto':
                radios_5_channel = 0
            else:
                radios_5_channel = int(dev.device_ap.radios_5_channel)

            if dev.device_ap.radios_5_ht == 'auto':
                radios_5_ht = 'auto'
            else:
                radios_5_ht = 'HT' + dev.device_ap.radios_5_ht

            if dev.device_ap.radios_5_power == 'auto':
                radios_5_power = 0
            elif dev.device_ap.radios_5_power == 'high':
                radios_5_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_5_txp'])
            elif dev.device_ap.radios_5_power == 'medium':
                radios_5_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_5_txp']) - 5
            elif dev.device_ap.radios_5_power == 'low':
                radios_5_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_5_txp']) - 10
            else:
                radios_5_power = int(dev.device_ap.radios_5_power)

            if dev.device_ap.radios_5_com == 'auto':
                radios_5_com = '11ac'
            else:
                radios_5_com = dev.device_ap.radios_5_com
        else:
            radios_2_channel = 0
            radios_2_ht = 'auto'
            radios_2_power = 0
            radios_2_com = '11ng'
            radios_5_channel = 0
            radios_5_ht = 'auto'
            radios_5_power = 0
            radios_5_com = '11ac'
    except Exception as e:
        print e,'xwwwwww'
        radios_2_channel = 0
        radios_2_ht = 'auto'
        radios_2_power = 0
        radios_2_com = '11ng'
        radios_5_channel = 0
        radios_5_ht = 'auto'
        radios_5_power = 0
        radios_5_com = '11ac'

    django_settings = Setting.objects.get(SIGN = 1)
    countryCode = django_settings.countryCode or 'CN'
    radio2_str = {
        "channel": radios_2_channel,
        "country": countryCode,
        "disabled": 0,
        "htmode": radios_2_ht,
        "hwmode": radios_2_com,
        "mode": "2.4G",
        "txpower": radios_2_power,
    }
    radio5_str =  {
        "channel": radios_5_channel,
        "country": countryCode,
        "disabled": 0,
        "htmode": radios_5_ht,
        "hwmode": radios_5_com,
        "mode": "5G",
        "txpower": radios_5_power
    }

    print 'radio2_str:{},radio5_str:{}'.format(radio2_str,radio5_str)

    apwlan2 = []
    apwlan5 = []
    if Group_wlan.objects.filter(group_id_id = dev.group_id).count() == 0:
        apwlan2 = []
        apwlan5 = []
    else:
        gp_wlan = Group_wlan.objects.filter(group_id_id = dev.group_id)
        for i in gp_wlan:
            if Device_wlan.objects.filter(device_id = dev.pk,wlan_id = i.wlan_id).exists():
                dev_wlan = Device_wlan.objects.get(device_id = dev.pk,wlan_id = i.wlan_id)
            else:
                dev_wlan = False

            if dev_wlan != False:
                i.wlan_service = dev_wlan.wlan_service if dev_wlan.wlan_service != "" else i.wlan_service
                i.upload_speed = dev_wlan.upload_speed if dev_wlan.upload_speed != "" else i.upload_speed
                i.download_speed = dev_wlan.download_speed if dev_wlan.download_speed != "" else i.download_speed
                i.passphrase = dev_wlan.passphrase if dev_wlan.passphrase != "" else i.passphrase
                i.wlan_ssid = dev_wlan.wlan_ssid if dev_wlan.wlan_ssid != "" else i.wlan_ssid
                i.radios_enable = dev_wlan.radios_enable if dev_wlan.radios_enable != "" else i.radios_enable
                i.vlan_enabled = dev_wlan.vlan_enabled if dev_wlan.vlan_enabled != "" else i.vlan_enabled
                i.vlan = dev_wlan.vlan if dev_wlan.vlan != "" else i.vlan

            wlan_service = 0 if i.wlan_service == 'on' else 1
            if i.upload_speed == 'Unlimited' and i.download_speed == 'Unlimited':
                peruserrate = 0
                upload_speed = 0
                download_speed = 0
            else:
                peruserrate = 1
                try:
                    if i.upload_speed == 'Unlimited':
                        upload_speed = 0
                    elif 'K' in i.upload_speed :
                        upload_speed = int(i.upload_speed.replace('K',''))
                    elif 'M' in i.upload_speed :
                        upload_speed = int(i.upload_speed.replace('M','')) * 1024
                    else:
                        upload_speed = int(i.upload_speed)

                    if i.download_speed == 'Unlimited':
                        download_speed = 0
                    elif 'K' in i.download_speed :
                        download_speed = int(i.download_speed.replace('K',''))
                    elif 'M' in i.download_speed :
                        download_speed = int(i.download_speed.replace('M','')) * 1024
                    else:
                        download_speed = int(i.download_speed)

                except Exception as e:
                    print e
                    peruserrate = 0
                    upload_speed = 0
                    download_speed = 0

            if i.sec_type == 'psk2' and (i.encry_type == 'tkip' or i.encry_type == 'ccmp'):
                encryption = i.sec_type + '+' + i.encry_type
            elif i.sec_type == 'eap':
                encryption = 'wpa2' + '+' + i.encry_type
            else:
                encryption = 'none'
            hidden_ssid = 1 if i.hidden_ssid == 'on' else 0
            if i.guest_enabled == 'on':
                portal = Public_function.protal_type_num_dict[i.auth_type]
            else:
                portal = 0
            if i.vlan_enabled == "on":
                vlan = i.vlan
            else:
                vlan = 0
            wl = {
                "apidx": int(i.wlan_id),
                "disabled": wlan_service,
                "downperuserrate": download_speed,
                "encryption": encryption,
                "hidden": hidden_ssid,
                "key": i.passphrase,
                "peruserrate": peruserrate,
                "portal": portal,
                "ssid": i.wlan_ssid,
                "upperuserrate": upload_speed,
                "vlan":int(vlan),
                "auth_port":i.eapPort,
                "auth_server":i.eapIP,
            }
            if i.radios_enable == "both":
                apwlan2.append(wl)
                apwlan5.append(wl)
            elif i.radios_enable == "2g":
                apwlan2.append(wl)
            elif i.radios_enable == "5g":
                apwlan5.append(wl)

    if radios_type == '0':
        data = "{'wifi':[{'radio':%s,'ap':%s},{'radio':%s,'ap':%s}]}"%(json.dumps(radio2_str ,sort_keys=True , ensure_ascii=False),json.dumps(apwlan2,sort_keys=True , ensure_ascii=False), json.dumps(radio5_str ,sort_keys=True , ensure_ascii=False), json.dumps(apwlan5 ,sort_keys=True , ensure_ascii=False))
    elif radios_type == '1':
        data = "{'wifi':[{'radio':%s,'ap':%s}]}"%(json.dumps(radio2_str ,sort_keys=True , ensure_ascii=False),json.dumps(apwlan2,sort_keys=True , ensure_ascii=False))
    print 'peizhi',data.replace(' ','').replace('\'','\"')
    return data.replace(' ','').replace('\'','\"')

@login_required
def ap_group(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)
    group = ''
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        group = Probe_group.objects.filter(group_type = '1')
    elif request.user.administrator_permission < 4 and request.user.group_status == 1:
        group = Probe_group.objects.filter(account_group_name = request.user.groupname,group_type = '1')
    for i in group:
        if i.group_name == "DefaultGroup" and i.account_group_name == "admin" and i.group_type == '1':
            i.device_count = Device.objects.filter(support_mode = '1').filter(Q(account_group_name = i.account_group_name,group_id = i.pk) |Q(account_group_name = '',group_id = 0)).count()
        else:
            i.device_count = Device.objects.filter(support_mode = '1').filter(group_id = i.id).filter(account_group_name = i.account_group_name).count()
        i.wlan_count = str(i.group_wlan_set.filter(wlan_service = 'on').count()) + '/' +str(i.group_wlan_set.count())

    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'ap-group.html',{'admin_error_json':admin_error_json,'admin_error':admin_error,'group':group,'error':error,"error_json":error_json,'plimit':plimit,'oemlimit':oemlimit})

@login_required
def ap_add_ajax(request):
    device = {'data':[]}
    if request.user.administrator_permission == 6 or request.user.administrator_permission == 5 or request.user.administrator_permission == 0:
        a = Probe_group.objects.get(group_name = 'DefaultGroup',account_group_name = 'admin',group_type = '1')
        device_quaryset = Device.objects.filter(support_mode = '1').filter(Q(account_group_name = '') | Q(account_group_name = 'admin')).filter(Q(group_id = 0) | Q(group_id = a.pk))
    if request.user.administrator_permission == 3 or request.user.administrator_permission == 2 :
        a = Probe_group.objects.get(group_name = request.user.username,account_group_name = request.user.username,group_type = '1')
        device_quaryset = Device.objects.filter(support_mode = '1').filter(account_group_name = request.user.username).filter(Q(group_id = 0) | Q(group_id = a.pk))
    if device_quaryset.count() != 0:
        for i in device_quaryset:
            if i.name != '':
                name = i.name
            else:
                name = i.model + '_' + i.mac[-6:]
            mac = i.mac.upper()[:2] + '-' + i.mac.upper()[2:4] + '-' + i.mac.upper()[4:6] + '-' + i.mac.upper()[6:8] + '-' + i.mac.upper()[8:10] + '-' + i.mac.upper()[10:12]
            device['data'].append({'name':name,'mac':mac,'action':''})
    return JsonResponse(device,safe = False)

@login_required
def add_ap_group_ajax(request):
    error = {"error_type":'',"error_msg":""}
    if request.method == "POST":
        add_group_device_list = json.loads(request.POST.get('add_group_device_list'))
        groupname = add_group_device_list['groupname']
        mac = add_group_device_list['mac']
        if request.user.groupname != '' :
            user_area = request.user.groupname
        else:
            user_area = 'admin'
        if Probe_group.objects.filter(account_group_name = user_area).filter(group_name = groupname,group_type = '1').exists():
            error = {"error_type":'failed',"error_msg":_(u"该组已存在")}
        else:
            try:
                pg = Probe_group()
                pg.group_name = groupname
                pg.account_group_name = user_area
                pg.group_type = '1'
                pg.save()
                pgid = Probe_group.objects.get(account_group_name = user_area,group_name = groupname,group_type = '1').pk
                if user_area == 'admin':
                    old_pgid = Probe_group.objects.get(account_group_name = user_area,group_name = 'DefaultGroup',group_type = '1').pk
                else:
                    old_pgid = Probe_group.objects.get(account_group_name = user_area,group_name = user_area,group_type = '1').pk
                if compare_group_wlan_settions(pgid,old_pgid):
                    for i in mac:
                        m = change_mac(i)
                        Device.objects.filter(support_mode = '1').filter(mac = m).update(group_id = pgid,account_group_name = user_area)
                        Device.objects.get(mac = m).device_wlan_set.all().delete()
                else:
                    for i in mac:
                        m = change_mac(i)
                        Device.objects.filter(support_mode = '1').filter(mac = m).update(group_id = pgid,account_group_name = user_area)
                        Device.objects.get(mac = m).device_wlan_set.all().delete()
                        create_wireless_json(Device.objects.get(mac = m))
                        #audit config str
                        # data_str = set_wlan_str(m)
                        # DataQuery.DQSetUpdateConfig('wireless',m,data_str)
                gp_str = portal_str(request.user.groupname)
                task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac,'mac_list')
                policy_str = user_policy_str(request.user.groupname)
                task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,mac,'mac_list')
                error = {"error_type":'success',"error_msg":_(u"添加成功")}
            except Exception as e:
                print e
                error = {"error_type":'failed',"error_msg":_(u"添加失败")}
        if error['error_type'] == "success":
            request.session['error_dict'] = error
        return JsonResponse(error,safe = False)

def change_mac(mac):
    return ''.join(mac.lower().split('-'))

@login_required
def ap_group_del(request):
    group_id = request.GET.get('group_id')
    pg = Probe_group.objects.get(pk = group_id)
    try:
        if pg.account_group_name == "admin":
            default_pg = Probe_group.objects.get(account_group_name = 'admin',group_name = "DefaultGroup",group_type = '1')
        else:
            default_pg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '1')

        a = Device.objects.filter(support_mode = '1').filter(group_id = pg.pk)
        mac_list = []
        for i in a :
            mac_list.append(i.mac)
            i.device_wlan_set.all().delete()
        Device.objects.filter(support_mode = '1').filter(group_id = pg.pk).update(group_id = default_pg.pk)
        if compare_group_wlan_settions(pg.pk,default_pg.pk):
            pass
        else:
            for i in a:
                create_wireless_json(i)
                #audit config str
                # data_str = set_wlan_str(i.mac)
                # DataQuery.DQSetUpdateConfig('wireless',i.mac,data_str)
        gp_str = portal_str(request.user.groupname)
        task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')
        policy_str = user_policy_str(request.user.groupname)
        task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,mac_list,'mac_list')
        # else:
        #     default_pg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '1')
        #     a = Device.objects.filter(support_mode = '1').filter(group_id = pg.pk)
        #     for i in a :
        #         i.device_wlan_set.all().delete()
        #     if compare_group_wlan_settions(pg.pk,default_pg.pk):
        #         pass
        #     else:
        #         for i in a:
        #             #audit config str
        #             pass
        #     Device.objects.filter(support_mode = '1').filter(group_id = pg.pk).update(group_id = default_pg.pk)

        pg.delete()
        error = {"error_type":'success',"error_msg":_(u"删除成功")}
    except Exception as e:
        print e
        error = {"error_type":'failed',"error_msg":_(u"删除失败")}

    request.session['error_dict'] = error
    return redirect('/ap/ap_group/')


@login_required
def remove_ap_ajax(request):
    group_id = request.GET.get("group_id")
    pg = Probe_group.objects.get(pk = group_id)
    device = {'data':[]}
    if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
        device_quaryset = Device.objects.filter(support_mode = '1').filter(Q(account_group_name = pg.account_group_name,group_id = pg.pk) |Q(account_group_name = '',group_id = 0))
    else:
        device_quaryset = Device.objects.filter(support_mode = '1').filter(account_group_name = pg.account_group_name,group_id = pg.pk)
    if device_quaryset.count() != 0:
        for i in device_quaryset:
            if i.name != '':
                name = i.name
            else:
                name = i.model + '_' + i.mac[-6:]
            mac = i.mac.upper()[:2] + '-' + i.mac.upper()[2:4] + '-' + i.mac.upper()[4:6] + '-' + i.mac.upper()[6:8] + '-' + i.mac.upper()[8:10] + '-' + i.mac.upper()[10:12]

            if request.user.administrator_permission == 3 or request.user.administrator_permission == 2 or request.user.administrator_permission == 1:
                if pg.group_name == pg.account_group_name :
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'off'})
                else:
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'on'})
            else:
                if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'off'})
                else:
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'on'})
    return JsonResponse(device,safe = False)


@login_required
def add_ap_ajax(request):
    group_id = request.GET.get("group_id")
    pg = Probe_group.objects.get(pk = group_id)
    device = {'data':[]}
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        dpg = Probe_group.objects.get(group_name = "DefaultGroup" ,account_group_name = "admin",group_type = '1')
        if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
            device_quaryset = ""
        elif pg.group_name != "DefaultGroup" and pg.account_group_name == "admin":
            device_quaryset = Device.objects.filter(support_mode = '1').filter(Q(account_group_name = '')|Q(account_group_name = 'admin') , Q(group_id = dpg.pk)|Q(group_id = 0))
        elif pg.account_group_name != "admin" and pg.group_name == pg.account_group_name:
            ptdpg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '1')
            device_quaryset = Device.objects.filter(support_mode = '1').filter(Q(account_group_name = '')|Q(account_group_name = 'admin') , Q(group_id = dpg.pk)|Q(group_id = 0))
        elif pg.account_group_name != "admin" and pg.group_name != pg.account_group_name:
            ptdpg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '1')
            device_quaryset = Device.objects.filter(support_mode = '1').filter(account_group_name = ptdpg.account_group_name,group_id = ptdpg.pk)
    elif request.user.administrator_permission != 0 and request.user.administrator_permission < 4:
        ptdpg = Probe_group.objects.get(Q(account_group_name = pg.account_group_name) & Q(group_name = pg.account_group_name) & Q(group_type = '1'))
        if pg.pk == ptdpg.pk:
            device_quaryset = ""
        elif pg.pk != ptdpg.pk:
            device_quaryset = Device.objects.filter(support_mode = '1').filter(account_group_name = ptdpg.account_group_name,group_id = ptdpg.pk)
    if device_quaryset != "":
        if device_quaryset.count() != 0:
            for i in device_quaryset:
                if i.name != '':
                    name = i.name
                else:
                    name = i.model + '_' + i.mac[-6:]
                mac = i.mac.upper()[:2] + '-' + i.mac.upper()[2:4] + '-' + i.mac.upper()[4:6] + '-' + i.mac.upper()[6:8] + '-' + i.mac.upper()[8:10] + '-' + i.mac.upper()[10:12]
                if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'off'})
                else:
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'on'})
    return JsonResponse(device,safe = False)

@login_required
def add_wlan_in_group_ajax(request):
    group_id = request.GET.get("group_id")
    pg = Probe_group.objects.get(pk = group_id)
    device = {'data':[]}
    pgwl = pg.group_wlan_set.all()
    for i in pgwl:
        device['data'].append({'name':i.wlan_ssid,'sec_type':i.sec_type,'guest_enabled':i.guest_enabled,'wlan_service':i.wlan_service,'pk':i.pk})

    compare_group_wlan_settions(group_id,group_id)
    return JsonResponse(device,safe = False)

@login_required
def ap_setting_ajax(request):
    group_id = request.GET.get('group_id')
    pg = Probe_group.objects.get(pk = group_id)
    return JsonResponse(model_to_dict(pg),safe = False)

@login_required
def modify_ap_group_ajax(request):
    sign = 0
    error = {"error_type":'',"error_msg":""}
    if request.method == "POST":
        modify_device = json.loads(request.POST.get('modify_device'))
        group_id = int(modify_device['id'])
        groupname = modify_device['groupname']
        area_name = modify_device['area_name']
        remove_device_mac = modify_device['remove_device_mac']
        add_device_mac = modify_device['add_device_mac']
        add_wlan = modify_device['add_wlan']
        if Probe_group.objects.exclude(pk = group_id).filter(Q(account_group_name = area_name) & Q(group_name = groupname)&Q(group_type = '1')).exists():
            error = {"error_type":'failed',"error_msg":_(u"组名已存在")}
        else:
            try:
                pg = Probe_group.objects.get(pk = group_id , account_group_name = area_name,group_type = '1')
                pg.group_name = groupname
                pg.save()
                if add_wlan == {}:
                    sign = 0
                else:
                    for k,v in add_wlan.iteritems():
                        # try:
                        i = pg.group_wlan_set.get(pk = int(k))
                        i.wlan_service = v
                        i.save()
                        # except Exception as e:
                        #     print e
                    sign = 1
                if remove_device_mac == {}:
                    pass
                else:
                    if request.user.administrator_permission == 1 or request.user.administrator_permission == 2 or request.user.administrator_permission == 3:
                        pgid = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '1').pk
                        if compare_group_wlan_settions(pg.pk,pgid):
                            for i in remove_device_mac:
                                m = change_mac(i)
                                Device.objects.get(mac = m).device_wlan_set.all().delete()
                                Device.objects.filter(support_mode = '1').filter(mac = m).update(group_id = pgid)
                        else:
                            for i in remove_device_mac:
                                m = change_mac(i)
                                Device.objects.get(mac = m).device_wlan_set.all().delete()
                                Device.objects.filter(support_mode = '1').filter(mac = m).update(group_id = pgid)
                                create_wireless_json(Device.objects.get(mac = m))
                                # data_str = set_wlan_str(m)
                                # DataQuery.DQSetUpdateConfig('wireless',m,data_str)
                                #audit config str

                        gp_str = portal_str(pg.account_group_name)
                        task_for_global_config_ajax.delay(pg.account_group_name,gp_str,remove_device_mac,'mac_list')
                        policy_str = user_policy_str(pg.account_group_name)
                        task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,remove_device_mac,'mac_list')
                    else:
                        if pg.account_group_name == "admin" or pg.group_name == pg.account_group_name:
                            pgid = Probe_group.objects.get(account_group_name = "admin",group_name = "DefaultGroup",group_type = '1').pk
                            if compare_group_wlan_settions(pg.pk,pgid):
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.get(mac = m).device_wlan_set.all().delete()
                                    Device.objects.filter(support_mode = '1').filter(mac = m).update(group_id = pgid,account_group_name = "admin")
                                    update_device2authpuppy(m,'admin')
                            else:
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.get(mac = m).device_wlan_set.all().delete()
                                    Device.objects.filter(support_mode = '1').filter(mac = m).update(group_id = pgid,account_group_name = "admin")

                                    update_device2authpuppy(m,'admin')
                                    create_wireless_json(Device.objects.get(mac = m))
                                    #audit config str
                                    # data_str = set_wlan_str(m)
                                    # DataQuery.DQSetUpdateConfig('wireless',m,data_str)

                            gp_str = portal_str(pg.account_group_name)
                            task_for_global_config_ajax.delay(pg.account_group_name,gp_str,remove_device_mac,'mac_list')
                            policy_str = user_policy_str(pg.account_group_name)
                            task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,remove_device_mac,'mac_list')
                        else:
                            pgid = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '1').pk
                            if compare_group_wlan_settions(pg.pk,pgid):
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.get(mac = m).device_wlan_set.all().delete()
                                    Device.objects.filter(support_mode = '1').filter(mac = m).update(group_id = pgid)
                            else:
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.get(mac = m).device_wlan_set.all().delete()
                                    Device.objects.filter(support_mode = '1').filter(mac = m).update(group_id = pgid)
                                    create_wireless_json(Device.objects.get(mac = m))
                                    #audit config str
                                    # data_str = set_wlan_str(m)
                                    # DataQuery.DQSetUpdateConfig('wireless',m,data_str)

                            gp_str = portal_str(pg.account_group_name)
                            task_for_global_config_ajax.delay(pg.account_group_name,gp_str,remove_device_mac,'mac_list')
                            policy_str = user_policy_str(pg.account_group_name)
                            task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,remove_device_mac,'mac_list')
                if add_device_mac == {}:
                    pass
                else:
                    for i in add_device_mac:
                        m = change_mac(i)
                        p = Device.objects.get(mac = m).group_id
                        Device.objects.get(mac = m).device_wlan_set.all().delete()
                        Device.objects.filter(support_mode = '1').filter(mac = m).update(group_id = pg.pk,account_group_name = pg.account_group_name)
                        update_device2authpuppy(m,pg.account_group_name)
                        if compare_group_wlan_settions(pg.pk,p):
                            pass
                        else:
                            create_wireless_json(Device.objects.get(mac = m))
                            # audit config str
                            # data_str = set_wlan_str(m)
                            # DataQuery.DQSetUpdateConfig('wireless',m,data_str)

                    gp_str = portal_str(pg.account_group_name)
                    task_for_global_config_ajax.delay(pg.account_group_name,gp_str,add_device_mac,'mac_list')
                    policy_str = user_policy_str(pg.account_group_name)
                    task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,add_device_mac,'mac_list')

                if sign == 0:
                    pass
                else:
                    task_for_modify_ap_group_ajax.delay(pg.pk)
                    gp_str = portal_str(pg.account_group_name)
                    mac_list = []
                    for i in Device.objects.filter(group_id = pg.pk):
                        mac_list.append(i.mac)
                    task_for_global_config_ajax.delay(pg.account_group_name,gp_str,mac_list,'mac_list')
                    policy_str = user_policy_str(pg.account_group_name)
                    task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,mac_list,'mac_list')
                    # for i in Device.objects.filter(support_mode = '1').filter(group_id = pg.pk):
                    #     m = change_mac(i.mac)

                    #      #audit config str
                    #     data_str = set_wlan_str(m)
                    #     DataQuery.DQSetUpdateConfig('wireless',m,data_str)
                error = {"error_type":'success',"error_msg":_(u"编辑成功")}
                if pg.account_group_name == "admin" and pg.group_name == 'DefaultGroup':
                    default_config_issued.delay()
            except Exception as e:
                print e
                error = {"error_type":'failed',"error_msg":_(u"编辑失败")}
        if error['error_type'] == "success":
            request.session['error_dict'] = error
        return JsonResponse(error,safe = False)

def compare_group_audit_setting(gp1,gp2):
    try:
        g1 = model_to_dict(Probe_group.objects.get(pk = int(gp1)))
        g2 = model_to_dict(Probe_group.objects.get(pk = int(gp2)))
        g1.pop('id')
        g1.pop('group_name')
        g1.pop('account_group_name')
        g1.pop('device_count')
        g1.pop('group_type')
        g2.pop('id')
        g2.pop('group_name')
        g2.pop('account_group_name')
        g2.pop('device_count')
        g2.pop('group_type')

        return g1 == g2
    except Exception as e:
        print e
        return False

def compare_group_wlan_settions(gp1,gp2):
    try:
        g1 = {}
        g2 = {}
        for i in Group_wlan.objects.filter(group_id_id = int(gp1)).order_by('wlan_id'):
            g1[str(i.wlan_id)] = model_to_dict(i)
            g1[str(i.wlan_id)].pop('group_id')
            g1[str(i.wlan_id)].pop('id')
        for i in Group_wlan.objects.filter(group_id_id = int(gp2)).order_by('wlan_id'):
            g2[str(i.wlan_id)] = model_to_dict(i)
            g2[str(i.wlan_id)].pop('group_id')
            g2[str(i.wlan_id)].pop('id')
        print g1 == g2 ,'(test)'
        return g1 == g2
    except Exception as e:
        print e
        return False


@login_required
def ap_wlan(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)

    wlan = ''
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        wlan = Group_wlan.objects.filter(group_id__group_type = '1')
    elif request.user.administrator_permission < 4 and request.user.group_status == 1:
        wlan = Group_wlan.objects.filter(group_id__account_group_name = request.user.groupname,group_id__group_type = '1')
    for wl in wlan:
        if wl.guest_enabled == "off":
            wl.auth_type = 'off'


    plimit = get_plimit()
    return render(request,'ap-wlan.html',{'admin_error_json':admin_error_json,'admin_error':admin_error,'wlan':wlan,'error':error,"error_json":error_json,'plimit':plimit})

@login_required
def choose_ap_group_ajax(request):
    gp = []
    # {'group_id':'','group_account_group_name':'','group_name':'',}
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        group = Probe_group.objects.filter(group_type = "1")
    elif request.user.administrator_permission < 4 and request.user.group_status == 1:
        group = Probe_group.objects.filter(group_type = "1",account_group_name = request.user.groupname)
    for i in group:
        if i.group_wlan_set.all().count() >= 4:
            gp.append({'group_id':i.pk,'group_account_group_name':i.account_group_name,'group_name':i.group_name,'disabled':'disabled'})
        else:
            gp.append({'group_id':i.pk,'group_account_group_name':i.account_group_name,'group_name':i.group_name,'disabled':'enabled'})
    return JsonResponse(gp,safe=False)

@login_required
def add_wlan_ajax(request):
    errors = {'sign':'false','message':''}
    if request.method == 'POST':
        logger.info('start add wlan ->run add wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        group_id = int(request.POST.get('group_id'))
        a = Probe_group.objects.get(id = group_id)
        if not a.group_wlan_set.filter(wlan_id = '0').exists():
            wlan_id = '0'
        elif not a.group_wlan_set.filter(wlan_id = '1').exists():
            wlan_id = '1'
        elif not a.group_wlan_set.filter(wlan_id = '2').exists():
            wlan_id = '2'
        elif not a.group_wlan_set.filter(wlan_id = '3').exists():
            wlan_id = '3'
        wlan_ssid = request.POST.get('ssid')
        if request.POST.get('wlan_service') == None:
            wlan_service = "off"
        else:
            wlan_service = request.POST.get('wlan_service')
        if request.POST.get('safe') == 'open':
            sec_type = 'open'
            encry_type = ''
        elif request.POST.get('safe') == 'tkip':
            sec_type = 'psk2'
            encry_type = 'tkip'
        elif request.POST.get('safe') == 'ccmp':
            sec_type = 'psk2'
            encry_type = 'ccmp'
        elif request.POST.get('safe') == 'eap':
            sec_type = 'eap'
            encry_type = 'ccmp'
        passphrase = request.POST.get('passphrase')
        eapIP = request.POST.get('eapIP')
        eapPort = request.POST.get('eapPort')
        if request.POST.get('guest_enabled') == None:
            guest_enabled = "off"
        else:
            guest_enabled = request.POST.get('guest_enabled')
        if request.POST.get('vlan_enabled') == None:
            vlan_enabled = "off"
        else:
            vlan_enabled = request.POST.get('vlan_enabled')
        if request.POST.get('vlan') == '':
            vlan = '0'
        else:
            vlan = request.POST.get('vlan')
        if request.POST.get('hidden_ssid') == None:
            hidden_ssid = "off"
        else:
            hidden_ssid = request.POST.get('hidden_ssid')
        upload_speed = request.POST.get('upload_speed')
        download_speed = request.POST.get('download_speed')
        radios_enable = request.POST.get('radios_enable')
        auth_type = request.POST.get('auth_type')

        wechat_appid = request.POST.get('wechat_appid')
        wechat_appkey = request.POST.get('wechat_appkey')
        wechat_shopid = request.POST.get('wechat_shopid')
        wechat_secretkey = request.POST.get('wechat_secretkey')
        if request.POST.get('wechat_forcefollow') == None:
            wechat_forcefollow = "false"
        else:
            wechat_forcefollow = 'true'

        # auth_server_hostname = request.POST.get('group_id')
        auth_server_loginurl = request.POST.get('auth_server_loginurl')
        auth_server_portalurl = request.POST.get('auth_server_portalurl')
        a = auth_server_loginurl.split('://')
        if len(a) >= 2:
            auth_server_hostname = a[1].split('/')[0]
        else:
            auth_server_hostname = ''
        # print group_id,"###",wlan_ssid,"###",wlan_service,"###",safe,"###",passphrase,"###",guest_enabled,"###",vlan_enabled,"###",vlan,"###",hidden_ssid,"###",upload_speed,"###",download_speed,"###",radios_enable,"###",auth_type

        try:
            gw = Group_wlan()
            gw.group_id_id = group_id
            gw.wlan_id = wlan_id
            gw.wlan_ssid = wlan_ssid
            gw.wlan_service = wlan_service
            gw.sec_type = sec_type
            gw.encry_type = encry_type
            gw.passphrase = passphrase

            gw.eapIP = eapIP
            gw.eapPort = eapPort

            gw.guest_enabled = guest_enabled
            gw.vlan_enabled = vlan_enabled
            gw.vlan = vlan
            gw.hidden_ssid = hidden_ssid
            gw.upload_speed = upload_speed
            gw.download_speed = download_speed
            gw.radios_enable = radios_enable
            gw.auth_type = auth_type

            gw.wechat_appid = wechat_appid
            gw.wechat_appkey = wechat_appkey
            gw.wechat_shopid = wechat_shopid
            gw.wechat_secretkey = wechat_secretkey
            gw.wechat_forcefollow = wechat_forcefollow
            gw.auth_server_hostname = auth_server_hostname
            gw.auth_server_loginurl = auth_server_loginurl
            gw.auth_server_portalurl = auth_server_portalurl

            gw.save()
            errors['message'] = _(u'创建成功！')
            errors['sign'] = "true"
            logger.info('start add wlan ->end add wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

            task_for_add_wlan_ajax.delay(group_id)
            gp_str = portal_str(request.user.groupname)
            mac_list = []
            for i in Device.objects.filter(group_id = group_id):
                mac_list.append(i.mac)
            task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')

            pg = Probe_group.objects.get(id = group_id)
            if pg.account_group_name == "admin" and pg.group_name == 'DefaultGroup':
                default_config_issued.delay()
            # for i in Device.objects.filter(group_id = group_id):
            #     m = change_mac(i.mac)
            #     data_str = set_wlan_str(m)
            #     DataQuery.DQSetUpdateConfig('wireless',m,data_str)
            logger.info('start add wlan ->end DQ add wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        except Exception as e:
            print e
            errors['message'] = _(u'创建失败！')
            errors['sign'] = "false"


    return JsonResponse(errors,safe=False)


@login_required
def ap_wlan_del(request):
    wlan_id = request.GET.get('wlan_id')
    try:
        logger.info('start del wlan ->run get wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        wl = Group_wlan.objects.get(pk = wlan_id)
        ###########################################
        ###删除组中的wlan
        #########################################
        logger.info('start del wlan ->end get wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        dev = Device.objects.filter(group_id = wl.group_id_id)
        logger.info('start del wlan ->end get dev --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        for d in dev:
            d.device_wlan_set.filter(wlan_id = wl.wlan_id).delete()
        logger.info('start del wlan ->end del device_wlan --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        logger.info('start del wlan ->print sql --- sql:{}'.format(connection.queries))
        wl.delete()
        logger.info('start del wlan ->end del wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        task_for_ap_del_wlan.delay(wl.group_id_id)
        gp_str = portal_str(request.user.groupname)
        mac_list = []
        for i in Device.objects.filter(group_id = wl.group_id_id):
            mac_list.append(i.mac)
        task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')

        pg = Probe_group.objects.get(id = wl.group_id_id)
        if pg.account_group_name == "admin" and pg.group_name == 'DefaultGroup':
            default_config_issued.delay()
        # for i in dev:
        #     m = change_mac(i.mac)
        #     data_str = set_wlan_str(m)
        #     DataQuery.DQSetUpdateConfig('wireless',m,data_str)
        logger.info('start del wlan ->end DQconfig --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        error = {"error_type":'success',"error_msg":_(u"删除成功")}
    except Exception as e:
        print e
        error = {"error_type":'failed',"error_msg":_(u"删除失败")}

    request.session['error_dict'] = error
    return redirect('/ap/ap_wlan/')

@login_required
def ap_wlan_info_ajax(request):
    wl_id = int(request.GET.get('id'))
    wlan = Group_wlan.objects.get(pk = wl_id)
    gp = {'gp':'','wl':''}
    gp['gp'] = {'group_id':wlan.group_id_id,'group_account_group_name':wlan.group_id.account_group_name,'group_name':wlan.group_id.group_name}
    gp['wl'] = model_to_dict(wlan)
    return JsonResponse(gp,safe=False)

@login_required
def modify_wlan_ajax(request):
    errors = {'sign':'false','message':''}
    if request.method == 'POST':
        wlan_ssid = request.POST.get('ssid')
        if request.POST.get('wlan_service') == None:
            wlan_service = "off"
        else:
            wlan_service = request.POST.get('wlan_service')
        print request.POST.get('safe'),123321
        if request.POST.get('safe') == 'open':
            sec_type = 'open'
            encry_type = ''
        elif request.POST.get('safe') == 'tkip':
            sec_type = 'psk2'
            encry_type = 'tkip'
        elif request.POST.get('safe') == 'ccmp':
            sec_type = 'psk2'
            encry_type = 'ccmp'
        elif request.POST.get('safe') == 'eap':
            sec_type = 'eap'
            encry_type = 'ccmp'
        passphrase = request.POST.get('passphrase')
        eapIP = request.POST.get('eapIP')
        eapPort = request.POST.get('eapPort')
        if request.POST.get('guest_enabled') == None:
            guest_enabled = "off"
        else:
            guest_enabled = request.POST.get('guest_enabled')
        if request.POST.get('vlan_enabled') == None:
            vlan_enabled = "off"
        else:
            vlan_enabled = request.POST.get('vlan_enabled')
        if request.POST.get('vlan') == '':
            vlan = '0'
        else:
            vlan = request.POST.get('vlan')
        if request.POST.get('hidden_ssid') == None:
            hidden_ssid = "off"
        else:
            hidden_ssid = request.POST.get('hidden_ssid')
        upload_speed = request.POST.get('upload_speed')
        download_speed = request.POST.get('download_speed')
        radios_enable = request.POST.get('radios_enable')
        auth_type = request.POST.get('auth_type')

        wechat_appid = request.POST.get('wechat_appid')
        wechat_appkey = request.POST.get('wechat_appkey')
        wechat_shopid = request.POST.get('wechat_shopid')
        wechat_secretkey = request.POST.get('wechat_secretkey')
        if request.POST.get('wechat_forcefollow') == None:
            wechat_forcefollow = "false"
        else:
            wechat_forcefollow = "true"

        # auth_server_hostname = request.POST.get('group_id')
        auth_server_loginurl = request.POST.get('auth_server_loginurl')
        auth_server_portalurl = request.POST.get('auth_server_portalurl')
        a = auth_server_loginurl.split('://')
        if len(a) >= 2:
            auth_server_hostname = a[1].split('/')[0]
        else:
            auth_server_hostname = ''

        pk = int(request.GET.get('pk'))
        try:
            gw = Group_wlan.objects.get(pk = pk)
            gw.wlan_ssid = wlan_ssid
            gw.wlan_service = wlan_service
            gw.sec_type = sec_type
            gw.encry_type = encry_type
            gw.passphrase = passphrase

            gw.eapIP = eapIP
            gw.eapPort = eapPort

            gw.guest_enabled = guest_enabled
            gw.vlan_enabled = vlan_enabled
            gw.vlan = vlan
            gw.hidden_ssid = hidden_ssid
            gw.upload_speed = upload_speed
            gw.download_speed = download_speed
            gw.radios_enable = radios_enable
            gw.auth_type = auth_type

            gw.wechat_appid = wechat_appid
            gw.wechat_appkey = wechat_appkey
            gw.wechat_shopid = wechat_shopid
            gw.wechat_secretkey = wechat_secretkey
            gw.wechat_forcefollow = wechat_forcefollow
            gw.auth_server_hostname = auth_server_hostname
            gw.auth_server_loginurl = auth_server_loginurl
            gw.auth_server_portalurl = auth_server_portalurl

            gw.save()
            errors['message'] = _(u'修改成功！')
            errors['sign'] = 'true'

            task_for_modify_wlan_ajax.delay(gw.group_id_id)
            gp_str = portal_str(request.user.groupname)
            mac_list = []
            for i in Device.objects.filter(group_id = gw.group_id_id):
                mac_list.append(i.mac)
            task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')

            pg = Probe_group.objects.get(id = gw.group_id_id)
            if pg.account_group_name == "admin" and pg.group_name == 'DefaultGroup':
                default_config_issued.delay()
            # for i in Device.objects.filter(group_id = gw.group_id_id):
            #     m = change_mac(i.mac)
            #     data_str = set_wlan_str(m)
            #     DataQuery.DQSetUpdateConfig('wireless',m,data_str)
        except Exception as e:
            print e
            errors['message'] = _(u'修改失败！')
            errors['sign'] = 'false'

    return JsonResponse(errors,safe=False)


@login_required
def ap_wlan_list(request):
    ret = {'data':[]}
    ap_id = int(request.GET.get('id'))
    try:
        wlan = Probe_group.objects.get(pk = Device.objects.get(id = ap_id).group_id).group_wlan_set.all()
        for i in wlan:
            a = model_to_dict(i)
            x = ""
            try:
                apwl = Device_wlan.objects.get(device_id = ap_id,wlan_id = i.wlan_id)
                if apwl.wlan_ssid != i.wlan_ssid and apwl.wlan_ssid != "":
                    x = "SSID:" + apwl.wlan_ssid + ","
                    a['wlan_ssid'] = apwl.wlan_ssid
                if apwl.wlan_service != i.wlan_service and apwl.wlan_service != "":
                    if apwl.wlan_service == "on":
                        x = x + "Enable" + ","
                    if apwl.wlan_service == "off":
                        x = x + "Disable" + ","
                if apwl.passphrase != i.passphrase and apwl.passphrase != "":
                    x = x + "PSK2" + ","
                if apwl.vlan_enabled != i.vlan_enabled and apwl.vlan_enabled != "":
                    if apwl.vlan_enabled == "on":
                        x = x + "VLAN:Enable" + ","
                    if apwl.vlan_enabled == "off":
                        x = x + "VLAN:Disable" + ","
                if apwl.vlan != i.vlan and apwl.vlan != "":
                    x = x + "VLANID:" + apwl.vlan +","
                if apwl.upload_speed != i.upload_speed and apwl.upload_speed != "":
                    x = x + "UPLOAD SPEED:" + apwl.upload_speed + ","
                if apwl.download_speed != i.download_speed and apwl.download_speed != "":
                    x = x + "DOWNLOAD SPEED:" + apwl.download_speed + ","
                if apwl.radios_enable != i.radios_enable and apwl.radios_enable != "":
                    if apwl.radios_enable == "both":
                        x = x + "APPLY:BOTH" + ","
                    if apwl.radios_enable == "5g":
                        x = x + "APPLY:5G" + ","
                    if apwl.radios_enable == "2g":
                        x = x + "APPLY:2G" + ","
                a["change"] = x
            except Exception as e:
                print e
                a["change"] = ""
            a['ap_id'] = ap_id
            ret['data'].append(a)
    except Exception as e:
        print e

    return JsonResponse(ret,safe=False)

@login_required
def wlan_show_detail(request):
    wl_id = int(request.GET.get('id'))
    ap_id = int(request.GET.get('ap_id'))
    wlan = Group_wlan.objects.get(pk = wl_id)
    ret = model_to_dict(wlan)
    if Device.objects.get(pk = ap_id).device_wlan_set.filter(wlan_id = wlan.wlan_id).exists():
        dw = Device.objects.get(pk = ap_id).device_wlan_set.get(wlan_id = wlan.wlan_id)
        if dw.wlan_ssid != wlan.wlan_ssid and dw.wlan_ssid != "":
            ret['wlan_ssid'] = dw.wlan_ssid
        if dw.wlan_service != wlan.wlan_service and dw.wlan_service != "":
            ret['wlan_service'] = dw.wlan_service
        if dw.passphrase != wlan.passphrase and dw.passphrase != "":
            ret['passphrase'] = dw.passphrase
        if dw.vlan_enabled != wlan.vlan_enabled and dw.vlan_enabled != "":
            ret['vlan_enabled'] = dw.vlan_enabled
        if dw.vlan != wlan.vlan and dw.vlan != "":
            ret['vlan'] = dw.vlan
        if dw.upload_speed != wlan.upload_speed and dw.upload_speed != "":
            ret['upload_speed'] = dw.upload_speed
        if dw.download_speed != wlan.download_speed and dw.download_speed != "":
            ret['download_speed'] = dw.download_speed
        if dw.radios_enable != wlan.radios_enable and dw.radios_enable != "":
            ret['radios_enable'] = dw.radios_enable

    if wlan.encry_type == "":
        ret['pass_sign'] = "0"
    else:
        ret['pass_sign'] = "1"
    return JsonResponse(ret,safe=False)

@login_required
def apply_change_wlan(request):
    wl_id = int(request.GET.get('id'))
    ap_id = int(request.GET.get('ap_id'))
    errors = ""
    if request.method == 'POST':
        wlan_ssid = request.POST.get('ssid')
        if request.POST.get('wlan_service') == None:
            wlan_service = "off"
        else:
            wlan_service = request.POST.get('wlan_service')
        passphrase = request.POST.get('passphrase')
        if request.POST.get('vlan_enabled') == None:
            vlan_enabled = "off"
        else:
            vlan_enabled = request.POST.get('vlan_enabled')
        if request.POST.get('vlan') == '':
            vlan = '0'
        else:
            vlan = request.POST.get('vlan')
        upload_speed = request.POST.get('upload_speed')
        download_speed = request.POST.get('download_speed')
        radios_enable = request.POST.get('radios_enable')
        gw = Group_wlan.objects.get(pk = wl_id)

        wlan_ssid = "" if wlan_ssid == gw.wlan_ssid else wlan_ssid
        wlan_service = "" if wlan_service == gw.wlan_service else wlan_service
        passphrase = "" if passphrase == gw.passphrase else passphrase
        vlan_enabled = "" if vlan_enabled == gw.vlan_enabled else vlan_enabled
        vlan = "" if vlan == gw.vlan else vlan
        upload_speed = "" if upload_speed == gw.upload_speed else upload_speed
        download_speed = "" if download_speed == gw.download_speed else download_speed
        radios_enable = "" if radios_enable == gw.radios_enable else radios_enable

        try:
            if Device_wlan.objects.filter(device_id = ap_id,wlan_id = gw.wlan_id).exists():
                dw = Device_wlan.objects.get(device_id = ap_id,wlan_id = gw.wlan_id)
                if (wlan_ssid != "" or dw.wlan_ssid != "") or (wlan_service != "" or dw.wlan_service != "") or (passphrase != "" or dw.passphrase != "") or (vlan_enabled != "" or dw.vlan_enabled != "") or (vlan != "" or dw.vlan != "") or (wlan_ssid != "" or dw.wlan_ssid != "") or (upload_speed != "" or dw.upload_speed != "") or (download_speed != "" or dw.download_speed != "") or (radios_enable != "" or dw.radios_enable != "") :
                    dw.device_id = ap_id
                    dw.wlan_id = gw.wlan_id
                    dw.wlan_ssid = wlan_ssid
                    dw.wlan_service = wlan_service
                    dw.passphrase = passphrase
                    dw.vlan_enabled = vlan_enabled
                    dw.vlan = vlan
                    dw.upload_speed = upload_speed
                    dw.download_speed = download_speed
                    dw.radios_enable = radios_enable
                    dw.save()
            else:
                dw = Device_wlan()
                if wlan_ssid != "" or wlan_service != "" or passphrase != "" or vlan_enabled != "" or vlan != "" or wlan_ssid != "" or upload_speed != "" or download_speed != "" or radios_enable != "" :
                    dw.device_id = ap_id
                    dw.wlan_id = gw.wlan_id
                    dw.wlan_ssid = wlan_ssid
                    dw.wlan_service = wlan_service
                    dw.passphrase = passphrase
                    dw.vlan_enabled = vlan_enabled
                    dw.vlan = vlan
                    dw.upload_speed = upload_speed
                    dw.download_speed = download_speed
                    dw.radios_enable = radios_enable
                    dw.save()
                    # errors = _(u'创建成功！')
            m = Device.objects.get(pk = ap_id).mac
            create_wireless_json(Device.objects.get(mac = m))
            # data_str = set_wlan_str(m)
            # DataQuery.DQSetUpdateConfig('wireless',m,data_str)

            gp_str = portal_str(request.user.groupname)
            mac_list = []
            mac_list.append(m)
            task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')
        except Exception as e:
            print e
            errors = _(u'修改失败！')
    return JsonResponse(errors,safe=False)

@login_required
def reduction_change_wlan(request):
    wl_id = int(request.POST.get('id'))
    ap_id = int(request.POST.get('ap_id'))
    errors = ""
    try:
        wlan = Group_wlan.objects.get(pk = wl_id)
        if Device_wlan.objects.filter(device_id = ap_id,wlan_id = wlan.wlan_id).exists():
            apwl = Device_wlan.objects.get(device_id = ap_id,wlan_id = wlan.wlan_id)
            apwl.delete()
            m = Device.objects.get(pk = ap_id).mac
            create_wireless_json(Device.objects.get(pk = ap_id))
            # data_str = set_wlan_str(m)
            # DataQuery.DQSetUpdateConfig('wireless',m,data_str)

            gp_str = portal_str(request.user.groupname)
            mac_list = []
            mac_list.append(m)
            task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')
        else:
            pass
    except Exception as e:
        errors = _(u'还原失败！')
    return JsonResponse(errors,safe=False)

@login_required
def ap_auto_accept(request):
    accept = request.POST.get('accept')
    error = {'error':''}
    if accept == "open":
        try:
            ss = Setting.objects.get(SIGN = 1)
            ss.ap_auto_accept = 'open'
            ss.save()
            error['error'] = _(u'修改成功')
        except Exception as e:
            print e
            error['error'] = _(u'修改失败')
    elif accept == "get":
        try:
            ss = Setting.objects.get(SIGN = 1)
            result = ss.ap_auto_accept
        except Exception as e:
            print e
            result = ""
        return JsonResponse(result,safe = False)
    else:
        try:
            ss = Setting.objects.get(SIGN = 1)
            ss.ap_auto_accept = 'close'
            ss.save()
            error['error'] = _(u'修改成功')
        except Exception as e:
            print e
            error['error'] = _(u'修改失败')
    return JsonResponse(error,safe= False)


@login_required
def ap_auto_update(request):
    error = {'error':''}
    if request.method == 'GET':
        get_type = request.GET.get('get_type')
        if get_type == 'option':
            result = []
            pg = Probe_group.objects.filter(account_group_name = request.user.groupname,group_type = '1')
            for i in pg:
                result.append({
                    'id': str(i.pk),
                    'text': str(i.group_name),
                })
            return JsonResponse({'results':result},safe = False)
        elif get_type == 'load_result':
            if Auto_Update_Rule.objects.filter(groupname = request.user.groupname,support_mode = '1').exists():
                auto_update_rule = Auto_Update_Rule.objects.get(groupname = request.user.groupname,support_mode = '1')
                result = auto_update_rule.rule
            else:
                result = json.dumps([])
            return JsonResponse(result,safe = False)

    elif request.method == 'POST':
        rule = request.POST.getlist('rule[]')
        if not rule:
            rule = []
        rule = json.dumps(rule)
        print rule
        if Auto_Update_Rule.objects.filter(groupname = request.user.groupname,support_mode = '1').exists():
            auto_update_rule = Auto_Update_Rule.objects.get(groupname = request.user.groupname,support_mode = '1')
        else:
            auto_update_rule = Auto_Update_Rule()
            auto_update_rule.groupname = request.user.groupname
            auto_update_rule.support_mode = '1'
        auto_update_rule.rule = rule
        try:
            auto_update_rule.save()
            error['error'] = _(u'修改成功')
        except Exception as e:
            print e
            error['error'] = _(u'修改失败')
        return JsonResponse(error,safe= False)

@login_required
def blacklist_add(request):
    blacklist = json.loads(request.POST.get('bl'))
    error = {'error':'','su':'false'}
    if blacklist != []:
        try:
            for i in blacklist:
                bl = ApBlackList()
                bl.mac = i
                bl.support_mode = '1'
                bl.save()
            error['error'] = _(u'执行成功')
            error['su'] = 'true'
        except Exception as e:
            print e
            error['error'] = _(u'执行失败')
    return JsonResponse(error,safe=False)

@login_required
def black_list_add_ajax(request):
    # dev = Device.objects.filter(support_mode = "1")
    dev = Device.objects.filter(support_mode = 1)
    data = []
    result = {}
    for d in dev:
        if ApBlackList.objects.filter(mac = d.mac).exists():
            pass
        else:
            if d.name == "":
                name = d.model + "_" + d.mac[:6]
            else:
                name = d.name
            data.append({'name':name,'mac':d.mac})
    result = {'data':data}
    return JsonResponse(result,safe = False)

@login_required
def black_list_table(request):
    apb = ApBlackList.objects.filter(support_mode = '1')
    data = []
    for i in apb:
        if Device.objects.filter(mac = i.mac).exists():
            d = Device.objects.get(mac = i.mac)
        else:
            d = ''
        if d == "":
            name = i.mac[:2] +":"+ i.mac[2:4] +":"+ i.mac[4:6] +":"+ i.mac[6:8] +":"+ i.mac[8:10] +":"+ i.mac[10:]
        else:
            if d.name == "":
                name = d.model + "_" + d.mac[:6]
            else:
                name = d.name
        ap = model_to_dict(i)
        ap['name'] = name
        data.append(ap)
    return JsonResponse({'data':data},safe=False)

@login_required
def blacklist_remove(request):
    mac = request.POST.get('mac')
    error = {'error':''}
    try:
        ApBlackList.objects.get(mac = mac).delete()
        error['error'] = _(u'操作成功')
    except Exception as e:
        print e
        error['error'] = _(u'操作失败')
    return JsonResponse(error,safe=False)

@login_required
def customer_list(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)

    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,"ap-customer.html",{'admin_error_json':admin_error_json,'admin_error':admin_error,'error':error,"error_json":error_json,'plimit':plimit,'oemlimit':oemlimit})

@login_required
def customer_table(request):
    # print 2131231,request.GET
    draw = int(request.GET.get('draw'))
    start = int(request.GET.get('start'))
    length = int(request.GET.get('length'))
    order_id = request.GET.get('order[0][column]')
    order = request.GET.get('columns['+order_id+'][data]')
    customer_type_receive = request.GET.get('customer_type')

    # radio id  0-> 2G 1-> 5G
    if customer_type_receive == '2G':
        customer_type_1 = 0
        customer_type_2 = 0
    elif customer_type_receive == '5G':
        customer_type_1 = 1
        customer_type_2 = 1
    else:
        customer_type_1 = 0
        customer_type_2 = 1

    if request.GET.get('order[0][dir]') == 'asc':
        order_type = ''
    elif request.GET.get('order[0][dir]') == 'desc':
        order_type = '-'
    search_value_old = request.GET.get('search[value]')
    search_value = search_value_old

    online_time = int(time.time()) - 900
    p_list = []
    if request.user.administrator_permission >= 4:
        p_list = [i.mac for i in Device.objects.filter(support_mode = 1)]

    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        p_list = [i.mac for i in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = 1))]
    if len(p_list) == 0:
        recordsFiltered = 0
        recordsTotal = 0
        customer_list = []
    else:
        # # 解决单元素tuple 末尾,问题
        # if len(p_list) == 1:
        #     p_list = p_list + p_list
        recordsTotal = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list)&(Q(radioid = customer_type_1)|Q(radioid = customer_type_2))).count()
        if search_value == '':
            recordsFiltered = recordsTotal
            customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list)&(Q(radioid = customer_type_1)|Q(radioid = customer_type_2))).order_by(order_type+order)[start:start+length]
        else:
            recordsFiltered = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list)&(Q(radioid = customer_type_1)|Q(radioid = customer_type_2))).filter(Q(ip__icontains=search_value)|Q(mac__icontains=search_value)|Q(ap__icontains=search_value)|Q(uphone__icontains=search_value)).count()
            customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list)&(Q(radioid = customer_type_1)|Q(radioid = customer_type_2))).filter(Q(ip__icontains=search_value)|Q(mac__icontains=search_value)|Q(ap__icontains=search_value)|Q(uphone__icontains=search_value))[start:start+length]


    ret = {'draw':draw,'recordsTotal':recordsTotal,'recordsFiltered':recordsFiltered,'data':[]}
    timenow = timezone.now()
    for al in customer_list:

        # al.last_heart_time = time.localtime(al.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")
        al.livetime = sec_to_hour(al.livetime)
        al.uptime = sec_to_hour(al.uptime)
        data = model_to_dict(al)
        if Customer_name.objects.filter(mac = al.mac).exists():
            cus = Customer_name.objects.get(mac = al.mac).name
        else:
            cus = ""
        if cus != "":
            name = cus
        elif al.hostname != "":
            name = al.hostname
        elif al.devtype != "":
            name = al.devtype +'_'+ al.mac
        else:
            name = al.mac[:2] +":"+ al.mac[2:4] +":"+ al.mac[4:6] +":"+ al.mac[6:8] +":"+ al.mac[8:10] +":"+ al.mac[10:]
        data['name'] = name

        if Device.objects.filter(mac = al.ap).exists():
            dev = Device.objects.get(mac = al.ap)
            if Device_wlan.objects.filter(device_id = dev.pk,wlan_id = str(al.wlanid-1)).exists():
                wlan = Device_wlan.objects.get(device_id = dev.pk,wlan_id = str(al.wlanid-1)).wlan_ssid
                if wlan == "":
                    if Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).exists():
                        wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).wlan_ssid
            elif Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).exists():
                wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).wlan_ssid
            else:
                wlan = ""
            if dev.name != "":
                ap_dev = dev.name
            else:
                ap_dev = dev.model + "_" + dev.mac

            a = int((timenow - dev.last_heart_time).total_seconds())
            dev.last_heart_time = timezone.localtime(dev.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")
            rss = dev.reboot_sign.split(",")
            if dev.last_heart_time == "1970-1-1 08:00:00":
                state = _(u"离线")
            elif a >= 86400:
                state = _(u"退服")
            elif a >= 3600:
                state = _(u"离线")
            elif a >= 900 :
                state = _(u"超时")
            elif a < 900:
                if rss[0] != '0' and rss[0] != '':
                    if rss[0] == '1':
                        state = _(u"重启")
                    elif rss[0] == '2':
                        state = _(u"升级")
                else:
                    state = _(u"在线")
            dev_id = dev.pk
        else:
            wlan = ""
            ap_dev = ""
            state = _(u"退服")
            dev_id = 0
        data['wlan'] = wlan
        data['ap_dev'] = ap_dev
        data['state'] = state
        data['dev_id'] = dev_id
        ret['data'].append(data)

    return JsonResponse(ret,safe = False)

@login_required
def ap_customer_tab_table(request):
    mac = request.GET.get('mac')
    ret = {'data':[]}

    online_time = int(time.time()) - 900

    customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap = mac))

    print 'qwqwqwqw',len(customer_list),'23232323232',customer_list
    for al in customer_list:
        # al.last_heart_time = time.localtime(al.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")
        # al.livetime = sec_to_hour(al.livetime)
        data = model_to_dict(al)
        if Customer_name.objects.filter(mac = al.mac).exists():
            cus = Customer_name.objects.get(mac = al.mac).name
        else:
            cus = ""
        if cus != "":
            name = cus
        elif al.hostname != "":
            name = al.hostname
        else:
            name = al.devtype +'_'+ al.mac
        data['name'] = name
        if Device.objects.filter(mac = al.ap).exists():
            dev = Device.objects.get(mac = al.ap)
            if Device_wlan.objects.filter(device_id = dev.pk,wlan_id = str(al.wlanid-1)).exists():
                wlan = Device_wlan.objects.get(device_id = dev.pk,wlan_id = al.wlanid-1).wlan_ssid
                if wlan == "":
                    if Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).exists():
                        wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).wlan_ssid
            elif Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).exists():
                wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).wlan_ssid
            else:
                wlan = ""
            if dev.name != "":
                ap_dev = dev.name
            else:
                ap_dev = dev.model + "_" + dev.mac
        else:
            wlan = ""
            ap_dev = ""
        data['wlan'] = wlan
        data['ap_dev'] = ap_dev

        ret['data'].append(data)
    return JsonResponse(ret,safe = False)

@login_required
def customer_detail(request):
    ret = {}
    pk = request.GET.get('id')
    if Customer.objects.filter(pk = pk).exists():
        ret = model_to_dict(Customer.objects.get(pk = pk))
    else:
        ret = {}
    if ret != {}:
        ret['last_heart_time'] = time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(ret['last_heart_time']))
        ret['livetime'] = sec_to_hour(ret['livetime'])
        ret['uptime'] = sec_to_hour(ret['uptime'])
        if Customer_name.objects.filter(mac = ret['mac']).exists():
            cus = Customer_name.objects.get(mac = ret['mac']).name
        else:
            cus = ""
        if cus != "":
            name = cus
        elif ret['hostname'] != "":
            name = ret['hostname']
        elif ret['devtype'] != "":
            name = ret['devtype'] +'_'+ ret['mac']
        else:
            name = ret['mac'][:2] +":"+ ret['mac'][2:4] +":"+ret['mac'][4:6] +":"+ ret['mac'][6:8] +":"+ ret['mac'][8:10] +":"+ ret['mac'][10:]
        ret['name'] = name

        if Device.objects.filter(mac = ret['ap']).exists():
            dev = Device.objects.get(mac = ret['ap'])
            if Device_wlan.objects.filter(device_id = dev.pk,wlan_id = str(ret['wlanid']-1)).exists():
                wlan = Device_wlan.objects.get(device_id = dev.pk,wlan_id = str(ret['wlanid']-1)).wlan_ssid
                if wlan == "":
                    if Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(ret['wlanid']-1)).exists():
                        wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(ret['wlanid']-1)).wlan_ssid
            elif Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(ret['wlanid']-1)).exists():
                wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(ret['wlanid']-1)).wlan_ssid
            else:
                wlan = ""
            if dev.name != "":
                ap_dev = dev.name
            else:
                ap_dev = dev.model + "_" + dev.mac
            timenow = timezone.now()
            a = int((timenow - dev.last_heart_time).total_seconds())
            dev.last_heart_time = timezone.localtime(dev.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")
            rss = dev.reboot_sign.split(",")
            if dev.last_heart_time == "1970-1-1 08:00:00":
                state = _(u"离线")
            elif a >= 86400:
                state = _(u"退服")
            elif a >= 3600:
                state = _(u"离线")
            elif a >= 900 :
                state = _(u"超时")
            elif a < 900:
                if rss[0] != '0' and rss[0] != '':
                    if rss[0] == '1':
                        state = _(u"重启")
                    elif rss[0] == '2':
                        state = _(u"升级")
                else:
                    state = _(u"在线")
            dev_id = dev.pk

            # if hasattr(dev,'device_ap'):
            #     if ret['radioid'] == 0:
            #         channel = dev.device_ap.radios_2_channel
            #     else:
            #         channel = dev.device_ap.radios_5_channel
            # else:
            #     channel = 'auto'
        else:
            wlan = ""
            ap_dev = ""
            state = _(u"退服")
            dev_id = 0
            # channel = 'auto'
        ret['wlan'] = wlan
        ret['ap_dev'] = ap_dev
        ret['state'] = state
        ret['dev_id'] = dev_id
        # ret['channel'] = channel

    return JsonResponse(ret,safe = False)

@login_required
def custome_detail_table(request):
    mac = request.GET.get('mac')
    ret = {'data':[]}

    customer_list = Customer_history.objects.filter(mac = mac).order_by('-last_heart_time')
    # al_last = {}
    for al in customer_list:
        if al.online_sign == 'offline' and al.uptime >= 300:
            data = {
                'livetime':sec_to_hour(al.uptime),
                'last_heart_time':time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(al.last_heart_time)),
                'tx_bytes':al.down,
                'rx_bytes':al.up,
            }
            ret['data'].append(data)
        # if al_last == {}:
        #     al_last = model_to_dict(al)
        # else:
        #     if al_last['online_sign'] == 'offline':
                # if al.online_sign == "online":
                #     data = {
                #         'livetime':sec_to_hour(abs(al_last['last_heart_time'] - al.last_heart_time)),
                #         'last_heart_time':time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(al_last['last_heart_time'])),
                #         'tx_bytes':al_last['down'],
                #         'rx_bytes':al_last['up'],
                #     }
                #     ret['data'].append(data)
                #     al_last = model_to_dict(al)
                # else :
                #     al_last = model_to_dict(al)
                #     al_last['online_sign'] = "offline"

            # else:
            #     al_last = model_to_dict(al)


    return JsonResponse(ret,safe = False)

@login_required
def change_customer_name(request):
    mac = request.POST.get('mac')
    name = request.POST.get('name')
    errors = {'sign':'false','mes':''}

    if Customer_name.objects.filter(mac = mac).exists():
        cus = Customer_name.objects.get(mac = mac)
    else:
        cus = Customer_name()
        cus.mac = mac
    cus.name = name
    try:
        cus.save()
        errors['sign'] = 'true'
        errors['mes'] = _(u'修改成功')
    except Exception as e:
        print e
        errors['sign'] = 'false'
        errors['mes'] = _(u'修改失败')

    return JsonResponse(errors,safe = False)


def sec_to_hour(value):
    try:
        if value != '' and value != 0:
            a = int(value)
            b = a/(24*3600)
            c = a%(24*3600)/3600
            d = a%(24*3600)%3600/60
            e = a%(24*3600)%3600%60
            if b == 0 and c == 0 and d == 0:
                return "%ds"%(e)
            elif b == 0 and c == 0:
                return "%dm %ds"%(d,e)
            elif b == 0:
                return "%dh %dm %ds"%(c,d,e)
            else:
                return "%dd %dh %dm %ds"%(b,c,d,e)
        else:
            return value
    except Exception as e:
        print e
        return value

@login_required
def gpon(request):
    mac = request.GET.get('mac')
    # mac = "001122334455"
    try:
        gpon = Gpon.objects.get(ap_mac = mac)
        gpon = model_to_dict(gpon)
        print '1232312321'
        gpon['null'] = 'false'
    except Exception as e:
        print e
        gpon = {'null':'true'}
    return JsonResponse(gpon,safe = False)


@login_required
def gpon_config(request):
    mac = request.POST.get('mac')
    set_type = request.POST.get('type')
    error = {'sign':''}
    try:
        gpon = Gpon.objects.get(ap_mac = mac)
    except Exception as e:
        print e,11111
        error['sign'] = 'false'
        error['mes'] = _(u'修改失败')
    if error['sign'] != 'false':
        param = ''
        try:
            # if set_type == 'sn':
            #     gpon.gpon_sn = value
            #     param = 'sn;' + str(value)
            if set_type == 'pwd':
                value = request.POST.get('value')
                gpon.gpon_pwd = value
                param = 'pwd;' + str(value)
            # if set_type == 'oui':
            #     gpon.manufacturer_oui = value
            #     param = 'oui;' + str(value)
            if set_type == 'update':
                value = json.loads(request.POST.get('value'))
                gpon.update_link = value[0]
                gpon.update_file = value[1]
            data = {'action':'gpcmd','mac':mac,'param':param}

            DataQuery.DQProcess(data)

            gpon.save()
            #触发下发
            error['sign'] = 'true'
            error['mes'] = _(u'修改成功')
        except Exception as e:
            print e,2222222
            error['sign'] = 'false'
            error['mes'] = _(u'修改失败')
    return JsonResponse(error,safe = False)

@login_required
def gpon_action(request):
    error = {'mes':'','sign':'false'}
    mac = request.POST.get('mac')
    action_type = request.POST.get('type')
    param = ''
    try:
        if action_type == 'reboot':
            param = 'reboot'
            data = {'action':'gpcmd','mac':mac,'param':param}
        if action_type == 'update':
            update_link = ''
            update_file = ''

            if hasattr(Setting.objects.get(SIGN = 1),'setting_gpon'):
                gp_st = Setting_gpon.objects.get(setting_id = Setting.objects.get(SIGN = 1))
                update_link = gp_st.update_link if gp_st.update_link != "" else update_link
                update_file = gp_st.update_file if gp_st.update_file != "" else update_file
            if Group_gpon.objects.filter(group_id = Device.objects.get(mac = mac).group_id).exists():
                gp_g = Group_gpon.objects.get(group_id = Device.objects.get(mac = mac).group_id)
                update_link = gp_g.update_link if gp_g.update_link != "" else update_link
                update_file = gp_g.update_file if gp_g.update_file != "" else update_file
            if Gpon.objects.filter(ap_mac = mac).exists():
                gp = Gpon.objects.get(ap_mac = mac)
                update_link = gp.update_link if gp.update_link != "" else update_link
                update_file = gp.update_file if gp.update_file != "" else update_file

            param = {
                "upgradeserver": str(update_link),
                "upgradefile": str(update_file),
            }
            data = {'action':'gppcmd','mac':mac,'param':param}
        # if action_type == 'save_data':
        #     param = 'save'
        # if action_type == 'recover':
        #     param = 'restore'
        DataQuery.DQProcess(data)
        error['sign'] = 'true'
        error['mes'] = _(u'操作成功')
    except Exception as e:
        print e,'asasasaasasa'
        error['sign'] = 'false'
        error['mes'] = _(u'操作失败')

    return JsonResponse(error,safe = False)

@login_required
def gpon_config_group(request):
    gp = request.POST.get('gp')
    set_type = request.POST.get('type')
    value = json.loads(request.POST.get('value'))
    error = {'sign':''}
    try:
        group_gpon = Group_gpon.objects.get(group_id = int(gp))
    except Exception as e:
        print e
        group_gpon = Group_gpon()
        group_gpon.group_id = int(gp)
    try:
        if set_type == 'update':
            group_gpon.update_link = value[0]
            group_gpon.update_file = value[1]
            group_gpon.save()
            error['sign'] = 'true'
            error['mes'] = _(u'修改成功')
    except Exception as e:
        print e
        error['sign'] = 'false'
        error['mes'] = _(u'修改失败')

    return JsonResponse(error,safe = False)

@login_required
def gpon_config_system(request):
    set_type = request.POST.get('type')
    value = json.loads(request.POST.get('value'))
    error = {'sign':''}
    try:
        ss = Setting.objects.get(SIGN = 1)
        try:
            sys_gpon = Setting_gpon.objects.get(setting_id = ss.pk)
        except Exception as e:
            print e
            sys_gpon = Setting_gpon()
            sys_gpon.setting_id = ss.pk
        try:
            if set_type == 'update':
                sys_gpon.update_link = value[0]
                sys_gpon.update_file = value[1]
                sys_gpon.save()
                error['sign'] = 'true'
                error['mes'] = _(u'修改成功')
        except Exception as e:
            print e
            error['sign'] = 'false'
            error['mes'] = _(u'修改失败')
    except Exception as e:
        print e
        error['sign'] = 'false'
        error['mes'] = _(u'修改失败')
    return JsonResponse(error,safe = False)

def show_gpon_config(request):
    req_type = request.GET.get('type')
    data = {}
    if req_type == "sys":
        try:
            ss = Setting.objects.get(SIGN = 1)
            sys_gpon = Setting_gpon.objects.get(setting_id = ss.pk)
            update_link = sys_gpon.update_link
            update_file = sys_gpon.update_file
        except Exception as e:
            print e
            update_link = ''
            update_file = ''
    if req_type == "group":
        gp = request.GET.get('gp')
        try:
            group_gpon = Group_gpon.objects.get(group_id = int(gp))
            update_link = group_gpon.update_link
            update_file = group_gpon.update_file
        except Exception as e:
            print e
            try:
                ss = Setting.objects.get(SIGN = 1)
                sys_gpon = Setting_gpon.objects.get(setting_id = ss.pk)
                update_link = sys_gpon.update_link
                update_file = sys_gpon.update_file
            except Exception as e:
                print e
                update_link = ''
                update_file = ''
    if req_type == "dev":
        mac = request.GET.get('mac')
        try:
            gpon = Gpon.objects.get(ap_mac = mac)
            update_link = gpon.update_link
            update_file = gpon.update_file
        except Exception as e:
            print e
            try:
                group_gpon = Group_gpon.objects.get(group_id = int(gp))
                update_link = group_gpon.update_link
                update_file = group_gpon.update_file
            except Exception as e:
                print e
                try:
                    ss = Setting.objects.get(SIGN = 1)
                    sys_gpon = Setting_gpon.objects.get(setting_id = ss.pk)
                    update_link = sys_gpon.update_link
                    update_file = sys_gpon.update_file
                except Exception as e:
                    print e
                    update_link = ''
                    update_file = ''
    data['update_link'] = update_link
    data['update_file'] = update_file
    return JsonResponse(data,safe = False)


@login_required
def ap_index(request):
    admin_error = {"error_type":'',"error_msg":''}
    p_all = ''
    onldevs = 0
    ofldevs = 0
    p_list = []
    logger.info('start run sql --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    dev_online_time = timezone.now() - datetime.timedelta(minutes = 15)
    if request.user.administrator_permission >= 4 :
        p_all = Device.objects.filter(support_mode = 1)
        logger.info('end run sql and start run list --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        p_list = [i.mac for i in p_all]
        logger.info('end run list --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        ##########xuyaoxiugai##################
        # onldevs = DataQuery.DQGetOnlinedDevs(p_list)
        # ofldevs = DataQuery.DQGetOfflinedDevs(p_list)
        onldevs = Device.objects.filter(Q(support_mode = 1) & Q(last_heart_time__gte = dev_online_time)).count()
        logger.info('end run onldevs --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        ofldevs = Device.objects.filter(Q(support_mode = 1) & Q(last_heart_time__lt = dev_online_time)).count()
        logger.info('end run ofldevs --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        p_all = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1'))
        p_list = [i.mac for i in p_all]
        ##########xuyaoxiugai##################
        # onldevs = DataQuery.DQGetOnlinedDevs(p_list)
        # ofldevs = DataQuery.DQGetOfflinedDevs(p_list)
        onldevs = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = 1) & Q(last_heart_time__gte = dev_online_time)).count()
        ofldevs = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = 1) & Q(last_heart_time__lt = dev_online_time)).count()

    online_time = int(time.time()) - 900
    logger.info('start run usercountsql --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    # pc_user = DataQuery.DQGetAPCache_nojson(request.user.groupname+"cache_pc_user")
    pc_user = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = "online")&Q(ap__in = p_list)&Q(devtype__in = ['PC','Mac PC'])).count()
    logger.info('end run pcuser --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    # phone_user = DataQuery.DQGetAPCache_nojson(request.user.groupname+"cache_phone_user")
    # phone_user = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = "online")&Q(ap__in = p_list)&Q(devtype__in = ['iphone','ipad','android'])).count()
    phone_user = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = "online")&Q(ap__in = p_list)).count() - pc_user
    logger.info('end run phoneuser --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    ap = {'online':onldevs,'offline':ofldevs,'pc_user':pc_user,'phone_user':phone_user}

    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'ap_index.html',{'ap':ap,'admin_error_json':admin_error_json,'admin_error':admin_error,'plimit':plimit,'oemlimit':oemlimit})

def ap_sort_type_change(request):
    sort_type = request.GET.get('type')
    top = []
    p_list = []
    # if request.user.administrator_permission >= 4 :
    #     p_list = [i.mac for i in Device.objects.filter(support_mode = '1')]
    # elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
    #     # p_all = Device.objects.filter(account_group_name = request.user.groupname)
    #     p_list = [i.mac for i in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1'))]
    #########xuyaoxiugai##################
    if sort_type == '1':
        # top1 = {'mac':'a4e6b1300005','hour':'1h30min22s'}
        # top2 = {'mac':'fcad0f0344b1','hour':'1h30min25s'}
        # top = [top1,top2]
        # top = DataQuery.DQGetTopItem("all", "onlinetime", p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_onlinetime")
    elif sort_type == '2':
        # top2 = {'mac':'a4e6b1300005','upload':'11111mb'}
        # top1 = {'mac':'fcad0f0344b1','upload':'111121mb'}
        # top = [top1,top2]
        # top = DataQuery.DQGetTopItem("all", "upload", p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_upload")
    elif sort_type == '3':
        # top2 = {'mac':'a4e6b1300005','upload':'11111mb'}
        # top1 = {'mac':'fcad0f0344b1','upload':'111121mb'}
        # top = [top1,top2]
        # top = DataQuery.DQGetTopItem("all", "download", p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_download")
    elif sort_type == '4':
        # top2 = {'mac':'a4e6b1300005','upload':'11111mb'}
        # top1 = {'mac':'fcad0f0344b1','upload':'111121mb'}
        # top = [top1,top2]
        # top = DataQuery.DQGetTopItem("all", "download", p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_immediateFlow")
    try:
        timenow = timezone.now()
        for t in top:
            name = ''
            dev_id = ''
            state = ''
            if Device.objects.filter(mac = t['mac']).exists():
                al = Device.objects.get(mac = t['mac'])
                name = al.name
                dev_id = al.pk
                state = state_check(t['mac'])
            if name == '':
                name = Device.objects.get(mac = t['mac']).model+ '_' + t['mac'][6:12]
            t['name'] = name
            t['dev_id'] = dev_id
            t['state'] = state
    except Exception as e:
        print e

    return JsonResponse(top,safe = False)

def ap_chartajax(request):
    net_type = request.GET.get('type')
    p_list = []
    data = {
        'type':'3G',
        'mac':[],
        'upload':[],
        'download':[],
        'name':[],
    }
    if net_type == '3g':
    #     data = {
    #     'type':'3G',
    #     'mac':[],
    #     'upload':[],
    #     'download':[],
    #     'name':[],
    # }
        data = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_net_type_3g")
        data['type'] = '3G'
    elif net_type == '4g':
    #     data = {
    #     'type':'4G',
    #     'mac':[],
    #     'upload':[],
    #     'download':[],
    #     'name':[],
    # }
        data = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_net_type_4g")
        data['type'] = '4G'
    elif net_type == 'car':
    #     data = {
    #     'type':_(u'车载'),
    #     'mac':[],
    #     'upload':[],
    #     'download':[],
    #     'name':[],
    # }
        data = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_net_type_car")
        data['type'] = _(u'车载')
    elif net_type == 'normal':
    #     data = {
    #     'type':_(u'普通'),
    #     'mac':[],
    #     'upload':[],
    #     'download':[],
    #     'name':[],
    # }
        data = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_net_type_normal")
        data['type'] = _(u'普通')
    # if request.user.administrator_permission >= 4 :
    #     p_list = [i.mac for i in Device.objects.filter(support_mode = '1')]
    # elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
    #     p_list = [i.mac for i in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1'))]
        ##########xuyaoxiugai##################
    # DataQuery.DQGetTopItemByDtype(data, net_type, p_list)


    try:
        for t in data['mac']:
            name = ''
            if Device.objects.filter(mac = t).exists():
                name = Device.objects.get(mac = t).name
            if name == '':
                name = Device.objects.get(mac = t).model+ '_' + t[6:12]
            #print 11111
            #print name
            #print 22222
            data['name'].append(name)
    except Exception as e:
        print e

    #print json.dumps(data)
    return JsonResponse(data)

def ap_user_chartajax(request):
    # net_type = request.GET.get('type')
    data = {
        'mac':[],
        'num':[],
        'name':[],
        'upload':[],
        'download':[],
    }
    p_list = []
    # if request.user.administrator_permission >= 4 :
    #     p_list = [i.mac for i in Device.objects.filter(support_mode = '1')]
    # elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
    #     p_list = [i.mac for i in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1'))]
        ##########xuyaoxiugai##################
    # data = DataQuery.DQGetTopAPUsers(data, p_list)
    data = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_user_chart")
    try:
        for t in data['mac']:
            name = ''
            if Device.objects.filter(mac = t).exists():
                name = Device.objects.get(mac = t).name
            if name == '':
                name = Device.objects.get(mac = t).model+ '_' + t[6:12]
            #print 11111
            #print name
            #print 22222
            data['name'].append(name)
        for uflow in data['uflow']:
            if uflow['z'] == 0:
                uflow['z'] = _(u'小于1')
            if uflow['w'] == 0:
                uflow['w'] = _(u'小于1')

    except Exception as e:
        print e
    print data,'xxxxxxxxxxxx'
    #print json.dumps(data)
    return JsonResponse(data)

def ap_ssid_select(request):
        group_list = []
        for i in Probe_group.objects.filter(account_group_name = request.user.groupname, group_type = 1):
            # print request.user.username,'ssssss',request.user.pk,  'GPGPGPGPGPGPGPGPGPGPGP', request.user.groupname
            # print dir(request.user), 'GPGPGPGPGPGPGPGPGPGPGP'
            group_list.append(i.pk)
        # print group_list, 'GGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGG'
        # for i in group_list:
        #     print i, 'GPGPGPGPGPGPGPGPGPGPGP'
        wlan_list = []
        if len(group_list) != 0:
            for i in Group_wlan.objects.filter(group_id__in = group_list):
                wlan_list.append({'group_id':i.group_id.pk,'wlan_id':i.wlan_id,'wlan_ssid':i.wlan_ssid,'group_name':i.group_id.group_name})
        print wlan_list, 'WWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW'
        return JsonResponse(wlan_list,safe = False)

def ap_user_counterajax(request):
    net_type = request.GET.get('type')
    print 'NNNNNNNNNNNNN', net_type, 'NNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNN'

    gpwl = net_type.split(';')
    if len(gpwl) == 2:
        groupId = gpwl[0]
        wlanId = gpwl[1]
    else:
        # defalt group id, wlan id
        groupId = '5'
        wlanId = '1'

    data = {
        'name':[],
        'time':[],
        'num':[],
        'upload':[],
        'download':[],
        'uflow':[],
    }
    p_list = []
    # data = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_user_chart")
    # try:
    #     for t in data['mac']:
    #         name = ''
    #         if Device.objects.filter(mac = t).exists():
    #             name = Device.objects.get(mac = t).name
    #         if name == '':
    #             name = Device.objects.get(mac = t).model+ '_' + t[6:12]
    #         data['name'].append(name)
    # except Exception as e:
    #     print e



    # data['time'] = ['17:00', '18:00', '19:00', '20:00', '21:00', '22:00', '23:00']
    # if Account_Group.objects.filter(groupname = request.user.groupname).exists():
    #     account_group = Account_Group.objects.get(groupname = request.user.groupname)
    # group_list = []
    # for i in Probe_group.objects.filter(account_group_name = request.user.groupname, group_type = 3):
    #     group_list.append(i.pk)
    # # print group_list, 'GGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGG'
    # # for i in group_list:
    # #     print i, 'GPGPGPGPGPGPGPGPGPGPGP'
    # wlan_list = []
    # if len(group_list) != 0:
    #     for i in Group_wlan.objects.filter(group_id__in = group_list):
    #         wlan_list.append({'group_id':i.group_id.pk,'wlan_id':i.wlan_id,'wlan_ssid':i.wlan_ssid,'group_name':i.group_id.group_name})

    # print wlan_list, 'WWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW'

    # groupId = '5'
    # wlanId = '3'
    timstamp = int(time.time())
    timstamp = timstamp - timstamp%3600
    uflow = []
    uflowdic = {}
    n = range(24)
    for i in reversed(n):
        timetmp = timstamp - 3600*i
        # print i, time.strftime("%H:%M", time.localtime(timetmp))
        data['time'].append(time.strftime("%H:%M", time.localtime(timetmp)))
        # data['num'].append(DataQuery.DQGetUsersByGroupID(groupId, wlanId, timetmp))
        # [
        #     {'y': 100, 'z': 1000},
        #     {'y': 1020, 'z': 1000}
        # ]
        # data['upload'].append(float(DataQuery.DQGetUsersDownLoadByGroupID(groupId, wlanId, timetmp))/1024/1024)
        # data['download'].append(float(DataQuery.DQGetUsersUpLoadByGroupID(groupId, wlanId, timetmp))/1024/1024)
        uflowdic['y'] = int(DataQuery.DQGetUsersByGroupID(groupId, wlanId, timetmp))

        uflowdic['z'] = int(float(DataQuery.DQGetUsersUpLoadByGroupID(groupId, wlanId, timetmp))/1024/1024)

        uflowdic['w'] = int(float(DataQuery.DQGetUsersDownLoadByGroupID(groupId, wlanId, timetmp))/1024/1024)
        uflow.append(uflowdic)
        uflowdic = {}

    # data['num'] = [123, 435, 123, 3457, 2456, 212, 1123]
    # data['upload'] = [456, 435, 3456, 3457, 3456, 2112, 3123]
    # data['download'] = [658, 435, 769, 3457, 3246, 2112, 1223]
    # data['name'] = ['witrusty']
    data['uflow'] = uflow
    for i in data['uflow']:
        if i['y'] != 0:
            if i['z'] == 0:
                i['z'] = _(u'小于1')
            if i['w'] == 0:
                i['w'] = _(u'小于1')
    print data,'PPPPPPPPPPPPPPPPPPPPPPPPPPPP'
    return JsonResponse(data)

def ap_warningajax(request):
    warning_type = request.GET.get('type')
    top = []
    # if request.user.administrator_permission >= 4 :
    #     p_list = [i.mac for i in Device.objects.filter(support_mode = '1')]
    # elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
    #     # p_all = Device.objects.filter(account_group_name = request.user.groupname)
    #     p_list = [i.mac for i in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '1'))]
        ##########xuyaoxiugai##################

    if warning_type == 'cpu':
        # top1 = {'mac':'00224f00f3aa','type':'CPU','num':'90%','name':'','ip':'192.168.1.79:36401','version':'3.04.16.171100','time':'2017-04-26 18:12:39'}
        # top2 = {'mac':'00224f00f3aa','type':'CPU','num':'90%','name':'','ip':'192.168.1.79:36401','version':'3.04.16.171100','time':'2017-04-26 18:12:39'}
        # top = [top1,top2]
        # top = DataQuery.DQGetWarnByDtype('CPU', p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_cpu")
    elif warning_type == 'flash':
        # top1 = {'mac':'a4e6b1300005','type':'FLASH','num':'90%','name':'','ip':'192.168.1.79:36401','version':'3.04.16.171100','time':'2017-04-26 18:12:39'}
        # top2 = {'mac':'fcad0f0344b1','type':'FLASH','num':'90%','name':'','ip':'192.168.1.79:36401','version':'3.04.16.171100','time':'2017-04-26 18:12:39'}
        # top = [top1,top2]
        # top = DataQuery.DQGetWarnByDtype('FLASH', p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_flash")
    elif warning_type == 'memory':
        # top1 = {'mac':'a4e6b1300005','type':'MEMORY','num':'90%','name':'','ip':'192.168.1.79:36401','version':'3.04.16.171100','time':'2017-04-26 18:12:39'}
        # top2 = {'mac':'fcad0f0344b1','type':'MEMORY','num':'90%','name':'','ip':'192.168.1.79:36401','version':'3.04.16.171100','time':'2017-04-26 18:12:39'}
        # top = [top1,top2]
        # top = DataQuery.DQGetWarnByDtype('MEMORY', p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"cache_memory")

    try:
        for t in top:
            name = ''
            state = ''
            dev_id = ''
            if Device.objects.filter(mac = t['mac']).exists():
                name = Device.objects.get(mac = t['mac']).name
                dev_id = Device.objects.get(mac = t['mac']).pk
            state = state_check(t['mac'])
            if name == '':
                name = Device.objects.get(mac = t['mac']).model+ '_' + t['mac'][6:12]
            t['name'] = name
            t['state'] = state
            t['dev_id'] = dev_id
    except Exception as e:
        print e
    return JsonResponse(top,safe = False)

@login_required
def customer_black_list_table(request):
    cusb = Customer_black_list.objects.filter(groupname = request.user.groupname)
    data = []
    for i in cusb:
        if Customer_name.objects.filter(mac = i.mac).exists():
            name = Customer_name.objects.get(mac = i.mac).name
        else:
            name = ""
        if Customer.objects.filter(mac = i.mac).exists():
            d = Customer.objects.get(mac = i.mac)
        else:
            d = ""
        if name != "":
            pass
        elif d == "":
            name = i.mac[:2] +":"+ i.mac[2:4] +":"+ i.mac[4:6] +":"+ i.mac[6:8] +":"+ i.mac[8:10] +":"+ i.mac[10:]
        elif d.hostname != "":
            name = d.hostname
        elif d.devtype != "":
            name = d.devtype +'_'+ d.mac
        else:
            name = d.mac[:2] +":"+ d.mac[2:4] +":"+ d.mac[4:6] +":"+ d.mac[6:8] +":"+ d.mac[8:10] +":"+ d.mac[10:]
        cus = model_to_dict(i)
        cus['name'] = name
        data.append(cus)
    return JsonResponse({'data':data},safe=False)

@login_required
def customer_black_list_add_ajax(request):
    online_time = int(time.time()) - 900

    p_list = [i.mac for i in Device.objects.filter(Q(support_mode = 1)|Q(support_mode = 3),Q(account_group_name = request.user.groupname))]

    if len(p_list) == 0:
        customer_list = []
    else:
        customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list))

    cus = customer_list
    data = []
    result = {}
    for d in cus:
        if Customer_black_list.objects.filter(Q(mac = d.mac)&Q(groupname = request.user.groupname)).exists():
            pass
        else:
            if Customer_name.objects.filter(mac = d.mac).exists():
                name = Customer_name.objects.get(mac = d.mac).name
            else:
                name = ""
            if name != "":
                pass
            elif d.hostname != "":
                name = d.hostname
            elif d.devtype != "":
                name = d.devtype +'_'+ d.mac
            else:
                name = d.mac[:2] +":"+ d.mac[2:4] +":"+ d.mac[4:6] +":"+ d.mac[6:8] +":"+ d.mac[8:10] +":"+ d.mac[10:]
            data.append({'name':name,'mac':d.mac})
    result = {'data':data}
    return JsonResponse(result,safe = False)

@login_required
def customer_blacklist_add(request):
    blacklist = json.loads(request.POST.get('bl'))
    input_mac = json.loads(request.POST.get('input_mac'))
    error = {'error':_(u'输入不合法'),'su':'false'}
    sign = 0
    print input_mac
    try:
        if blacklist != []:
            for i in blacklist:
                if Customer_black_list.objects.filter(Q(mac = i)&Q(groupname = request.user.groupname)).exists():
                    pass
                else:
                    bl = Customer_black_list()
                    bl.mac = i
                    bl.groupname = request.user.groupname
                    bl.save()
                    sign = 1
        for i in input_mac:
            if i.strip() == "":
                pass
            else:
                i = i.strip()
                i = re.sub(r'[:-]','',i)
                if len(i) != 12:
                    pass
                else:
                    if re.search(r'[^0-9a-fA-F]',i):
                        pass
                    else:
                        if Customer_black_list.objects.filter(Q(mac = i)&Q(groupname = request.user.groupname)).exists():
                            pass
                        else:
                            bl = Customer_black_list()
                            bl.mac = i
                            bl.groupname = request.user.groupname
                            bl.save()
                            sign = 1
        policy_str = user_policy_str(request.user.groupname)
        task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,[],'all')
        default_config_issued.delay()
        nonoperate_default_config_issued.delay()
    except Exception as e:
        print e
        sign = 2

    if sign == 1:
        error['error'] = _(u'执行成功')
        error['su'] = 'true'
    elif sign == 2:
        error['error'] = _(u'执行失败')

    return JsonResponse(error,safe=False)

@login_required
def customer_blacklist_remove(request):
    mac = request.POST.get('mac')
    error = {'error':''}
    try:
        Customer_black_list.objects.get(Q(mac = mac)&Q(groupname = request.user.groupname)).delete()
        policy_str = user_policy_str(request.user.groupname)
        task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,[],'all')
        default_config_issued.delay()
        nonoperate_default_config_issued.delay()
        error['error'] = _(u'操作成功')
    except Exception as e:
        print e
        error['error'] = _(u'操作失败')
    return JsonResponse(error,safe=False)

@login_required
def customer_white_list_table(request):
    cusb = Customer_white_list.objects.filter(groupname = request.user.groupname)
    data = []
    for i in cusb:
        if Customer_name.objects.filter(mac = i.mac).exists():
            name = Customer_name.objects.get(mac = i.mac).name
        else:
            name = ""
        if Customer.objects.filter(mac = i.mac).exists():
            d = Customer.objects.get(mac = i.mac)
        else:
            d = ""
        if name != "":
            pass
        elif d == "":
            name = i.mac[:2] +":"+ i.mac[2:4] +":"+ i.mac[4:6] +":"+ i.mac[6:8] +":"+ i.mac[8:10] +":"+ i.mac[10:]
        elif d.hostname != "":
            name = d.hostname
        elif d.devtype != "":
            name = d.devtype +'_'+ d.mac
        else:
            name = d.mac[:2] +":"+ d.mac[2:4] +":"+ d.mac[4:6] +":"+ d.mac[6:8] +":"+ d.mac[8:10] +":"+ d.mac[10:]
        cus = model_to_dict(i)
        cus['name'] = name
        data.append(cus)
    return JsonResponse({'data':data},safe=False)

@login_required
def customer_white_list_add_ajax(request):
    online_time = int(time.time()) - 900

    p_list = [i.mac for i in Device.objects.filter(Q(support_mode = 1)|Q(support_mode = 3),Q(account_group_name = request.user.groupname))]

    if len(p_list) == 0:
        customer_list = []
    else:
        customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list))

    cus = customer_list
    data = []
    result = {}
    for d in cus:
        if Customer_white_list.objects.filter(Q(mac = d.mac)&Q(groupname = request.user.groupname)).exists():
            pass
        else:
            if Customer_name.objects.filter(mac = d.mac).exists():
                name = Customer_name.objects.get(mac = d.mac).name
            else:
                name = ""
            if name != "":
                pass
            elif d.hostname != "":
                name = d.hostname
            elif d.devtype != "":
                name = d.devtype +'_'+ d.mac
            else:
                name = d.mac[:2] +":"+ d.mac[2:4] +":"+ d.mac[4:6] +":"+ d.mac[6:8] +":"+ d.mac[8:10] +":"+ d.mac[10:]
            data.append({'name':name,'mac':d.mac})
    result = {'data':data}
    return JsonResponse(result,safe = False)

@login_required
def customer_whitelist_add(request):
    whitelist = json.loads(request.POST.get('wl'))
    input_mac = json.loads(request.POST.get('input_mac'))
    error = {'error':_(u'输入不合法'),'su':'false'}
    sign = 0
    print input_mac
    try:
        if whitelist != []:
            for i in whitelist:
                if Customer_white_list.objects.filter(Q(mac = i)&Q(groupname = request.user.groupname)).exists():
                    pass
                else:
                    bl = Customer_white_list()
                    bl.mac = i
                    bl.groupname = request.user.groupname
                    bl.save()
                    sign = 1
        for i in input_mac:
            if i.strip() == "":
                pass
            else:
                i = i.strip()
                i = re.sub(r'[:-]','',i)
                if len(i) != 12:
                    pass
                else:
                    if re.search(r'[^0-9a-fA-F]',i):
                        pass
                    else:
                        if Customer_white_list.objects.filter(Q(mac = i)&Q(groupname = request.user.groupname)).exists():
                            pass
                        else:
                            bl = Customer_white_list()
                            bl.mac = i
                            bl.groupname = request.user.groupname
                            bl.save()
                            sign = 1
        policy_str = user_policy_str(request.user.groupname)
        task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,[],'all')
        default_config_issued.delay()
        nonoperate_default_config_issued.delay()
    except Exception as e:
        print e
        sign = 2

    if sign == 1:
        error['error'] = _(u'执行成功')
        error['su'] = 'true'
    elif sign == 2:
        error['error'] = _(u'执行失败')

    return JsonResponse(error,safe=False)

@login_required
def customer_whitelist_remove(request):
    mac = request.POST.get('mac')
    error = {'error':''}
    try:
        Customer_white_list.objects.get(Q(mac = mac)&Q(groupname = request.user.groupname)).delete()
        policy_str = user_policy_str(request.user.groupname)
        task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,[],'all')
        default_config_issued.delay()
        nonoperate_default_config_issued.delay()
        error['error'] = _(u'操作成功')
    except Exception as e:
        print e
        error['error'] = _(u'操作失败')
    return JsonResponse(error,safe=False)


@login_required
def customer_black_white_switch(request):
    if request.method == 'GET':
        mode = request.GET.get('mode')
        if mode == "get":
            switch = ""
            if Customer_black_white_switch.objects.filter(groupname = request.user.groupname).exists():
                switch = Customer_black_white_switch.objects.get(groupname = request.user.groupname).switch
            if switch == "":
                switch = "none"
            return JsonResponse({'switch':switch},safe = False)

    if request.method == "POST":
        mode = request.POST.get('mode')
        switch = request.POST.get('switch')
        error = {'error':"false",'msg':''}
        if mode == "post":
            try:
                if Customer_black_white_switch.objects.filter(groupname = request.user.groupname).exists():
                    customer_switch = Customer_black_white_switch.objects.get(groupname = request.user.groupname)
                else:
                    customer_switch = Customer_black_white_switch()
                    customer_switch.groupname = request.user.groupname
                print switch,type(switch)
                customer_switch.switch = switch
                customer_switch.save()

                policy_str = user_policy_str(request.user.groupname)
                task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,[],'all')
                default_config_issued.delay()
                nonoperate_default_config_issued.delay()
                error['error'] = 'true'
                error['msg'] = _(u'修改成功')
            except Exception as e:
                print e
                error['error'] = 'false'
                error['msg'] = _(u'修改失败')
            return JsonResponse(error,safe = False)






def ap_wlan_create_get_portal_config(request):
    mode = request.GET.get('mode')
    data = {'auth_server':''}
    if mode == 'open_form':
        groupname = Account_Group.objects.get(groupname = request.user.groupname)
        if Guest_policy.objects.filter(domainname_id = groupname.pk).exists():
            try:
                auth_server = json.loads(Guest_policy.objects.get(domainname_id = groupname.pk).auth_server)
            except Exception as e:
                print e
                auth_server = ''
            data['auth_server'] = auth_server
    return JsonResponse(data,safe = False)








@login_required
def ap_user_policy_config(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)

    wlan = ''
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        wlan = Group_wlan.objects.filter(group_id__group_type = '1')
    elif request.user.administrator_permission < 4 and request.user.group_status == 1:
        wlan = Group_wlan.objects.filter(group_id__account_group_name = request.user.groupname,group_id__group_type = '1')

    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'ap-user-policy-config.html',{'admin_error_json':admin_error_json,'admin_error':admin_error,'wlan':wlan,'error':error,"error_json":error_json,'plimit':plimit,'oemlimit':oemlimit})


@login_required
def ap_user_policy_config_ajax(request):
    ag = Account_Group.objects.get(groupname = request.user.groupname)
    ret = {}
    if User_policy_config.objects.filter(domainname = ag).exists():
        upc = User_policy_config.objects.get(domainname = ag)
        ret = model_to_dict(upc)
    else:
        ret = user_policy_config
    return JsonResponse(ret,safe = False)

@login_required
def policy_config_ajax(request):
    error = ''
    if request.method == 'POST':
        dom = Account_Group.objects.get(groupname = request.user.groupname)
        domainname = dom
        dual_max_user = int(request.POST.get('dual_max_user'))
        single_max_user = int(request.POST.get('single_max_user'))
        rssi_threshold = int(request.POST.get('rssi_threshold'))
        reject_max = int(request.POST.get('reject_max'))
        if request.POST.get('access_policy') == 'on':
            access_policy = 1
        else:
            access_policy = 0
        if request.POST.get('load_balance') == 'on':
            load_balance = 1
        else:
            load_balance = 0
        if request.POST.get('l2_isolation') == 'on':
            l2_isolation = 1
        else:
            l2_isolation = 0
        if request.POST.get('band_steering') == 'on':
            band_steering = 1
        else:
            band_steering = 0
        if request.POST.get('roaming_policy') == 'on':
            roaming_policy = 1
        else:
            roaming_policy = 0
        roaming_assoc_rssi = int(request.POST.get('roaming_assoc_rssi'))
        roaming_unassoc_rssi = int(request.POST.get('roaming_assoc_rssi'))

        if User_policy_config.objects.filter(domainname = domainname).exists():
            upc = User_policy_config.objects.get(domainname = domainname)
        else:
            upc = User_policy_config()
            upc.domainname = domainname
        try:
            upc.dual_max_user = dual_max_user
            upc.single_max_user = single_max_user
            upc.rssi_threshold = rssi_threshold
            upc.reject_max = reject_max
            upc.access_policy = access_policy
            upc.load_balance = load_balance
            upc.l2_isolation = l2_isolation
            upc.band_steering = band_steering
            upc.roaming_policy = roaming_policy
            upc.roaming_assoc_rssi = roaming_assoc_rssi
            upc.roaming_unassoc_rssi = roaming_unassoc_rssi
            upc.save()

            policy_str = user_policy_str(request.user.groupname)
            task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,[],'all')
            default_config_issued.delay()
            nonoperate_default_config_issued.delay()
            error = _(u'修改成功！')
        except Exception as e:
            print e
            error = _(u'修改失败！')

    return JsonResponse(error,safe = False)






def accessed_device2authpuppy(mac):
    try:
        db = MySQLdb.connect("localhost", settings_py.DATABASES['default']['USER'], settings_py.DATABASES['default']['PASSWORD'], "authpuppy" )
    except MySQLdb.Error, e:
        try:
            sqlError =  "Error %d:%s" % (e.args[0], e.args[1])
        except IndexError:
            print "MySQL Error:%s" % str(e)
        return False

    cursor = db.cursor()

    try:
        mac = (mac[:2] + ':' + mac[2:4] + ':' + mac[4:6] + ':' + mac[6:8] + ':' + mac[8:10] + ':' + mac[10:]).upper()
        sql = "INSERT into nodes(name,gw_id,created_at,updated_at,domainname) values ('','{}','0000-00-00 00:00:00','0000-00-00 00:00:00','admin')".format(mac)

    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.execute(sql)
    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print sql
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.close()
        db.commit()
    except Exception as e:
        print e
        cursor.close()
        db.rollback()

    db.close()


def del_device2authpuppy(mac):
    try:
        db = MySQLdb.connect("localhost", settings_py.DATABASES['default']['USER'], settings_py.DATABASES['default']['PASSWORD'], "authpuppy" )
    except MySQLdb.Error, e:
        try:
            sqlError =  "Error %d:%s" % (e.args[0], e.args[1])
        except IndexError:
            print "MySQL Error:%s" % str(e)
        return False

    cursor = db.cursor()

    try:
        mac = (mac[:2] + ':' + mac[2:4] + ':' + mac[4:6] + ':' + mac[6:8] + ':' + mac[8:10] + ':' + mac[10:]).upper()
        # sql = "INSERT into nodes(name,gw_id,created_at,updated_at,domainname) values ('','{}','0000-00-00 00:00:00','0000-00-00 00:00:00','admin')".format(mac)
        sql = "DELETE from nodes where gw_id = '{}'".format(mac)

    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.execute(sql)
    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print sql
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.close()
        db.commit()
    except Exception as e:
        print e
        cursor.close()
        db.rollback()

    db.close()

def update_device2authpuppy(mac,domainname):
    try:
        db = MySQLdb.connect("localhost", settings_py.DATABASES['default']['USER'], settings_py.DATABASES['default']['PASSWORD'], "authpuppy" )
    except MySQLdb.Error, e:
        try:
            sqlError =  "Error %d:%s" % (e.args[0], e.args[1])
        except IndexError:
            print "MySQL Error:%s" % str(e)
        return False

    cursor = db.cursor()

    try:
        mac = (mac[:2] + ':' + mac[2:4] + ':' + mac[4:6] + ':' + mac[6:8] + ':' + mac[8:10] + ':' + mac[10:]).upper()
        # sql = "INSERT into nodes(name,gw_id,created_at,updated_at,domainname) values ('','{}','0000-00-00 00:00:00','0000-00-00 00:00:00','admin')".format(mac)
        sql = "UPDATE nodes set domainname = '{1}' where gw_id = '{0}'".format(mac,domainname)

    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.execute(sql)
    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print sql
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.close()
        db.commit()
    except Exception as e:
        print e
        cursor.close()
        db.rollback()

    db.close()










@login_required
def ap_guest_policy(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)

    wlan = ''
    # if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
    #     wlan = Group_wlan.objects.all()
    # elif request.user.administrator_permission < 4 and request.user.group_status == 1:
    #     wlan = Group_wlan.objects.filter(group_id__account_group_name = request.user.groupname,group_id__group_type = '1')
    wlan = Group_wlan.objects.filter(group_id__account_group_name = request.user.groupname,group_id__group_type = '1')
    plimit = get_plimit()
    return render(request,'ap-guest-policy.html',{'admin_error_json':admin_error_json,'admin_error':admin_error,'wlan':wlan,'error':error,"error_json":error_json,'plimit':plimit})


@login_required
def auth_server_ajax(request):
    error = ''
    if request.method == "POST":
        login = request.POST.get('login')
        portal = request.POST.get('portal')
        try:
            ag = Account_Group.objects.get(groupname = request.user.groupname)
            a = login.split('://')
            if a[0] == "https" :
                sslenable = 'true'
                httporsslport = 443
            elif a[0] == "http":
                sslenable = 'false'
                httporsslport = 80
            else:
                sslenable = ''
                httporsslport = 80
            path = '/'
            if len(a) >= 2:
                hostname = a[1].split('/')[0]
            else:
                hostname = ''
            if hasattr(ag,'guest_policy'):
                gp = Guest_policy.objects.get(domainname = ag)
            else:
                gp = Guest_policy()
                gp.domainname = ag
            auth_server = {}
            auth_server['hostname'] = hostname
            auth_server['sslenable'] = sslenable
            auth_server['httporsslport'] = httporsslport
            auth_server['path'] = path
            auth_server['login'] = login
            auth_server['portal'] = portal
            gp.auth_server = json.dumps(auth_server)
            gp.save()
            error = _(u'修改成功！')
            gp_str = portal_str(request.user.groupname)
            task_for_global_config_ajax.delay(request.user.groupname,gp_str)
            default_config_issued.delay()
            nonoperate_default_config_issued.delay()
        except Exception as e:
            print e
            error = _(u'修改失败！')
    return JsonResponse(error,safe = False)

@login_required
def auth_server_show_ajax(request):
    ag = Account_Group.objects.get(groupname = request.user.groupname)
    if hasattr(ag,'guest_policy'):
        gp = Guest_policy.objects.get(domainname = ag)
        sh = model_to_dict(gp)
    else:
        sh = guest_policy_config
    print sh['white_list'],type(sh['white_list'])
    return JsonResponse(sh,safe = False)

@login_required
def global_config_ajax(request):
    error = ''
    if request.method == 'POST':
        bypass = request.POST.get('bypass')
        wechatallowed = request.POST.get('wechatallowed')
        auth_validate_timeout = request.POST.get('auth_validate_timeout')
        client_timeout = request.POST.get('client_timeout')
        white_list = request.POST.get('white_list')
        trusted_mac_list = request.POST.get('trusted_mac_list')
        auth_validate_timeout_0_num = request.POST.get('auth_validate_timeout_0_num')
        auth_validate_timeout_0_suffix = request.POST.get('auth_validate_timeout_0_suffix')
        client_timeout_0_num = request.POST.get('client_timeout_0_num')
        client_timeout_0_suffix = request.POST.get('client_timeout_0_suffix')

        if bypass == None:
            bypass = "false"
        elif bypass == 'on':
            bypass = "true"
        if wechatallowed == None:
            wechatallowed = "false"
        elif wechatallowed == 'on':
            wechatallowed = "true"
        if auth_validate_timeout == '0':
            auth_validate_timeout = str(int(auth_validate_timeout_0_num) * int(auth_validate_timeout_0_suffix) * 60)
        else:
            auth_validate_timeout = str(int(auth_validate_timeout) * 60)
        if client_timeout == '0':
            client_timeout = str(int(client_timeout_0_num) * int(client_timeout_0_suffix))
        else:
            client_timeout = client_timeout

        try:
            ag = Account_Group.objects.get(groupname = request.user.groupname)
            if hasattr(ag,'guest_policy'):
                gp = Guest_policy.objects.get(domainname = ag)
                ju_gp1 = model_to_dict(gp)
            else:
                gp = Guest_policy()
                gp.domainname = ag
                ju_gp1 = ""
            gp.bypass = bypass
            gp.wechatallowed = wechatallowed
            gp.auth_validate_timeout = auth_validate_timeout
            gp.client_timeout = client_timeout
            gp.white_list = white_list
            gp.trusted_mac_list = trusted_mac_list
            gp.save()
            ju_gp2 = model_to_dict(gp)

            gp_str = portal_str(request.user.groupname)
            if ju_gp1 != "":
                if judge_global_portal_conf(ju_gp1,ju_gp2):
                    pass
                else:
                    task_for_global_config_ajax.delay(request.user.groupname,gp_str)
            else:
                task_for_global_config_ajax.delay(request.user.groupname,gp_str)
            default_config_issued.delay()
            nonoperate_default_config_issued.delay()
            error = _(u'修改成功！')
        except Exception as e:
            print e
            error = _(u'修改失败！')

    return JsonResponse(error,safe = False)

def judge_global_portal_conf(gp1,gp2):
    try:
        del gp1['id'],gp2['id'],gp1['domainname'],gp2['domainname'],gp1['auth_server'],gp2['auth_server']
        if gp1 == gp2:
            return True
        else:
            return False
    except Exception as e:
        print e
        return False


def portal_str(groupname):
    if Account_Group.objects.filter(groupname = groupname).exists():
        ag = Account_Group.objects.get(groupname = groupname)
        if hasattr(ag,'guest_policy'):
            gp = model_to_dict(Guest_policy.objects.get(domainname = ag))
        else:
            gp = guest_policy_config
    else:
        gp = guest_policy_config

    guest_policy = gp


    white_list = json.loads(guest_policy['white_list'])
    trusted_mac_list = json.loads(guest_policy['trusted_mac_list'])
    if guest_policy['wechatallowed'] == 'true' :
        white_list = list(set(white_list)) + ['long.weixin.qq.com','short.weixin.qq.com','szlong.weixin.qq.com','szshort.weixin.qq.com','mp.weixin.qq.com','res.wx.qq.com']
    # 保序去重
    # new_list = list(set(white_list))
    # new_list.sort(key = white_list.index)
    gp_str = {"bypass": guest_policy['bypass'], "validatetime": int(guest_policy['auth_validate_timeout']), "clienttimeout": int(guest_policy['client_timeout']), "whitelist": list(set(white_list)), "trustedlist": list(set(trusted_mac_list)), "wireless":[],"personaltype":"normal"}
    gp_str = json.dumps(gp_str,sort_keys = True).replace(' ','')
    print gp_str
    return gp_str


def user_policy_str(groupname):
    if Account_Group.objects.filter(groupname = groupname).exists():
        ag = Account_Group.objects.get(groupname = groupname)
        if hasattr(ag,'user_policy_config'):
            up = model_to_dict(User_policy_config.objects.get(domainname = ag))
        else:
            up = user_policy_config
    else:
        up = user_policy_config

    user_policy = up

    rssi_max = str(int(float(user_policy['rssi_threshold'])) + 95)
    policy = collections.OrderedDict([("access_policy",str(user_policy['access_policy'])),("band_steer",str(user_policy['band_steering'])),("dual_max_user",str(user_policy['dual_max_user'])),("l2_isolation",str(user_policy['l2_isolation'])),("load_balance",str(user_policy['load_balance'])),("reject_max",str(user_policy['reject_max'])),("roaming_assoc_rssi",str(user_policy['roaming_assoc_rssi'])),("roaming_detect",str(user_policy['roaming_policy'])),("roaming_unassoc_rssi",str(user_policy['roaming_unassoc_rssi'])),("rssi_max",rssi_max),("single_max_user",str(user_policy['single_max_user'])),("thredhold_5g",'0'),("thredhold_5g_rssi",'0')])
    mac_list = []
    if Customer_black_white_switch.objects.filter(groupname = groupname).exists():
        wb_type = Customer_black_white_switch.objects.get(groupname = groupname).switch
    else:
        wb_type = "none"
    if wb_type == "":
        wb_type = "none"
        mac_list = []
    elif wb_type == "black":
        wb_type = "blacklist"
        for i in Customer_black_list.objects.filter(groupname = groupname):
            i.mac = (i.mac[:2] + ':' + i.mac[2:4] + ':' + i.mac[4:6] + ':' + i.mac[6:8] + ':' + i.mac[8:10] + ':' + i.mac[10:]).lower()
            mac_list.append(i.mac)
    elif wb_type == "white":
        wb_type = "whitelist"
        for i in Customer_white_list.objects.filter(groupname = groupname):
            i.mac = (i.mac[:2] + ':' + i.mac[2:4] + ':' + i.mac[4:6] + ':' + i.mac[6:8] + ':' + i.mac[8:10] + ':' + i.mac[10:]).lower()
            mac_list.append(i.mac)
    wb_list = collections.OrderedDict([('maclist',mac_list),('type',wb_type)])

    if Timing_Policy.objects.filter(groupname = groupname).exists():
        time_pilocy = model_to_dict(Timing_Policy.objects.get(groupname = groupname))
    else:
        time_pilocy = {'gpon_light_switch':'off'}
    print 'xxxxx'
    if time_pilocy['gpon_light_switch'] == 'on':
        switch = 'on'
        timer = []
        timer_origin = json.loads(time_pilocy['gpon_timer'])
        print timer_origin,type(timer_origin)
        for i in timer_origin:
            timer.append(collections.OrderedDict([('end',i['end']),('start',i['start']),('weekday',','.join(i['weekday']))]))
    elif time_pilocy['gpon_light_switch'] == 'off':
        switch = 'off'
        timer = ''
    # timer = json.dumps(timer)
    ledtimer = collections.OrderedDict([('switch',switch),('timer',timer)])
    # policy_str = json.dumps({'policy':policy,'list':wb_list,'ledtimer':ledtimer},sort_keys = True).replace(' ','')
    policy_str = json.dumps(collections.OrderedDict([('ledtimer',ledtimer),('list',wb_list),('policy',policy)])).replace(' ','')
    print policy_str,321321
    return policy_str


def timing_policy_ajax(request):
    """Use ajax get and post the Timing_Policy info.

    if GET that use ajax get the time_pilocy info for the front page to show them,
    if POST that use ajax post the info into the Timing_Policy table to save them.

    Arguments:
        request {[request]} -- A request object for all of the info from front.
    Returns:
        Get: return the Timing_Policy info
        Post: return the result for create/update Timing_Policy
    """
    if request.method == 'GET':
        groupname = request.user.groupname
        if Timing_Policy.objects.filter(groupname = groupname).exists():
            timingpolicy = model_to_dict(Timing_Policy.objects.get(groupname = groupname))
        else:
            timingpolicy = {}
        return JsonResponse(timingpolicy,safe = False)
    elif request.method == 'POST':
        message = {}
        groupname = request.user.groupname
        light_switch = request.POST.get('light_switch')
        policy_weekday_1 = request.POST.getlist('policy_weekday_1[]')
        policy_time_start_1 = request.POST.get('policy_time_start_1')
        policy_time_end_1 = request.POST.get('policy_time_end_1')
        if Timing_Policy.objects.filter(groupname = groupname).exists():
            timingpolicy = Timing_Policy.objects.get(groupname = groupname)
        else:
            timingpolicy = Timing_Policy()
            timingpolicy.groupname = groupname


        if light_switch == 'on' and policy_weekday_1 and policy_time_start_1 and policy_time_end_1 :
            timingpolicy.gpon_light_switch = light_switch
            timer = json.dumps([
                        {
                            'weekday': policy_weekday_1,
                            'start': policy_time_start_1,
                            'end': policy_time_end_1
                        },
                    ])
        else:
            timingpolicy.gpon_light_switch = 'off'
            timer = ''
        timingpolicy.gpon_timer = timer
        try:
            timingpolicy.save()
            policy_str = user_policy_str(request.user.groupname)
            task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,[],'all')
            message = {'result':'success','mes':_(u'修改成功')}
        except Exception as e:
            print e
            message = {'result':'failed','mes':_(u'修改失败')}
            raise e
        return JsonResponse(message,safe = False)










DEFAULT_DOMAIN_NAME = "admin"
@login_required
def auth_proxy(request):
    AUTH_ROOT = ""
    AUTH_ADMIN = ""
    global DEFAULT_DOMAIN_NAME
    authAdminAccount = "admin"
    authAdminPasswd = "citrus"
    # $AUTH_ROOT = "http://".$_SERVER["SERVER_NAME"].":9090/"
    AUTH_ROOT = "http://localhost:9090/"
    # AUTH_ROOT = "http://192.168.1.254:9090/"
    AUTH_ADMIN = "signin%5Busername%5D=" + authAdminAccount + "&signin%5Bpassword%5D=" + authAdminPasswd + "&signin%5B_csrf_token%5D=CSRFSecret"
    # if request.POST.get("json") != None:
    #     json_x = json.loads(request.POST.get("json"))
    # else:
    #     json_x = ""
    json_x = request.POST
    if json_x != "" and 'list' in json_x:
        voucher = json_x['list']
    else:
        voucher = ''
    if json_x != "" and 'cmd' in json_x:
        cmd = json_x['cmd']
    else:
        cmd = ''
    if json_x != "" and 'code' in json_x:
        code = json_x['code']
    else:
        code = ''

    # json = isset($_REQUEST["json"]) ? json_decode($_REQUEST["json"]) : ""
    # voucher = (is_object($json) && isset($json->list)) ?$json->list:""
    # cmd = (is_object($json) && isset($json->cmd)) ?$json->cmd:""
    # code = (is_object($json) && isset($json->code)) ?$json->code:""
    action = ""
    data = ""
    data2 = ""

    if request.POST.get("action") != None:
        action = request.POST.get("action")
    else:
        action = ""


    if action != "" or voucher != "" or cmd != "":

        if request.POST.get("data") !=None:
            data = request.POST.get("data")
        else:
            data = ""

        if action == "authAddUser":
            data = request.POST.get('username')
            data2 = request.POST.get('password')
            data3 = request.POST.get('expiretime') or "0"
            if data3 != "0":
                data3 = str(int(time.time()) + int(data3) * 86400)
        if action == "platformConfig":
            if request.POST.get("platform") != None:
                platform = request.POST.get("platform")
            else:
                platform = ""
        if cmd == "voucherDelete":
            action = cmd
            code = code.replace('-','')
        if voucher == "voucherList":
            action = voucher
        if action == "authAddVoucherConfig":
            if request.POST.get("essid") !=None:
                essid = request.POST.get("essid")
            else:
                essid = ""
            if request.POST.get("code") !=None:
                code = request.POST.get("code")
            else:
                code = ""
            if request.POST.get("num") !=None:
                num = request.POST.get("num")
            else:
                num = ""
            if request.POST.get("duration") !=None:
                duration = request.POST.get("duration")
            else:
                duration = ""
            if request.POST.get("byte_quota") !=None:
                byte_quota = request.POST.get("byte_quota")
            else:
                byte_quota = ""
            if request.POST.get("user_support") !=None:
                user_support = request.POST.get("user_support")
            else:
                user_support = ""
            if request.POST.get("remarks") !=None:
                remarks = request.POST.get("remarks")
            else:
                remarks = ""

        if action == "authAddSmsConfig":
            if request.POST.get("appkey") !=None:
                appkey = request.POST.get("appkey")
            else:
                appkey = ""
            if request.POST.get("secret") !=None:
                secret = request.POST.get("secret")
            else:
                secret = ""
            if request.POST.get("signName") !=None:

                signName = request.POST.get("signName")

            else:
                signName = ""
            if request.POST.get("templateNamecn") !=None:
                templateNamecn = request.POST.get("templateNamecn")
            else:
                templateNamecn = ""
            if request.POST.get("templateNameen") !=None:
                templateNameen = request.POST.get("templateNameen")
            else:
                templateNameen = ""

        if action == "authAddSmsIhuyi":
            if request.POST.get("account") !=None:
                account = request.POST.get("account")
            else:
                account = ""
            if request.POST.get("password") !=None:
                password = request.POST.get("password")
            else:
                password = ""
            if request.POST.get("sms_cn") !=None:
                sms_cn = request.POST.get("sms_cn")
            else:
                sms_cn = ""
            if request.POST.get("sms_en") !=None:
                sms_en = request.POST.get("sms_en")
            else:
                sms_en = ""

        if action == "authAddSmsConfigAliyun":
            accessId = request.POST.get("accessId") if request.POST.get("accessId") != None else ""
            accessKey = request.POST.get("accessKey") if request.POST.get("accessKey") != None else ""
            signName = request.POST.get("signName") if request.POST.get("signName") != None else ""
            templateCodecn = request.POST.get("templateCodecn") if request.POST.get("templateCodecn") != None else ""
            templateCodeen = request.POST.get("templateCodeen") if request.POST.get("templateCodeen") != None else ""
            endPoint = request.POST.get("endPoint") if request.POST.get("endPoint") != None else ""
            topicName = request.POST.get("topicName") if request.POST.get("topicName") != None else ""
            topicName = request.POST.get("topicName") if request.POST.get("topicName") != None else ""
            paramKey = request.POST.get("paramKey") if request.POST.get("paramKey") != None else ""

        if action == "authAddSmsYunpian":
            if request.POST.get("apikey") !=None:
                apikey = request.POST.get("apikey")
            else:
                apikey = ""
            if request.POST.get("text") !=None:
                text = request.POST.get("text")
            else:
                text = ""

        if action == "saveRadiusSetting":
            radiusip = request.POST.get('radiusip') or ''
            radiusport = request.POST.get('radiusport') or ''
            radiusaccount = request.POST.get('radiusaccount') or ''
            radiuspasswd = request.POST.get('radiuspasswd') or ''
    if request.user.groupname !=None:
        domainname = request.user.groupname
    else:
        domainname = ""
    return_result = ""

    if action == 'platformConfig':
        platformConfig(domainname,AUTH_ADMIN, AUTH_ROOT,platform)
    if action == 'voucherDelete':
        return_result = voucherDelete(domainname,AUTH_ADMIN, AUTH_ROOT,code)
    if action == 'voucherList':
        return_result = voucherList(domainname,AUTH_ADMIN, AUTH_ROOT)
    if action == 'authAddVoucherConfig':
        return_result = authAddVoucherConfig(domainname,AUTH_ADMIN, AUTH_ROOT,essid ,code,num,duration,byte_quota,user_support,remarks)

    if action == 'authAddSmsConfig':

        return_result = authAddSmsConfig(domainname,AUTH_ADMIN, AUTH_ROOT, appkey, secret,signName,templateNamecn,templateNameen)

    if action == 'authAddSmsIhuyi':
        return_result = authAddSmsIhuyi(domainname,AUTH_ADMIN, AUTH_ROOT, account,password,sms_cn,sms_en)

    if action == 'authAddSmsConfigAliyun':
        return_result = authAddSmsConfigAliyun(domainname,AUTH_ADMIN, AUTH_ROOT, accessId,accessKey,signName,templateCodecn,templateCodeen,endPoint,topicName,paramKey)

    if action == 'authAddSmsYunpian':
        return_result = authAddSmsYunpian(domainname,AUTH_ADMIN, AUTH_ROOT, apikey,text)

    if action == 'authSwitchOffAllPlugins':
        authSwitchOffAllPlugins(AUTH_ADMIN, AUTH_ROOT)
    if action == 'authSwitchApAuthLocalUserPlugin':
        authSwitchApAuthLocalUserPlugin(AUTH_ADMIN, AUTH_ROOT, data)
    if action == 'authSwitchApAuthLocalUserPlugin_setAppid':
        print "12345qwert"
        authSwitchApAuthLocalUserPlugin_setWeChatArg(domainname, AUTH_ADMIN, AUTH_ROOT, "appid", data)
        print 9876543
    if action == 'authSwitchApAuthLocalUserPlugin_setAppSecret':
        authSwitchApAuthLocalUserPlugin_setWeChatArg(domainname, AUTH_ADMIN, AUTH_ROOT, "appsecret", data)
    if action == 'authSwitchApAuthLocalUserPlugin_setShopid':
        authSwitchApAuthLocalUserPlugin_setWeChatArg(domainname, AUTH_ADMIN, AUTH_ROOT, "shopid", data)
    if action == 'authSwitchApAuthLocalUserPlugin_setSecretKey':
        authSwitchApAuthLocalUserPlugin_setWeChatArg(domainname, AUTH_ADMIN, AUTH_ROOT, "secretkey", data)
    if action == 'authSwitchApAuthLocalUserPlugin_setForceFollow':
        authSwitchApAuthLocalUserPlugin_setWeChatArg(domainname, AUTH_ADMIN, AUTH_ROOT, "forcefollow", data)
    if action == 'authSwitchApAuthSplashOnlyPlugin':
        authSwitchApAuthSplashOnlyPlugin(AUTH_ADMIN, AUTH_ROOT)
    if action == 'authSetAuthEndTimeminute':
        authSetAuthEndTimeminute(AUTH_ADMIN, AUTH_ROOT, data)

    if action == 'authClosePortalPage':
        authClosePortalPage(domainname, AUTH_ADMIN, AUTH_ROOT, data)
    if action == 'authSetPortalPageWithCheck':
        authSetPortalPageWithCheck(domainname, AUTH_ADMIN, AUTH_ROOT, data)
    if action == 'authSetPortalPage':
        authSetPortalPage(domainname, AUTH_ADMIN, AUTH_ROOT, data)

    if action == 'authShowUsers':
        return_result = authShowUsers(domainname, AUTH_ADMIN, AUTH_ROOT)
    # if action == 'authSearchUser':
        # authSearchUser(domainname, AUTH_ADMIN, AUTH_ROOT, data)
    if action == 'authAddUser':
        return_result = authAddUser(domainname, AUTH_ADMIN, AUTH_ROOT, data, data2,data3)
    if action == 'authDeleteUser':
        return_result = authDeleteUser(domainname, AUTH_ADMIN, AUTH_ROOT, data)
    if action == 'authGetUserId':
        authGetUserId(domainname, AUTH_ADMIN, AUTH_ROOT, data)
    if action == 'authGetconfig':
        return_result = authGetconfig(domainname, AUTH_ADMIN, AUTH_ROOT)
    # if action == 'authGetGlobalConfig':
        # authGetGlobalConfig(domainname)
    # if action == 'authSetGlobalConfig':
        # authSetGlobalConfig(domainname, data, data2)
    if action == 'authStartAllPlugins':
        authStartAllNeededPlugins(AUTH_ADMIN, AUTH_ROOT)

    if action == 'saveRadiusSetting':
        saveRadiusSetting(domainname, AUTH_ADMIN, AUTH_ROOT,radiusip,radiusport,radiusaccount,radiuspasswd)

    return JsonResponse(return_result,safe = False)

def authSwitchOffAllPlugins(authInfo, authRoot):
    action = authRoot + "plugin/manage/save"
    postData = ""
    result = httpPost(authInfo, action, postData)

def authStartAllNeededPlugins(authInfo, authRoot):
    print 'asdfgh'
    action = authRoot + "plugin/manage/save"
    print 'poiuyt[]'
    postData = "apPluginManagerList%5BapExternalCMSPlugin%5D=on&apPluginManagerList%5BapAuthSplashOnlyPlugin%5D=on&apPluginManagerList%5BapAuthLocalUserPlugin%5D=on"
    print ',mnbvc'
    httpPost(authInfo, action, postData)
    print 'zxcvbn'
    authSetAuthEndTimeminute(authInfo, authRoot, 30)

def authSwitchApAuthLocalUserPlugin(authInfo, authRoot, auth_type):
    authStartAllNeededPlugins(authInfo, authRoot)
    action = authRoot + "authlocaluser/authway"
    if auth_type != "phone_auth" and auth_type != "user_auth" and auth_type != "wechat_auth" and auth_type != "wechat_phone_auth":
        auth_type = 'none_auth'
    postData = "apAuthLocalUserPlugin_authentication_way=" + auth_type
    httpPost(authInfo, action, postData)

def authSwitchApAuthLocalUserPlugin_setWeChatArg(domainname, authInfo,authRoot, argName, argValue):
    print '12345poiuytre'
    authStartAllNeededPlugins(authInfo, authRoot)
    print 'qiwgqu'
    action = authRoot + "authlocaluser/setwechatarg"

    postData = argName + "=" + argValue + "&domainname=" + domainname
    httpPost(authInfo, action, postData)
def platformConfig(domainname,authInfo, authRoot,platform):
    action = authRoot + "authlocaluser/setsmsarg"
    postData = urllib.urlencode({'platform':platform,'domainname':domainname})
    # postData = "platform=" + platform + "&domainname=" + domainname
    httpPost(authInfo, action, postData)

def voucherDelete(domainname,authInfo, authRoot,code):
    action = authRoot + "authlocaluser/deletevoucherarg"
    postData = "domainname=" + domainname + "&code=" + code
    httpPost(authInfo, action, postData)
    return "1"

def voucherList(domainname,authInfo, authRoot):
    action = authRoot + "authlocaluser/getvoucherarg"
    postData = "domainname=" + domainname
    httpPost(authInfo, action, postData)
    result = json.dumps(json.loads(httpPost(authInfo, action, postData))["data"])
    return result

def authAddVoucherConfig(domainname,authInfo, authRoot,essid,code ,num,duration,byte_quota,user_support,remarks):
    if essid != "" and num != "" and duration != "" and byte_quota != "" and user_support != "":
        if int(duration) != 8:
            duration = str(int(duration) * 24)
        byte_quota = str(int(byte_quota) * 1024 * 1024)
        i = 0
        for i in range(0,int(num)):
            action = authRoot + "authlocaluser/setvoucherarg"
            postData = urllib.urlencode({'essid':essid,'duration':duration,'byte_quota':byte_quota,'user_support':user_support,'remarks':remarks,'domainname':domainname,'code':code})
            # postData = "essid=" + essid + "&duration=" + duration + "&byte_quota=" + byte_quota + "&user_support=" + user_support + "&remarks=" + remarks + "&domainname=" + domainname + "&code=" + code
            httpPost(authInfo, action, postData)
        return "1"
    else:
        return "0"

def authAddSmsConfig(domainname,authInfo, authRoot, appkey, secret,signName,templateNamecn,templateNameen):

    if appkey != "" and secret != "" and signName != "" and (templateNamecn != "" or templateNameen != "" ):

        action = authRoot + "authlocaluser/setsmsarg"

        postData = urllib.urlencode({"alidayu_appkey":appkey,"alidayu_secret":secret,"alidayu_signName":signName,"alidayu_templateName_cn":templateNamecn,"alidayu_templateName_en":templateNameen,"domainname":domainname})
        # postData = "alidayu_appkey=" + appkey + "&alidayu_secret=" + secret + "&alidayu_signName=" + signName + "&alidayu_templateName_cn=" + templateNamecn + "&alidayu_templateName_en=" + templateNameen + "&domainname=" + domainname
        httpPost(authInfo, action, postData)

        platformConfig(domainname, authInfo, authRoot, "alidayu")

        return "1"
    else:

        return "0"

def authAddSmsIhuyi(domainname,authInfo, authRoot, account,password,sms_cn,sms_en):
    if account != "" and password != "" and (sms_en != "" or sms_cn != "" ):
        action = authRoot + "authlocaluser/setsmsarg"
        postData = urllib.urlencode({'smsaccount':account,'smspassword':password,'smstemplates_cn':sms_cn,'smstemplates_en':sms_en,'domainname':domainname})
        # postData = "smsaccount=" + account + "&smspassword=" + password + "&smstemplates_cn=" + sms_cn + "&smstemplates_en=" + sms_en + "&domainname=" + domainname
        httpPost(authInfo, action, postData)
        platformConfig(domainname, authInfo, authRoot, "ihuyi")
        return "1"
    else:
        return "0"

def authAddSmsConfigAliyun(domainname,authInfo, authRoot,
    accessId, accessKey,signName,templateCodecn,templateCodeen, endPoint, topicName, paramKey):
    if accessId !="" and accessKey !="" and signName !="" and endPoint != "" and topicName != "" and (templateCodecn !="" or templateCodeen !=""):
        action = authRoot + "authlocaluser/setsmsarg"
        postData = urllib.urlencode({'aliyun_accessId':accessId,'aliyun_accessKey':accessKey,'aliyun_signName':signName,'aliyun_templateCode_cn':templateCodecn,'aliyun_templateCode_en':templateCodeen,'aliyun_endPoint':endPoint,'aliyun_topicName':topicName,'aliyun_paramKey':paramKey,'domainname':domainname})
        # postData = "aliyun_accessId=" + accessId + "&aliyun_accessKey=" + accessKey + "&aliyun_signName=" + signName + "&aliyun_templateCode_cn=" + templateCodecn + "&aliyun_templateCode_en=" + templateCodeen + "&aliyun_endPoint=" + endPoint + "&aliyun_topicName=" + topicName + "&aliyun_paramKey=" + paramKey + "&domainname=" + domainname
        httpPost(authInfo, action, postData)
        platformConfig(domainname, authInfo, authRoot, "aliyun")
        return "1"
    else:
        return "0"

def authAddSmsYunpian(domainname,authInfo, authRoot, apikey,text):
    if apikey != "" and text != "" :
        action = authRoot + "authlocaluser/setsmsarg"
        postData = urllib.urlencode({'yunpian_apikey':apikey,'yunpian_text':text,'domainname':domainname})
        # postData = "smsaccount=" + account + "&smspassword=" + password + "&smstemplates_cn=" + sms_cn + "&smstemplates_en=" + sms_en + "&domainname=" + domainname
        httpPost(authInfo, action, postData)
        platformConfig(domainname, authInfo, authRoot, "yunpian")
        return "1"
    else:
        return "0"

def authSwitchApAuthSplashOnlyPlugin(authInfo, authRoot):
    action = authRoot + "plugin/manage/save"
    postData = "apPluginManagerList%5BapExternalCMSPlugin%5D=on&apPluginManagerList%5BapAuthSplashOnlyPlugin%5D=on"
    httpPost(authInfo, action, postData)

def authSetAuthEndTimeminute(authInfo, authRoot, expiryTime):
    print 'lkjhgfd'
    action = authRoot + "configure"
    print 'lkjhgf1223456'
    postData = urllib.urlencode({"submit[Save]":"Save","apconfigure[connection_expiry]":expiryTime})
    # postData = "submit%5BSave%5D=Save&apconfigure%5Bconnection_expiry%5D=" + expiryTime
    print '=-09876'
    # //."&apconfigure%5Bsite_name%5D=AuthSite"
    httpPost(authInfo, action, postData)

def authClosePortalPage(domainname, authInfo, authRoot, portalUrl):

    action = authRoot + "plugin/configure/apExternalCMSPlugin"
    prefix = ""
    global DEFAULT_DOMAIN_NAME
    if domainname != "" and domainname != DEFAULT_DOMAIN_NAME:
        prefix = domainname + ":"
    postData = "apExternalCMS%5B" + prefix + "portal_page%5D=" + portalUrl
    httpPost(authInfo, action, postData)

def authSetPortalPageWithCheck(domainname, authInfo, authRoot, portalUrl):

    action = authRoot + "plugin/configure/apExternalCMSPlugin"
    prefix = ""
    global DEFAULT_DOMAIN_NAME
    if domainname != "" and domainname != DEFAULT_DOMAIN_NAME:
        prefix = domainname + ":"
    postData = "apExternalCMS%5B" + prefix + "redirect_portal_page%5D=1&apExternalCMS%5B" + prefix + "portal_page%5D=" + portalUrl
    httpPost(authInfo, action, postData)

def authSetPortalPage(domainname, authInfo, authRoot, portalUrl):

    action = authRoot + "plugin/configure/apExternalCMSPlugin"
    prefix = ""
    global DEFAULT_DOMAIN_NAME
    if domainname != "" and domainname != DEFAULT_DOMAIN_NAME:
        prefix = domainname + ":"
    postData = "apExternalCMS%5B" + prefix + "portal_page%5D=" + portalUrl
    httpPost(authInfo, action, postData)

def authShowUsers(domainname, authInfo, authRoot):
    action = authRoot + "authlocaluser"
    postData = "domainname=" + domainname + "&ap_user_filters%5Bemail%5D%5Btext%5D=@"
    return httpPost(authInfo, action, postData)

# def authSearchUser(domainname, authInfo, authRoot):
#     action = authRoot + "authlocaluser/search"
#     postData = "domainname=" + domainname
#     print httpPost(authInfo, action, postData)

def saveRadiusSetting(domainname, authInfo, authRoot,radiusip,radiusport,radiusaccount,radiuspasswd):
    if radiusip != '' and radiusport != '' and radiusaccount != '' and radiuspasswd != '':
        action = authRoot + "authradiussetting/save"
        postData = urllib.urlencode({'radiusip':radiusip,'radiusport':radiusport,'radiusaccount':radiusaccount,'radiuspasswd':radiuspasswd,'domainname':domainname})
        httpPost(authInfo, action, postData)
        return '1'
    else:
        return '0'


def authAddUser(domainname, authInfo, authRoot, username, passwd,expiretime):
    userId = authGetUserId(domainname, authInfo, authRoot,username)
    newusername = ""
    if userId != "":
        return "0"
    action = authRoot + "authlocaluser/create"
    email = str(int(time.time())) + "@witrusty.com"
    status = '1'
    global DEFAULT_DOMAIN_NAME
    if domainname == DEFAULT_DOMAIN_NAME:
        newusername = "" + username
    else:
        newusername = domainname + ":" + username
    postData = "ap_user%5B_csrf_token%5D=CSRFSecret" + "&ap_user%5Bemail%5D=" + email + "&ap_user%5Bpassword%5D=" + passwd + "&ap_user%5Busername%5D=" + newusername + "&ap_user%5Bstatus%5D=" + status + "&ap_user%5Bexpiretime%5D=" + str(expiretime)
    # postData = urllib.urlencode({"ap_user[_csrf_token]":"CSRFSecret","ap_user[email]":email,"ap_user[password]":passwd,"ap_user[username]":newusername,"ap_user[status]":status})
    httpPost(authInfo, action, postData)
    userId = authGetUserId(domainname, authInfo, authRoot, username)
    if userId != "":
        return "1"
    return "2"

def authDeleteUser(domainname, authInfo, authRoot, username):
    userId = authGetUserId(domainname, authInfo, authRoot, username)
    if userId == "":
        return "0"
    action = authRoot + "authlocaluser/delete/" + userId
    postData = "sf_method=delete"
    res = httpPost(authInfo, action, postData)
    userId = authGetUserId(domainname, authInfo, authRoot, username)
    if userId == "":
        return "1"
    return "2"

def authGetUserId(domainname, authInfo, authRoot, username):
    action = authRoot + "authlocaluser/getid"
    global DEFAULT_DOMAIN_NAME
    if domainname == DEFAULT_DOMAIN_NAME:
        postData = "username=" + "" + username
    else:
        postData = "username=" + domainname + ":" + username
    res = httpPost(authInfo, action, postData)
    return res

def authGetconfig(domainname, authInfo, authRoot):
    action = authRoot + "node/getconfig"
    postData = "domainname=" + domainname
    return httpPost(authInfo, action, postData)

def httpPost(authInfo, url, postData=""):
    # PostUrl = "http://192.168.1.253:9090/"
    cookiefile = "/var/lib/php5/authpuppyCookie.txt"
    if not os.path.exists(cookiefile):
        os.system("touch {}".format(cookiefile))
        cookie = cookielib.MozillaCookieJar(cookiefile)
        # cookie.save(cookiefile)
    else:
        cookie = cookielib.MozillaCookieJar(cookiefile)
        cookie.load(ignore_discard = True,ignore_expires = True)
    # cookie = cookielib.CookieJar()

    handler = urllib2.HTTPCookieProcessor(cookie)
    opener = urllib2.build_opener(handler)
    request = urllib2.Request(url, authInfo,)

    try:
        response = opener.open(request)
        cookie.save(ignore_discard = True,ignore_expires = True)
        output = response.read()

    except Exception as e:
        print 'authpuppy httpPost2',e
    if postData != "":

        request = urllib2.Request(url, postData,)
        try:
            response = opener.open(request)
            output = response.read()
        except Exception as e:
            print 'authpuppy httpPost3',e

    return output.strip()


def voucherprint(request):
    AUTH_ROOT = ""
    AUTH_ADMIN = ""
    global DEFAULT_DOMAIN_NAME
    authAdminAccount = "admin"
    authAdminPasswd = "citrus"
    # $AUTH_ROOT = "http://".$_SERVER["SERVER_NAME"].":9090/"
    AUTH_ROOT = "http://localhost:9090/"
    # AUTH_ROOT = "http://192.168.1.254:9090/"
    AUTH_ADMIN = "signin%5Busername%5D=" + authAdminAccount + "&signin%5Bpassword%5D=" + authAdminPasswd + "&signin%5B_csrf_token%5D=CSRFSecret"
    action = ""
    data = ""

    if request.user.groupname !=None:
        domainname = request.user.groupname
    else:
        domainname = ""
    # $domainname = $_SESSION['groupname']
    # $htmlclass = languageIsChinese() ? "" : ' class="en_US" '
    code = ""
    if request.GET.get('code') != None:
        code = request.GET.get('code')
    row = ""
    if code == "":
        row = voucherPrintMany(domainname,AUTH_ADMIN, AUTH_ROOT)
    else:
        data = json.loads(voucherPrintSingle(domainname,AUTH_ADMIN, AUTH_ROOT,code))
    print row,1111,data
    return render(request,"voucherprint.html",{'data':data,'row':row})


def voucherPrintMany(domainname,authInfo, authRoot):
    action = authRoot + "authlocaluser/getvouchermangarg"
    postData = "domainname=" + domainname
    httpPost(authInfo, action, postData)
    row = json.dumps(json.loads(httpPost(authInfo, action, postData))["data"])
    # result = httpPost(authInfo, action, postData)
    # result = json_decode(result)
    # $row = $result->data
    return row

def voucherPrintSingle(domainname,authInfo, authRoot,code):
    action = authRoot + "authlocaluser/getvouchersinglearg"
    postData = "domainname=" + domainname + "&code=" + code
    httpPost(authInfo, action, postData)
    data = json.dumps(json.loads(httpPost(authInfo, action, postData))["data"])
    return data

# portal-master
def config_php():
    DB_HOST = 'localhost'
    DB_USER = 'authpuppy'
    DB_PWD = 'citrus'
    # DB_USER = settings_py.DATABASES['default']['USER']
    # DB_PWD = settings_py.DATABASES['default']['PASSWORD']
    DB_NAME = 'authpuppy'
    DB_CHARSET = 'utf8'
    try:
        db = MySQLdb.connect(DB_HOST,DB_USER,DB_PWD,DB_NAME,charset='utf8')
        cursor = db.cursor()
        # cursor.execute("SET "+ DB_NAME + ' ' + DB_CHARSET)
        return db
    except Exception as e:
        print e
        return '1'



def list_template_php(request):
    db = config_php()
    if db == '1':
        return JsonResponse('Connect to db failed!',safe=False)
    else:
        title = request.GET.get('title')
        groupname = request.GET.get('groupname')
        try:
            cursor = db.cursor()
            sql = "select configs,templatename from portal_custom where title = '%s' and groupname = '%s' and state_template ='1'"%(title,groupname)
            cursor.execute(sql)
            result = cursor.fetchone()
            if result and cursor.rowcount > 0:
                configs = result[0]
            else:
                sql = "select configs,templatename from portal_custom where title = '%s' and groupname = 'root' and state_template ='1'"%(title)
                cursor.execute(sql)
                result = cursor.fetchone()
                configs = result[0]
            cursor.close()
            db.close()
            print configs
            return JsonResponse(configs,safe = False)
        except Exception as e:
            print e
            cursor.close()
            db.close()
            return JsonResponse('Connect to db failed!',safe = False)

def list_portal_php(request):
    db = config_php()
    print 1111111111111
    if db == '1':
        return JsonResponse('Connect to db failed!',safe=False)
    else:
        print 2222222222222
        data = []
        template = "template"
        try:
            print 33333333
            cursor = db.cursor()
            sql = "select theme from template_portal where template = '%s'"%(template)
            cursor.execute(sql)
            result = cursor.fetchall()
            for row in result:
                data.append({'name':row[0]})
            cursor.close()
            db.close()
            print 44444444444
            print data
            return JsonResponse(data,safe = False)
        except Exception as e:
            raise e
            cursor.close()
            db.close()
            return JsonResponse('Connect to db failed!',safe=False)

def list_page_php(request):
    db = config_php()
    print 555555555555
    if db == '1':
        return JsonResponse('Connect to db failed!',safe=False)
    else:
        print 6666666666
        title = request.GET.get('title')
        groupname = request.GET.get('groupname')
        try:
            print 7777777777777
            cursor = db.cursor()
            sql = "select templatename from portal_custom where title = '%s' and groupname = '%s' and state_template ='1'"%(title,groupname)
            cursor.execute(sql)
            result = cursor.fetchone()
            print 888888888888
            if result and cursor.rowcount > 0:
                templatename = result[0]
            else:
                sql = "select templatename from portal_custom where title = '%s' and groupname = 'root' and state_template ='1'"%(title)
                cursor.execute(sql)
                result = cursor.fetchone()
                templatename = result[0]
            cursor.close()
            db.close()
            print 99999999999
            print templatename
            return JsonResponse(templatename,safe = False)
        except Exception as e:
            print e
            cursor.close()
            db.close()
            return JsonResponse('Connect to db failed!',safe=False)

# def functions_php(request):
#     if request.method == "POST":    # 请求方法为POST时，进行处理
#         upfile =request.FILES.get("file", None)    # 获取上传的文件，如果没有文件，则默认为None
#         if not upfile:
#             error = _(u"没有选择文件")
#         else:
#             if upfile.content_type != 'image/jpeg' and upfile.content_type != 'image/gif' and upfile.content_type != 'image/png' and upfile.content_type != 'image/jpg':
#                 error = _(u"非图片文件")
#             elif upfile.name.split('.')[-1] != "jpeg" and upfile.name.split('.')[-1] != "jpg" and upfile.name.split('.')[-1] != "gif" and upfile.name.split('.')[-1] != "png":
#                 error = _(u"非图片文件")
#             elif upfile.size > 2097152:
#                 error = _(u"文件大小超过2M")
#             else:
#                 try:
#                     nn = str(int(time.time())) + str(random.randint(1000,9999)) + '.' + upfile.name.split('.')[-1]
#                     destination = open("./statics/portal_pic/" + nn,'wb+')    # 打开特定的文件进行二进制的写操作
#                     try:
#                         for chunk in upfile.chunks():      # 分块写入文件
#                             destination.write(chunk)
#                     except Exception as e:
#                         print e
#                         error = _(u"上传文件错误")
#                     finally:
#                         destination.close()
#                 except Exception as e:
#                     print e
#                     error = _(u"上传文件错误")
#         return 1


def configs_php(request):
    try:
        db = config_php()
        res = ""
        if db == '1':
            return JsonResponse('Connect to db failed!',safe=False)
        else:
            data = {}
            r_json = request.POST.get('portal')
            regl = '""(\w+)""'
            r_json = re.sub(regl,r'"\1"',r_json)
            r_json = json.loads(r_json)
            groupname = r_json['groupname'] if 'groupname' in r_json else ""
            templatename = r_json['templatename'] if 'templatename' in r_json else ""
            title = r_json['title'] if 'title' in r_json else ""
            status = 1
            config = r_json['config']
            prebgfile = config[0]['prebgfile'] if len(config) > 0 and 'prebgfile' in config[0] else ""
            prelogofile = config[0]['prelogofile'] if len(config) > 0 and 'prelogofile' in config[0] else ""

            config = json.dumps(r_json['config'],ensure_ascii=False)
            print config,'1x1x1x1x1x122222',r_json['config']
            config = config.replace("'","\\'")
            config = config.replace('"','\\"')
            cursor = db.cursor(cursorclass = MySQLdb.cursors.DictCursor)
            sql = "select state from portal_custom where title = '{}' and groupname = '{}'".format(title,groupname)
            cursor.execute(sql)
            result = cursor.fetchone()
            if result and cursor.rowcount > 0:
                print 't2t2t2t2t'
                rows = result
                state = rows['state']
                print config,'1x1x1x1x1x1'
                sql = "update portal_custom set configs='{}', status = '{}' where title = '{}' and groupname = '{}'".format(config,status,title,groupname)
                cursor.execute(sql)
                db.commit()
                if cursor.rowcount == 0 :
                    print " affected rows 0 , maybe not change any thing "
                print state ,'x1x1x1x11x1'
                # if state == '1':
                sql = "select configs,templatename,title from portal_custom where groupname = '{}' and status = '1'".format(groupname)
                # sql = "select configs,templatename,title from portal_custom where groupname = '{}'".format(groupname)
                cursor.execute(sql)
                if cursor.rowcount > 0:
                    row = cursor.fetchone()
                    # data.append({'configs':row['configs'],'templatename':row['templatename'],'title':row['title']})
                    print 'xt1tx1tx1tx'
                    data = {'configs':row['configs'],'templatename':row['templatename'],'title':row['title']}
                    configs = json.loads(data['configs'])
                    base = configs[0]['base']
                    bb = base['config']['background'] if 'background' in base['config'] else ""
                    logo = base['type']['logo'] if 'logo' in base['type'] else ""
                    hdheight = base['headheight']['headheight']['default'] if "default" in base['headheight']['headheight'] else ""
                    argue = base['argue']['argue'] if 'argue' in base['argue'] else ""
                    logo_left = base['configs']['margin-left']['default'] if 'default' in base['configs']['margin-left'] else ""
                    print 'xt1tx1tx1tx2'
                    logo_top = base['configs']['margin-top']['default'] if 'default' in base['configs']['margin-top'] else ""
                    logo_height = base['configs']['height']['default'] if 'default' in base['configs']['height'] else ""
                    logo_width = base['configs']['width']['default'] if 'default' in base['configs']['width'] else ""
                    back_height = base['config_b']['height']['default'] if 'default' in base['config_b']['height'] else ""

                    r_type = configs[0]['comps'][0]['type']
                    config = configs[0]['comps'][0]['config']
                    print 'xt1tx1tx1tx3'
                    text_one = config['text-one'] if 'text-one' in config else ""
                    text_two = config['text-two'] if 'text-two' in config else ""
                    text_three = config['text-three'] if 'text-three' in config else ""
                    text_four = config['text-four'] if 'text-four' in config else ""
                    text_five = config['text-five'] if 'text-five' in config else ""
                    print configs,'x2x2x2x2'
                    configc = configs[0]['comps'][0]['configs']
                    button_width = configc['width'] if 'width' in configc else ""
                    colorOne = configc['color-one'] if 'color-one' in configc else ""
                    colorTwo = configc['color-two'] if 'color-two' in configc else ""
                    colorThree = configc['color-three'] if 'color-three' in configc else ""

                    Suffix = 'css'
                    path = './statics/portal-master/portalcss/'
                    filename = data['templatename'] + '_' + data['title'] + '.' + Suffix
                    try:
                        myfile = open(path + filename,'wb')
                        print 'xt1tx1tx1tx4'
                        if data['title'] == "onekey":
                            print 1111
                            height_back = 300 - int(hdheight)
                            print hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo,height_back
                            txt = ".head{{height:{}px;}}\n.head .headline{{color:{};}}\n.head .Logo{{background:url({}) no-repeat; height:{}px; width:{}px; margin-left:{}px; margin-top:{}px;background-size:{};}}\n#mobile div.back_pic{{background: url({}) no-repeat;height:{}px;background-size:{};}}\n#onekey_login {{background:{};width: {};}}\ndiv.lineblock input {{color:{};}}\n#mobile div.advertise_pic {{height:{}px;}}".format(hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo,height_back)
                            print txt
                        elif data['title'] == 'wechat':
                            print 2222222222
                            txt = ".head{{height:{}px;}}\n.head .headline{{color:{};}}\n.head .Logo{{background:url({}) no-repeat; height:{}px; width:{}px; margin-left:{}px; margin-top:{}px;background-size:{};}}\n#mobile div.back_pic{{background: url({}) no-repeat;height:{}px;background-size:{};}}\na.mod-simple-follow-page__attention-btn{{background-color:{};width:{};}}\n.mod-simple-follow-page__attention-txt{{color:{};}}".format(hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width +'%',colorTwo)
                        elif data['title'] == 'voucher':
                            print 333333333
                            txt = ".head{{height:{}px;}}\n.head .headline{{color:{};}}\n.head .Logo{{background:url({}) no-repeat; height:{}px; width:{}px; margin-left:{}px; margin-top:{}px;background-size:{};}}\n#mobile div.back_pic{{background: url({}) no-repeat;height:{}px;background-size:{};}}\na#volume_auth_a{{background-color:{};width:{};}}\n.mod-simple-follow-page__attention input {{ color:{};}}".format(hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width + '%',colorTwo)
                        else:
                            print 34343
                            height_back = 300 - int(hdheight)
                            print 4444444444
                            txt = ".head{{height:{}px;}}\n.head .headline{{color:{};}}\n.head .Logo{{background:url({}) no-repeat; height:{}px; width:{}px; margin-left:{}px; margin-top:{}px;background-size:{};}}\n#mobile div.back_pic{{background: url({}) no-repeat;height:{} px;background-size:{};}}\ndiv.logbut a.login {{background:{};width:{};}}\ndiv.lineblock input {{color: {};}}\n#mobile div.advertise_pic {{height:{}px;}}".format(hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo,height_back)

                        imagesdir = "./media/portal_master/"
                        lndir = "/var/www/authpuppy/web/images/portal_master/"
                        lndir_source = "/opt/venv/wyun/DevCloud/media/portal_master/"
                        my_createdir(lndir)
                        print 'xt1tx1tx1tx5'
                        if prebgfile != "" and os.path.exists(imagesdir + prebgfile):
                            bgfile = prebgfile.replace("pre", "curr")
                            oldfile = groupname + data["title"] + "bg.jpg"
                            oldfileprefix = groupname + data["title"] + "currbg*jpg"

                            os.system(" sh -c 'cd " + imagesdir + ";rm -f " + oldfile + " " + oldfileprefix + ";mv -f " + prebgfile + " " + bgfile + "'")
                            os.system("cp " + lndir_source + bgfile + " " + lndir + bgfile)

                        if prelogofile != "" and os.path.exists(imagesdir + prelogofile):
                            logofile = prelogofile.replace("pre", "curr")
                            oldfile = groupname + data["title"] + "logo.jpg"
                            oldfileprefix = groupname + data["title"] + "currlogo*jpg"
                            os.system(" sh -c 'cd " + imagesdir +";rm -f " + oldfile + " " + oldfileprefix + ";mv -f " + prelogofile + " " + logofile + "'")
                            os.system("cp " + lndir_source + logofile + " " + lndir + logofile)
                        print 'xt1tx1tx1txxxxxx5'
                        myfile.write(txt)
                        myfile.close()
                        os.system(" chmod 755 " + path + filename)
                    except Exception as e:
                        print e
                        print "Unable to open file!"
                    suffix = 'txt'
                    print 'xt1tx1tx1tx6'
                    filenames = data["templatename"] + '_' + data['title'] + '.' + suffix
                    try:
                        myfiles = open(path + filenames,'wb')
                        txts = "{}\n{}\n{}\n{}\n{}".format(text_one,text_two,text_three,text_four,text_five)
                        myfiles.write(txts)
                        myfiles.close()
                        os.system(" chmod 755 " + path + filenames)
                    except Exception as e:
                        print e
                        print "Unable to open file!"
                    print 'xt1tx1tx1tx7'
                    name = data["templatename"] + "_" + title
                    filepath = './statics/portal-master/portalcss/'
                    filepaths = './statics/portal-master/txt/'
                    path = '/opt/run/portal/'
                    my_createdir(path)
                    filename = name + '.' + Suffix
                    filenames  = name + '.' + suffix
                    txtname = argue
                    nameargue = 'agreement'
                    cname = groupname + title + '.' + Suffix
                    tname = groupname + title + '.' + suffix
                    aname = groupname + title + nameargue + '.' + suffix
                    if state == '1':
                        os.system('cp ' + filepath + filename + '  ' + path + cname + ' ')
                        os.system('cp ' + filepath + filenames + '  ' + path + tname + ' ')
                        os.system('chmod 777 ' + path + '*')
                        print 'xt1tx1tx1tx8'
                        if argue != "":
                            os.system("cp " + filepaths + txtname + "  " + path + aname + " ")
                            os.system('chmod 777 ' + path + aname + ' ')
                            pathargue = '/statics/portal-master/template/'
                            name_phone = 'arguement.php'
                            name_phones = groupname + name_phone
                            name_account = 'arguementphone.php'
                            name_accounts = groupname + name_account
                            if data['title'] == 'phonelogin':
                                os.system('cp' + pathargue + name_phone + '  ' + path + name_phones + ' ')
                            elif data['title'] == 'accountlogin':
                                os.system('cp' + pathargue + name_account + '  ' + path + name_accounts + ' ')
                    templatename = data['templatename']
                    title = data['title']
                    sql = "update portal_custom set status = '0' where templatename = '{}' and title = '{}'".format(templatename,title)
                    cursor.execute(sql)
                    db.commit()
                    res = 333
                    print 't1t1t1t1t1t1t'

                res = '333'
                print 't1t1t1t1t1t1t2'
            else:
                print 't2t2t2t2t3'
                sql = "select state from portal_custom where title = '{}' and groupname = 'root'".format(title)
                cursor.execute(sql)
                rows = cursor.fetchone()
                print rows
                state = rows['state']
                print 'xt1tx1tx1tx10'
                templatenames = groupname
                state_template = 1
                sql = "insert into portal_custom(groupname,templatename,title,configs,state,status,state_template) value('{}','{}','{}','{}','{}','{}','{}')".format(groupname,templatenames,title,config,state,status,state_template)
                cursor.execute(sql)
                db.commit()
                print 'x1x1x1x11x1x'
                if cursor.rowcount > 0:
                    print 'x1x1x1x11x1x3'
                    print state
                    # if state == "1":
                    sql = "select configs,templatename,title from portal_custom where groupname = '{}' and status = '1'".format(groupname)
                    # sql = "select configs,templatename,title from portal_custom where groupname = '{}'".format(groupname)
                    cursor.execute(sql)
                    row = cursor.fetchone()
                    print 'x1x1x1x11x1x2'
                    if row and cursor.rowcount > 0:
                        data = {'configs':row['configs'],'templatename':row['templatename'],'title':row['title']}
                        configs = json.loads(data['configs'])
                        bb = configs[0]['base']['config']['background'] if 'background' in configs[0]['base']['config'] else ""
                        logo = configs[0]['base']['type']['logo'] if 'logo' in configs[0]['base']['type'] else ""
                        hdheight = configs[0]['base']['headheight']['headheight']['default'] if 'default' in configs[0]['base']['headheight']['headheight'] else ""
                        print 'xt1tx1tx1tx11'
                        argue = configs[0]['base']['argue']['argue'] if 'argue' in configs[0]['base']['argue'] else ""
                        logo_left = configs[0]['base']['configs']['margin-left']['default'] if 'default' in configs[0]['base']['configs']['margin-left'] else ""
                        logo_top = configs[0]['base']['configs']['margin-top']['default'] if 'default' in configs[0]['base']['configs']['margin-top'] else ""
                        logo_height = configs[0]['base']['configs']['height']['default'] if 'default' in configs[0]['base']['configs']['height'] else ""
                        logo_width = configs[0]['base']['configs']['width']['default'] if 'default' in configs[0]['base']['configs']['width'] else ""
                        back_height = configs[0]['base']['config_b']['height']['default'] if 'default' in configs[0]['base']['config_b']['height'] else ""
                        r_type = configs[0]['comps'][0]['type']
                        text_one = configs[0]['comps'][0]['config']['text-one'] if 'text-one' in configs[0]['comps'][0]['config'] else ""
                        text_two = configs[0]['comps'][0]['config']['text-two'] if 'text-two' in configs[0]['comps'][0]['config'] else ""
                        text_three = configs[0]['comps'][0]['config']['text-three'] if 'text-three' in configs[0]['comps'][0]['config'] else ""
                        text_four = configs[0]['comps'][0]['config']['text-four'] if 'text-four' in configs[0]['comps'][0]['config'] else ""
                        text_five = configs[0]['comps'][0]['config']['text-five'] if 'text-five' in configs[0]['comps'][0]['config'] else ""
                        button_width = configs[0]['comps'][0]['configs']['width'] if 'width' in configs[0]['comps'][0]['configs'] else ""
                        colorOne = configs[0]['comps'][0]['configs']['color-one'] if 'color-one' in configs[0]['comps'][0]['configs'] else ""
                        colorTwo = configs[0]['comps'][0]['configs']['color-two'] if 'color-two' in configs[0]['comps'][0]['configs'] else ""
                        colorThree = configs[0]['comps'][0]['configs']['color-three'] if 'color-three' in configs[0]['comps'][0]['configs'] else ""
                        Suffix = 'css'
                        path = '/statics/portal-master/portalcss/'
                        my_createdir(path)
                        print 'xt1tx1tx1tx12'
                        filename = data['templatename'] + "_" + data['title'] + "." + Suffix
                        try:
                            myfile = open(path + filename,'wb')
                            if data['title'] == "onekey":
                                height_back = 300-int(hdheight)
                                txt = ".head{{height:{}px;}}\n.head .headline{{color:{};}}\n.head .Logo{{background:url({}) no-repeat; height:{}px; width:{}px; margin-left:{}px; margin-top:{}px;background-size:{};}}\n#mobile div.back_pic{{background: url({}) no-repeat;height:{}px;background-size: {};}}\n#onekey_login {{background:{};width:{};}}\ndiv.lineblock input {{color:{};}}\n#mobile div.advertise_pic {{height:{}px;}}".format(hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo,height_back)
                            elif data["title"] == 'wechat':
                                txt = ".head{{height:{}px;}}\n.head .headline{{color:{};}}\n.head .Logo{{background:url({}) no-repeat; height:{}px; width:{}px; margin-left:{}px; margin-top:{}px;background-size:{};}}\n#mobile div.back_pic{{background: url({}) no-repeat;height:{}px;background-size: {};}}\na.mod-simple-follow-page__attention-btn{{background-color:{};width:{};}}\n.mod-simple-follow-page__attention-txt{{color:{};}}".format(hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo)
                            elif data["title"] == 'voucher':
                                txt = ".head{{height:{}px;}}\n.head .headline{{color:{};}}\n.head .Logo{{background:url({}) no-repeat; height:{}px; width:{}px; margin-left:{}px; margin-top:{}px;background-size:{};}}\n#mobile div.back_pic{{background: url({}) no-repeat;height:{}px;background-size: {};}}\na#volume_auth_a{{background-color:{};width:{};}}\n.mod-simple-follow-page__attention input {{ color:{};}}".format(hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo)
                            else:
                                height_back = 300-int(hdheight)
                                txt = ".head{{height:{}px;}}\n.head .headline{{color:{};}}\n.head .Logo{{background:url({}) no-repeat; height:{}px; width:{}px; margin-left:{}px; margin-top:{}px;background-size:{};}}\n#mobile div.back_pic{{background: url({}) no-repeat;height:{}px;background-size:{};}}\ndiv.logbut a.login {{background:{};width:{};}}\ndiv.lineblock input {{color:{};}}\n#mobile div.advertise_pic {{height:{}px;}}".format(hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo,height_back)

                            imagesdir = "./media/portal_master/"
                            lndir = "/var/www/authpuppy/web/images/portal_master/"
                            lndir_source = "/opt/venv/wyun/DevCloud/media/portal_master/"
                            my_createdir(lndir)
                            if prebgfile != "" and os.path.exists(imagesdir + prebgfile):
                                bgfile = prebgfile.replace("pre", "curr")
                                oldfile = groupname + data["title"] + "bg.jpg"
                                oldfileprefix = groupname + data["title"] + "currbg*jpg"

                                os.system(" sh -c 'cd " + imagesdir + ";rm -f " + oldfile + " " + oldfileprefix + ";mv -f " + prebgfile + " " + bgfile + "'")
                                os.system("cp " + lndir_source + bgfile + " " + lndir + bgfile)

                            if prelogofile != "" and os.path.exists(imagesdir + prelogofile):
                                logofile = prelogofile.replace("pre", "curr")
                                oldfile = groupname + data["title"] + "logo.jpg"
                                oldfileprefix = groupname + data["title"] + "currlogo*jpg"
                                os.system(" sh -c 'cd " + imagesdir +";rm -f " + oldfile + " " + oldfileprefix + ";mv -f " + prelogofile + " " + logofile + "'")
                                os.system("cp " + lndir_source + logofile + " " + lndir + logofile)
                            print 'xt1tx1tx1tx13'
                            myfile.write(txt)
                            myfile.close()
                            os.system(" chmod 755 " + path + filename)
                        except Exception as e:
                            print e
                            print "Unable to open file!"
                        suffix = 'txt'
                        filenames = data['templatename'] + "_" + data['title'] + "." + suffix
                        try:
                            myfiles = open(path + filenames,"wb")
                            txts = "{}\n{}\n{}\n{}\n{}".format(text_one,text_two,text_three,text_four,text_five)
                            myfiles.write(txts)
                            myfiles.close()
                            os.system(" chmod 755 " + path + filenames)
                        except Exception as e:
                            print e
                            print "Unable to open file!"
                        print 'xt1tx1tx1tx14'
                        name = data['templatename'] + '_' + title
                        filepath = './statics/portal-master/portalcss/'
                        filepaths = './statics/portal-master/txt/'
                        path = '/opt/run/portal/'
                        my_createdir(path)
                        filename = name + '.' + Suffix
                        filenames  = name + '.' + suffix
                        txtname = argue
                        nameargue = 'agreement'
                        cname = groupname + title + '.' + Suffix
                        tname = groupname + title + '.' + suffix
                        aname = groupname + title + nameargue + '.' + suffix
                        if state == '1':
                            os.system('cp ' + filepath + filename + '  ' + path + cname + ' ')
                            os.system('cp ' + filepath + filenames + '  ' + path + tname + ' ')
                            print 'xt1tx1tx1tx15'
                            if argue != "":
                                os.system("cp " + filepaths + txtname + "  " + path + aname +" ")
                                os.system("chmod 777 " + path + aname + " ")
                                pathargue = '/var/www/authpuppy/template/'
                                name_phone = 'arguement.php'
                                name_phones = groupname + name_phone
                                name_account = 'arguementphone.php'
                                name_accounts = groupname + name_account
                                if data['title'] == 'phonelogin':
                                    os.system("cp " + pathargue + name_phone  + "  " + path + name_phones + " ")
                                elif data['title'] == 'accountlogin':
                                    os.system("cp " + pathargue + name_account + "  " + path + name_accounts + " ")
                        templatename = data['templatename']
                        title = data["title"]
                        sql = "update portal_custom set status = '0' where templatename = '{}' and title = '{}'".format(templatename,title)
                        cursor.execute(sql)
                        db.commit()
                        res = '666'
                        print 't1t1t1t1t1t1t4'

                else:
                    res = "555"
                    print 't1t1t1t1t1t1t5'
            cursor.close()
            db.close()
            print '111111'
            return JsonResponse(res,safe = False)
    except Exception as e:
        print e




def update_file_php(request):
    try:
        db = config_php()
        res = '1'
        fileName = request.GET.get('filename')
        timestamp = request.GET.get('ts')
        upfile = request.FILES.get('image',None)
        if not upfile:
            error = _(u"没有选择文件")
        else:
            try:
                # path = './statics/portal-master/images/'
                path = "./media/portal_master/"
                Suffix = "jpg"
                name = fileName + '_' + timestamp + '.' + Suffix
                os.system("rm -f " + path + fileName + "_*.jpg")
                destination = open(path + name,'wb+')    # 打开特定的文件进行二进制的写操作
                try:
                    for chunk in upfile.chunks():      # 分块写入文件
                        destination.write(chunk)
                    res = '"' + name + '"'
                except Exception as e:
                    print e
                    error = _(u"上传文件错误")
                finally:
                    destination.close()
            except Exception as e:
                print e
                error = _(u"上传文件错误")
        db.close()
    except Exception as e:
        print e
        db.close()
    return JsonResponse(res,safe = False)

def statement_php(request):
    try:
        res = '1'
        db = config_php()
        fileName = request.GET.get('filename')
        upfile = request.FILES.get('txt',None)
        if not upfile:
                error = _(u"没有选择文件")
        else:
            try:
                path = './statics/portal-master/txt/'
                # path = '/var/www/portal-master/web/client/txt/'
                cmd = "mkdir -p " + path + "; chmod 777 " + path
                os.system(cmd)
                Suffix = os.path.splitext(upfile.name)[1]
                name = fileName + Suffix
                destination = open(path + name,'wb+')    # 打开特定的文件进行二进制的写操作
                try:
                    for chunk in upfile.chunks():      # 分块写入文件
                        destination.write(chunk)
                    res = name
                except Exception as e:
                    print e
                    error = _(u"上传文件错误")
                finally:
                    destination.close()
            except Exception as e:
                print e
                error = _(u"上传文件错误")
        db.close()
    except Exception as e:
        print e
        db.close()
    return JsonResponse(res,safe = False)

def show_type_php(request):
    db = config_php()
    if db == '1':
        return JsonResponse('Connect to db failed!',safe=False)
    else:
        template = request.GET.get('template')
        try:
            cursor = db.cursor()
            sql = "select type from portal_custom where template = '%s'"%(template)
            cursor.execute(sql)
            result = cursor.fetchone()
            if result and cursor.rowcount > 0:
                r_type = result[0]
            cursor.close()
            db.close()
            return JsonResponse(r_type,safe = False)
        except Exception as e:
            print e
            cursor.close()
            db.close()
            return JsonResponse('Connect to db failed!',safe=False)

def show_page_php(request):
    db = config_php()
    if db == '1':
        return JsonResponse('Connect to db failed!',safe=False)
    else:
        template = request.GET.get('template')
        try:
            cursor = db.cursor()
            sql = "select configs from portal_custom where template = '%s'"%(template)
            cursor.execute(sql)
            result = cursor.fetchone()
            if result and cursor.rowcount > 0:
                configs = result[0]
            cursor.close()
            db.close()
            return configs
        except Exception as e:
            print e
            cursor.close()
            db.close()
            return JsonResponse('Connect to db failed!',safe=False)

def save_page_php(request):
    db = config_php()
    if db == '1':
        return JsonResponse('Connect to db failed!',safe=False)
    else:
        r_json = json.loads(request.POST.get('portal'))
        username = "22"
        groupname = ''
        templatename = r_json['templatename']
        theme = r_json['name']
        r_type = r_json['type']
        config = json.dumps(r_json['config'])
        config = config.replace("'","\\'")
        config = config.replace('"','\\"')
        try:
            cursor = db.cursor()
            sql = "select * from template_portal where type = '%s'" %(r_type)
            cursor.excute(sql)
            result = cursor.fetchall()
            if result and cursor.rowcount > 0:
                sql = "update template_portal set configs='%s' where type = '%s' and template = '%s'"%(config,r_type,templatename)
                cursor.excute(sql)
                db.commit()
                if cursor.rowcount > 0 :
                    res = '333'
                else:
                    res = '222'
            else:
                sql = "insert into template_portal(template,theme,type,configs) value('%s','%s','%s','%s')"%(templatename,theme,r_type,config)
                cursor.excute(sql)
                db.commit()
                if cursor.rowcount > 0:
                    res = '666'
                else:
                    res = '555'
            cursor.close()
            db.close()
            return JsonResponse(res,safe = False)
        except Exception as e:
            print e
            cursor.close()
            db.close()
            return JsonResponse('Connect to db failed!',safe=False)

def operation_php(request):
    db = config_php()
    if db == '1':
        return JsonResponse('Connect to db failed!',safe=False)
    else:
        id_de = request.GET.get('id')
        id_ap = request.GET.get('ida')
        if id_de != '':
            try:
                cursor = db.cursor()
                sql = "delete from portal_custom where id = '%s'"%(id_de)
                cursor.execute(sql)
                db.commit()
                if cursor.rowcount > 0 :
                    data = "delete"
            except Exception as e:
                print e
                cursor.close()
                db.close()
                return JsonResponse('Connect to db failed!',safe=False)
        elif id_ap != 0 :
            try:
                cursor = db.cursor()
                sql = "select id from portal_custom where status = '1'"
                cursor.execute(sql)
                result = cursor.fetchone()
                if result and cursor.rowcount > 0 :
                    r_id = result[0]
                    sql = "update portal_custom set status = '0' where id = '%s'"%(r_id)
                    cursor.execute(sql)
                    db.commit()
                    if cursor.rowcount > 0:
                        sql = "update portal_custom set status = '1' where id = '%s'"%(id_ap)
                        cursor.execute(sql)
                        db.commit()
                        if cursor.rowcount > 0:
                            data = 'apply'
                else:
                    sql = "update portal_custom set status = '1' where id = '%s'"%(id_ap)
                    cursor.execute(sql)
                    db.commit()
                    if cursor.rowcount > 0 :
                        data = "apply"
            except Exception as e:
                print e
                cursor.close()
                db.close()
                return JsonResponse('Connect to db failed!',safe=False)
        cursor.close()
        db.close()
        return JsonResponse(data,safe = False)

def my_createdir(path):
    os.system(" mkdir -p " + path +" ")
    os.system(" chmod 777 " + path + " -R ")
    return True



def create_portal_css_php(request):
    db = config_php()
    if db == '1':
        return JsonResponse({'meta':'Connect to db failed!'},safe=False)
    else:
        groupname = request.user.groupname
        sql = "select configs,templatename,title from portal_custom where (groupname = '%s' or groupname = 'root') and status = '1'" %(groupname)
        cursor = db.cursor(cursorclass = MySQLdb.cursors.DictCursor)
        cursor.execute(sql)
        result = cursor.fetchall()

        if cursor.rowcount > 0 and result:
            for row in result:
                data = {"configs":row['configs'],"templatename":row['templatename'],"title":row['title']}
                # print data
                configs = json.loads(data["configs"])
                base = configs[0]['base']
                if 'background' in base['config']:
                    bb = base['config']['background']
                else:
                    bb = ''
                if 'logo' in base['type']:
                    logo = base['type']['logo']
                else:
                    logo = ''
                if 'default' in base['headheight']['headheight']:
                    hdheight = base['headheight']['headheight']['default']
                else:
                    hdheight = ''
                if 'argue' in base['argue']:
                    argue = base['argue']['argue']
                else:
                    argue = ''
                if 'default' in base['configs']['margin-left']:
                    logo_left = base['configs']['margin-left']['default']
                else:
                    logo_left = ''
                if 'default' in base['configs']['margin-top']:
                    logo_top = base['configs']['margin-top']['default']
                else:
                    logo_top = ''
                if 'default' in base['configs']['height']:
                    logo_height = base['configs']['height']['default']
                else:
                    logo_height = ''
                if 'default' in base['configs']['width']:
                    logo_width = base['configs']['width']['default']
                else:
                    logo_width = ''
                if 'default' in base['config_b']['height']:
                    back_height = base['configs']['width']['default']
                else:
                    back_height = ''

                comps = configs[0]['comps']
                # $type[$i] = $comps[$i][0]->type

                if 'text-one' in comps[0]['config']:
                    text_one = comps[0]['config']['text-one']
                else:
                    text_one = ''
                if 'text-two' in comps[0]['config']:
                    text_two = comps[0]['config']['text-two']
                else:
                    text_two = ''
                if 'text-three' in comps[0]['config']:
                    text_three = comps[0]['config']['text-three']
                else:
                    text_three = ''
                if 'text-four' in comps[0]['config']:
                    text_four = comps[0]['config']['text-four']
                else:
                    text_four = ''
                if 'text-five' in comps[0]['config']:
                    text_five = comps[0]['config']['text-five']
                else:
                    text_five = ''

                if 'width' in comps[0]['configs']:
                    button_width = comps[0]['configs']['width']
                else:
                    button_width = ''
                if 'color-one' in comps[0]['configs']:
                    colorOne = comps[0]['configs']['color-one']
                else:
                    colorOne = ''
                if 'color-two' in comps[0]['configs']:
                    colorTwo = comps[0]['configs']['color-two']
                else:
                    colorTwo = ''
                if 'color-three' in comps[0]['configs']:
                    colorThree = comps[0]['configs']['color-three']
                else:
                    colorThree = ''

                Suffix = 'css'
                path = '/var/www/was/library/portalcss/'
                my_createdir(path)
                filename = data['templatename'] + '_' + data["title"] + '.' + Suffix
                try:
                    myfile = open(path + filename,'wb')
                    txt = ""
                    if data['title'] == 'onekey':
                        height_back = 300-int(hdheight)
                        txt = ".head{height: %spx;}\n.head .headline{color:%s;}\n.head .Logo{background:url(%s) no-repeat; height:%spx; width:%spx; margin-left:%spx; margin-top:%spx;background-size: %s ;}\n#mobile div.back_pic{background: url(%s) no-repeat;height:%spx;background-size: %s ;}\n#onekey_login {background:%s;width:%s;}\ndiv.lineblock input {color:%s;}\n#mobile div.advertise_pic {height:%spx;}" % (hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo,height_back)
                    elif data['title'] == 'wechat':
                        txt = ".head{height:%spx;}\n.head .headline{color:%s;}\n.head .Logo{background:url(%s) no-repeat; height:%spx; width:%spx; margin-left:%spx; margin-top:%spx;background-size: %s ;}\n#mobile div.back_pic{background: url(%s) no-repeat;height:%spx;background-size: %s ;}\na.mod-simple-follow-page__attention-btn{background-color:%s;width:%s;}\n.mod-simple-follow-page__attention-txt{color:%s;}" % (hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo)
                    elif data['title'] == 'voucher':
                        txt = ".head{height:%spx;}\n.head .headline{color:%s;}\n.head .Logo{background:url(%s) no-repeat; height:%spx; width:%spx; margin-left:%spx; margin-top:%spx;background-size: %s ;}\n#mobile div.back_pic{background: url(%s) no-repeat;height:%spx;background-size: %s ;}\na#volume_auth_a{background-color:%s;width:%s;}\n.mod-simple-follow-page__attention input { color: %s; }" % (hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo)
                    else:
                        height_back = 300-int(hdheight)
                        txt = ".head{height:%spx;}\n.head .headline{color:%s;}\n.head .Logo{background:url(%s) no-repeat; height:%spx; width:%spx; margin-left:%spx; margin-top:%spx;background-size: %s ;}\n#mobile div.back_pic{background: url(%s) no-repeat;height:%spx;background-size: %s ;}\ndiv.logbut a.login {background:%s;width:%s;}\ndiv.lineblock input {color:%s;}\n#mobile div.advertise_pic {height:%spx;}" % (hdheight,colorOne,logo,logo_height,logo_width,logo_left,logo_top,'100%',bb,back_height,'100%',colorThree,button_width+'%',colorTwo,height_back)
                    myfile.write(txt)
                    myfile.close()
                    os.system("chmod " + " 755 " + path + filename)
                except Exception as e:
                    print e
                suffix = 'txt'
                filenames = data["templatename"]+'_'+data["title"]+'.'+suffix
                try:
                    myfiles = open(path + filenames, "wb")
                    txts = "%s\n%s\n%s\n%s\n%s"%(text_one,text_two,text_three,text_four,text_five)
                    myfiles.write(txts)
                    myfiles.close()
                    os.system("chmod " +" 755 " + path + filenames )
                except Exception as e:
                    print e
                templatename = data["templatename"]
                title = data["title"]
                sql = "update portal_custom set status = '0' where templatename = '%s' and title = '%s'"%(templatename,title)
                result1 = cursor.execute(sql)
                db.commit()
        cursor.close()
        db.close()
        return JsonResponse({'meta':"ok"},safe=False)



def list_edit_portal_php(request):
    db = config_php()
    if db == '1':
        return JsonResponse('Connect to db failed!',safe=False)
    else:
        try:
            data = []
            meta={ 'rc' : 'ok' }
            groupname = request.user.groupname
            cursor = db.cursor(cursorclass = MySQLdb.cursors.DictCursor)
            sql = "select templatename,title,state,groupname from portal_custom where (groupname = '{}' or groupname = 'root') and state_template = '1'".format(groupname)
            cursor.execute(sql)
            result = cursor.fetchall()
            templates = {}
            default_template = {}
            print result
            print '22222222s2'
            for row in result:
                if row['groupname'] == "root":
                    del row["templatename"]
                    default_template[row['title']] = row
                else:
                    if row['templatename'] not in templates:
                        templates[row['templatename']] = { row['title'] : row }
                    else:
                        templates[row['templatename']][row['title']] = row
                i = 0
                print 'ttttt',templates
                for key,value in templates.items():
                    print 't1t1t1t1t',default_template
                    for dkey,dvalue in default_template.items():
                        print '22222222s23'
                        if dkey not in templates[key]:
                            templates[key][dkey] = dvalue
                        print '22222222s24'
                    print data,i
                    data.append({'name':key})
                    print 't2t2t2t2t2'
                    for authtype,content in templates[key].items():
                        print authtype,content
                        # del content["templatename"]

                        # if authtype == "onekey" or authtype == "accountlogin" or authtype == "voucher" or authtype == "phonelogin" or authtype == "wechat":
                        data[i][authtype] = content
                    i += 1
            print data,1111
            if i == 0:
                data.append(default_template)
                data[0]["name"] = "template"
            print data,222
            result= {'data':data,'meta':meta}
            cursor.close()
            db.close()
            return JsonResponse(result,safe=False)
        except Exception as e:
            print e


def portal_operation_php(request):
    db = config_php()
    if db == '1':
        return JsonResponse('Connect to db failed!',safe=False)
    else:
        try:
            # print 'x1x1x1x1x1x101010'
            # json_str = request.POST.get('json')
            # print 'x1x1x1x1x1x4444',json_str
            # json_str = json.loads(json_str)
            # print 'x1x1x1x1x1x222'
            # cmd = json_str['cmd'] if 'cmd' in json_str else ""
            # print 'x1x1x1x1x1x333'
            cmd = request.POST.get('cmd') if request.POST.get('cmd')  else ""
            groupname = request.user.groupname
            meta = {}
            print 'x1x1x1x1x1x'
            if cmd == "apply":
                title = request.POST.get('title') if request.POST.get('title')  else ""
                # title= json_str['title'] if 'title' in json_str else ""
                state = 1
                cursor = db.cursor(cursorclass = MySQLdb.cursors.DictCursor)
                print 'x1x1x1x1x1xx12'
                sql = "update portal_custom set state = '{}' where title = '{}' and state_template ='1' and (groupname = '{}' or groupname = 'root')".format(state,title,groupname)
                result = cursor.execute(sql)
                db.commit()
                print 'x1x1x1x1x1x2'
                if result and cursor.rowcount > 0:
                    sql = "select templatename,configs,groupname from portal_custom where title = '{}' and state_template ='1' and groupname = '{}'".format(title,groupname)
                    cursor.execute(sql)
                    print cursor.rowcount
                    print 'x1x1x1x1x1x3'
                    if cursor.rowcount == 0:
                        sql = "select templatename,configs,groupname from portal_custom where title = '{}' and state_template ='1' and groupname = 'root'".format(title)
                        print 'y1y1y1y1yxx'
                        cursor.execute(sql)
                        print 'y1y1y1y1yqq'
                        print cursor.rowcount,12312321321
                    print 'y1y1y1y1y'
                    row = cursor.fetchone()
                    print row
                    print 'y1y1y1y1yewq'
                    configs = row['configs']
                    print 'y1y1y1y1yxxqq'
                    configs = json.loads(configs)
                    print 'y1y1y1y1y321'
                    base = configs[0]['base']
                    print 'y1y1y1y1y123'
                    argues= base['argue']
                    print 'y1y1y1y1y2'
                    argue = argues['argue'] if 'argue' in argues else ""
                    name = row['templatename'] + '_' + title
                    # filepath='/var/www/was/library/portalcss/'
                    # filepaths='/var/www/portal-master/web/client/txt/'
                    filepath = './statics/portal-master/portalcss/'
                    filepaths = './statics/portal-master/txt/'
                    Suffix = 'css'
                    suffix = 'txt'
                    print 'y1y1y1y1y3'
                    filename = name + "." + Suffix
                    filenames = name + "." + suffix
                    txtname = argue
                    print 'y1y1y1y1y4'
                    path = '/opt/run/portal/'
                    my_createdir(path)
                    nameargue = 'agreement'
                    print 'x1x1x1x1x1x4'
                    if row['groupname'] == groupname:
                        cname = groupname + title + "." + Suffix
                        tname = groupname + title + "." + suffix
                        aname = groupname + title + nameargue + "." + suffix
                    else:
                        cname = ""
                        tname = ""
                        aname = ""
                    os.system(" cp " + filepath + filename + "  " + path + cname + " ")
                    os.system(" cp " + filepath + filenames + "  " + path + tname + " ")
                    os.system(" chmod 777 " + path + "*")
                    print 'x1x1x1x1x1x5'
                    if argue  != "":
                        os.system(" cp " + filepaths + txtname + "  " + path + aname + " ")
                        os.system(" chmod 777 " + path + aname + " ")
                        pathargue = '/var/www/authpuppy/template/'
                        name_phone = 'arguement.php'
                        name_phones = groupname + name_phone
                        name_account = 'arguementphone.php'
                        name_accounts = groupname + name_account
                        if title == 'phonelogin':
                            os.system(" cp " + pathargue + name_phone + "  " + path + name_phones + " ")
                        elif title == 'accountlogin':
                            os.system(" cp " + pathargue + name_account + "  " + path + name_accounts + " ")
                        os.system(" chmod 777 " + pathargue + "*")
                    meta["rc"] = "ok"
                    meta["msg"] = "api.err.OperationSuccess"
                    meta["allow_success_msg"] = True
                    print 'x1x1x1x1x1x6'
                else:
                    meta["rc"] = "error"
                    meta["msg"] = "api.err.OperationFailed"

            elif cmd == "cancel":
                print 'x1x1x1x1x1x7'
                title = request.POST.get('title') if request.POST.get('title')  else ""
                state = 0
                cursor = db.cursor(cursorclass = MySQLdb.cursors.DictCursor)
                sql = "update portal_custom set state = '{}' where title = '{}' and (groupname = '{}' or groupname = 'root') and state_template ='1'".format(state,title,groupname)
                result = cursor.execute(sql)
                db.commit()
                print 'x1x1x1x1x1x8'
                if cursor.rowcount == 0:
                    meta["rc"] = "error"
                    meta["msg"] = "api.err.OperationFailed"
                else:
                    sql = "select groupname from portal_custom where title = '{}' and state_template ='1'".format(title)
                    result = cursor.execute(sql)
                    row = cursor.fetchone()
                    path = '/opt/run/portal/'
                    Suffix = 'css'
                    suffix = 'txt'
                    nameargue = 'agreement'
                    filename = groupname + title + "." + Suffix
                    filenames = groupname + title + "." + suffix
                    aname = groupname + title + nameargue + "." + suffix
                    os.system(" rm -f " + path + filename + " ")
                    os.system(" rm -f " + path + filenames + " ")
                    os.system(" rm -f " + path + aname + " ")
                    name_phone = 'arguement.php'
                    name_phones = groupname + name_phone
                    name_account = 'arguementphone.php'
                    name_accounts = groupname + name_account
                    print 'x1x1x1x1x1x9'
                    if title == 'phonelogin':
                        os.system(" rm " + path + name_phones + " ")
                    elif title == 'accountlogin':
                        os.system(" rm " + path + name_accounts + " ")
                    meta["rc"] = "ok"
                    meta["msg"] = "api.err.OperationSuccess"
                    meta["allow_success_msg"] = True

            result = {'meta':meta}
            cursor.close()
            db.close()
            return JsonResponse(result,safe = False)
        except Exception as e:
            print e


def authAccountInfoApi(request):
    db = config_php()
    if db != '1':
        userGroupName = request.user.groupname or ''
        if userGroupName != '':
            if request.method == 'GET':
                mode = request.GET.get('mode')
                if mode == 'export':
                    try:
                        cursor = db.cursor()
                        if userGroupName == 'admin':
                            sql = "SELECT username,password,expiretime from ap_user where username not like 'auto_%' and username not like 'wechat_%' and username not like '%:%'"
                        else:
                            sql = "SELECT username,password,expiretime from ap_user where username not like 'auto_%' and username not like 'wechat_%' and username like '%{}:%'".format(userGroupName)
                        cursor.execute(sql)
                        result = cursor.fetchall()

                        wb = Workbook()
                        sheet = wb.active
                        sheet.title = "Account Info"
                        # sheet['A1'].value = "ID"
                        sheet['A1'].value = "Account"
                        sheet['B1'].value = "Password"
                        sheet['C1'].value = "Expiration time"
                        excel_row = 1
                        for i in result:
                            # print i
                            if userGroupName == 'admin':
                                username = i[0]
                            else:
                                username = i[0].split(':')[1]
                            password = i[1]
                            expiretime = i[2]
                            if expiretime:
                                try:
                                    expiretime = expiretime.strftime('%Y-%m-%d %H:%M:%S')
                                except Exception as e:
                                    print e
                                    expiretime = None
                            excel_row += 1
                            # sheet["A%d" % (excel_row)].value = excel_row - 1
                            sheet["A%d" % (excel_row)].value = username
                            sheet["B%d" % (excel_row)].value = password
                            sheet["C%d" % (excel_row)].value = expiretime

                        filename = "./authAccountInfo.xlsx"
                        exist_file = os.path.exists("./authAccountInfo.xlsx")
                        if exist_file:
                            os.remove(filename)
                        wb.save(filename)
                        wrapper = FileWrapper(open(filename, 'rb'))
                        response = HttpResponse(wrapper, content_type='application/vnd.ms-excel')
                        response['Content-Length'] = os.path.getsize(filename)
                        response['Content-Disposition'] = 'attachment; filename=AuthAccountInfo.xlsx'
                        cursor.close()
                        db.close()

                        return response
                    except Exception as e:
                        print e
                        cursor.close()
                        db.close()

            if request.method == 'POST':
                upfile =request.FILES.get("file", None)    # 获取上传的文件，如果没有文件，则默认为None
                error = ""
                if not upfile:
                    error = _(u"没有选择文件")
                else:
                    if os.path.exists("./statics/upload/"+ request.user.groupname):
                        pass
                    else:
                        os.makedirs("./statics/upload/"+ request.user.groupname)
                    destination = open("./statics/upload/"+ request.user.groupname +"/"+upfile.name,'wb+')    # 打开特定的文件进行二进制的写操作
                    try:
                        for chunk in upfile.chunks():      # 分块写入文件
                            destination.write(chunk)
                    except Exception as e:
                        print e
                        error = _(u"文件写入失败")
                    finally:
                        destination.close()
                    if error == "":
                        insert = 0
                        update = 0
                        wrong = 0
                        furl = ''
                        try:
                            wb = load_workbook(filename="./statics/upload/"+ request.user.groupname +"/"+upfile.name)
                            ws = wb.active
                            rows = ws.rows

                            writewb = Workbook()
                            writews = writewb.active
                            writews.append(['account','password','expiretime','error'])
                            if (ws.cell('A1').value == "Account" and ws.cell('B1').value == "Password" and ws.cell('C1').value == "Expiration time") :
                                for i in rows:
                                    sign = 0
                                    if ws.cell('A1').value == "Account":
                                        account = i[0].value or ''
                                        account = str(account).strip()
                                        password = i[1].value or ''
                                        password = str(password).strip()
                                        expiretime = i[2].value or ''
                                        expiretime = str(expiretime).strip()
                                    elif ws.cell('B1').value == "Account":
                                        account = i[1].value or ''
                                        account = str(account).strip()
                                        password = i[2].value or ''
                                        password = str(password).strip()
                                        expiretime = i[3].value or ''
                                        expiretime = str(expiretime).strip()
                                    if account == 'Account' and password == 'Password' and expiretime == 'Expiration time':
                                        pass
                                    else:
                                        if len(account) == 0 or len(password) == 0 :
                                            wrong = wrong + 1
                                            sign = 1
                                            writews.append([account,password,expiretime,_(u'Account/Password 为空')])
                                        elif len(account) < 4  or len(password) < 4:
                                            wrong = wrong + 1
                                            sign = 1
                                            writews.append([account,password,expiretime,_(u'Account/Password 不能少于4个字符')])
                                        elif len(''.join(re.findall(r'[^0-9a-zA-Z_]',account))) != 0:
                                            wrong = wrong + 1
                                            sign = 1
                                            writews.append([account,password,expiretime,_(u'Account 只允许0-9a-zA-Z和_')])
                                        elif len(''.join(re.findall(r'[^0-9a-zA-Z]',password))) != 0:
                                            wrong = wrong + 1
                                            sign = 1
                                            writews.append([account,password,expiretime,_(u'Password 只允许0-9a-zA-Z')])
                                        elif expiretime != '' and expiretime != 'None':
                                            try:
                                                if expiretime != datetime.datetime.strptime(expiretime,'%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S'):
                                                    raise ValueError
                                            except ValueError:
                                                wrong = wrong + 1
                                                sign = 1
                                                writews.append([account,password,expiretime,_(u'Expiration time 内容或格式错误')])

                                        if sign == 0:
                                            if userGroupName == 'admin':
                                                usernameLower = account.lower()
                                            else:
                                                account = '{}:{}'.format(userGroupName,account)
                                                usernameLower = '{}:{}'.format(userGroupName,account.lower())
                                            if expiretime == '' or expiretime == 'None':
                                                expiretime = 'NULL'

                                            try:
                                                cursor = db.cursor()
                                                sql = "SELECT count(*) from ap_user where username_lower = '{}'".format(usernameLower)
                                                cursor.execute(sql)
                                                result = cursor.fetchone()
                                                if result[0] == 0:
                                                    accountEmail = '{}{}@witrusty.com'.format(int(time.time()),random.randint(0,9999))
                                                    if expiretime != 'NULL':
                                                        sql = "INSERT into ap_user (username,password,email,registered_on,status,username_lower,phone,expiretime) values ('{username}','{password}','{email}','{registered_on}','{status}','{username_lower}','{phone}','{expiretime}')".format(username = account,password = password,email = accountEmail,registered_on = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),status = 1,username_lower = usernameLower,phone = '',expiretime = expiretime)
                                                    else:
                                                        sql = "INSERT into ap_user (username,password,email,registered_on,status,username_lower,phone,expiretime) values ('{username}','{password}','{email}','{registered_on}','{status}','{username_lower}','{phone}',{expiretime})".format(username = account,password = password,email = accountEmail,registered_on = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),status = 1,username_lower = usernameLower,phone = '',expiretime = expiretime)
                                                    insert = insert + 1
                                                else:
                                                    if expiretime != 'NULL':
                                                        sql = "UPDATE ap_user set username = '{username}',password = '{password}', expiretime = '{expiretime}' where username_lower = '{username_lower}'".format(username = account,password = password,username_lower = usernameLower,expiretime = expiretime)
                                                    else:
                                                        sql = "UPDATE ap_user set username = '{username}',password = '{password}', expiretime = {expiretime} where username_lower = '{username_lower}'".format(username = account,password = password,username_lower = usernameLower,expiretime = expiretime)
                                                    update = update + 1

                                                cursor.execute(sql)
                                                db.commit()
                                            except Exception as e:
                                                print e
                                                db.rollback()
                                                wrong = wrong + 1
                                                writews.append([account,password,expiretime,_(u'设备存储失败')])

                                cursor.close()
                                db.close()
                                if os.path.exists("./statics/download/"+ request.user.groupname):
                                    pass
                                else:
                                    os.makedirs("./statics/download/"+ request.user.groupname)
                                if wrong != 0 :
                                    try:
                                        furl = "download/" + request.user.groupname + "/" + str(int(time.time())) + "-authInfoImportResult.xlsx"
                                        fn = "./statics/"+ furl
                                        writewb.save(filename = fn)
                                    except Exception as e:
                                        print e
                                        error = _(u'错误列表写入失败')
                            else:
                                error = _(u'文件内容不合法')
                        except Exception as e:
                            print e
                            error = _(u'文件读取失败')

                if error == "":
                    error = '{insert}{insertNum}{item},{update}{updateNum}{item},{wrong}{wrongNum}{item}'.format(insert = _(u"插入"),insertNum = insert,update = _(u"更新"),updateNum = update,wrong = _(u"错误"),wrongNum = wrong,item = _(u"项"))
                return JsonResponse({'error':error,'furl':furl},safe = False)
    return redirect('/ap-list/ap_guest_policy/')



def deviceLocateApi(request):
    status = ''
    if request.method == 'POST':
        mac = request.POST.get('mac') or ''
        state = request.POST.get('state') or 'off'
        if Device.objects.filter(mac = mac).exists():
            device = Device.objects.get(mac = mac)
        else:
            device = None
        if device:
            try:
                data = {'action':'leddetect','mac':mac,'param':state}
                DataQuery.DQProcess(data)
                device.locateState = state
                device.save()
                status = 'success'
            except Exception as e:
                print e
                status = 'failed'

    return JsonResponse({'status':status})


def protalCustomPage(request):
    return render(request, 'protalCustomerIndex.html')
