#coding=utf-8
from django.utils.translation import ugettext as _
from django.utils.translation import ugettext_lazy as lazy_
from django.shortcuts import render
from django.contrib import auth
from django.contrib.auth import login,authenticate
from account.models import NewUser,Account_Group,Auto_Update_Rule
from device.models import Device,Device_probe,Probe_config,Probe_audit_basic_status,Probe_audit_dev_status,Probe_audit_place_status,Probe_event,Probe_group
from ap.models import Device_ap,Ap_user_policy_config_extend
from django.contrib.auth.decorators import login_required
import models
import json
from django.http import JsonResponse
from django.forms.models import model_to_dict
from django.http import HttpResponse,HttpResponseRedirect
# from forms import
from django.utils import timezone
import datetime
from django.core.urlresolvers import reverse
from django.shortcuts import redirect
from django.core.servers.basehttp import FileWrapper
import time
import pytz
import urllib2
import hashlib
import re
# Create your views here.
import sys
sys.path.append("DataQuery")
import DataQuery
import os
import os.path
from system.models import Setting,Page_limit,Oem_limit
from django.core.files import File
import Public_function
import logging
import MySQLdb
import MySQLdb.cursors
from django.db import connection
from django.conf import settings as settings_py
from django.utils.timezone import is_naive, make_aware, utc, is_aware
from django.db.models import Q
from django.db import connection
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


from device.tasks import default_config_issued

def get_plimit():
    if Page_limit.objects.filter(pk = 1).exists():
        plimit = model_to_dict(Page_limit.objects.get(pk = 1))
    else:
        plimit = Public_function.page_limit_dict
    return plimit

def get_oemlimit():
    if Oem_limit.objects.filter(pk = 1).exists():
        oemlimit = model_to_dict(Oem_limit.objects.get(pk = 1))
    else:
        oemlimit = Public_function.oem_limit_dict
    return oemlimit


def upgrade_file_list():
    django_settings = Setting.objects.get(SIGN = 1)
    try:
        # url = "http://115.28.241.216:81/version/byz/"
        url = django_settings.VERSION_SERVER_URL+"version/byz/"
        website = urllib2.urlopen(url)

        html = website.read()

        # links = re.findall('href="(.*)">byzoro', html)
        links = re.findall('href="(.*)">[a-z]', html)
        # print links
        a = []
        for i in links:
            if i[-4:] == '.bin':
                if 'kernel' in i or 'rootfs' in i:
                    pass
                else:
                    a.append(i)
        # print a
        return a
    except Exception as e:
        return False

def upgrade_file(dev_type,fl):
    r = dev_type_change(dev_type)
    result = ''
    print r
    if r == False or fl == False:
        pass
    else:
        for i in fl:
            if i == r :
                result = i
        if result != "":
            return result

    return False

def dev_type_change(dev_type):
    try:
        f = open('./statics/upgrade/device_upgrade_dict.info','r')
        try:
            data = json.load(f,'utf-8')
            return data[dev_type]
        except Exception as e:
            return False
        finally:
            f.close()
    except Exception as e:
        return False



def compare_ver(file1,ver1):
    if file1 == False:
        return False
    return file1 != str(ver1)

def probe_event_save(pe):
    probe_event = Probe_event()
    probe_event.event_time = pe["event_time"]
    probe_event.event = pe['event']
    probe_event.msg = pe['msg']
    probe_event.admin_username = pe['admin_username']
    probe_event.probe_mac = pe['probe_mac']
    try:
        probe_event.save()
        return True
    except Exception as e:
        print e
        return False
@login_required
def probe_list(request):
    # global ugfl
    DataQuery.ugfl = upgrade_file_list()
    errors = ''
    # 未准入的探针相关 #
    # probe1 = {'mac':'112233445566','model':'X-300','privateip':'192.168.1.101','lastip':'10.200.10.1:20000','version':'2.00000001','last_heart_time':'2017-04-07 15:30:00'}
    # probe2 = {'mac':'112233445566','model':'X-300','privateip':'192.168.1.101','lastip':'10.200.10.1:20000','version':'2.00000001','last_heart_time':'2017-04-07 15:30:00'}
    # probe3 = {'mac':'112233445566','model':'X-300','privateip':'192.168.1.101','lastip':'10.200.10.1:20000','version':'2.00000001','last_heart_time':'2017-04-07 15:30:00'}
    # wait_access_list = [probe1,probe2,probe3]
    #
    # wait_access_list = DataQuery.DQGetAccess("probe")
    #
    # -------------- #
    # 已准入的探针相关 #

    # already_access_list = models.Device.objects.all()
    # DataQuery.DQUpdateAccessed("probe", already_access_list)
    # timenow = timezone.now()

    # for al in already_access_list:
    #     # if type(al.last_heart_time) == unicode:
    #     #     a = int(time.mktime(time.localtime(time.time()))) - int(al.last_heart_time)
    #     #     # print a
    #     #     # print al.last_heart_time
    #     #     # al.last_heart_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(al.last_heart_time)))
    #     #     al.last_heart_time = datetime.datetime.utcfromtimestamp(float(al.last_heart_time))
    #     #     #print al.last_heart_time
    #     # else:
    #     a = int((timenow - al.last_heart_time).total_seconds())
    #         # print type(a)
    #     al.upgrade_button = ''
    #     rss = al.reboot_sign.split(",")
    #     if rss[0] != '0':
    #         if rss[0] == '1':
    #             al.state = _(u"重启")
    #         elif rss[0] == '2':
    #             al.state = _(u"升级")
    #     else:
    #         if a >= 86400:
    #             al.state = _(u"退服")
    #         elif a >= 3600:
    #             al.state = _(u"离线")
    #         elif a >= 900 :
    #             al.state = _(u"超时")
    #         elif a < 900:
    #             al.state = _(u"在线")
    #     if compare_ver(''.join(re.findall("[0-9\.]",''.join(re.findall("(\d\.\d+\.\d+\.\w\d+)",dev_type_change(al.own_model)[:-4])))),al.version) == True:
    #         al.upgrade_button = True
    #         al.upgrade_version = dev_type_change(al.own_model)
    # -------------- #
    try:
        errors = request.session.get('error_message')
        # print errors
        del request.session['error_message']
    except Exception as e:
        pass
    errors_json = json.dumps(errors)
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'probe-list.html',{'errors_json':errors_json,'errors':errors,'admin_error_json':admin_error_json,'admin_error':admin_error,'issue_config_switch':issue_config_switch,'plimit':plimit,'oemlimit':oemlimit})


@login_required
def probe_access(request):
    errors = ''
    # 接受到的mac
    r_mac = request.GET.get('mac')
    # 数据信息
    # data = {'mac':r_mac,'name':'','model':'X-300','sn':'00000001','lastip':'10.200.10.1:20000','privateip':'192.168.1.101','currstanum':'0','totalstanum':'1','currhotnum':'0','totalhotnum':'1','placecode':'0010001','devicecode':'0020002','manufacturer':'百卓','manufacturer_ip':'115.28.241.216','version':'2.0001','last_heart_time':'2017-04-07 15:30:00','upload':'0','download':'0','own_model':'WitFi-P550E'}
    data = DataQuery.DQGetAccessDev("probe", r_mac)
    if models.Device.objects.filter(Q(mac = r_mac)&Q(support_mode = '2')).exists():
        errors = _(u"设备已存在！")
    if len(errors) == 0:
        if models.Device.objects.filter(Q(mac = r_mac)&~Q(support_mode = '2')).exists() :
            models.Device.objects.get(Q(mac = r_mac)&~Q(support_mode = '2')).delete()
            try:
                models.Probe_config.objects.get(mac = r_mac).delete()
            except Exception as e:
                print e
            try:
                models.Probe_audit_basic_status.objects.get(mac = r_mac).delete()
            except Exception as e:
                print e
            try:
                models.Probe_audit_dev_status.objects.get(mac = r_mac).delete()
            except Exception as e:
                print e
            try:
                models.Probe_audit_place_status.objects.get(mac = r_mac).delete()
            except Exception as e:
                print e

        probe = Device()
        probe.mac = data['mac']
        if data['name'] == '':
            probe.name = data['model'] + '_' + data['mac'][-6:]
        else:
            probe.name = data['name']
        probe.model = data['model']
        probe.sn = data['sn']
        probe.lastip = data['lastip']
        probe.privateip = data['privateip']
        probe.version = data['version']
        probe.last_heart_time = data['last_heart_time']
        probe.upload = data['upload']
        probe.download = data['download']
        probe.own_model = data['own_model']
        probe.support_mode = '2'

        # probe.currstanum = data['currstanum']
        # probe.totalstanum = data['totalstanum']
        # probe.currhotnum = data['currhotnum']
        # probe.totalhotnum = data['totalhotnum']
        # probe.placecode = data['placecode']
        # probe.devicecode = data['devicecode']
        # probe.manufacturer = data['manufacturer']
        # probe.manufacturer_ip = data['manufacturer_ip']

        if probe is not None:
            try:
                probe.save()

                p = models.Device.objects.get(mac = r_mac)
                d = Device_probe()
                d.device = p
                d.currstanum = data['currstanum']
                d.totalstanum = data['totalstanum']
                d.currhotnum = data['currhotnum']
                d.totalhotnum = data['totalhotnum']
                d.placecode = data['placecode']
                d.devicecode = data['devicecode']
                d.manufacturer = data['manufacturer']
                d.manufacturer_ip = data['manufacturer_ip']
                d.save()

                DataQuery.DQSetAccessed("probe", r_mac)
                errors = _(u'准入成功！')
                event = {
                    'event_time' : timezone.now(),
                    'event' : 'PROBE_ADMITED_BY_ADMIN',
                    'msg' : 'Device['+probe.mac+'] was admited by Admin['+request.user.username+']',
                    'admin_username' : request.user.username,
                    'probe_mac' : probe.mac,
                }
                try:
                    probe_event_save(event)
                except Exception as e:
                    print e
            except Exception as e:
                errors = _(u'准入失败！')

    request.session['error_message'] = errors
    return redirect('/probe-list/')

def wait_access_ajax(request):
    wait_access_list = ''
    if request.user.administrator_permission > 4 or request.user.administrator_permission == 0 :
        p_all = Device.objects.filter(support_mode = '2')
        wait_access_list = DataQuery.DQGetAccess("probe")
        # ap1 = {'mac':'112233445566','model':'X-300','privateip':'192.168.1.101','lastip':'10.200.10.1:20000','version':'2.00000001','last_heart_time':'2017-04-07 15:30:00'}
        # wait_access_list = [ap1]

    ret = {'data':wait_access_list}
    return JsonResponse(ret,safe = False)

def already_access_ajax(request):
    # print 2131231,request.GET
    draw = int(request.GET.get('draw'))
    start = int(request.GET.get('start'))
    length = int(request.GET.get('length'))
    order_id = request.GET.get('order[0][column]')
    order = request.GET.get('columns['+order_id+'][data]')

    device_type_receive = request.GET.get('device_type')
    dev_online_time = timezone.now() - datetime.timedelta(minutes = 15)
    if device_type_receive == 'online':
        time_condition = Q(last_heart_time__gte = dev_online_time)
    elif device_type_receive == 'offline':
        time_condition = Q(last_heart_time__lt = dev_online_time)
    else:
        time_condition = Q(last_heart_time__gte = dev_online_time)|~Q(last_heart_time__gte = dev_online_time)

    if request.GET.get('order[0][dir]') == 'asc':
        order_type = ''
    elif request.GET.get('order[0][dir]') == 'desc':
        order_type = '-'

    if order == 'state':
        order = 'last_heart_time'

    search_value_old = request.GET.get('search[value]')
    search_value = search_value_old

    # search_value = ''
    # for s in search_value_old:
    #     if s>= u"\u4e00" and s<= u"\u9fa5":
    #         pass
    #     else:
    #         search_value = search_value + s
    if request.user.administrator_permission >= 4 or request.user.administrator_permission == 0 :
        recordsTotal = Device.objects.filter(support_mode = '2').filter(time_condition).count()
        if search_value == '':
            recordsFiltered = recordsTotal
            already_access_list = Device.objects.filter(support_mode = '2').filter(time_condition).order_by(order_type+order)[start:start+length]
        else:
            if re.match(r'[\u4e00 -\u9fa5]+',search_value) == None :
                if search_value in [u'退服',u'离线',u'超时',u'在线']:
                    sv = {
                        u'在线':'online',
                        u'超时':'time out',
                        u'离线':'offline',
                        u'退服':'retired',
                    }
                    already_access_list = models.Device.objects.order_by(order_type+order).filter(support_mode = '2').filter(time_condition).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = models.Device.objects.order_by(order_type+order).filter(support_mode = '2').filter(time_condition).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value)).count()
                else:
                    already_access_list = models.Device.objects.order_by(order_type+order).filter(support_mode = '2').filter(time_condition).filter(Q(state__contains=search_value)|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = models.Device.objects.order_by(order_type+order).filter(support_mode = '2').filter(time_condition).filter(Q(state__contains=search_value)|Q(name__contains=search_value)).count()
            else:
                already_access_list = models.Device.objects.order_by(order_type+order).filter(support_mode = '2').filter(time_condition).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-'))))[start:start+length]
                recordsFiltered = models.Device.objects.order_by(order_type+order).filter(support_mode = '2').filter(time_condition).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-')))).count()
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        recordsTotal = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2')).filter(time_condition).count()
        if search_value == '':
            recordsFiltered = recordsTotal
            already_access_list = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2')).filter(time_condition).order_by(order_type+order)[start:start+length]
        else:
            if re.match(r'[\u4e00 -\u9fa5]+',search_value) == None :
                if search_value in [u'退服',u'离线',u'超时',u'在线']:
                    sv = {
                        u'在线':'online',
                        u'超时':'time out',
                        u'离线':'offline',
                        u'退服':'retired',
                    }
                    already_access_list = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value)).count()
                else:
                    already_access_list = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2')).filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2')).filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value)).count()
            else:
                already_access_list = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-'))))[start:start+length]
                recordsFiltered = models.Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-')))).count()

    DataQuery.DQUpdateAccessed("probe", already_access_list)
    timenow = timezone.now()


    issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
    ret = {'draw':draw,'recordsTotal':recordsTotal,'recordsFiltered':recordsFiltered,'data':[]}
    for al in already_access_list:
        # if type(al.last_heart_time) == unicode:
        #     a = int(time.mktime(time.localtime(time.time()))) - int(al.last_heart_time)
        #     # print a
        #     # print al.last_heart_time
        #     # al.last_heart_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(al.last_heart_time)))
        #     al.last_heart_time = datetime.datetime.utcfromtimestamp(float(al.last_heart_time))
        #     #print al.last_heart_time
        # else:
        a = int((timenow - al.last_heart_time).total_seconds())

        al.last_heart_time = timezone.localtime(al.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")

        al.upgrade_button = ''
        rss = al.reboot_sign.split(",")

        if al.last_heart_time == "1970-1-1 08:00:00":
            al.state = _(u"离线")
        elif a >= 86400:
            al.state = _(u"退服")
        elif a >= 3600:
            al.state = _(u"离线")
        elif a >= 900 :
            al.state = _(u"超时")
        elif a < 900:
            if rss[0] != '0' and rss[0] != '':
                if rss[0] == '1':
                    al.state = _(u"重启")
                elif rss[0] == '2':
                    al.state = _(u"升级")
            else:
                al.state = _(u"在线")

        try:
            if compare_ver(''.join(re.findall("[0-9\.]",''.join(re.findall("(\d\.\d+\.\d+\.\w\d+)",dev_type_change(al.own_model)[:-4])))),al.version) == True:
                al.upgrade_button = True
                al.upgrade_version = dev_type_change(al.own_model)
        except Exception as e:
            print e
            al.upgrade_button = False
            al.upgrade_version = ''

        admin_power_control = ""
        if request.user.administrator_permission >= 5:
            if al.account_group_name == 'admin' or al.account_group_name == '':
                admin_power_control = ""
            else:
                admin_power_control = "ban"


        aall = model_to_dict(al)
        aall['issue_config_switch'] = issue_config_switch
        aall['admin_power_control'] = admin_power_control
        cpuInfo = DataQuery.DQGetAPCpu("probe",al.mac) or ""
        aall['cpu'] = cpuInfo

        try:
            gpn = models.Probe_group.objects.get(pk = aall['group_id']).group_name
        except Exception as e:
            print e
            gpn = "DefaultGroup"
        if aall['account_group_name'] == "":
            acpn = "admin"
        else:
            acpn = aall['account_group_name']
        aall['group_id'] = acpn + "/" + gpn
        ret['data'].append(aall)

    return JsonResponse(ret,safe = False)


@login_required
def probe_reboot(request):
    errors = ''
    r_mac = request.POST.get('mac')
    data = {'action':'reboot','mac':r_mac}
    #执行动作
    try:
        #在此调用函数
        if Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch == "on":
            DataQuery.DQProcess(data)
        errors = _(u'执行成功！')
        event = {
            'event_time' : timezone.now(),
            'event' : 'PROBE_RESTARTED_BY_ADMIN',
            'msg' : 'Device['+r_mac+'] was restarted by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'probe_mac' : r_mac,
        }
        try:
            nu = Device.objects.get(mac = r_mac)
            nu.reboot_sign = '1'+','+str(int(time.mktime(datetime.datetime.now().timetuple())))
            nu.save()
        except Exception as e:
            print e
        try:
            # DataQuery.LOGGING_INIT()
            # logging.info('Device['+r_mac+'] was restarted by Admin['+request.user.username+']')
            # logging.info('GGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGG')
            # print 'Device['+r_mac+'] was restarted by Admin['+request.user.username+']'
            probe_event_save(event)
        except Exception as e:
            print e
    except Exception as e:
        print e
        errors = _(u'执行失败！')
    return JsonResponse(errors,safe=False)


@login_required
def probe_update(request):
    django_settings = Setting.objects.get(SIGN = 1)
    errors = ''
    r_mac = request.POST.get('mac')
    #data = {'action':'upgrade','mac':r_mac, 'param':'http://115.28.241.216/img/witfios3.01.03.r191710.bin'}
    u = upgrade_file(Device.objects.get(mac = r_mac).own_model,DataQuery.ugfl)
    if u == False:
        errors = _(u'没有此型号对应的版本文件')
    else:
        data = {'action':'upgrade','mac':r_mac,'param':django_settings.VERSION_SERVER_URL+"version/byz/"+u}
        #执行动作
        try:
            #在此调用函数
            if Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch == "on":
                DataQuery.DQProcess(data)
            errors = _(u'执行成功！')
            if u != False:
                event = {
                    'event_time' : timezone.now(),
                    'event' : 'PROBE_UPGRADE_BY_ADMIN_TO_VERSION',
                    'msg' : 'Device['+r_mac+'] was upgrade to version['+u+'] by Admin['+request.user.username+']',
                    'admin_username' : request.user.username,
                    'probe_mac' : r_mac,
                }
                try:
                    nu = Device.objects.get(mac = r_mac)
                    nu.reboot_sign = '2'+','+str(int(time.mktime(datetime.datetime.now().timetuple())))
                    nu.save()
                except Exception as e:
                    print e
                try:
                    probe_event_save(event)
                except Exception as e:
                    print e

        except Exception as e:
            print e
            errors = _(u'执行失败！')
    return JsonResponse(errors,safe=False)

@login_required
def probe_del(request):
    errors = ''
    r_mac = request.POST.get('mac')
    data = {'action':'del','mac':r_mac}
    #执行动作
    try:
        #在此调用函数
        models.Device.objects.filter(mac = r_mac).delete()
        try:
            models.Probe_config.objects.filter(mac = r_mac).delete()
        except Exception as e:
            print e
        try:
            models.Probe_audit_basic_status.objects.filter(mac = r_mac).delete()
        except Exception as e:
            print e
        try:
            models.Probe_audit_dev_status.objects.filter(mac = r_mac).delete()
        except Exception as e:
            print e
        try:
            models.Probe_audit_place_status.objects.filter(mac = r_mac).delete()
        except Exception as e:
            print e
        errors = _(u'执行成功！')
        event = {
            'event_time' : timezone.now(),
            'event' : 'PROBE_FORGOTEN_BY_ADMIN',
            'msg' : 'Device['+r_mac+'] was forgoten by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'probe_mac' : r_mac,
        }
        try:
            DataQuery.DQDelAccessed("probe", r_mac)
            if Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch == "on":
                data = {'action':'reset','mac':r_mac}
                DataQuery.DQProcess(data)

            probe_event_save(event)
        except Exception as e:
            print e
    except Exception as e:
        print e
        errors = _(u'执行失败！')
    return JsonResponse(errors,safe = False)

@login_required
def probe_detail(request):
    r_mac = request.GET.get('mac')
    al = models.Device.objects.get(mac = r_mac)
    name = al.name
    vpn = al.vpn
    vpnip = al.vpnip
    try:
        d = models.Probe_config.objects.get(mac = r_mac)
        # print d,1111111111
        ac_address = d.ac_address
        log_address = d.log_address
        ip_model = d.ip_model
        ip_address = d.ip_address
        subnet_mask = d.subnet_mask
        gateway = d.gateway
        preferred_dns = d.preferred_dns
        alternative_dns = d.alternative_dns
        preferred_ntp = d.preferred_ntp
        alternative_ntp = d.alternative_ntp
    except Exception as e:
        print e
        ac_address = ''
        log_address = ''
        ip_model = ''
        ip_address = ''
        subnet_mask = ''
        gateway = ''
        preferred_dns = ''
        alternative_dns = ''
        preferred_ntp = ''
        alternative_ntp = ''

    admin_power_control = ""
    if request.user.administrator_permission >= 5:
        if al.account_group_name == 'admin' or al.account_group_name == '':
            admin_power_control = ""
        else:
            admin_power_control = "ban"
    data = {'mac':r_mac,'sn':'','model':'','version':'','place':'','ip':'','running_time':'','name':name,'last_time':'','cpu':'','memory':'','flash':'','ac_address':ac_address,'log_address':log_address,'ip_model':ip_model,'ip_address':ip_address,'subnet_mask':subnet_mask,'gateway':gateway,'preferred_dns':preferred_dns,'alternative_dns':alternative_dns,'vpn':vpn,'vpnip':vpnip,'preferred_ntp':preferred_ntp,'alternative_ntp':alternative_ntp,}
    DataQuery.DQUpdateAccessedDev('probe', r_mac, data)
    data["admin_power_control"] = admin_power_control
    data["locateState"] = al.locateState

    return JsonResponse(data)

@login_required
def probe_config_modify(request):
    errors = ''
    m_model = request.GET.get('modify_model')
    #print request.GET,request.POST
    if request.method == 'POST':
        mac = request.POST.get('mac')
        if m_model == '1':
            name = request.POST.get('name')
            d = models.Device.objects.get(mac = mac)
            d.name = name
            try:
                d.save()
                errors = _(u'修改成功！')
            except Exception as e:
                errors = _(u'修改失败！')
                print e
        if m_model == '2':
            ac_address = request.POST.get('ac_address')
            try:
                c = models.Probe_config.objects.get(mac = mac)
            except Exception as e:
                print e
                c = Probe_config()
                c.mac = mac

            c.ac_address = ac_address
            try:
                c.save()
                errors = _(u'修改成功！')
                # set ac ip address
                if len(ac_address.split(':')) == 1:
                    ac_address = '{}:20018'.format(ac_address.replace(':',''))
                data = {'action':'acip','mac':mac,'param':ac_address}
                if Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch == "on":
                    DataQuery.DQProcess(data)
            except Exception as e:
                errors = _(u'修改失败！')
                print e
        if m_model == '3':
            log_address = request.POST.get('log_address')
            try:
                a = models.Probe_config.objects.get(mac = mac)
            except Exception as e:
                print e
                a = Probe_config()
                a.mac = mac

            a.log_address = log_address
            try:
                a.save()
                errors = _(u'修改成功！')
                # set log ip address
                if len(log_address) == 0:
                    data = {'action':'logip','mac':mac,'param':'stop'}
                else:
                    data = {'action':'logip','mac':mac,'param':log_address}
                if Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch == "on":
                    DataQuery.DQProcess(data)
            except Exception as e:
                errors = _(u'修改失败！')
                print e
        if m_model == '4':
            ip_model = request.POST.get('ip_model')
            ip_address = request.POST.get('ip_address')
            subnet_mask = request.POST.get('subnet_mask')
            gateway = request.POST.get('gateway')
            preferred_dns = request.POST.get('preferred_dns')
            alternative_dns = request.POST.get('alternative_dns')
            try:
                b = models.Probe_config.objects.get(mac = mac)
            except Exception as e:
                print e
                b = Probe_config()
                b.mac = mac
            if ip_model == None:
                ip_model = ''
            b.ip_model = ip_model
            b.ip_address = ip_address
            b.subnet_mask = subnet_mask
            b.gateway = gateway
            b.preferred_dns = preferred_dns
            b.alternative_dns = alternative_dns
            try:
                b.save()
                errors = _(u'修改成功！')
            except Exception as e:
                errors = _(u'修改失败！')
                print e
        if m_model == '5':
            preferred_ntp = request.POST.get('preferred_ntp')
            alternative_ntp = request.POST.get('alternative_ntp')
            try:
                f = models.Probe_config.objects.get(mac = mac)
            except Exception as e:
                print e
                f = Probe_config()
                f.mac = mac
            f.preferred_ntp = preferred_ntp
            f.alternative_ntp = alternative_ntp
            try:
                f.save()
                if len(preferred_ntp) == 0:
                    preferred_ntp = 'stop'
                if len(alternative_ntp) == 0:
                    alternative_ntp = 'stop'

                ntpstr = preferred_ntp + ';' + alternative_ntp
                data = {'action':'ntp','mac':mac,'param':ntpstr}
                if Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch == "on":
                    DataQuery.DQProcess(data)
                errors = _(u'修改成功！')
            except Exception as e:
                errors = _(u'修改失败！')
                print e

        if m_model == 'user':
            max_user = request.POST.get('max_user')
            de = Device.objects.get(mac = mac)
            if Ap_user_policy_config_extend.objects.filter(device = de).exists():
                au = Ap_user_policy_config_extend.objects.get(device = de)
            else:
                au = Ap_user_policy_config_extend()
                au.device = de
            au.max_user = max_user
            try:
                au.save()
                errors = _(u'修改成功！')
            except Exception as e:
                print e
                errors = _(u'修改失败！')
    return JsonResponse(errors,safe=False)

@login_required
def probe_vpn(request):
    model = request.POST.get('model')
    mac = request.POST.get('mac')
    errors = ''
    if model == 'on':
        serverip = '115.28.241.216'
        timeval = str(int(time.time()))
        m = hashlib.md5()
        m.update(timeval+mac.upper())
        token = m.hexdigest()
        urls = "http://"+serverip+"/vpnserver/vpn_address_asign.php?t="+timeval+"&token="+token+"&mac="+mac.upper()
        result = json.loads(urllib2.urlopen(urls).read())
        try:
            if 'rc' in result['meta'] and result['meta']['rc'] == 'ok' and 'ip' in result['data'] and 'mac' in result['data'] and result['data']['mac'] == mac.upper():

                ip = result['data']['ip']
                #add
                #
                errors = result['meta']['rc']
                ret = {'result':'on','ip':ip}

                event = {
                    'event_time' : timezone.now(),
                    'event' : 'PROBE_START_REMOTE_CONNECT_BY_ADMIN',
                    'msg' : 'Device['+mac+'] was open remote connect by Admin['+request.user.username+']',
                    'admin_username' : request.user.username,
                    'probe_mac' : mac,
                }
                try:
                    probe_event_save(event)
                except Exception as e:
                    print e
            elif 'rc' in result['meta'] and result['meta']['rc'] == 'error':
                errors = result['meta']['rc']
                ret = {'result':'off','ip':''}
        except Exception as e:
            print e
            ret = {'result':'off','ip':''}
            ip = ''

    elif model == 'off':
        ret = {'result':'off','ip':''}
        ip = ''
        event = {
            'event_time' : timezone.now(),
            'event' : 'PROBE_STOP_REMOTE_CONNECT_BY_ADMIN',
            'msg' : 'Device['+mac+'] was close remote connect by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'probe_mac' : mac,
        }
        try:
            probe_event_save(event)
        except Exception as e:
            print e

    try:
        v = models.Device.objects.get(mac = mac)
        v.vpn = ret['result']
        v.vpnip = ip
        v.save()

        #send message to devs
        if model == 'on':
            ip = result['data']['ip']
            param = "115.28.241.216;1194;" + ip
            data = {'action':'vpn','mac':mac,'param':param}
        else:
            data = {'action':'vpn','mac':mac,'param':'stop'}

        if Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch == "on":
            DataQuery.DQProcess(data)
    except Exception as e:
        print e
    return JsonResponse(ret)

@login_required
def probe_audit_status(request):
    mac = request.GET.get('mac')

    try:
        basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
    except Exception as e:
        print e
        basic = {}

    try:
        dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
    except Exception as e:
        print e
        dev = {}

    try:
        place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
    except Exception as e:
        print e
        place = {}

    # if request.GET.get('type') == 'nonoperate':
    #     manu = Public_function.nonoperate_manu
    # else:
        # manu = Public_function.manu
    manu = ''
    # dev_type = Public_function.dev_type
    # place_type = Public_function.place_type
    # nature = Public_function.nature
    # id_type = Public_function.id_type
    dev_type = ''
    place_type = ''
    nature = ''
    id_type = ''

    return JsonResponse({'basic':basic,'dev':dev,'place':place,'manu':manu,'dev_type':dev_type,'place_type':place_type,'nature':nature,'id_type':id_type})


@login_required
def probe_audit_modify(request):
    errors = ''
    if request.method == 'POST':
        mac = request.POST.get('mac')
        audit_corp = request.POST.get('audit_corp')

        ftp_name = request.POST.get('ftp_name')
        ftp_passwd = request.POST.get('ftp_passwd')
        ftp_port = request.POST.get('ftp_port')

        audit_ip = request.POST.get('audit_ip')
        audit_port = request.POST.get('audit_port')
        location_encode = request.POST.get('location_encode')
        device_encode = request.POST.get('device_encode')
        longitude = request.POST.get('longitude')
        latitude =request.POST.get('latitude')
        collection_radius = request.POST.get('collection_radius')
        equipment_type = request.POST.get('equipment_type')
        equipment_name = request.POST.get('equipment_name')
        equipment_address = request.POST.get('equipment_address')
        software_orgcode = request.POST.get('software_orgcode')
        software_orgname = request.POST.get('software_orgname')
        software_address =request.POST.get('software_address')
        contactor =request.POST.get('contactor')
        contactor_tel =request.POST.get('contactor_tel')
        contactor_mail =request.POST.get('contactor_mail')
        place_name = request.POST.get('place_name')
        site_address = request.POST.get('site_address')
        netsite_type = request.POST.get('netsite_type')
        bussiness_nature = request.POST.get('bussiness_nature')
        law_principal_name = request.POST.get('law_principal_name')
        certificate_type = request.POST.get('certificate_type')
        certificate_id =request.POST.get('certificate_id')
        relationship_account =request.POST.get('relationship_account')
        start_time =request.POST.get('start_time')
        end_time =request.POST.get('end_time')
        site_type =request.POST.get('site_type')
        police_station_code =request.POST.get('police_station_code')
        policeName = request.POST.get('policeName')
        districtCode = request.POST.get('districtCode')

        try:
            audit_setting = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
            audit_setting.pop('id')
            audit_setting.pop('mac')
            audit_setting.pop('ssid')
        except Exception as e:
            audit_setting = {'audit_corp':'','ftp_name':'','ftp_passwd':'','ftp_port':'','audit_ip':'','audit_port':'','location_encode':'','device_encode':'','longitude':'','latitude':''}
        try:
            asd = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
            asd.pop('id')
            asd.pop('mac')
            audit_setting.update(asd)
        except Exception as e:
            audit_setting.update({'collection_radius':'','collection_equipment_type':'','collection_equipment_name':'','collection_equipment_address':'','security_software_orgcode':'','security_software_orgname':'','security_software_address':'','contactor':'','contactor_tel':'','contactor_mail':''})
        try:
            asd = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
            asd.pop('id')
            asd.pop('mac')
            audit_setting.update(asd)
        except Exception as e:
            audit_setting.update({'place_name':'','site_address':'','netsite_type':'','bussiness_nature':'','law_principal_name':'','law_principal_certificate_type':'','law_principal_certificate_id':'','relationship_account':'','start_time':'','end_time':'','site_type':'','police_station_code':''})

        audit_setting_now = {'audit_corp':audit_corp,'ftp_name':ftp_name,'ftp_passwd':ftp_passwd,'ftp_port':ftp_port,'audit_ip':audit_ip,'audit_port':audit_port,'location_encode':location_encode,'device_encode':device_encode,'longitude':longitude,'latitude':latitude,'collection_radius':collection_radius,'collection_equipment_type':equipment_type,'collection_equipment_name':equipment_name,'collection_equipment_address':equipment_address,'security_software_orgcode':software_orgcode,'security_software_orgname':software_orgname,'security_software_address':software_address,'contactor':contactor,'contactor_tel':contactor_tel,'contactor_mail':contactor_mail,'place_name':place_name,'site_address':site_address,'netsite_type':netsite_type,'bussiness_nature':bussiness_nature,'law_principal_name':law_principal_name,'law_principal_certificate_type':certificate_type,'law_principal_certificate_id':certificate_id,'relationship_account':relationship_account,'start_time':start_time,'end_time':end_time,'site_type':site_type,'police_station_code':police_station_code}
        for i in audit_setting_now:
            if audit_setting_now[i] == None:
                audit_setting_now[i] = ""
        sign = 0
        print audit_setting,123,audit_setting_now,audit_setting == audit_setting_now
        if audit_setting == audit_setting_now:
            sign = 0
        else:
            sign = 1
        if sign == 1:
            try:
                basic = models.Probe_audit_basic_status.objects.get(mac = mac)
            except Exception as e:
                print e
                basic = Probe_audit_basic_status()
                basic.mac = mac
            if audit_corp == None:
                audit_corp = ''
            basic.audit_corp = audit_corp
            basic.ftp_name = ftp_name
            basic.ftp_passwd = ftp_passwd
            basic.ftp_port = ftp_port
            basic.audit_ip = audit_ip
            basic.audit_port = audit_port
            basic.location_encode = location_encode
            basic.device_encode = device_encode
            basic.longitude = longitude
            basic.latitude = latitude
            try:
                basic.save()
                errors = _(u'修改成功！')
            except Exception as e:
                errors = _(u'修改失败！')
                print e

            try:
                dev = models.Probe_audit_dev_status.objects.get(mac = mac)
            except Exception as e:
                print e
                dev = Probe_audit_dev_status()
                dev.mac = mac
            if equipment_type == None:
                equipment_type = ''
            dev.collection_radius = collection_radius
            dev.collection_equipment_type = equipment_type
            dev.collection_equipment_name = equipment_name
            dev.collection_equipment_address = equipment_address
            dev.security_software_orgcode = software_orgcode
            dev.security_software_orgname = software_orgname
            dev.security_software_address = software_address
            dev.contactor = contactor
            dev.contactor_tel = contactor_tel
            dev.contactor_mail = contactor_mail

            try:
                dev.save()
                errors = _(u'修改成功！')
            except Exception as e:
                errors = _(u'修改失败！')
                print e

            try:
                place = models.Probe_audit_place_status.objects.get(mac = mac)
            except Exception as e:
                print e
                place = Probe_audit_place_status()
                place.mac = mac
            if netsite_type == None:
                netsite_type = ''
            if bussiness_nature == None:
                bussiness_nature = ''
            if certificate_type == None:
                certificate_type = ''
            if site_type == None:
                site_type = ''
            if police_station_code == None:
                police_station_code = ''
            if policeName == None:
                policeName = ''
            if districtCode == None:
                districtCode = ''
            place.place_name = place_name
            place.site_address = site_address
            place.netsite_type = netsite_type
            place.bussiness_nature = bussiness_nature
            place.law_principal_name = law_principal_name
            place.law_principal_certificate_type = certificate_type
            place.law_principal_certificate_id = certificate_id
            place.relationship_account = relationship_account
            place.start_time = start_time
            place.end_time = end_time
            place.site_type = site_type
            place.police_station_code = police_station_code
            place.policeName = policeName
            place.districtCode = districtCode



            try:
                place.save()
                errors = _(u'修改成功！')
            except Exception as e:
                errors = _(u'修改失败！')
                print e

            if Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch == "on":
                #audit config str
                astr = '{"audit":{'
                print mac

                # dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                dev = audit_setting_compare(mac,1,"dev")
                dev.pop('mac')
                dev.pop('id')
                DataQuery.DQClearNullValeForDict(dev)
                dstr = json.dumps(dev, sort_keys=True, ensure_ascii=False)
                dstr = dstr.replace(" ", '')
                #print dstr
                astr = astr + '"device":' + dstr + ','

                # place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                place = audit_setting_compare(mac,1,"place")
                place.pop('mac')
                place.pop('id')
                DataQuery.DQClearNullValeForDict(place)
                pstr = json.dumps(place, sort_keys=True, ensure_ascii=False)
                pstr = pstr.replace(" ", '')
                #print pstr
                astr = astr + '"site":' + pstr + ','

                # basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                basic = audit_setting_compare(mac,1,"basic")
                basic.pop('mac')
                basic.pop('id')
                DataQuery.DQClearNullValeForDict(basic)
                bstr = json.dumps(basic, sort_keys=True, ensure_ascii=False)
                bstr = bstr.replace(" ", '')
                #print bstr
                astr = astr + '"system":' + bstr + "}}"

                # print "ZZZZZZZZZZ"
                # print type(astr.encode("utf-8"))
                print 'astr: {} '.format(astr)
                DataQuery.DQSetUpdateConfig("audit", mac, astr.encode("utf-8"))
                # print astr
                #basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                #dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                #place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
        else:
            errors = _(u'修改成功！')
    return JsonResponse(errors,safe = False)

@login_required
def eventajax(request):
    event_type = request.GET.get('type')
    event_time_type = request.GET.get('time')
    seconds = int(event_time_type) * 3600
    t = timezone.now() - datetime.timedelta(hours = int(event_time_type))

    event = {'data':[]}
    p_list = []
    if request.user.administrator_permission >= 4 or request.user.administrator_permission == 0 :
        probe_event = Probe_event.objects.filter(event_time__gte = t).order_by('-event_time')

        if event_type == "dev":
            # event = {'data':[{'event_time':'2017/05/04 12:05:35','msg':'Device[A4E6B1300005] was onlined.','action':_(u'存档'),'probe_mac':'A4E6B1300005','event':u'','state':state_check('A4E6B1300005')},{'event_time':'2017/05/04 12:04:44','msg':'Device[8c8401163ae0] was offlined.','action':_(u"存档"),'probe_mac':'8c8401163ae0','event':u'PROBE_WAS_OFFLINED','state':state_check('8c8401163ae0')}]}
            p_list = [a.mac for a in Device.objects.filter(support_mode = '2')]
            event = DataQuery.DQGetEventFromRedis(seconds, p_list)
            for evd in event['data']:
                evd['probe_mac'] = evd['probe_mac'].lower()
                evd['state'] = state_check(evd['probe_mac'].lower())

        if event_type == "admin":
            if probe_event.count() != 0:
                for i in probe_event:
                    state = state_check(i.probe_mac)
                    event['data'].append({'event_time':timezone.localtime(i.event_time).strftime("%Y-%m-%d %H:%M:%S"),'msg':i.msg,'probe_mac':i.probe_mac,'admin_username':i.admin_username,'event':i.event,'state':state})
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        p_all = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2'))
        u_all = [a.username for a in NewUser.objects.filter(groupname = request.user.groupname)]
        p_list = [a.mac for a in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '2'))]
        ##########xuyaoxiugai##################
        probe_event = Probe_event.objects.filter(event_time__gte = t).order_by('-event_time')

        if event_type == "dev":
            # event = {'data':[{'event_time':'2017/05/04 12:05:35','msg':'Device[A4E6B1300005] was onlined.','action':_(u'存档'),'probe_mac':'A4E6B1300005','event':u'','state':state_check('A4E6B1300005')},{'event_time':'2017/05/04 12:04:44','msg':'Device[8c8401163ae0] was offlined.','action':_(u"存档"),'probe_mac':'8c8401163ae0','event':u'PROBE_WAS_OFFLINED','state':state_check('8c8401163ae0')}]}
            event = DataQuery.DQGetEventFromRedis(seconds, p_list)
            for evd in event['data']:
                evd['probe_mac'] = evd['probe_mac'].lower()
                evd['state'] = state_check(evd['probe_mac'].lower())

        if event_type == "admin":
            if probe_event.count() != 0:
                for i in probe_event:
                    if i.admin_username in u_all:
                        state = state_check(i.probe_mac)
                        event['data'].append({'event_time':timezone.localtime(i.event_time).strftime("%Y-%m-%d %H:%M:%S"),'msg':i.msg,'probe_mac':i.probe_mac,'admin_username':i.admin_username,'event':i.event,'state':state})

    return JsonResponse(event,safe = False)


def state_check(mac):
    state = ''
    try:
        al = models.Device.objects.get(mac = mac)

        a = int((timezone.now()-al.last_heart_time).total_seconds())
        # print type(a)
        if a >= 86400:
            al.state = _(u"退服")
        elif a >= 3600:
            al.state = _(u"离线")
        elif a >= 900 :
            al.state = _(u"超时")
        elif a < 900:
            al.state = _(u"在线")
        state = al.state
    except Exception as e:
        print e
        state = _(u'已移除')

    return state

@login_required
def device_group(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)
    group = ''
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        group = Probe_group.objects.filter(group_type = '2')
    elif request.user.administrator_permission < 4 and request.user.group_status == 1:
        group = Probe_group.objects.filter(account_group_name = request.user.groupname,group_type = '2')
    for i in group:
        if i.group_name == "DefaultGroup" and i.account_group_name == "admin" and i.group_type == '2':
            i.device_count = Device.objects.filter(support_mode = '2').filter(Q(account_group_name = i.account_group_name,group_id = i.pk) |Q(account_group_name = '',group_id = 0)).count()
        else:
            i.device_count = Device.objects.filter(support_mode = '2').filter(group_id = i.id).filter(account_group_name = i.account_group_name).count()
    issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'probe-group.html',{'admin_error_json':admin_error_json,'admin_error':admin_error,'group':group,'error':error,"error_json":error_json,"issue_config_switch":issue_config_switch,'plimit':plimit,'oemlimit':oemlimit})

@login_required
def device_add_ajax(request):
    device = {'data':[]}
    if request.user.administrator_permission == 6 or request.user.administrator_permission == 5 or request.user.administrator_permission == 0:
        a = Probe_group.objects.get(group_name = 'DefaultGroup',account_group_name = 'admin',group_type = '2')
        device_quaryset = Device.objects.filter(support_mode = '2').filter(Q(account_group_name = '') | Q(account_group_name = 'admin')).filter(Q(group_id = 0) | Q(group_id = a.pk))
    if request.user.administrator_permission == 3 or request.user.administrator_permission == 2 :
        a = Probe_group.objects.get(group_name = request.user.username,account_group_name = request.user.username,group_type = '2')
        device_quaryset = Device.objects.filter(support_mode = '2').filter(account_group_name = request.user.username).filter(Q(group_id = 0) | Q(group_id = a.pk))
    if device_quaryset.count() != 0:
        for i in device_quaryset:
            if i.name != '':
                name = i.name
            else:
                name = i.model + '_' + i.mac[-6:]
            mac = i.mac.upper()[:2] + '-' + i.mac.upper()[2:4] + '-' + i.mac.upper()[4:6] + '-' + i.mac.upper()[6:8] + '-' + i.mac.upper()[8:10] + '-' + i.mac.upper()[10:12]
            device['data'].append({'name':name,'mac':mac,'action':''})
    return JsonResponse(device,safe = False)


@login_required
def add_device_group_ajax(request):
    error = {"error_type":'',"error_msg":""}
    if request.method == "POST":
        add_group_device_list = json.loads(request.POST.get('add_group_device_list'))
        groupname = add_group_device_list['groupname']
        mac = add_group_device_list['mac']
        if request.user.groupname != '' :
            user_area = request.user.groupname
        else:
            user_area = 'admin'
        if Probe_group.objects.filter(account_group_name = user_area).filter(group_name = groupname,group_type = '2').exists():
            error = {"error_type":'failed',"error_msg":_(u"该组已存在")}
        else:
            try:
                pg = Probe_group()
                pg.group_name = groupname
                pg.account_group_name = user_area
                pg.group_type = '2'
                pg.save()
                pgid = Probe_group.objects.get(account_group_name = user_area,group_name = groupname,group_type = '2').pk
                if user_area == 'admin':
                    old_pgid = Probe_group.objects.get(account_group_name = user_area,group_name = 'DefaultGroup',group_type = '2').pk
                else:
                    old_pgid = Probe_group.objects.get(account_group_name = user_area,group_name = user_area,group_type = '2').pk
                if compare_group_audit_setting(pgid,old_pgid):
                    for i in mac:
                        m = change_mac(i)
                        Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid,account_group_name = user_area)
                else:
                    issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                    for i in mac:
                        m = change_mac(i)
                        Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid,account_group_name = user_area)
                        if issue_config_switch == 'on':
                            #audit config str

                            astr = '{"audit":{'
                            print m
                            # dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                            dev = audit_setting_compare(m,1,"dev")
                            dev.pop('mac')
                            dev.pop('id')
                            DataQuery.DQClearNullValeForDict(dev)

                            dstr = json.dumps(dev, sort_keys=True, ensure_ascii=False)
                            dstr = dstr.replace(" ", '')
                            #print dstr
                            astr = astr + '"device":' + dstr + ','

                            # place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                            place = audit_setting_compare(m,1,"place")
                            place.pop('mac')
                            place.pop('id')
                            DataQuery.DQClearNullValeForDict(place)
                            pstr = json.dumps(place, sort_keys=True, ensure_ascii=False)
                            pstr = pstr.replace(" ", '')
                            #print pstr
                            astr = astr + '"site":' + pstr + ','
                            # basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                            basic = audit_setting_compare(m,1,"basic")
                            basic.pop('mac')
                            basic.pop('id')
                            DataQuery.DQClearNullValeForDict(basic)
                            bstr = json.dumps(basic, sort_keys=True, ensure_ascii=False)
                            bstr = bstr.replace(" ", '')
                            #print bstr
                            astr = astr + '"system":' + bstr + "}}"
                            # print "ZZZZZZZZZZ"
                            # print type(astr.encode("utf-8"))
                            DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                            # print astr
                            #basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                            #dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                            #place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))

                error = {"error_type":'success',"error_msg":_(u"添加成功")}
            except Exception as e:
                print e
                error = {"error_type":'failed',"error_msg":_(u"添加失败")}
        if error['error_type'] == "success":
            request.session['error_dict'] = error
        return JsonResponse(error,safe = False)

def change_mac(mac):
    return ''.join(mac.lower().split('-'))


@login_required
def device_group_del(request):
    group_id = request.GET.get('group_id')
    pg = Probe_group.objects.get(pk = group_id)
    try:
        if pg.account_group_name == "admin":
            default_pg = Probe_group.objects.get(account_group_name = 'admin',group_name = "DefaultGroup",group_type = '2')
        else:
            default_pg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '2')
        a = Device.objects.filter(support_mode = '2').filter(group_id = pg.pk)
        mac_list = []
        for i in a:
            mac_list.append(i.mac)
        Device.objects.filter(support_mode = '2').filter(group_id = pg.pk).update(group_id = default_pg.pk)
        if compare_group_audit_setting(pg.pk,default_pg.pk):
            pass
        else:
            issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
            for i in mac_list:
                if issue_config_switch == 'on':
                    #audit config str

                    astr = '{"audit":{'
                    print i

                    # dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                    dev = audit_setting_compare(i,1,"dev")
                    dev.pop('mac')
                    dev.pop('id')
                    DataQuery.DQClearNullValeForDict(dev)
                    dstr = json.dumps(dev, sort_keys=True, ensure_ascii=False)
                    dstr = dstr.replace(" ", '')
                    #print dstr
                    astr = astr + '"device":' + dstr + ','

                    # place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                    place = audit_setting_compare(i,1,"place")
                    place.pop('mac')
                    place.pop('id')
                    DataQuery.DQClearNullValeForDict(place)
                    pstr = json.dumps(place, sort_keys=True, ensure_ascii=False)
                    pstr = pstr.replace(" ", '')
                    #print pstr
                    astr = astr + '"site":' + pstr + ','

                    # basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                    basic = audit_setting_compare(i,1,"basic")
                    basic.pop('mac')
                    basic.pop('id')
                    DataQuery.DQClearNullValeForDict(basic)
                    bstr = json.dumps(basic, sort_keys=True, ensure_ascii=False)
                    bstr = bstr.replace(" ", '')
                    #print bstr
                    astr = astr + '"system":' + bstr + "}}"

                    # print "ZZZZZZZZZZ"
                    # print type(astr.encode("utf-8"))
                    DataQuery.DQSetUpdateConfig("audit", i, astr.encode("utf-8"))

                    # print astr
                    #basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                    #dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                    #place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
        pg.delete()
        error = {"error_type":'success',"error_msg":_(u"删除成功")}
    except Exception as e:
        print e
        error = {"error_type":'failed',"error_msg":_(u"删除失败")}

    request.session['error_dict'] = error
    return redirect('/device/device_group/')

@login_required
def remove_device_ajax(request):
    group_id = request.GET.get("group_id")
    pg = Probe_group.objects.get(pk = group_id)
    device = {'data':[]}
    if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
        device_quaryset = Device.objects.filter(support_mode = '2').filter(Q(account_group_name = pg.account_group_name,group_id = pg.pk) |Q(account_group_name = '',group_id = 0))
    else:
        device_quaryset = Device.objects.filter(support_mode = '2').filter(account_group_name = pg.account_group_name,group_id = pg.pk)
    if device_quaryset.count() != 0:
        for i in device_quaryset:
            if i.name != '':
                name = i.name
            else:
                name = i.model + '_' + i.mac[-6:]
            mac = i.mac.upper()[:2] + '-' + i.mac.upper()[2:4] + '-' + i.mac.upper()[4:6] + '-' + i.mac.upper()[6:8] + '-' + i.mac.upper()[8:10] + '-' + i.mac.upper()[10:12]

            if request.user.administrator_permission == 3 or request.user.administrator_permission == 2 or request.user.administrator_permission == 1:
                if pg.group_name == pg.account_group_name :
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'off'})
                else:
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'on'})
            else:
                if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'off'})
                else:
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'on'})
    return JsonResponse(device,safe = False)

@login_required
def add_device_ajax(request):
    group_id = request.GET.get("group_id")
    pg = Probe_group.objects.get(pk = group_id)
    device = {'data':[]}
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        dpg = Probe_group.objects.get(group_name = "DefaultGroup" ,account_group_name = "admin",group_type = '2')
        if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
            device_quaryset = ""
        elif pg.group_name != "DefaultGroup" and pg.account_group_name == "admin":
            device_quaryset = Device.objects.filter(support_mode = '2').filter(Q(account_group_name = '')|Q(account_group_name = 'admin') , Q(group_id = dpg.pk)|Q(group_id = 0))
        elif pg.account_group_name != "admin" and pg.group_name == pg.account_group_name:
            ptdpg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '2')
            device_quaryset = Device.objects.filter(support_mode = '2').filter(Q(account_group_name = '')|Q(account_group_name = 'admin') , Q(group_id = dpg.pk)|Q(group_id = 0))
        elif pg.account_group_name != "admin" and pg.group_name != pg.account_group_name:
            ptdpg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '2')
            device_quaryset = Device.objects.filter(support_mode = '2').filter(account_group_name = ptdpg.account_group_name,group_id = ptdpg.pk)
    elif request.user.administrator_permission != 0 and request.user.administrator_permission < 4:
        ptdpg = Probe_group.objects.get(Q(account_group_name = pg.account_group_name) & Q(group_name = pg.account_group_name) & Q(group_type = '2'))
        if pg.pk == ptdpg.pk:
            device_quaryset = ""
        elif pg.pk != ptdpg.pk:
            device_quaryset = Device.objects.filter(support_mode = '2').filter(account_group_name = ptdpg.account_group_name,group_id = ptdpg.pk)
    if device_quaryset != "":
        if device_quaryset.count() != 0:
            for i in device_quaryset:
                if i.name != '':
                    name = i.name
                else:
                    name = i.model + '_' + i.mac[-6:]
                mac = i.mac.upper()[:2] + '-' + i.mac.upper()[2:4] + '-' + i.mac.upper()[4:6] + '-' + i.mac.upper()[6:8] + '-' + i.mac.upper()[8:10] + '-' + i.mac.upper()[10:12]
                if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'off'})
                else:
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'on'})
    return JsonResponse(device,safe = False)

@login_required
def device_setting_ajax(request):
    group_id = request.GET.get('group_id')
    pg = Probe_group.objects.get(pk = group_id)
    admin_power_control = ""
    if request.user.administrator_permission >= 5:
        if pg.account_group_name == 'admin' or pg.account_group_name == '':
            admin_power_control = ""
        else:
            admin_power_control = "ban"
    pgd =  model_to_dict(pg)
    pgd['admin_power_control'] = admin_power_control
    return JsonResponse(pgd,safe = False)


@login_required
def modify_device_group_ajax(request):
    sign = 0
    error = {"error_type":'',"error_msg":""}
    if request.method == "POST":
        modify_device = json.loads(request.POST.get('modify_device'))
        group_id = int(modify_device['id'])
        groupname = modify_device['groupname']
        area_name = modify_device['area_name']
        remove_device_mac = modify_device['remove_device_mac']
        add_device_mac = modify_device['add_device_mac']
        setting = modify_device['setting']

        if Probe_group.objects.exclude(pk = group_id).filter(Q(account_group_name = area_name) & Q(group_name = groupname)&Q(group_type = '2')).exists():
            error = {"error_type":'failed',"error_msg":_(u"组名已存在")}
        else:
            try:
                pg = Probe_group.objects.get(pk = group_id , account_group_name = area_name,group_type = '2')
                pg.group_name = groupname

                ss = model_to_dict(pg)
                ss.pop('id')
                ss.pop('group_name')
                ss.pop('account_group_name')
                ss.pop('device_count')
                ss.pop('group_type')
                ss.pop('wlan_count')

                for k,v in setting.items():
                    if v == None:
                        setting[k] = ""
                print setting
                if ss == setting:
                    sign = 0
                    pass
                else:
                    pg.audit_corp = setting['audit_corp']
                    pg.audit_ip = setting['audit_ip']
                    pg.ftp_name = setting['ftp_name']
                    pg.ftp_passwd = setting['ftp_passwd']
                    pg.ftp_port = setting['ftp_port']
                    pg.audit_port = setting['audit_port']
                    pg.location_encode = setting['location_encode']
                    pg.device_encode = setting['device_encode']
                    pg.longitude = setting['longitude']
                    pg.latitude = setting['latitude']
                    pg.collection_radius = setting['collection_radius']
                    pg.collection_equipment_type = setting['collection_equipment_type']
                    pg.collection_equipment_name = setting['collection_equipment_name']
                    pg.collection_equipment_address = setting['collection_equipment_address']
                    pg.security_software_orgcode = setting['security_software_orgcode']
                    pg.security_software_orgname = setting['security_software_orgname']
                    pg.security_software_address = setting['security_software_address']
                    pg.contactor = setting['contactor']
                    pg.contactor_tel = setting['contactor_tel']
                    pg.contactor_mail = setting['contactor_mail']

                    pg.place_name = setting['place_name']
                    pg.site_address = setting['site_address']
                    pg.netsite_type = setting['netsite_type']
                    pg.bussiness_nature = setting['bussiness_nature']
                    pg.law_principal_name = setting['law_principal_name']
                    pg.law_principal_certificate_type = setting['law_principal_certificate_type']
                    pg.law_principal_certificate_id = setting['law_principal_certificate_id']
                    pg.relationship_account = setting['relationship_account']
                    pg.start_time = setting['start_time']
                    pg.end_time = setting['end_time']
                    pg.site_type = setting['site_type']
                    pg.police_station_code = setting['police_station_code']

                    pg.ssid = setting['ssid']
                    sign = 1
                pg.save()


                if remove_device_mac == {}:
                    pass
                else:
                    if request.user.administrator_permission == 1 or request.user.administrator_permission == 2 or request.user.administrator_permission == 3:
                        pgid = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '2').pk
                        if compare_group_audit_setting(pg.pk,pgid):
                            for i in remove_device_mac:
                                m = change_mac(i)
                                Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid)
                        else:
                            issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                            for i in remove_device_mac:
                                m = change_mac(i)
                                Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid)
                                if issue_config_switch == "on":
                                    #audit config str

                                    astr = '{"audit":{'
                                    print m

                                    # dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                                    dev = audit_setting_compare(m,1,"dev")
                                    dev.pop('mac')
                                    dev.pop('id')
                                    DataQuery.DQClearNullValeForDict(dev)
                                    dstr = json.dumps(dev, sort_keys=True, ensure_ascii=False)
                                    dstr = dstr.replace(" ", '')
                                    #print dstr
                                    astr = astr + '"device":' + dstr + ','

                                    # place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                                    place = audit_setting_compare(m,1,"place")
                                    place.pop('mac')
                                    place.pop('id')
                                    DataQuery.DQClearNullValeForDict(place)
                                    pstr = json.dumps(place, sort_keys=True, ensure_ascii=False)
                                    pstr = pstr.replace(" ", '')
                                    #print pstr
                                    astr = astr + '"site":' + pstr + ','

                                    # basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                                    basic = audit_setting_compare(m,1,"basic")
                                    basic.pop('mac')
                                    basic.pop('id')
                                    DataQuery.DQClearNullValeForDict(basic)
                                    bstr = json.dumps(basic, sort_keys=True, ensure_ascii=False)
                                    bstr = bstr.replace(" ", '')
                                    #print bstr
                                    astr = astr + '"system":' + bstr + "}}"

                                    # print "ZZZZZZZZZZ"
                                    # print 123
                                    # print type(astr.encode("utf-8"))
                                    DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                                    print 'astr: {} '.format(astr)
                                    # print astr
                                    #basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                                    #dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                                    #place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                    else:
                        if pg.account_group_name == "admin" or pg.group_name == pg.account_group_name:
                            pgid = Probe_group.objects.get(account_group_name = "admin",group_name = "DefaultGroup",group_type = '2').pk
                            if compare_group_audit_setting(pg.pk,pgid):
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid,account_group_name = "admin")
                            else:
                                issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid,account_group_name = "admin")
                                    if issue_config_switch == "on":
                                        #audit config str

                                        astr = '{"audit":{'
                                        print m

                                        # dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                                        dev = audit_setting_compare(m,1,"dev")
                                        dev.pop('mac')
                                        dev.pop('id')
                                        DataQuery.DQClearNullValeForDict(dev)
                                        dstr = json.dumps(dev, sort_keys=True, ensure_ascii=False)
                                        dstr = dstr.replace(" ", '')
                                        #print dstr
                                        astr = astr + '"device":' + dstr + ','

                                        # place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                                        place = audit_setting_compare(m,1,"place")
                                        place.pop('mac')
                                        place.pop('id')
                                        DataQuery.DQClearNullValeForDict(place)
                                        pstr = json.dumps(place, sort_keys=True, ensure_ascii=False)
                                        pstr = pstr.replace(" ", '')
                                        #print pstr
                                        astr = astr + '"site":' + pstr + ','

                                        # basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                                        basic = audit_setting_compare(m,1,"basic")
                                        basic.pop('mac')
                                        basic.pop('id')
                                        DataQuery.DQClearNullValeForDict(basic)
                                        bstr = json.dumps(basic, sort_keys=True, ensure_ascii=False)
                                        bstr = bstr.replace(" ", '')
                                        #print bstr
                                        astr = astr + '"system":' + bstr + "}}"

                                        # print "ZZZZZZZZZZ"
                                        # print 456
                                        # print type(astr.encode("utf-8"))
                                        DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                                        print 'astr: {} '.format(astr)
                                        # print astr
                                        #basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                                        #dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                                        #place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                        else:
                            pgid = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '2').pk
                            if compare_group_audit_setting(pg.pk,pgid):
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid)
                            else:
                                issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid)
                                    if issue_config_switch == "on":
                                        #audit config str

                                        astr = '{"audit":{'
                                        print m

                                        # dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                                        dev = audit_setting_compare(m,1,"dev")
                                        dev.pop('mac')
                                        dev.pop('id')
                                        DataQuery.DQClearNullValeForDict(dev)
                                        dstr = json.dumps(dev, sort_keys=True, ensure_ascii=False)
                                        dstr = dstr.replace(" ", '')
                                        #print dstr
                                        astr = astr + '"device":' + dstr + ','

                                        # place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                                        place = audit_setting_compare(m,1,"place")
                                        place.pop('mac')
                                        place.pop('id')
                                        DataQuery.DQClearNullValeForDict(place)
                                        pstr = json.dumps(place, sort_keys=True, ensure_ascii=False)
                                        pstr = pstr.replace(" ", '')
                                        #print pstr
                                        astr = astr + '"site":' + pstr + ','

                                        # basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                                        basic = audit_setting_compare(m,1,"basic")
                                        basic.pop('mac')
                                        basic.pop('id')
                                        DataQuery.DQClearNullValeForDict(basic)
                                        bstr = json.dumps(basic, sort_keys=True, ensure_ascii=False)
                                        bstr = bstr.replace(" ", '')
                                        #print bstr
                                        astr = astr + '"system":' + bstr + "}}"

                                        # print "ZZZZZZZZZZ"
                                        # print 789
                                        # print type(astr.encode("utf-8"))
                                        DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                                        print 'astr: {} '.format(astr)
                                        # print astr
                                        #basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                                        #dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                                        #place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))

                if add_device_mac == {}:
                    pass
                else:
                    issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                    for i in add_device_mac:
                        m = change_mac(i)
                        p = Device.objects.get(mac = m).group_id
                        Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pg.pk,account_group_name = pg.account_group_name)
                        if compare_group_audit_setting(pg.pk,p):
                            pass
                        else:
                            if issue_config_switch == 'on':
                                #audit config str
                                astr = '{"audit":{'
                                print m

                                # dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                                dev = audit_setting_compare(m,1,"dev")
                                dev.pop('mac')
                                dev.pop('id')
                                DataQuery.DQClearNullValeForDict(dev)
                                dstr = json.dumps(dev, sort_keys=True, ensure_ascii=False)
                                dstr = dstr.replace(" ", '')
                                #print dstr
                                astr = astr + '"device":' + dstr + ','

                                # place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                                place = audit_setting_compare(m,1,"place")
                                place.pop('mac')
                                place.pop('id')
                                DataQuery.DQClearNullValeForDict(place)
                                pstr = json.dumps(place, sort_keys=True, ensure_ascii=False)
                                pstr = pstr.replace(" ", '')
                                #print pstr
                                astr = astr + '"site":' + pstr + ','

                                # basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                                basic = audit_setting_compare(m,1,"basic")
                                basic.pop('mac')
                                basic.pop('id')
                                DataQuery.DQClearNullValeForDict(basic)
                                bstr = json.dumps(basic, sort_keys=True, ensure_ascii=False)
                                bstr = bstr.replace(" ", '')
                                #print bstr
                                astr = astr + '"system":' + bstr + "}}"

                                # print "ZZZZZZZZZZ"
                                # print 101
                                # print type(astr.encode("utf-8"))
                                DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                                print 'astr: {} '.format(astr)
                                # print astr
                                #basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                                #dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                                #place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                if sign == 0:
                    pass
                else:
                    issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                    for i in Device.objects.filter(support_mode = '2').filter(group_id = pg.pk):
                        m = change_mac(i.mac)

                        if issue_config_switch == "on":
                            #audit config str
                            astr = '{"audit":{'
                            print m

                            # dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                            dev = audit_setting_compare(m,1,"dev")
                            dev.pop('mac')
                            dev.pop('id')
                            DataQuery.DQClearNullValeForDict(dev)
                            dstr = json.dumps(dev, sort_keys=True, ensure_ascii=False)
                            dstr = dstr.replace(" ", '')
                            #print dstr
                            astr = astr + '"device":' + dstr + ','

                            # place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                            place = audit_setting_compare(m,1,"place")
                            place.pop('mac')
                            place.pop('id')
                            DataQuery.DQClearNullValeForDict(place)
                            pstr = json.dumps(place, sort_keys=True, ensure_ascii=False)
                            pstr = pstr.replace(" ", '')
                            #print pstr
                            astr = astr + '"site":' + pstr + ','

                            # basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                            basic = audit_setting_compare(m,1,"basic")
                            basic.pop('mac')
                            basic.pop('id')
                            DataQuery.DQClearNullValeForDict(basic)
                            bstr = json.dumps(basic, sort_keys=True, ensure_ascii=False)
                            bstr = bstr.replace(" ", '')
                            #print bstr
                            astr = astr + '"system":' + bstr + "}}"

                            # print "ZZZZZZZZZZ"
                            # print 102
                            # print type(astr.encode("utf-8"))
                            DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                            print 'astr: {} '.format(astr)
                            # print astr
                            #basic = model_to_dict(models.Probe_audit_basic_status.objects.get(mac = mac))
                            #dev = model_to_dict(models.Probe_audit_dev_status.objects.get(mac = mac))
                            #place = model_to_dict(models.Probe_audit_place_status.objects.get(mac = mac))
                error = {"error_type":'success',"error_msg":_(u"编辑成功")}
                if pg.account_group_name == "admin" and pg.group_name == 'DefaultGroup':
                    default_config_issued.delay()
            except Exception as e:
                print e
                error = {"error_type":'failed',"error_msg":_(u"编辑失败")}
        if error['error_type'] == "success":
            request.session['error_dict'] = error
        return JsonResponse(error,safe = False)



def audit_setting_compare(mac,model_type,kind):
    model_type = int(model_type)
    audit_setting = ''
    dev = Device.objects.get(mac = mac)
    if model_type == 1:
        if dev.group_id == 0:
            if dev.support_mode == '3':
                group = Probe_group.objects.get(group_name = 'DefaultGroup',account_group_name = 'admin',group_type = '3')
            else:
                group = Probe_group.objects.get(group_name = 'DefaultGroup',account_group_name = 'admin',group_type = '2')
        else:
            group = Probe_group.objects.get(pk = dev.group_id)

        if kind == 'dev':
            try:
                audit_setting = models.Probe_audit_dev_status.objects.get(mac = mac)

                audit_setting.collection_radius = group.collection_radius if audit_setting.collection_radius == '' else audit_setting.collection_radius
                audit_setting.collection_equipment_type = group.collection_equipment_type if audit_setting.collection_equipment_type == '' else audit_setting.collection_equipment_type
                audit_setting.collection_equipment_name = group.collection_equipment_name if audit_setting.collection_equipment_name == '' else audit_setting.collection_equipment_name
                audit_setting.collection_equipment_address = group.collection_equipment_address if audit_setting.collection_equipment_address == '' else audit_setting.collection_equipment_address
                audit_setting.security_software_orgcode = group.security_software_orgcode if audit_setting.security_software_orgcode == '' else audit_setting.security_software_orgcode
                audit_setting.security_software_orgname = group.security_software_orgname if audit_setting.security_software_orgname == '' else audit_setting.security_software_orgname
                audit_setting.security_software_address = group.security_software_address if audit_setting.security_software_address == '' else audit_setting.security_software_address
                audit_setting.contactor = group.contactor if audit_setting.contactor == '' else audit_setting.contactor
                audit_setting.contactor_tel = group.contactor_tel if audit_setting.contactor_tel == '' else audit_setting.contactor_tel
                audit_setting.contactor_mail = group.contactor_mail if audit_setting.contactor_mail == '' else audit_setting.contactor_mail
                auds = model_to_dict(audit_setting)
                for k in auds:
                    if auds[k] == None:
                        auds[k] = ""
                return auds
            except Exception as e:
                print e
                audit_setting = {'collection_radius':'','collection_equipment_type':'','collection_equipment_name':'','collection_equipment_address':'','security_software_orgcode':'','security_software_orgname':'','security_software_address':'','contactor':'','contactor_tel':'','contactor_mail':'','mac':'','id':''}

                audit_setting['collection_radius'] = group.collection_radius if audit_setting['collection_radius'] == '' else audit_setting['collection_radius']
                audit_setting['collection_equipment_type'] = group.collection_equipment_type if audit_setting['collection_equipment_type'] == '' else audit_setting['collection_equipment_type']
                audit_setting['collection_equipment_name'] = group.collection_equipment_name if audit_setting['collection_equipment_name'] == '' else audit_setting['collection_equipment_name']
                audit_setting['collection_equipment_address'] = group.collection_equipment_address if audit_setting['collection_equipment_address'] == '' else audit_setting['collection_equipment_address']
                audit_setting['security_software_orgcode'] = group.security_software_orgcode if audit_setting['security_software_orgcode'] == '' else audit_setting['security_software_orgcode']
                audit_setting['security_software_orgname'] = group.security_software_orgname if audit_setting['security_software_orgname'] == '' else audit_setting['security_software_orgname']
                audit_setting['security_software_address'] = group.security_software_address if audit_setting['security_software_address'] == '' else audit_setting['security_software_address']
                audit_setting['contactor'] = group.contactor if audit_setting['contactor'] == '' else audit_setting['contactor']
                audit_setting['contactor_tel'] = group.contactor_tel if audit_setting['contactor_tel'] == '' else audit_setting['contactor_tel']
                audit_setting['contactor_mail'] = group.contactor_mail if audit_setting['contactor_mail'] == '' else audit_setting['contactor_mail']

                auds = audit_setting
                for k in auds:
                    if auds[k] == None:
                        auds[k] = ""
                return auds


        if kind == 'place':
            print 'xt1tx1'
            if models.Probe_audit_basic_status.objects.filter(mac = mac).exists():
                place_audit_corp = models.Probe_audit_basic_status.objects.get(mac = mac).audit_corp
            else:
                place_audit_corp = group.audit_corp
            print 'xt2xt2',place_audit_corp
            try:
                audit_setting = models.Probe_audit_place_status.objects.get(mac = mac)

                audit_setting.place_name = group.place_name if audit_setting.place_name == '' else audit_setting.place_name
                audit_setting.site_address = group.site_address if audit_setting.site_address == '' else audit_setting.site_address
                audit_setting.netsite_type = group.netsite_type if audit_setting.netsite_type == '' else audit_setting.netsite_type
                audit_setting.bussiness_nature = group.bussiness_nature if audit_setting.bussiness_nature == '' else audit_setting.bussiness_nature
                audit_setting.law_principal_name = group.law_principal_name if audit_setting.law_principal_name == '' else audit_setting.law_principal_name
                audit_setting.law_principal_certificate_type = group.law_principal_certificate_type if audit_setting.law_principal_certificate_type == '' else audit_setting.law_principal_certificate_type
                audit_setting.law_principal_certificate_id = group.law_principal_certificate_id if audit_setting.law_principal_certificate_id == '' else audit_setting.law_principal_certificate_id
                audit_setting.relationship_account = group.relationship_account if audit_setting.relationship_account == '' else audit_setting.relationship_account
                audit_setting.start_time = group.start_time if audit_setting.start_time == '' else audit_setting.start_time
                audit_setting.end_time = group.end_time if audit_setting.end_time == '' else audit_setting.end_time

                audit_setting.site_type = group.site_type if audit_setting.site_type == '' else audit_setting.site_type
                audit_setting.police_station_code = group.police_station_code if audit_setting.police_station_code == '' else audit_setting.police_station_code

                a_s = model_to_dict(audit_setting)
                print 'tttttt11111',a_s
                if place_audit_corp == '27':
                    a_s['policestation'] = a_s['police_station_code']
                    a_s['area'] = a_s['site_type']
                if place_audit_corp == '29':
                    a_s['policestation'] = a_s['police_station_code']
                del a_s['site_type']
                del a_s['police_station_code']

                auds = a_s
                for k in auds:
                    if auds[k] == None:
                        auds[k] = ""
                print 'tttttt22222',a_s
                return auds

            except Exception as e:
                print e,
                print 'xt3xt3'
                audit_setting = {'place_name':'','site_address':'','netsite_type':'','bussiness_nature':'','law_principal_name':'','law_principal_certificate_type':'','law_principal_certificate_id':'','relationship_account':'','start_time':'','end_time':'','mac':'','id':'','site_type':'','police_station_code':''}

                audit_setting['place_name'] = group.place_name if audit_setting['place_name'] == '' else audit_setting['place_name']
                audit_setting['site_address'] = group.site_address if audit_setting['site_address'] == '' else audit_setting['site_address']
                audit_setting['netsite_type'] = group.netsite_type if audit_setting['netsite_type'] == '' else audit_setting['netsite_type']
                audit_setting['bussiness_nature'] = group.bussiness_nature if audit_setting['bussiness_nature'] == '' else audit_setting['bussiness_nature']
                audit_setting['law_principal_name'] = group.law_principal_name if audit_setting['law_principal_name'] == '' else audit_setting['law_principal_name']
                audit_setting['law_principal_certificate_type'] = group.law_principal_certificate_type if audit_setting['law_principal_certificate_type'] == '' else audit_setting['law_principal_certificate_type']
                audit_setting['law_principal_certificate_id'] = group.law_principal_certificate_id if audit_setting['law_principal_certificate_id'] == '' else audit_setting['law_principal_certificate_id']
                audit_setting['relationship_account'] = group.relationship_account if audit_setting['relationship_account'] == '' else audit_setting['relationship_account']
                audit_setting['start_time'] = group.start_time if audit_setting['start_time'] == '' else audit_setting['start_time']
                audit_setting['end_time'] = group.end_time if audit_setting['end_time'] == '' else audit_setting['end_time']
                audit_setting['site_type'] = group.site_type if audit_setting['site_type'] == '' else audit_setting['site_type']
                audit_setting['police_station_code'] = group.police_station_code if audit_setting['police_station_code'] == '' else audit_setting['police_station_code']

                if place_audit_corp == '27':
                    audit_setting['policestation'] = audit_setting['police_station_code']
                    audit_setting['area'] = audit_setting['site_type']
                if place_audit_corp == '29':
                    audit_setting['policestation'] = audit_setting['police_station_code']
                del audit_setting['site_type']
                del audit_setting['police_station_code']

                auds = audit_setting
                for k in auds:
                    if auds[k] == None:
                        auds[k] = ""
                return auds


        if kind == 'basic':
            try:
                audit_setting = models.Probe_audit_basic_status.objects.get(mac = mac)

                audit_setting.audit_corp = group.audit_corp if audit_setting.audit_corp == '' else audit_setting.audit_corp

                audit_setting.ftp_name = group.ftp_name if audit_setting.ftp_name == '' else audit_setting.ftp_name
                audit_setting.ftp_passwd = group.ftp_passwd if audit_setting.ftp_passwd == '' else audit_setting.ftp_passwd
                audit_setting.ftp_port = group.ftp_port if audit_setting.ftp_port == '' else audit_setting.ftp_port

                audit_setting.audit_ip = group.audit_ip if audit_setting.audit_ip == '' else audit_setting.audit_ip
                audit_setting.audit_port = group.audit_port if audit_setting.audit_port == '' else audit_setting.audit_port
                audit_setting.location_encode = group.location_encode if audit_setting.location_encode == '' else audit_setting.location_encode
                audit_setting.device_encode = group.device_encode if audit_setting.device_encode == '' else audit_setting.device_encode
                audit_setting.longitude = group.longitude if audit_setting.longitude == '' else audit_setting.longitude
                audit_setting.latitude = group.latitude if audit_setting.latitude == '' else audit_setting.latitude
                audit_setting.ssid = json.loads(group.ssid)
                # print audit_setting.ssid,'QWEQWEQWEWQEQ'
                a_s = model_to_dict(audit_setting)
                if a_s['ssid'] == "":
                    a_s['ssid'] = []
                if dev.own_model != 'WitMAX-P550E-L':
                    del a_s['ssid']

                # renzixing,wangbo,chongqingaisi,byzoro feijing
                if audit_setting.audit_corp == '50':
                    pass
                # elif audit_setting.audit_corp == '3' or audit_setting.audit_corp == '15' or  audit_setting.audit_corp == '22' or  audit_setting.audit_corp == '27' or  audit_setting.audit_corp == '28':
                elif audit_setting.audit_corp in Public_function.auditCropLimitList:
                    del a_s['ftp_port']
                else:
                    del a_s['ftp_port']
                    del a_s['ftp_name']
                    del a_s['ftp_passwd']

                for k in a_s:
                    if a_s[k] == None:
                        a_s[k] = ""
                print a_s
                return a_s

                # return model_to_dict(audit_setting)

            except Exception as e:
                print e
                audit_setting = {'audit_corp':'','ftp_name':'', 'ftp_passwd':'', 'ftp_port':'', 'audit_ip':'','audit_port':'','location_encode':'','device_encode':'','longitude':'','latitude':'','mac':'','id':'','ssid':''}

                audit_setting['audit_corp'] = group.audit_corp if audit_setting['audit_corp'] == '' else audit_setting['audit_corp']
                # renzixing,wangbo,chongqingaisi,byzoro feijing
                # if audit_setting['audit_corp'] == '3' or audit_setting['audit_corp'] == '15' ||  audit_setting['audit_corp'] == '22' || audit_setting['audit_corp'] == '50':
                audit_setting['ftp_name'] = group.ftp_name if audit_setting['ftp_name'] == '' else audit_setting['ftp_name']
                audit_setting['ftp_passwd'] = group.ftp_passwd if audit_setting['ftp_passwd'] == '' else audit_setting['ftp_passwd']
                audit_setting['ftp_port'] = group.ftp_port if audit_setting['ftp_port'] == '' else audit_setting['ftp_port']

                audit_setting['audit_ip'] = group.audit_ip if audit_setting['audit_ip'] == '' else audit_setting['audit_ip']
                audit_setting['audit_port'] = group.audit_port if audit_setting['audit_port'] == '' else audit_setting['audit_port']
                audit_setting['location_encode'] = group.location_encode if audit_setting['location_encode'] == '' else audit_setting['location_encode']
                audit_setting['device_encode'] = group.device_encode if audit_setting['device_encode'] == '' else audit_setting['device_encode']
                audit_setting['longitude'] = group.longitude if audit_setting['longitude'] == '' else audit_setting['longitude']
                audit_setting['latitude'] = group.latitude if audit_setting['latitude'] == '' else audit_setting['latitude']
                audit_setting['ssid'] = json.loads(group.ssid)
                # print audit_setting['ssid'],'ASDASDADS'

                if audit_setting['ssid'] == "":
                    audit_setting['ssid'] = []
                if dev.own_model != 'WitMAX-P550E-L':
                    del audit_setting['ssid']

                # renzixing,wangbo,chongqingaisi,byzoro feijing
                if audit_setting['audit_corp'] == '50':
                    pass
                # elif audit_setting['audit_corp'] == '3' or audit_setting['audit_corp'] == '15' or audit_setting['audit_corp'] == '22' or audit_setting['audit_corp'] == '27' or audit_setting['audit_corp'] == '28':
                elif audit_setting['audit_corp'] in Public_function.auditCropLimitList:
                    del audit_setting['ftp_port']
                else:
                    del audit_setting['ftp_name']
                    del audit_setting['ftp_passwd']
                    del audit_setting['ftp_port']

                auds = audit_setting
                for k in auds:
                    if auds[k] == None:
                        auds[k] = ""
                print auds
                return auds
                # return audit_setting



def compare_group_audit_setting(gp1,gp2):
    try:
        g1 = model_to_dict(Probe_group.objects.get(pk = int(gp1)))
        g2 = model_to_dict(Probe_group.objects.get(pk = int(gp2)))
        g1.pop('id')
        g1.pop('group_name')
        g1.pop('account_group_name')
        g1.pop('device_count')
        g1.pop('group_type')
        g2.pop('id')
        g2.pop('group_name')
        g2.pop('account_group_name')
        g2.pop('device_count')
        g2.pop('group_type')

        for i in g1:
            if g1[i] == None:
                g1[i] = ""
        for i in g2:
            if g2[i] == None:
                g2[i] = ""
        return g1 == g2
    except Exception as e:
        print e
        return False

def exportdevicebutton_ajax(request):
    # mode = request.POST.get('mode')
    mode = request.GET.get('mode')
    # all device 1, online 2, other 3,
    extype = request.GET.get('extype')

    # admin
    if request.user.administrator_permission >= 4 or request.user.administrator_permission == 0 :
        already_access_list = Device.objects.filter(support_mode = mode)
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        already_access_list = Device.objects.filter(support_mode = mode,account_group_name = request.user.groupname)

    timenow = timezone.now()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Devices Status"
    sheet['A1'].value = "ID"
    sheet['B1'].value = "Device Name"
    sheet['C1'].value = "Device MAC"
    sheet['D1'].value = "Device SN"
    sheet['E1'].value = "Group Name"
    sheet['F1'].value = "IP Address"
    sheet['G1'].value = "Status"
    sheet['H1'].value = "Version"

    excel_row = 1
    for al in already_access_list:
        a = int((timenow - al.last_heart_time).total_seconds())

        al.last_heart_time = timezone.localtime(al.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")

        if al.last_heart_time == "1970-1-1 08:00:00":
            al.state = _(u"离线")
        elif a >= 86400:
            al.state = _(u"退服")
        elif a >= 3600:
            al.state = _(u"离线")
        elif a >= 900 :
            al.state = _(u"超时")
        elif a < 900:
            al.state = _(u"在线")

        if extype == '1':
            pass
        else:
            if extype == '2' and al.state == _(u"在线"):
                # print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ ONLINE", al.name, al.mac, al.account_group_name, al.lastip, al.state, al.version
                pass
            else:
                if extype == '3' and al.state != _(u"在线"):
                    # print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ ONTHER", al.name, al.mac, al.account_group_name, al.lastip, al.state, al.version
                    pass
                else:
                    continue

        excel_row += 1
        sheet["A%d" % (excel_row)].value = excel_row - 1
        sheet["B%d" % (excel_row)].value = al.name
        sheet["C%d" % (excel_row)].value = al.mac
        sheet["D%d" % (excel_row)].value = al.sn
        if al.account_group_name == "":
            sheet["E%d" % (excel_row)].value = "admin"
        else:
            sheet["E%d" % (excel_row)].value = al.account_group_name
        sheet["F%d" % (excel_row)].value = al.lastip
        sheet["G%d" % (excel_row)].value = al.state
        sheet["H%d" % (excel_row)].value = al.version

    filename = "./tmp.xlsx"

    exist_file = os.path.exists("./tmp.xlsx")
    if exist_file:
        os.remove(filename)
    wb.save(filename)

    print "5555555555555555555555555555555555555555", mode
    wrapper = FileWrapper(open(filename, 'rb'))
    response = HttpResponse(wrapper, content_type='application/vnd.ms-excel')
    response['Content-Length'] = os.path.getsize(filename)
    response['Content-Disposition'] = 'attachment; filename=Devices.xlsx'
    return response

def adddevicebutton_ajax(request):
    name = request.POST.get('name')
    sn = request.POST.get('sn')
    mac = request.POST.get('mac').lower()
    mode = request.POST.get('mode')

    power = request.user.administrator_permission
    error = {'sign':'','mes':'','reason':''}
    gp_id = Probe_group.objects.get(group_name = 'DefaultGroup',account_group_name = 'admin',group_type = mode).pk
    if power == 6 or power == 5:
        if Device.objects.filter(Q(mac = mac) | Q(sn = sn)).exists() :
            error['sign'] = 'False'
            error['mes'] = _(u'设备已存在')
        else:
            try:
                dev = Device()
                dev.name = name
                dev.sn = sn
                dev.mac = mac
                dev.group_id = gp_id
                dev.account_group_name = 'admin'
                dev.support_mode = mode
                dev.last_heart_time = '1970-1-1 08:00:00'
                dev.save()
                #通知redis
                DataQuery.DQimportdev(mac)
                error['sign'] = 'True'
                error['mes'] = _(u'添加成功')
                if int(mode) == 1 or int(mode) == 3:
                    accessed_device2authpuppy(mac)
            except Exception as e:
                print e
                error['sign'] = 'False'
                error['mes'] = _(u'添加失败')
                error['reason'] = e
    else:
        if Device.objects.filter(Q(mac = mac) & Q(sn = sn) & Q(support_mode = mode) , Q(account_group_name = 'admin')|Q(account_group_name = ''), Q(group_id = gp_id)|Q(group_id = 0)).exists():

            group_id = Probe_group.objects.get(Q(group_name = request.user.groupname) & Q(account_group_name = request.user.groupname) & Q(group_type = mode)).pk
            try:
                dev = Device.objects.get(Q(mac = mac) & Q(sn = sn) & Q(support_mode = mode) , Q(account_group_name = 'admin')|Q(account_group_name = ''), Q(group_id = gp_id)|Q(group_id = 0))
                dev.name = name
                dev.account_group_name = request.user.groupname
                dev.group_id = group_id
                dev.save()
                error['sign'] = 'True'
                error['mes'] = _(u'添加成功')

                if int(mode) == 1 or int(mode) == 3:
                    update_device2authpuppy(mac,request.user.groupname)
            except Exception as e:
                print e
                error['sign'] = 'False'
                error['mes'] = _(u'添加失败')
                error['reason'] = e

        else:
            error['sign'] = 'False'
            error['mes'] = _(u'设备不存在')

    return JsonResponse(error)

def importdevice(request):
    mode = request.GET.get('mode')

    upfile =request.FILES.get("file", None)    # 获取上传的文件，如果没有文件，则默认为None
    error = ""
    right = 0
    wrong = 0
    if not upfile:
        error = _(u"没有选择文件")
    else:
        try:
            if os.path.exists("./statics/upload/"+ request.user.groupname):
                pass
            else:
                os.makedirs("./statics/upload/"+ request.user.groupname)

            destination = open("./statics/upload/"+ request.user.groupname +"/"+upfile.name,'wb+')    # 打开特定的文件进行二进制的写操作
            try:
                for chunk in upfile.chunks():      # 分块写入文件
                    destination.write(chunk)
            except Exception as e:
                print e
                error = _(u"文件写入失败")
            finally:
                destination.close()
        except Exception as e:
            print e
            error = _(u"上传文件失败")
        if error == "":
            right = 0
            wrong = 0
            furl = ''
            try:
                wb = load_workbook(filename="./statics/upload/"+ request.user.groupname +"/"+upfile.name)
                ws = wb.active
                rows = ws.rows

                writewb = Workbook()
                writews = writewb.active
                writews.append(['mac','error'])


                power = request.user.administrator_permission
                gp_id = Probe_group.objects.get(group_name = 'DefaultGroup',account_group_name = 'admin',group_type = mode).pk
                if (ws.cell('A1').value == "Device Name" and ws.cell('B1').value == "Device MAC" and ws.cell('C1').value == "Device SN") or (ws.cell('B1').value == "Device Name" and ws.cell('C1').value == "Device MAC" and ws.cell('D1').value == "Device SN"):
                    for i in rows:
                        sign = 0
                        if ws.cell('A1').value == "Device Name":
                            if i[0].value:
                                name = str(i[0].value).strip()
                            else:
                                name = ''
                            if i[2].value:
                                sn = str(i[2].value).strip()
                            else:
                                sn = ''
                            if i[1].value:
                                mac = str(i[1].value).strip()
                            else:
                                mac = ''
                        elif ws.cell('B1').value == "Device Name":
                            if i[1].value:
                                name = str(i[1].value).strip()
                            else:
                                name = ''
                            if i[3].value:
                                sn = str(i[3].value).strip()
                            else:
                                sn = ''
                            if i[2].value:
                                mac = str(i[2].value).strip()
                            else:
                                mac = ''
                        if name == "Device Name" and mac == "Device MAC" and sn == "Device SN":
                            pass
                        else:
                            if len(sn) == 0 or len(mac) == 0:
                                wrong = wrong + 1
                                sign = 1
                                writews.append([mac,_(u'MAC/SN 为空')])
                            else:
                                if len(''.join(re.findall(r'[0-9a-fA-F]',mac))) != 12:
                                    wrong = wrong + 1
                                    sign = 1
                                    writews.append([mac,_(u'MAC 不合法')])
                                else:
                                    mac = ''.join(re.findall(r'[0-9a-fA-F]',mac)).lower()
                            if sign == 0:
                                if power == 6 or power == 5:
                                    if Device.objects.filter(Q(mac = mac) | Q(sn = sn)).exists() :
                                        wrong = wrong + 1
                                        writews.append([mac,_(u'设备已存在')])
                                    else:
                                        try:
                                            dev = Device()
                                            dev.name = name
                                            dev.sn = sn
                                            dev.mac = mac
                                            dev.group_id = gp_id
                                            dev.account_group_name = 'admin'
                                            dev.support_mode = mode
                                            dev.last_heart_time = '1970-1-1 08:00:00'
                                            dev.save()
                                            #通知redis
                                            DataQuery.DQimportdev(mac)

                                            if int(mode) == 1 or int(mode) == 3:
                                                accessed_device2authpuppy(mac)
                                            right = right + 1
                                        except Exception as e:
                                            print e
                                            wrong = wrong + 1
                                            writews.append([mac,_(u'设备存储失败')])
                                else:
                                    if Device.objects.filter(Q(mac = mac) & Q(sn = sn) & Q(support_mode = mode) , Q(account_group_name = 'admin')|Q(account_group_name = ''), Q(group_id = gp_id)|Q(group_id = 0)).exists():

                                        group_id = Probe_group.objects.get(Q(group_name = request.user.groupname) & Q(account_group_name = request.user.groupname) & Q(group_type = mode)).pk
                                        try:
                                            dev = Device.objects.get(Q(mac = mac) & Q(sn = sn) & Q(support_mode = mode) , Q(account_group_name = 'admin')|Q(account_group_name = ''), Q(group_id = gp_id)|Q(group_id = 0))
                                            dev.name = name
                                            dev.account_group_name = request.user.groupname
                                            dev.group_id = group_id
                                            dev.save()

                                            if int(mode) == 1 or int(mode) == 3:
                                                update_device2authpuppy(mac,request.user.groupname)

                                            right = right + 1
                                        except Exception as e:
                                            print e
                                            wrong = wrong + 1
                                            writews.append([mac,_(u'设备存储失败')])
                                    else:
                                        wrong = wrong + 1
                                        writews.append([mac,_(u'设备不存在')])
                    if os.path.exists("./statics/download/"+ request.user.groupname):
                        pass
                    else:
                        os.makedirs("./statics/download/"+ request.user.groupname)
                    if wrong != 0 :
                        try:
                            furl = "download/" + request.user.groupname + "/" + str(int(time.time())) + ".xlsx"
                            fn = "./statics/"+ furl
                            writewb.save(filename = fn)
                        except Exception as e:
                            print e
                            error = _(u'错误列表写入失败')
                else:
                    error = _(u'文件内容不合法')
            except Exception as e:
                print e
                error = _(u'文件读取失败')

    if error == "":
        # error = _(u'导入成功%d项 , 导入失败%d项'%(right,wrong))
        error = '{0}{1}{2} , {3}{4}{2}'.format(_(u'导入成功'),right,_(u'项'),_(u'导入失败'),wrong )
    return JsonResponse({'error':error,'furl':furl},safe = False)



@login_required
def probe_auto_accept(request):
    accept = request.POST.get('accept')
    error = {'error':''}
    if accept == "open":
        try:
            ss = Setting.objects.get(SIGN = 1)
            ss.probe_auto_accept = 'open'
            ss.save()
            error['error'] = _(u'修改成功')
        except Exception as e:
            print e
            error['error'] = _(u'修改失败')
    elif accept == "get":
        try:
            ss = Setting.objects.get(SIGN = 1)
            result = ss.probe_auto_accept
        except Exception as e:
            print e
            result = ""
        return JsonResponse(result,safe = False)
    else:
        try:
            ss = Setting.objects.get(SIGN = 1)
            ss.probe_auto_accept = 'close'
            ss.save()
            error['error'] = _(u'修改成功')
        except Exception as e:
            print e
            error['error'] = _(u'修改失败')
    return JsonResponse(error,safe= False)


@login_required
def probe_auto_update(request):
    error = {'error':''}
    if request.method == 'GET':
        get_type = request.GET.get('get_type')
        if get_type == 'option':
            result = []
            pg = Probe_group.objects.filter(account_group_name = request.user.groupname,group_type = '2')
            for i in pg:
                result.append({
                    'id': str(i.pk),
                    'text': str(i.group_name),
                })
            return JsonResponse({'results':result},safe = False)
        elif get_type == 'load_result':
            if Auto_Update_Rule.objects.filter(groupname = request.user.groupname,support_mode = '2').exists():
                auto_update_rule = Auto_Update_Rule.objects.get(groupname = request.user.groupname,support_mode = '2')
                result = auto_update_rule.rule
            else:
                result = json.dumps([])
            return JsonResponse(result,safe = False)

    elif request.method == 'POST':
        rule = request.POST.getlist('rule[]')
        if not rule:
            rule = []
        rule = json.dumps(rule)
        print rule
        if Auto_Update_Rule.objects.filter(groupname = request.user.groupname,support_mode = '2').exists():
            auto_update_rule = Auto_Update_Rule.objects.get(groupname = request.user.groupname,support_mode = '2')
        else:
            auto_update_rule = Auto_Update_Rule()
            auto_update_rule.groupname = request.user.groupname
            auto_update_rule.support_mode = '2'
        auto_update_rule.rule = rule
        try:
            auto_update_rule.save()
            error['error'] = _(u'修改成功')
        except Exception as e:
            print e
            error['error'] = _(u'修改失败')
        return JsonResponse(error,safe= False)



def accessed_device2authpuppy(mac):
    try:
        db = MySQLdb.connect("localhost", settings_py.DATABASES['default']['USER'], settings_py.DATABASES['default']['PASSWORD'], "authpuppy" )
    except MySQLdb.Error, e:
        try:
            sqlError =  "Error %d:%s" % (e.args[0], e.args[1])
        except IndexError:
            print "MySQL Error:%s" % str(e)
        return False

    cursor = db.cursor()

    try:
        mac = (mac[:2] + ':' + mac[2:4] + ':' + mac[4:6] + ':' + mac[6:8] + ':' + mac[8:10] + ':' + mac[10:]).upper()
        sql = "INSERT into nodes(name,gw_id,created_at,updated_at,domainname) values ('','{}','0000-00-00 00:00:00','0000-00-00 00:00:00','admin')".format(mac)

    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.execute(sql)
    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print sql
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.close()
        db.commit()
    except Exception as e:
        print e
        cursor.close()
        db.rollback()

    db.close()


def update_device2authpuppy(mac,domainname):
    try:
        db = MySQLdb.connect("localhost", settings_py.DATABASES['default']['USER'], settings_py.DATABASES['default']['PASSWORD'], "authpuppy" )
    except MySQLdb.Error, e:
        try:
            sqlError =  "Error %d:%s" % (e.args[0], e.args[1])
        except IndexError:
            print "MySQL Error:%s" % str(e)
        return False

    cursor = db.cursor()

    try:
        mac = (mac[:2] + ':' + mac[2:4] + ':' + mac[4:6] + ':' + mac[6:8] + ':' + mac[8:10] + ':' + mac[10:]).upper()
        # sql = "INSERT into nodes(name,gw_id,created_at,updated_at,domainname) values ('','{}','0000-00-00 00:00:00','0000-00-00 00:00:00','admin')".format(mac)
        sql = "UPDATE nodes set domainname = '{1}' where gw_id = '{0}'".format(mac,domainname)

    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.execute(sql)
    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print sql
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.close()
        db.commit()
    except Exception as e:
        print e
        cursor.close()
        db.rollback()

    db.close()
