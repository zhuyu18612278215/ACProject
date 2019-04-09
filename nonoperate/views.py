#coding=utf-8
from django.utils.translation import ugettext as _
from django.utils.translation import ugettext_lazy as lazy_
from django.shortcuts import render
from django.contrib import auth
from django.contrib.auth import login,authenticate
from account.models import NewUser,Account_Group,Auto_Update_Rule
from django.contrib.auth.decorators import login_required
from device.models import Device,Probe_config,Probe_audit_basic_status,Probe_audit_dev_status,Probe_audit_place_status,Probe_event,Probe_group
from models import NONOPERATE_event
import json
from django.http import JsonResponse
from django.forms.models import model_to_dict
from django.http import HttpResponse,HttpResponseRedirect
from ap.models import Device_ap,AP_event,Group_wlan,Device_wlan,User_policy_config,Guest_policy,ApBlackList,Gpon,Group_gpon,Setting_gpon,Customer,Customer_name,Customer_black_list,Customer_history,Customer_white_list,Customer_black_white_switch,Timing_Policy
from django.utils import timezone
import datetime
from django.core.urlresolvers import reverse
from django.shortcuts import redirect
import time
import pytz
import urllib
import urllib2
import cookielib
import hashlib
import re
# Create your views here.
import sys
sys.path.append("DataQuery")
import DataQuery
import os
import os.path
from system.models import Setting,Page_limit,Oem_limit
from django.core.files import File
import Public_function
import logging
from django.utils.timezone import is_naive, make_aware, utc, is_aware
from django.db.models import Q
from ap.default_settions import user_policy_config,guest_policy_config
import random
import MySQLdb
import MySQLdb.cursors
from django.db import connection
from django.conf import settings as settings_py
reload(sys)
sys.setdefaultencoding('utf-8')
import logging
from nonoperate.tasks import task_for_nonoperate_del_wlan,task_for_modify_nonoperate_group_ajax,task_for_add_wlan_ajax,task_for_modify_wlan_ajax,task_for_global_config_ajax,task_for_user_policy_config_ajax,default_config_issued
import collections
from django.core.servers.basehttp import FileWrapper
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
# Create your views here.
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
fh = logging.FileHandler("./statics/log/nonoperatelog.log")
fh.setLevel(logging.DEBUG)
formater = logging.Formatter('[%(levelname)s][%(asctime)s]--## %(message)s')
fh.setFormatter(formater)
logger.addHandler(fh)

def get_plimit():
    if Page_limit.objects.filter(pk = 1).exists():
        plimit = model_to_dict(Page_limit.objects.get(pk = 1))
    else:
        plimit = Public_function.page_limit_dict
    return plimit

def get_oemlimit():
    if Oem_limit.objects.filter(pk = 1).exists():
        oemlimit = model_to_dict(Oem_limit.objects.get(pk = 1))
    else:
        oemlimit = Public_function.oem_limit_dict
    return oemlimit



def upgrade_file_list():
    django_settings = Setting.objects.get(SIGN = 1)
    try:
        # url = "http://115.28.241.216:81/version/byz/"
        url = django_settings.VERSION_SERVER_URL+"version/feijing/"
        website = urllib2.urlopen(url)

        html = website.read()

        # links = re.findall('href="(.*)">witfios', html)
        links = re.findall('href="(.*)">[a-z]', html)
        # print links
        a = []
        for i in links:
            if i[-4:] == '.bin':
                if 'kernel' in i or 'rootfs' in i:
                    pass
                else:
                    a.append(i)
        # print a
        return a
    except Exception as e:
        print e,'upgrade_file_list'
        return False

def upgrade_file(dev_type,fl):
    r = dev_type_change(dev_type)
    result = ''
    print r,dev_type,fl
    if r == False or fl == False:
        pass
    else:
        for i in fl:
            if i == r :
                result = i
        if result != "":
            return result

    return False

def dev_type_change(dev_type):
    try:
        f = open('./statics/upgrade/feijing_upgrade_dict.info','r')
        try:
            data = json.load(f,'utf-8')
            return data[dev_type]
        except Exception as e:
            return False
        finally:
            f.close()
    except Exception as e:
        return False

def compare_ver(file1,ver1):
    if file1 == False:
        return False
    return file1 != str(ver1)

def ap_event_save(pe):
    nonoperate_event = NONOPERATE_event()
    nonoperate_event.event_time = pe["event_time"]
    nonoperate_event.event = pe['event']
    nonoperate_event.msg = pe['msg']
    nonoperate_event.admin_username = pe['admin_username']
    nonoperate_event.ap_mac = pe['ap_mac']
    try:
        nonoperate_event.save()
        return True
    except Exception as e:
        print e
        return False


@login_required
def nonoperate_access(request):
    errors = ''
    # 接受到的mac
    r_mac = request.GET.get('mac')
    # 数据信息
    data = {'mac':r_mac,'name':'','model':'X-300','sn':'00000001','lastip':'10.200.10.1:20000','privateip':'192.168.1.101','currstanum':'0','version':'2.0001','last_heart_time':'2017-04-07 15:30:00','upload':'0','download':'0','radios_type':'0','own_model':'WitFi-DAP510E'}
    #
    # data = DataQuery.DQGetAccessDev("ap", r_mac)
    if Device.objects.filter(Q(mac = r_mac)&Q(support_mode = '3')).exists():
        errors = _(u"设备已存在！")
    if len(errors) == 0:
        if Device.objects.filter(Q(mac = r_mac)&~Q(support_mode = '3')).exists() :
            Device.objects.get(Q(mac = r_mac)&~Q(support_mode = '3')).delete()
            # try:
            #     Probe_config.objects.get(mac = r_mac).delete()
            # except Exception as e:
            #     print e
            # try:
            #     Probe_audit_basic_status.objects.get(mac = r_mac).delete()
            # except Exception as e:
            #     print e
            # try:
            #     Probe_audit_dev_status.objects.get(mac = r_mac).delete()
            # except Exception as e:
            #     print e
            # try:
            #     Probe_audit_place_status.objects.get(mac = r_mac).delete()
            # except Exception as e:
            #     print e
        ap = Device()
        ap.mac = data['mac']
        if data['name'] == '':
            ap.name = data['model'] + '_' + data['mac'][-6:]
        else:
            ap.name = data['name']
        ap.model = data['model']
        ap.sn = data['sn']
        ap.lastip = data['lastip']
        ap.privateip = data['privateip']
        ap.version = data['version']
        ap.last_heart_time = data['last_heart_time']
        ap.upload = data['upload']
        ap.download = data['download']
        ap.up_pkts = '0'#data['up_pkts']
        ap.down_pkts = '0'#data['down_pkts']
        ap.own_model = data['own_model']
        ap.support_mode = '3'

        if ap is not None:
            try:
                ap.save()

                p = Device.objects.get(mac = r_mac)
                d = Device_ap()
                d.device = p
                d.apusernum = 0#data['apusernum']
                d.guestsnum = 0#data['guestsnum']
                d.radios_type = '0'#data['radios_type']
                d.radios_2_channel = 'auto'#data['radios_2_channel']
                d.radios_2_ht = '20'#data['radios_2_ht']
                d.radios_2_power = 'auto'#data['radios_2_power']
                d.radios_2_com = 'auto'#data['radios_2_com']
                d.radios_5_channel = 'auto'#data['radios_5_channel']
                d.radios_5_ht = '20'#data['radios_5_ht']
                d.radios_5_power = 'auto'#data['radios_5_power']
                d.radios_2_currstanum = 0#data['radios_2_currstanum']
                d.radios_2_guestsnum = 0#data['radios_2_guestsnum']
                d.radios_5_currstanum = 0#data['radios_5_currstanum']
                d.radios_5_guestsnum = 0#data['radios_5_guestsnum']
                d.save()

                # DataQuery.DQSetAccessed("nonoperate", r_mac)
                DataQuery.DQSetAccessed("ap", r_mac)
                errors = _(u'准入成功！')
                event = {
                    'event_time' : timezone.now(),
                    'event' : 'NONOPERATE_ADMITED_BY_ADMIN',
                    'msg' : 'Device['+ap.mac+'] was admited by Admin['+request.user.username+']',
                    'admin_username' : request.user.username,
                    'ap_mac' : ap.mac,
                }

                accessed_device2authpuppy(r_mac)

                try:
                    ap_event_save(event)
                except Exception as e:
                    print e
            except Exception as e:
                errors = _(u'准入失败！')

    request.session['error_message'] = errors
    return redirect('/ap-list/')

@login_required
def nonoperate_list(request):
    global ap_ugfl
    DataQuery.ap_ugfl = upgrade_file_list()
    errors = ''

    try:
        errors = request.session.get('error_message')
        # print errors
        del request.session['error_message']
    except Exception as e:
        pass
    errors_json = json.dumps(errors)
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'non-operate-list.html',{'errors_json':errors_json,'errors':errors,'admin_error_json':admin_error_json,'admin_error':admin_error,'plimit':plimit,'oemlimit':oemlimit})


def nonoperate_wait_access_ajax(request):
    wait_access_list = ''
    if request.user.administrator_permission > 4 or request.user.administrator_permission == 0 :
        # p_all = Device.objects.filter(support_mode = '3')
        # wait_access_list = DataQuery.DQGetAccess("nonoperate")
        wait_access_list = DataQuery.DQGetAccess("ap")
        #

        # ap1 = {'mac':'112233445566','model':'X-300','privateip':'192.168.1.101','lastip':'10.200.10.1:20000','version':'2.00000001','last_heart_time':'2017-04-07 15:30:00'}
        # wait_access_list = [ap1]

    ret = {'data':wait_access_list}
    return JsonResponse(ret,safe = False)

def nonoperate_already_access_ajax(request):
    # print 2131231,request.GET
    draw = int(request.GET.get('draw'))
    start = int(request.GET.get('start'))
    length = int(request.GET.get('length'))
    order_id = request.GET.get('order[0][column]')
    order = request.GET.get('columns['+order_id+'][data]')

    device_type_receive = request.GET.get('device_type')
    dev_online_time = timezone.now() - datetime.timedelta(minutes = 15)
    if device_type_receive == 'online':
        time_condition = Q(last_heart_time__gte = dev_online_time)
    elif device_type_receive == 'offline':
        time_condition = Q(last_heart_time__lt = dev_online_time)
    else:
        time_condition = Q(last_heart_time__gte = dev_online_time)|~Q(last_heart_time__gte = dev_online_time)


    if request.GET.get('order[0][dir]') == 'asc':
        order_type = ''
    elif request.GET.get('order[0][dir]') == 'desc':
        order_type = '-'

    if order == 'state':
        order = 'last_heart_time'
    search_value_old = request.GET.get('search[value]')
    search_value = search_value_old
    # search_value = ''
    # for s in search_value_old:
    #     if s>= u"\u4e00" and s<= u"\u9fa5":
    #         pass
    #     else:
    #         search_value = search_value + s

    if request.user.administrator_permission >= 4 or request.user.administrator_permission == 0 :
        recordsTotal = Device.objects.filter(support_mode = '3').filter(time_condition).count()
        if search_value == '':
            recordsFiltered = recordsTotal
            already_access_list = Device.objects.filter(support_mode = '3').filter(time_condition).order_by(order_type+order)[start:start+length]
        else:
            if re.match(r'[\u4e00 -\u9fa5]+',search_value) == None :
                if search_value in [u'退服',u'离线',u'超时',u'在线']:
                    sv = {
                        u'在线':'online',
                        u'超时':'time out',
                        u'离线':'offline',
                        u'退服':'retired',
                    }
                    already_access_list = Device.objects.filter(support_mode = '3').filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = Device.objects.filter(support_mode = '3').filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value)).count()
                else:
                    already_access_list = Device.objects.filter(support_mode = '3').filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = Device.objects.filter(support_mode = '3').filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value)).count()
            else:
                already_access_list = Device.objects.filter(support_mode = '3').filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-'))))[start:start+length]
                recordsFiltered = Device.objects.filter(support_mode = '3').filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-')))).count()
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        recordsTotal = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3')).filter(time_condition).count()
        if search_value == '':
            recordsFiltered = recordsTotal
            already_access_list = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3')).filter(time_condition).order_by(order_type+order)[start:start+length]
        else:
            if re.match(r'[\u4e00 -\u9fa5]+',search_value) == None :
                if search_value in [u'退服',u'离线',u'超时',u'在线']:
                    sv = {
                        u'在线':'online',
                        u'超时':'time out',
                        u'离线':'offline',
                        u'退服':'retired',
                    }
                    already_access_list = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=sv[search_value])|Q(name__contains=search_value)).count()
                else:
                    already_access_list = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3')).filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value))[start:start+length]
                    recordsFiltered = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3')).filter(time_condition).order_by(order_type+order).filter(Q(state__contains=search_value)|Q(name__contains=search_value)).count()
            else:
                already_access_list = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-'))))[start:start+length]
                recordsFiltered = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3')).filter(time_condition).order_by(order_type+order).filter(Q(state__icontains=search_value)|Q(name__icontains=search_value)|Q(lastip__icontains=search_value)|Q(version__icontains=search_value)|Q(model__icontains=search_value)|Q(mac__icontains=search_value)|Q(mac__icontains=''.join(search_value.split(':')))|Q(mac__icontains=''.join(search_value.split('-')))).count()

    DataQuery.DQUpdateAccessed("ap", already_access_list)
    timenow = timezone.now()

    ret = {'draw':draw,'recordsTotal':recordsTotal,'recordsFiltered':recordsFiltered,'data':[]}
    for al in already_access_list:
        a = int((timenow - al.last_heart_time).total_seconds())

        al.last_heart_time = timezone.localtime(al.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")

        al.upgrade_button = ''
        rss = al.reboot_sign.split(",")
        if al.last_heart_time == "1970-1-1 08:00:00":
            al.state = _(u"离线")
        elif a >= 86400:
            al.state = _(u"退服")
        elif a >= 3600:
            al.state = _(u"离线")
        elif a >= 900 :
            al.state = _(u"超时")
        elif a < 900:
            if rss[0] != '0' and rss[0] != '':
                if rss[0] == '1':
                    al.state = _(u"重启")
                elif rss[0] == '2':
                    al.state = _(u"升级")
            else:
                al.state = _(u"在线")
        try:
            if compare_ver(''.join(re.findall("[0-9\.]",''.join(re.findall("(\d\.\d+\.\d+\.\w\d+)",dev_type_change(al.own_model)[:-4])))),al.version) == True:
                al.upgrade_button = True
                al.upgrade_version = dev_type_change(al.own_model)
        except Exception as e:
            print e
            al.upgrade_button = False
            al.upgrade_version = ''

        admin_power_control = ""
        if request.user.administrator_permission >= 5:
            if al.account_group_name == 'admin' or al.account_group_name == '':
                admin_power_control = ""
            else:
                admin_power_control = "ban"

        aall = model_to_dict(al)
        aall['admin_power_control'] = admin_power_control
        cpuInfo = DataQuery.DQGetAPCpu("ap",al.mac) or ""
        aall['cpu'] = cpuInfo

        try:
            gpn = Probe_group.objects.get(pk = aall['group_id']).group_name
        except Exception as e:
            print e
            gpn = "DefaultGroup"
        if aall['account_group_name'] == "":
            acpn = "admin"
        else:
            acpn = aall['account_group_name']
        aall['group_id'] = acpn + "/" + gpn
        ret['data'].append(aall)
    return JsonResponse(ret,safe = False)


@login_required
def nonoperate_reboot(request):
    errors = ''
    r_mac = request.POST.get('mac')
    data = {'action':'reboot','mac':r_mac}
    #执行动作
    try:
        #在此调用函数
        DataQuery.DQProcess(data)
        errors = _(u'执行成功！')
        event = {
            'event_time' : timezone.now(),
            'event' : 'NONOPERATE_RESTARTED_BY_ADMIN',
            'msg' : 'Device['+r_mac+'] was restarted by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'ap_mac' : r_mac,
        }
        try:
            nu = Device.objects.get(mac = r_mac)
            nu.reboot_sign = '1'+','+str(int(time.mktime(datetime.datetime.now().timetuple())))
            nu.save()
        except Exception as e:
            print e
        try:
            # DataQuery.LOGGING_INIT()
            # logging.info('Device['+r_mac+'] was restarted by Admin['+request.user.username+']')
            # logging.info('GGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGG')
            # print 'Device['+r_mac+'] was restarted by Admin['+request.user.username+']'
            ap_event_save(event)
        except Exception as e:
            print e
    except Exception as e:
        print e
        errors = _(u'执行失败！')
    return JsonResponse(errors,safe=False)

@login_required
def nonoperate_update(request):
    django_settings = Setting.objects.get(SIGN = 1)
    errors = ''
    r_mac = request.POST.get('mac')
    #data = {'action':'upgrade','mac':r_mac, 'param':'http://115.28.241.216/img/witfios3.01.03.r191710.bin'}
    u = upgrade_file(Device.objects.get(mac = r_mac).own_model,DataQuery.ap_ugfl)
    if u == False:
        errors = _(u'没有此型号对应的版本文件')
    else:
        data = {'action':'upgrade','mac':r_mac,'param':django_settings.VERSION_SERVER_URL+"version/feijing/"+u}
        #执行动作
        try:
            #在此调用函数
            DataQuery.DQProcess(data)
            errors = _(u'执行成功！')
            if u != False:
                event = {
                    'event_time' : timezone.now(),
                    'event' : 'NONOPERATE_UPGRADE_BY_ADMIN_TO_VERSION',
                    'msg' : 'Device['+r_mac+'] was upgrade to version['+u+'] by Admin['+request.user.username+']',
                    'admin_username' : request.user.username,
                    'ap_mac' : r_mac,
                }
                try:
                    nu = Device.objects.get(mac = r_mac)
                    nu.reboot_sign = '2'+','+str(int(time.mktime(datetime.datetime.now().timetuple())))
                    nu.save()
                except Exception as e:
                    print e
                try:
                    ap_event_save(event)
                except Exception as e:
                    print e

        except Exception as e:
            print e
            errors = _(u'执行失败！')
    return JsonResponse(errors,safe=False)

@login_required
def nonoperate_del(request):
    errors = ''
    r_mac = request.POST.get('mac')
    data = {'action':'del','mac':r_mac}
    #执行动作
    try:
        #在此调用函数
        Device.objects.filter(mac = r_mac).delete()
        try:
            Probe_config.objects.filter(mac = r_mac).delete()
            Gpon.objects.filter(ap_mac = r_mac).delete()

            Probe_audit_basic_status.objects.filter(mac = r_mac).delete()
            Probe_audit_dev_status.objects.filter(mac = r_mac).delete()
            Probe_audit_place_status.objects.filter(mac = r_mac).delete()
        except Exception as e:
            print e
        errors = _(u'执行成功！')
        event = {
            'event_time' : timezone.now(),
            'event' : 'NONOPERATE_FORGOTEN_BY_ADMIN',
            'msg' : 'Device['+r_mac+'] was forgoten by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'ap_mac' : r_mac,
        }
        try:
            # DataQuery.DQDelAccessed("nonoperate", r_mac)
            DataQuery.DQDelAccessed("ap", r_mac)
            data = {'action':'reset','mac':r_mac}
            DataQuery.DQProcess(data)

            del_device2authpuppy(r_mac)

            ap_event_save(event)
        except Exception as e:
            print e
    except Exception as e:
        print e
        errors = _(u'执行失败！')
    return JsonResponse(errors,safe = False)

def nonoperate_customers_get_by_mac(apmac):
    mac = apmac
    online_time = int(time.time()) - 900

    print "56756756756756756756756756757", mac
    customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap = mac)&Q(portal_enable = '0'))

    return len(customer_list)

def nonoperate_guests_get_by_mac(apmac):
    mac = apmac
    online_time = int(time.time()) - 900

    print "56756756756756756756756756757", mac
    customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap = mac)&Q(portal_enable = '1'))

    return len(customer_list)

@login_required
def nonoperate_detail(request):
    r_mac = request.GET.get('mac')
    r = Device.objects.get(mac = r_mac)
    name = r.name
    vpn = r.vpn
    vpnip = r.vpnip

    upload = r.upload
    download = r.download
    up_pkts = r.up_pkts
    down_pkts = r.down_pkts
    apusernum = 0
    guestsnum = 0
    if hasattr(r,'device_ap'):
        # apusernum = r.device_ap.apusernum
        # guestsnum = r.device_ap.guestsnum
        guestsnum = nonoperate_guests_get_by_mac(r_mac)
        apusernum = nonoperate_customers_get_by_mac(r_mac)
        radios_type = r.device_ap.radios_type
        radios_2_channel = r.device_ap.radios_2_channel
        radios_2_ht = r.device_ap.radios_2_ht
        radios_2_power = r.device_ap.radios_2_power
        radios_2_com = r.device_ap.radios_2_com
        radios_5_channel = r.device_ap.radios_5_channel
        radios_5_ht = r.device_ap.radios_5_ht
        radios_5_power = r.device_ap.radios_5_power
        radios_2_currstanum = r.device_ap.radios_2_currstanum
        radios_2_guestsnum = r.device_ap.radios_2_guestsnum
        radios_5_currstanum = r.device_ap.radios_5_currstanum
        radios_5_guestsnum = r.device_ap.radios_5_guestsnum
    else:
        # apusernum = 0
        # guestsnum = 0
        guestsnum = nonoperate_guests_get_by_mac(r_mac)
        apusernum = nonoperate_customers_get_by_mac(r_mac)
        try:
            radios_type = Public_function.device_type_info_dict[r.own_model]['radio_num']
        except Exception as e:
            print e
            radios_type = '0'
        radios_2_channel = 'auto'
        radios_2_ht = '20'
        radios_2_power = 'auto'
        radios_2_com = 'auto'
        radios_5_channel = 'auto'
        radios_5_ht = '20'
        radios_5_power = 'auto'
        radios_2_currstanum = 0
        radios_2_guestsnum = 0
        radios_5_currstanum = 0
        radios_5_guestsnum = 0

    try:
        d = Probe_config.objects.get(mac = r_mac)
        # print d,1111111111
        ac_address = d.ac_address
        log_address = d.log_address
        ip_model = d.ip_model
        ip_address = d.ip_address
        subnet_mask = d.subnet_mask
        gateway = d.gateway
        preferred_dns = d.preferred_dns
        alternative_dns = d.alternative_dns
        preferred_ntp = d.preferred_ntp
        alternative_ntp = d.alternative_ntp
    except Exception as e:
        print e
        ac_address = ''
        log_address = ''
        ip_model = ''
        ip_address = ''
        subnet_mask = ''
        gateway = ''
        preferred_dns = ''
        alternative_dns = ''
        preferred_ntp = ''
        alternative_ntp = ''

    admin_power_control = ""
    if request.user.administrator_permission >= 5:
        if r.account_group_name == 'admin' or r.account_group_name == '':
            admin_power_control = ""
        else:
            admin_power_control = "ban"
    data = {'mac':r_mac,'sn':'','model':'','version':'','ip':'','running_time':'','name':name,'last_time':'','cpu':'','memory':'','flash':'','apusernum':apusernum,'guestsnum':guestsnum,'upload':upload,'download':download,'up_pkts':up_pkts,'down_pkts':down_pkts,'radios_type':radios_type,'radios_2_channel':radios_2_channel,'radios_2_ht':radios_2_ht,'radios_2_power':radios_2_power,'radios_2_com':radios_2_com,'radios_5_channel':radios_5_channel,'radios_5_ht':radios_5_ht,'radios_5_power':radios_5_power,'radios_2_currstanum':radios_2_currstanum,'radios_2_guestsnum':radios_2_guestsnum,'radios_5_currstanum':radios_5_currstanum,'radios_5_guestsnum':radios_5_guestsnum,'ac_address':ac_address,'log_address':log_address,'ip_model':ip_model,'ip_address':ip_address,'subnet_mask':subnet_mask,'gateway':gateway,'preferred_dns':preferred_dns,'alternative_dns':alternative_dns,'vpn':vpn,'vpnip':vpnip,'preferred_ntp':preferred_ntp,'alternative_ntp':alternative_ntp,}
    DataQuery.DQUpdateAccessedDev('ap', r_mac, data)
    # DataQuery.DQUpdateAccessedDev('nonoperate', r_mac, data)

    data["admin_power_control"] = admin_power_control
    data["locateState"] = r.locateState

    # data_str = set_wlan_str(r_mac)
    # DataQuery.DQSetUpdateConfig('wireless',r_mac,data_str)
    return JsonResponse(data)

@login_required
def nonoperate_config(request):
    r_mac = request.GET.get('mac')
    r = Device.objects.get(mac = r_mac)

    try:
        radios_type = r.device_ap.radios_type
        radios_2_channel = r.device_ap.radios_2_channel
        radios_2_ht = r.device_ap.radios_2_ht
        radios_2_power = r.device_ap.radios_2_power
        radios_2_com = r.device_ap.radios_2_com
        radios_5_channel = r.device_ap.radios_5_channel
        radios_5_ht = r.device_ap.radios_5_ht
        radios_5_power = r.device_ap.radios_5_power
        radios_5_com = r.device_ap.radios_5_com
    except Exception as e:
        print e
        try:
            radios_type = Public_function.device_type_info_dict[r.own_model]['radio_num']
        except Exception as e:
            print e
            radios_type = '0'
        radios_2_channel = 'auto'
        radios_2_ht = '20'
        radios_2_power = 'auto'
        radios_2_com = 'auto'
        radios_5_channel = 'auto'
        radios_5_ht = '20'
        radios_5_power = 'auto'
        radios_5_com = 'auto'

    if hasattr(r,'ap_user_policy_config_extend'):
        max_user = r.ap_user_policy_config_extend.max_user
    else:
        if hasattr(Account_Group.objects.get(groupname = r.account_group_name),'user_policy_config'):
            if radios_type == '0' :
                max_user = Account_Group.objects.get(groupname = r.account_group_name).user_policy_config.dual_max_user
            else:
                max_user = Account_Group.objects.get(groupname = r.account_group_name).user_policy_config.single_max_user
        else:
            if radios_type == '0' :
                max_user = user_policy_config['dual_max_user']
            else:
                max_user = user_policy_config['single_max_user']

    try:
        r_2g_maxp = Public_function.device_type_info_dict[r.own_model]['radio_2_txp']
    except Exception as e:
        r_2g_maxp = '20'
    try:
        r_5g_maxp = Public_function.device_type_info_dict[r.own_model]['radio_5_txp']
    except Exception as e:
        r_5g_maxp = '20'

    data = {'radios_type':radios_type,'radios_2_channel':radios_2_channel,'radios_2_ht':radios_2_ht,'radios_2_power':radios_2_power,'radios_2_com':radios_2_com,'radios_5_channel':radios_5_channel,'radios_5_ht':radios_5_ht,'radios_5_power':radios_5_power,'radios_5_com':radios_5_com,'r_2g_maxp':r_2g_maxp,'r_5g_maxp':r_5g_maxp,'max_user':max_user}

    return JsonResponse(data)

@login_required
def nonoperate_radios_config(request):
    errors = ''
    m_model = request.GET.get('modify_model')
    if request.method == "POST":
        mac = request.POST.get('mac')
        if m_model == '6':
            radios_2_channel = request.POST.get('2g_channel')
            radios_2_ht = request.POST.get('2g_ht')
            radios_2_power = request.POST.get('2G_power')
            radios_2_com = request.POST.get('radios_2_com')
            radios_5_channel = request.POST.get('5g_channel')
            radios_5_ht = request.POST.get('5g_ht')
            radios_5_power = request.POST.get('5G_power')
            radios_2_customize = request.POST.get('2g_customize')
            radios_5_customize = request.POST.get('5g_customize')
            radios_5_com = request.POST.get('radios_5_com')
            try:
                dev = Device.objects.get(mac = mac)
                if hasattr(dev,'device_ap'):
                    rad = dev.device_ap
                else:
                    rad = Device_ap()
                    rad.device_id = dev.pk
                try:
                    rad.radios_type = Public_function.device_type_info_dict[dev.own_model]['radio_num']
                except Exception as e:
                    print e
                    rad.radios_type = '0'
                rad.radios_2_channel = radios_2_channel
                rad.radios_2_ht = radios_2_ht
                if radios_2_power == "customize":
                    rad.radios_2_power = radios_2_customize
                else:
                    rad.radios_2_power = radios_2_power
                if radios_5_power == "customize":
                    rad.radios_5_power = radios_5_customize
                else:
                    rad.radios_5_power = radios_5_power
                rad.radios_2_com = radios_2_com
                rad.radios_5_channel = radios_5_channel
                rad.radios_5_ht = radios_5_ht
                rad.radios_5_com = radios_5_com

                rad.save()

                create_wireless_json(dev)
                # data_str = set_wlan_str(mac)
                # DataQuery.DQSetUpdateConfig('wireless',mac,data_str)
                # ntpstr = preferred_ntp + ';' + alternative_ntp
                # data = {'action':'ntp','mac':mac,'param':ntpstr}
                # DataQuery.DQProcess(data)
                errors = _(u'修改成功！')
            except Exception as e:
                print e,'3213213212'
                errors = _(u'修改失败！')

    return JsonResponse(errors,safe = False)

@login_required
def nonoperate_vpn(request):
    model = request.POST.get('model')
    mac = request.POST.get('mac')
    errors = ''
    if model == 'on':
        serverip = '115.28.241.216'
        timeval = str(int(time.time()))
        m = hashlib.md5()
        m.update(timeval+mac.upper())
        token = m.hexdigest()
        urls = "http://"+serverip+"/vpnserver/vpn_address_asign.php?t="+timeval+"&token="+token+"&mac="+mac.upper()
        result = json.loads(urllib2.urlopen(urls).read())
        try:
            if 'rc' in result['meta'] and result['meta']['rc'] == 'ok' and 'ip' in result['data'] and 'mac' in result['data'] and result['data']['mac'] == mac.upper():

                ip = result['data']['ip']
                #add
                #
                errors = result['meta']['rc']
                ret = {'result':'on','ip':ip}

                event = {
                    'event_time' : timezone.now(),
                    'event' : 'AP_START_REMOTE_CONNECT_BY_ADMIN',
                    'msg' : 'Device['+mac+'] was open remote connect by Admin['+request.user.username+']',
                    'admin_username' : request.user.username,
                    'ap_mac' : mac,
                }
                try:
                    ap_event_save(event)
                except Exception as e:
                    print e
            elif 'rc' in result['meta'] and result['meta']['rc'] == 'error':
                errors = result['meta']['rc']
                ret = {'result':'off','ip':''}
        except Exception as e:
            print e
            ret = {'result':'off','ip':''}
            ip = ''

    elif model == 'off':
        ret = {'result':'off','ip':''}
        ip = ''
        event = {
            'event_time' : timezone.now(),
            'event' : 'AP_STOP_REMOTE_CONNECT_BY_ADMIN',
            'msg' : 'Device['+mac+'] was close remote connect by Admin['+request.user.username+']',
            'admin_username' : request.user.username,
            'ap_mac' : mac,
        }
        try:
            ap_event_save(event)
        except Exception as e:
            print e

    try:
        v = Device.objects.get(mac = mac)
        v.vpn = ret['result']
        v.vpnip = ip
        v.save()

        #send message to devs
        if model == 'on':
            ip = result['data']['ip']
            param = "115.28.241.216;1194;" + ip
            data = {'action':'vpn','mac':mac,'param':param}
        else:
            data = {'action':'vpn','mac':mac,'param':'stop'}

        DataQuery.DQProcess(data)
    except Exception as e:
        print e
    return JsonResponse(ret)

@login_required
def nonoperate_eventajax(request):
    event_type = request.GET.get('type')
    event_time_type = request.GET.get('time')
    seconds = int(event_time_type) * 3600
    t = timezone.now() - datetime.timedelta(hours = int(event_time_type))
    p_list = []
    event = {'data':[]}

    if request.user.administrator_permission >= 4 or request.user.administrator_permission == 0 :
        probe_event = NONOPERATE_event.objects.filter(event_time__gte = t).order_by('-event_time')

        if event_type == "dev":
            # event = {'data':[{'event_time':'2017/05/04 12:05:35','msg':'Device[A4E6B1300005] was onlined.','action':_(u'存档'),'probe_mac':'A4E6B1300005','event':u'','state':state_check('A4E6B1300005')},{'event_time':'2017/05/04 12:04:44','msg':'Device[8c8401163ae0] was offlined.','action':_(u"存档"),'probe_mac':'8c8401163ae0','event':u'PROBE_WAS_OFFLINED','state':state_check('8c8401163ae0')}]}
            p_list = [a.mac for a in Device.objects.filter(support_mode = '3')]
            event = DataQuery.DQGetEventFromRedis(seconds, p_list)
            for evd in event['data']:
                evd['probe_mac'] = evd['probe_mac'].lower()
                evd['state'] = state_check(evd['probe_mac'].lower())

        if event_type == "admin":
            if probe_event.count() != 0:
                for i in probe_event:
                    state = state_check(i.ap_mac)
                    event['data'].append({'event_time':timezone.localtime(i.event_time).strftime("%Y-%m-%d %H:%M:%S"),'msg':i.msg,'probe_mac':i.ap_mac,'admin_username':i.admin_username,'event':i.event,'state':state})
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        p_all = Device.objects.filter(Q(support_mode = '3')&Q(account_group_name = request.user.groupname))
        u_all = [a.username for a in NewUser.objects.filter(groupname = request.user.groupname)]
        p_list = [a.mac for a in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3'))]
        ##########xuyaoxiugai##################
        probe_event = NONOPERATE_event.objects.filter(event_time__gte = t).order_by('-event_time')

        if event_type == "dev":
            # event = {'data':[{'event_time':'2017/05/04 12:05:35','msg':'Device[A4E6B1300005] was onlined.','action':_(u'存档'),'probe_mac':'A4E6B1300005','event':u'','state':state_check('A4E6B1300005')},{'event_time':'2017/05/04 12:04:44','msg':'Device[8c8401163ae0] was offlined.','action':_(u"存档"),'probe_mac':'8c8401163ae0','event':u'PROBE_WAS_OFFLINED','state':state_check('8c8401163ae0')}]}

            event = DataQuery.DQGetEventFromRedis(seconds, p_list)
            for evd in event['data']:
                evd['probe_mac'] = evd['probe_mac'].lower()
                evd['state'] = state_check(evd['probe_mac'].lower())

        if event_type == "admin":
            if probe_event.count() != 0:
                for i in probe_event:
                    if i.admin_username in u_all:
                        state = state_check(i.ap_mac)
                        event['data'].append({'event_time':timezone.localtime(i.event_time).strftime("%Y-%m-%d %H:%M:%S"),'msg':i.msg,'probe_mac':i.ap_mac,'admin_username':i.admin_username,'event':i.event,'state':state})

    return JsonResponse(event,safe = False)

def state_check(mac):
    state = ''
    try:
        al = Device.objects.get(mac = mac)

        a = int((timezone.now()-al.last_heart_time).total_seconds())
        # print type(a)
        if a >= 86400:
            al.state = _(u"退服")
        elif a >= 3600:
            al.state = _(u"离线")
        elif a >= 900 :
            al.state = _(u"超时")
        elif a < 900:
            al.state = _(u"在线")
        state = al.state
    except Exception as e:
        print e
        state = _(u'已移除')

    return state

def create_wireless_json(i):
    if i.model == "KL0001" or i.model == "TPY3101-HR":
        pass
    else:
        m = change_mac(i.mac)
        data_str = set_wlan_str(m)
        DataQuery.DQSetUpdateConfig('wireless',m,data_str)

def set_wlan_str(mac):
    dev = Device.objects.get(mac = mac)
    try:
        radios_type = Public_function.device_type_info_dict[dev.own_model]['radio_num']
    except Exception as e:
        print e
        radios_type = '0'

    try:
        if hasattr(dev,'device_ap'):
            if dev.device_ap.radios_2_channel == 'auto':
                radios_2_channel = 0
            else:
                radios_2_channel = int(dev.device_ap.radios_2_channel)

            if dev.device_ap.radios_2_ht == 'auto':
                radios_2_ht = 'auto'
            else:
                radios_2_ht = 'HT' + dev.device_ap.radios_2_ht

            if dev.device_ap.radios_2_power == 'auto':
                radios_2_power = 0
            elif dev.device_ap.radios_2_power == 'high':
                radios_2_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_2_txp'])
            elif dev.device_ap.radios_2_power == 'medium':
                radios_2_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_2_txp']) - 5
            elif dev.device_ap.radios_2_power == 'low':
                radios_2_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_2_txp']) - 10
            else:
                radios_2_power = int(dev.device_ap.radios_2_power)

            if dev.device_ap.radios_2_com == 'auto':
                radios_2_com = '11ng'
            else:
                radios_2_com = dev.device_ap.radios_2_com



            if dev.device_ap.radios_5_channel == 'auto':
                radios_5_channel = 0
            else:
                radios_5_channel = int(dev.device_ap.radios_5_channel)

            if dev.device_ap.radios_5_ht == 'auto':
                radios_5_ht = 'auto'
            else:
                radios_5_ht = 'HT' + dev.device_ap.radios_5_ht

            if dev.device_ap.radios_5_power == 'auto':
                radios_5_power = 0
            elif dev.device_ap.radios_5_power == 'high':
                radios_5_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_5_txp'])
            elif dev.device_ap.radios_5_power == 'medium':
                radios_5_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_5_txp']) - 5
            elif dev.device_ap.radios_5_power == 'low':
                radios_5_power = int(Public_function.device_type_info_dict[dev.own_model]['radio_5_txp']) - 10
            else:
                radios_5_power = int(dev.device_ap.radios_5_power)

            if dev.device_ap.radios_5_com == 'auto':
                radios_5_com = '11ac'
            else:
                radios_5_com = dev.device_ap.radios_5_com
        else:
            radios_2_channel = 0
            radios_2_ht = 'auto'
            radios_2_power = 0
            radios_2_com = '11ng'
            radios_5_channel = 0
            radios_5_ht = 'auto'
            radios_5_power = 0
            radios_5_com = '11ac'
    except Exception as e:
        print e,'xwwwwww'
        radios_2_channel = 0
        radios_2_ht = 'auto'
        radios_2_power = 0
        radios_2_com = '11ng'
        radios_5_channel = 0
        radios_5_ht = 'auto'
        radios_5_power = 0
        radios_5_com = '11ac'

    radio2_str = {
        "channel": radios_2_channel,
        "country": "CN",
        "disabled": 0,
        "htmode": radios_2_ht,
        "hwmode": radios_2_com,
        "mode": "2.4G",
        "txpower": radios_2_power,
    }
    radio5_str =  {
        "channel": radios_5_channel,
        "country": "CN",
        "disabled": 0,
        "htmode": radios_5_ht,
        "hwmode": radios_5_com,
        "mode": "5G",
        "txpower": radios_5_power
    }

    apwlan2 = []
    apwlan5 = []
    if Group_wlan.objects.filter(group_id_id = dev.group_id).count() == 0:
        apwlan2 = []
        apwlan5 = []
    else:
        gp_wlan = Group_wlan.objects.filter(group_id_id = dev.group_id)
        for i in gp_wlan:
            if Device_wlan.objects.filter(device_id = dev.pk,wlan_id = i.wlan_id).exists():
                dev_wlan = Device_wlan.objects.get(device_id = dev.pk,wlan_id = i.wlan_id)
            else:
                dev_wlan = False

            if dev_wlan != False:
                i.wlan_service = dev_wlan.wlan_service if dev_wlan.wlan_service != "" else i.wlan_service
                i.upload_speed = dev_wlan.upload_speed if dev_wlan.upload_speed != "" else i.upload_speed
                i.download_speed = dev_wlan.download_speed if dev_wlan.download_speed != "" else i.download_speed
                i.passphrase = dev_wlan.passphrase if dev_wlan.passphrase != "" else i.passphrase
                i.wlan_ssid = dev_wlan.wlan_ssid if dev_wlan.wlan_ssid != "" else i.wlan_ssid
                i.radios_enable = dev_wlan.radios_enable if dev_wlan.radios_enable != "" else i.radios_enable
                i.vlan_enabled = dev_wlan.vlan_enabled if dev_wlan.vlan_enabled != "" else i.vlan_enabled
                i.vlan = dev_wlan.vlan if dev_wlan.vlan != "" else i.vlan


            wlan_service = 0 if i.wlan_service == 'on' else 1
            if i.upload_speed == 'Unlimited' and i.download_speed == 'Unlimited':
                peruserrate = 0
                upload_speed = 0
                download_speed = 0
            else:
                peruserrate = 1
                try:
                    if i.upload_speed == 'Unlimited':
                        upload_speed = 0
                    elif 'K' in i.upload_speed :
                        upload_speed = int(i.upload_speed.replace('K',''))
                    elif 'M' in i.upload_speed :
                        upload_speed = int(i.upload_speed.replace('M','')) * 1024
                    else:
                        upload_speed = int(i.upload_speed)

                    if i.download_speed == 'Unlimited':
                        download_speed = 0
                    elif 'K' in i.download_speed :
                        download_speed = int(i.download_speed.replace('K',''))
                    elif 'M' in i.download_speed :
                        download_speed = int(i.download_speed.replace('M','')) * 1024
                    else:
                        download_speed = int(i.download_speed)

                except Exception as e:
                    print e
                    peruserrate = 0
                    upload_speed = 0
                    download_speed = 0

            if i.sec_type == 'psk2' and (i.encry_type == 'tkip' or i.encry_type == 'ccmp'):
                encryption = i.sec_type + '+' + i.encry_type
            else:
                encryption = 'none'
            hidden_ssid = 1 if i.hidden_ssid == 'on' else 0
            if i.guest_enabled == 'on':
                portal = Public_function.protal_type_num_dict[i.auth_type]
            else:
                portal = 0
            if i.vlan_enabled == "on":
                vlan = i.vlan
            else:
                vlan = 0
            wl = {
                "apidx": int(i.wlan_id),
                "disabled": wlan_service,
                "downperuserrate": download_speed,
                "encryption": encryption,
                "hidden": hidden_ssid,
                "key": i.passphrase,
                "peruserrate": peruserrate,
                "portal": portal,
                "ssid": i.wlan_ssid,
                "upperuserrate": upload_speed,
                "vlan":int(vlan)
            }
            if i.radios_enable == "both":
                apwlan2.append(wl)
                apwlan5.append(wl)
            elif i.radios_enable == "2g":
                apwlan2.append(wl)
            elif i.radios_enable == "5g":
                apwlan5.append(wl)

    if radios_type == '0':
        data = "{'wifi':[{'radio':%s,'ap':%s},{'radio':%s,'ap':%s}]}"%(json.dumps(radio2_str ,sort_keys=True , ensure_ascii=False),json.dumps(apwlan2,sort_keys=True , ensure_ascii=False), json.dumps(radio5_str ,sort_keys=True , ensure_ascii=False), json.dumps(apwlan5 ,sort_keys=True , ensure_ascii=False))
    elif radios_type == '1':
        data = "{'wifi':[{'radio':%s,'ap':%s}]}"%(json.dumps(radio2_str ,sort_keys=True , ensure_ascii=False),json.dumps(apwlan2,sort_keys=True , ensure_ascii=False))
    print 'peizhi',data.replace(' ','').replace('\'','\"')
    return data.replace(' ','').replace('\'','\"')

@login_required
def nonoperate_group(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)
    group = ''
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        group = Probe_group.objects.filter(group_type = '3')
    elif request.user.administrator_permission < 4 and request.user.group_status == 1:
        group = Probe_group.objects.filter(account_group_name = request.user.groupname,group_type = '3')
    for i in group:
        if i.group_name == "DefaultGroup" and i.account_group_name == "admin" and i.group_type == '3':
            i.device_count = Device.objects.filter(support_mode = '3').filter(Q(account_group_name = i.account_group_name,group_id = i.pk) |Q(account_group_name = '',group_id = 0)).count()
        else:
            i.device_count = Device.objects.filter(support_mode = '3').filter(group_id = i.id).filter(account_group_name = i.account_group_name).count()
        i.wlan_count = str(i.group_wlan_set.filter(wlan_service = 'on').count()) + '/' +str(i.group_wlan_set.count())

    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'nonoperate-group.html',{'admin_error_json':admin_error_json,'admin_error':admin_error,'group':group,'error':error,"error_json":error_json,'plimit':plimit,'oemlimit':oemlimit})


@login_required
def nonoperate_add_ajax(request):
    device = {'data':[]}
    if request.user.administrator_permission == 6 or request.user.administrator_permission == 5 or request.user.administrator_permission == 0:
        a = Probe_group.objects.get(group_name = 'DefaultGroup',account_group_name = 'admin',group_type = '3')
        device_quaryset = Device.objects.filter(support_mode = '3').filter(Q(account_group_name = '') | Q(account_group_name = 'admin')).filter(Q(group_id = 0) | Q(group_id = a.pk))
    if request.user.administrator_permission == 3 or request.user.administrator_permission == 2 :
        a = Probe_group.objects.get(group_name = request.user.username,account_group_name = request.user.username,group_type = '3')
        device_quaryset = Device.objects.filter(support_mode = '3').filter(account_group_name = request.user.username).filter(Q(group_id = 0) | Q(group_id = a.pk))
    if device_quaryset.count() != 0:
        for i in device_quaryset:
            if i.name != '':
                name = i.name
            else:
                name = i.model + '_' + i.mac[-6:]
            mac = i.mac.upper()[:2] + '-' + i.mac.upper()[2:4] + '-' + i.mac.upper()[4:6] + '-' + i.mac.upper()[6:8] + '-' + i.mac.upper()[8:10] + '-' + i.mac.upper()[10:12]
            device['data'].append({'name':name,'mac':mac,'action':''})
    return JsonResponse(device,safe = False)

@login_required
def add_nonoperate_group_ajax(request):
    error = {"error_type":'',"error_msg":""}
    if request.method == "POST":
        add_group_device_list = json.loads(request.POST.get('add_group_device_list'))
        groupname = add_group_device_list['groupname']
        mac = add_group_device_list['mac']
        if request.user.groupname != '' :
            user_area = request.user.groupname
        else:
            user_area = 'admin'
        if Probe_group.objects.filter(account_group_name = user_area).filter(group_name = groupname,group_type = '3').exists():
            error = {"error_type":'failed',"error_msg":_(u"该组已存在")}
        else:
            try:
                pg = Probe_group()
                pg.group_name = groupname
                pg.account_group_name = user_area
                pg.group_type = '3'
                pg.save()
                pgid = Probe_group.objects.get(account_group_name = user_area,group_name = groupname,group_type = '3').pk
                if user_area == 'admin':
                    old_pgid = Probe_group.objects.get(account_group_name = user_area,group_name = 'DefaultGroup',group_type = '3').pk
                else:
                    old_pgid = Probe_group.objects.get(account_group_name = user_area,group_name = user_area,group_type = '3').pk

                if compare_group_wlan_settions(pgid,old_pgid):
                    for i in mac:
                        m = change_mac(i)
                        Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid,account_group_name = user_area)
                        Device.objects.get(mac = m).device_wlan_set.all().delete()
                else:
                    for i in mac:
                        m = change_mac(i)
                        Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid,account_group_name = user_area)
                        Device.objects.get(mac = m).device_wlan_set.all().delete()
                        create_wireless_json(Device.objects.get(mac = m))
                        #audit config str
                        # data_str = set_wlan_str(m)
                        # DataQuery.DQSetUpdateConfig('wireless',m,data_str)
                gp_str = portal_str(request.user.groupname)
                task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac,'mac_list')
                policy_str = user_policy_str(request.user.groupname)
                task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,mac,'mac_list')
                if compare_group_audit_setting(pgid,old_pgid):
                    pass
                    # for i in mac:
                    #     m = change_mac(i)
                    #     Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid,account_group_name = user_area)
                else:
                    # issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                    for i in mac:
                        m = change_mac(i)
                        # Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid,account_group_name = user_area)
                        # if issue_config_switch == 'on':
                        astr = audit_str(m)
                        DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                error = {"error_type":'success',"error_msg":_(u"添加成功")}
            except Exception as e:
                print e
                error = {"error_type":'failed',"error_msg":_(u"添加失败")}
        if error['error_type'] == "success":
            request.session['error_dict'] = error
        return JsonResponse(error,safe = False)

def change_mac(mac):
    return ''.join(mac.lower().split('-'))

@login_required
def nonoperate_group_del(request):
    group_id = request.GET.get('group_id')
    pg = Probe_group.objects.get(pk = group_id)
    try:
        if pg.account_group_name == "admin":
            default_pg = Probe_group.objects.get(account_group_name = 'admin',group_name = "DefaultGroup",group_type = '3')
        else:
            default_pg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '3')

        a = Device.objects.filter(support_mode = '3').filter(group_id = pg.pk)
        mac_list = []
        for i in a :
            mac_list.append(i.mac)
            i.device_wlan_set.all().delete()
        Device.objects.filter(support_mode = '3').filter(group_id = pg.pk).update(group_id = default_pg.pk)
        if compare_group_wlan_settions(pg.pk,default_pg.pk):
            pass
        else:
            for i in a:
                create_wireless_json(i)
                #audit config str
                # data_str = set_wlan_str(i.mac)
                # DataQuery.DQSetUpdateConfig('wireless',i.mac,data_str)
        gp_str = portal_str(request.user.groupname)
        task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')
        policy_str = user_policy_str(request.user.groupname)
        task_for_user_policy_config_ajax.delay(request.user.groupname,policy_str,mac_list,'mac_list')
        if compare_group_audit_setting(pg.pk,default_pg.pk):
            pass
        else:
            # issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
            for i in mac_list:
                # if issue_config_switch == 'on':
                    #audit config str
                astr = audit_str(i)
                DataQuery.DQSetUpdateConfig("audit", i, astr.encode("utf-8"))
        pg.delete()
        error = {"error_type":'success',"error_msg":_(u"删除成功")}
    except Exception as e:
        print e
        error = {"error_type":'failed',"error_msg":_(u"删除失败")}

    request.session['error_dict'] = error
    return redirect('/nonoperate/nonoperate_group/')


@login_required
def remove_nonoperate_ajax(request):
    group_id = request.GET.get("group_id")
    pg = Probe_group.objects.get(pk = group_id)
    device = {'data':[]}
    if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
        device_quaryset = Device.objects.filter(support_mode = '3').filter(Q(account_group_name = pg.account_group_name,group_id = pg.pk) |Q(account_group_name = '',group_id = 0))
    else:
        device_quaryset = Device.objects.filter(support_mode = '3').filter(account_group_name = pg.account_group_name,group_id = pg.pk)
    if device_quaryset.count() != 0:
        for i in device_quaryset:
            if i.name != '':
                name = i.name
            else:
                name = i.model + '_' + i.mac[-6:]
            mac = i.mac.upper()[:2] + '-' + i.mac.upper()[2:4] + '-' + i.mac.upper()[4:6] + '-' + i.mac.upper()[6:8] + '-' + i.mac.upper()[8:10] + '-' + i.mac.upper()[10:12]

            if request.user.administrator_permission == 3 or request.user.administrator_permission == 2 or request.user.administrator_permission == 1:
                if pg.group_name == pg.account_group_name :
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'off'})
                else:
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'on'})
            else:
                if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'off'})
                else:
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'on'})
    return JsonResponse(device,safe = False)


@login_required
def add_nonoperate_ajax(request):
    group_id = request.GET.get("group_id")
    pg = Probe_group.objects.get(pk = group_id)
    device = {'data':[]}
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        dpg = Probe_group.objects.get(group_name = "DefaultGroup" ,account_group_name = "admin",group_type = '3')
        if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
            device_quaryset = ""
        elif pg.group_name != "DefaultGroup" and pg.account_group_name == "admin":
            device_quaryset = Device.objects.filter(support_mode = '3').filter(Q(account_group_name = '')|Q(account_group_name = 'admin') , Q(group_id = dpg.pk)|Q(group_id = 0))
        elif pg.account_group_name != "admin" and pg.group_name == pg.account_group_name:
            ptdpg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '3')
            device_quaryset = Device.objects.filter(support_mode = '3').filter(Q(account_group_name = '')|Q(account_group_name = 'admin') , Q(group_id = dpg.pk)|Q(group_id = 0))
        elif pg.account_group_name != "admin" and pg.group_name != pg.account_group_name:
            ptdpg = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '3')
            device_quaryset = Device.objects.filter(support_mode = '3').filter(account_group_name = ptdpg.account_group_name,group_id = ptdpg.pk)
    elif request.user.administrator_permission != 0 and request.user.administrator_permission < 4:
        ptdpg = Probe_group.objects.get(Q(account_group_name = pg.account_group_name) & Q(group_name = pg.account_group_name) & Q(group_type = '3'))
        if pg.pk == ptdpg.pk:
            device_quaryset = ""
        elif pg.pk != ptdpg.pk:
            device_quaryset = Device.objects.filter(support_mode = '3').filter(account_group_name = ptdpg.account_group_name,group_id = ptdpg.pk)
    if device_quaryset != "":
        if device_quaryset.count() != 0:
            for i in device_quaryset:
                if i.name != '':
                    name = i.name
                else:
                    name = i.model + '_' + i.mac[-6:]
                mac = i.mac.upper()[:2] + '-' + i.mac.upper()[2:4] + '-' + i.mac.upper()[4:6] + '-' + i.mac.upper()[6:8] + '-' + i.mac.upper()[8:10] + '-' + i.mac.upper()[10:12]
                if pg.group_name == "DefaultGroup" and pg.account_group_name == "admin":
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'off'})
                else:
                    device['data'].append({'name':name,'mac':mac,'action':'','sign':'on'})
    return JsonResponse(device,safe = False)


@login_required
def add_nonoperate_wlan_in_group_ajax(request):
    group_id = request.GET.get("group_id")
    pg = Probe_group.objects.get(pk = group_id)
    device = {'data':[]}
    pgwl = pg.group_wlan_set.all()
    for i in pgwl:
        device['data'].append({'name':i.wlan_ssid,'sec_type':i.sec_type,'guest_enabled':i.guest_enabled,'wlan_service':i.wlan_service,'pk':i.pk})
    return JsonResponse(device,safe = False)

@login_required
def nonoperate_setting_ajax(request):
    group_id = request.GET.get('group_id')
    pg = Probe_group.objects.get(pk = group_id)
    return JsonResponse(model_to_dict(pg),safe = False)

@login_required
def modify_nonoperate_group_ajax(request):
    sign = 0
    error = {"error_type":'',"error_msg":""}
    if request.method == "POST":
        modify_device = json.loads(request.POST.get('modify_device'))
        group_id = int(modify_device['id'])
        groupname = modify_device['groupname']
        area_name = modify_device['area_name']
        remove_device_mac = modify_device['remove_device_mac']
        add_device_mac = modify_device['add_device_mac']
        setting = modify_device['setting']
        add_wlan = modify_device['add_wlan']
        if Probe_group.objects.exclude(pk = group_id).filter(Q(account_group_name = area_name) & Q(group_name = groupname)&Q(group_type = '3')).exists():
            error = {"error_type":'failed',"error_msg":_(u"组名已存在")}
        else:
            try:
                pg = Probe_group.objects.get(pk = group_id , account_group_name = area_name,group_type = '3')
                pg.group_name = groupname

                ss = model_to_dict(pg)
                ss.pop('id')
                ss.pop('group_name')
                ss.pop('account_group_name')
                ss.pop('device_count')
                ss.pop('group_type')
                ss.pop('wlan_count')

                for k,v in setting.items():
                    if v == None:
                        setting[k] = ""
                print setting
                if ss == setting and add_wlan == {}:
                    sign = 0
                    pass
                else:
                    pg.audit_corp = setting['audit_corp']
                    pg.audit_ip = setting['audit_ip']

                    pg.ftp_name = setting['ftp_name']
                    pg.ftp_passwd = setting['ftp_passwd']
                    pg.ftp_port = setting['ftp_port']

                    pg.audit_port = setting['audit_port']
                    pg.location_encode = setting['location_encode']
                    pg.device_encode = setting['device_encode']
                    pg.longitude = setting['longitude']
                    pg.latitude = setting['latitude']
                    pg.collection_radius = setting['collection_radius']
                    pg.collection_equipment_type = setting['collection_equipment_type']
                    pg.collection_equipment_name = setting['collection_equipment_name']
                    pg.collection_equipment_address = setting['collection_equipment_address']
                    pg.security_software_orgcode = setting['security_software_orgcode']
                    pg.security_software_orgname = setting['security_software_orgname']
                    pg.security_software_address = setting['security_software_address']
                    pg.contactor = setting['contactor']
                    pg.contactor_tel = setting['contactor_tel']
                    pg.contactor_mail = setting['contactor_mail']

                    pg.place_name = setting['place_name']
                    pg.site_address = setting['site_address']
                    pg.netsite_type = setting['netsite_type']
                    pg.bussiness_nature = setting['bussiness_nature']
                    pg.law_principal_name = setting['law_principal_name']
                    pg.law_principal_certificate_type = setting['law_principal_certificate_type']
                    pg.law_principal_certificate_id = setting['law_principal_certificate_id']
                    pg.relationship_account = setting['relationship_account']
                    pg.start_time = setting['start_time']
                    pg.end_time = setting['end_time']
                    pg.police_station_code = setting['police_station_code']

                    sign = 1

                    for k,v in add_wlan.iteritems():
                        # try:
                        i = pg.group_wlan_set.get(pk = int(k))
                        i.wlan_service = v
                        i.save()
                        # except Exception as e:
                        #     print e
                pg.save()



                if remove_device_mac == {}:
                    pass
                else:
                    if request.user.administrator_permission == 1 or request.user.administrator_permission == 2 or request.user.administrator_permission == 3:
                        pgid = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '3').pk
                        if compare_group_wlan_settions(pg.pk,pgid):
                            for i in remove_device_mac:
                                m = change_mac(i)
                                Device.objects.get(mac = m).device_wlan_set.all().delete()
                                Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid)
                        else:
                            for i in remove_device_mac:
                                m = change_mac(i)
                                Device.objects.get(mac = m).device_wlan_set.all().delete()
                                Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid)
                                create_wireless_json(Device.objects.get(mac = m))
                                # data_str = set_wlan_str(m)
                                # DataQuery.DQSetUpdateConfig('wireless',m,data_str)
                                #audit config str

                        gp_str = portal_str(pg.account_group_name)
                        task_for_global_config_ajax.delay(pg.account_group_name,gp_str,remove_device_mac,'mac_list')
                        policy_str = user_policy_str(pg.account_group_name)
                        task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,remove_device_mac,'mac_list')
                        if compare_group_audit_setting(pg.pk,pgid):
                            pass
                            # for i in remove_device_mac:
                            #     m = change_mac(i)
                            #     Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid)
                        else:
                            # issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                            for i in remove_device_mac:
                                m = change_mac(i)
                                # Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid)
                                # if issue_config_switch == "on":
                                    #audit config str
                                astr = audit_str(m)
                                DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                    else:
                        if pg.account_group_name == "admin" or pg.group_name == pg.account_group_name:
                            pgid = Probe_group.objects.get(account_group_name = "admin",group_name = "DefaultGroup",group_type = '3').pk
                            if compare_group_wlan_settions(pg.pk,pgid):
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.get(mac = m).device_wlan_set.all().delete()
                                    Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid,account_group_name = "admin")
                                    update_device2authpuppy(m,'admin')
                            else:
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.get(mac = m).device_wlan_set.all().delete()
                                    Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid,account_group_name = "admin")

                                    update_device2authpuppy(m,'admin')
                                    create_wireless_json(Device.objects.get(mac = m))
                                    #audit config str
                                    # data_str = set_wlan_str(m)
                                    # DataQuery.DQSetUpdateConfig('wireless',m,data_str)

                            gp_str = portal_str(pg.account_group_name)
                            task_for_global_config_ajax.delay(pg.account_group_name,gp_str,remove_device_mac,'mac_list')
                            policy_str = user_policy_str(pg.account_group_name)
                            task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,remove_device_mac,'mac_list')
                            if compare_group_audit_setting(pg.pk,pgid):
                                pass
                                # for i in remove_device_mac:
                                #     m = change_mac(i)
                                #     Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid,account_group_name = "admin")
                            else:
                                # issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    # Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid,account_group_name = "admin")
                                    # if issue_config_switch == "on":
                                        #audit config str
                                    astr = audit_str(m)
                                    DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                        else:
                            pgid = Probe_group.objects.get(account_group_name = pg.account_group_name,group_name = pg.account_group_name,group_type = '3').pk
                            if compare_group_wlan_settions(pg.pk,pgid):
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.get(mac = m).device_wlan_set.all().delete()
                                    Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid)
                            else:
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    Device.objects.get(mac = m).device_wlan_set.all().delete()
                                    Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid)
                                    create_wireless_json(Device.objects.get(mac = m))
                                    #audit config str
                                    # data_str = set_wlan_str(m)
                                    # DataQuery.DQSetUpdateConfig('wireless',m,data_str)

                            gp_str = portal_str(pg.account_group_name)
                            task_for_global_config_ajax.delay(pg.account_group_name,gp_str,remove_device_mac,'mac_list')
                            policy_str = user_policy_str(pg.account_group_name)
                            task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,remove_device_mac,'mac_list')
                            if compare_group_audit_setting(pg.pk,pgid):
                                pass
                                # for i in remove_device_mac:
                                #     m = change_mac(i)
                                #     Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pgid)
                            else:
                                # issue_config_switch = Account_Group.objects.get(groupname = request.user.groupname).issue_config_switch
                                for i in remove_device_mac:
                                    m = change_mac(i)
                                    # Device.objects.filter(support_mode = '2').filter(mac = m).update(group_id = pgid)
                                    # if issue_config_switch == "on":
                                        #audit config str
                                    astr = audit_str(m)
                                    DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                if add_device_mac == {}:
                    pass
                else:
                    for i in add_device_mac:
                        m = change_mac(i)
                        p = Device.objects.get(mac = m).group_id
                        Device.objects.get(mac = m).device_wlan_set.all().delete()
                        Device.objects.filter(support_mode = '3').filter(mac = m).update(group_id = pg.pk,account_group_name = pg.account_group_name)
                        update_device2authpuppy(m,pg.account_group_name)
                        if compare_group_wlan_settions(pg.pk,p):
                            pass
                        else:
                            create_wireless_json(Device.objects.get(mac = m))
                            # audit config str
                            # data_str = set_wlan_str(m)
                            # DataQuery.DQSetUpdateConfig('wireless',m,data_str)
                        if compare_group_audit_setting(pg.pk,p):
                            pass
                        else:
                            # if issue_config_switch == 'on':
                                #audit config str
                            astr = audit_str(m)
                            DataQuery.DQSetUpdateConfig("audit", m, astr.encode("utf-8"))
                    gp_str = portal_str(pg.account_group_name)
                    task_for_global_config_ajax.delay(pg.account_group_name,gp_str,add_device_mac,'mac_list')
                    policy_str = user_policy_str(pg.account_group_name)
                    task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,add_device_mac,'mac_list')

                if sign == 0:
                    pass
                else:
                    task_for_modify_nonoperate_group_ajax.delay(pg.pk)
                    gp_str = portal_str(pg.account_group_name)
                    mac_list = []
                    for i in Device.objects.filter(group_id = pg.pk):
                        mac_list.append(i.mac)
                    task_for_global_config_ajax.delay(pg.account_group_name,gp_str,mac_list,'mac_list')
                    policy_str = user_policy_str(pg.account_group_name)
                    task_for_user_policy_config_ajax.delay(pg.account_group_name,policy_str,mac_list,'mac_list')
                    # for i in Device.objects.filter(support_mode = '3').filter(group_id = pg.pk):
                    #     m = change_mac(i.mac)

                    #      #audit config str
                    #     data_str = set_wlan_str(m)
                    #     DataQuery.DQSetUpdateConfig('wireless',m,data_str)
                error = {"error_type":'success',"error_msg":_(u"编辑成功")}
                if pg.account_group_name == "admin" and pg.group_name == 'DefaultGroup':
                    default_config_issued.delay()
            except Exception as e:
                print e
                error = {"error_type":'failed',"error_msg":_(u"编辑失败")}
        if error['error_type'] == "success":
            request.session['error_dict'] = error
        return JsonResponse(error,safe = False)


def compare_group_audit_setting(gp1,gp2):
    try:
        g1 = model_to_dict(Probe_group.objects.get(pk = int(gp1)))
        g2 = model_to_dict(Probe_group.objects.get(pk = int(gp2)))
        g1.pop('id')
        g1.pop('group_name')
        g1.pop('account_group_name')
        g1.pop('device_count')
        g1.pop('group_type')
        g2.pop('id')
        g2.pop('group_name')
        g2.pop('account_group_name')
        g2.pop('device_count')
        g2.pop('group_type')

        for i in g1:
            if g1[i] == None:
                g1[i] = ""
        for i in g2:
            if g2[i] == None:
                g2[i] = ""
        return g1 == g2
    except Exception as e:
        print e
        return False

def compare_group_wlan_settions(gp1,gp2):
    try:
        g1 = {}
        g2 = {}
        for i in Group_wlan.objects.filter(group_id_id = int(gp1)).order_by('wlan_id'):
            g1[str(i.wlan_id)] = model_to_dict(i)
            g1[str(i.wlan_id)].pop('group_id')
            g1[str(i.wlan_id)].pop('id')
        for i in Group_wlan.objects.filter(group_id_id = int(gp2)).order_by('wlan_id'):
            g2[str(i.wlan_id)] = model_to_dict(i)
            g2[str(i.wlan_id)].pop('group_id')
            g2[str(i.wlan_id)].pop('id')
        print g1 == g2 ,'(test)'
        return g1 == g2
    except Exception as e:
        print e
        return False

def audit_str(mac):

    astr = '{"audit":{'
    print mac
    # dev = model_to_dict(Probe_audit_dev_status.objects.get(mac = mac))
    dev = audit_setting_compare(mac,1,"dev")
    dev.pop('mac')
    dev.pop('id')
    DataQuery.DQClearNullValeForDict(dev)
    dstr = json.dumps(dev, sort_keys=True, ensure_ascii=False)
    dstr = dstr.replace(" ", '')
    #print dstr
    astr = astr + '"device":' + dstr + ','

    # place = model_to_dict(Probe_audit_place_status.objects.get(mac = mac))
    place = audit_setting_compare(mac,1,"place")
    place.pop('mac')
    place.pop('id')
    DataQuery.DQClearNullValeForDict(place)
    pstr = json.dumps(place, sort_keys=True, ensure_ascii=False)
    pstr = pstr.replace(" ", '')
    #print pstr
    astr = astr + '"site":' + pstr + ','

    # basic = model_to_dict(Probe_audit_basic_status.objects.get(mac = mac))
    basic = audit_setting_compare(mac,1,"basic")
    basic.pop('mac')
    basic.pop('id')
    DataQuery.DQClearNullValeForDict(basic)
    bstr = json.dumps(basic, sort_keys=True, ensure_ascii=False)
    bstr = bstr.replace(" ", '')
    #print bstr
    astr = astr + '"system":' + bstr + "}}"

    # DataQuery.DQSetUpdateConfig("audit", mac, astr.encode("utf-8"))
    return astr

def audit_setting_compare(mac,model_type,kind):
    model_type = int(model_type)
    audit_setting = ''
    dev = Device.objects.get(mac = mac)
    if model_type == 1:
        if dev.group_id == 0:
            group = Probe_group.objects.get(group_name = 'DefaultGroup',account_group_name = 'admin',group_type = '3')
        else:
            group = Probe_group.objects.get(pk = dev.group_id)

        if kind == 'dev':
            try:
                audit_setting = Probe_audit_dev_status.objects.get(mac = mac)

                audit_setting.collection_radius = group.collection_radius if audit_setting.collection_radius == '' else audit_setting.collection_radius
                audit_setting.collection_equipment_type = group.collection_equipment_type if audit_setting.collection_equipment_type == '' else audit_setting.collection_equipment_type
                audit_setting.collection_equipment_name = group.collection_equipment_name if audit_setting.collection_equipment_name == '' else audit_setting.collection_equipment_name
                audit_setting.collection_equipment_address = group.collection_equipment_address if audit_setting.collection_equipment_address == '' else audit_setting.collection_equipment_address
                audit_setting.security_software_orgcode = group.security_software_orgcode if audit_setting.security_software_orgcode == '' else audit_setting.security_software_orgcode
                audit_setting.security_software_orgname = group.security_software_orgname if audit_setting.security_software_orgname == '' else audit_setting.security_software_orgname
                audit_setting.security_software_address = group.security_software_address if audit_setting.security_software_address == '' else audit_setting.security_software_address
                audit_setting.contactor = group.contactor if audit_setting.contactor == '' else audit_setting.contactor
                audit_setting.contactor_tel = group.contactor_tel if audit_setting.contactor_tel == '' else audit_setting.contactor_tel
                audit_setting.contactor_mail = group.contactor_mail if audit_setting.contactor_mail == '' else audit_setting.contactor_mail
                auds = model_to_dict(audit_setting)
            except Exception as e:
                print e
                audit_setting = {'collection_radius':'','collection_equipment_type':'','collection_equipment_name':'','collection_equipment_address':'','security_software_orgcode':'','security_software_orgname':'','security_software_address':'','contactor':'','contactor_tel':'','contactor_mail':'','mac':'','id':''}

                audit_setting['collection_radius'] = group.collection_radius if audit_setting['collection_radius'] == '' else audit_setting['collection_radius']
                audit_setting['collection_equipment_type'] = group.collection_equipment_type if audit_setting['collection_equipment_type'] == '' else audit_setting['collection_equipment_type']
                audit_setting['collection_equipment_name'] = group.collection_equipment_name if audit_setting['collection_equipment_name'] == '' else audit_setting['collection_equipment_name']
                audit_setting['collection_equipment_address'] = group.collection_equipment_address if audit_setting['collection_equipment_address'] == '' else audit_setting['collection_equipment_address']
                audit_setting['security_software_orgcode'] = group.security_software_orgcode if audit_setting['security_software_orgcode'] == '' else audit_setting['security_software_orgcode']
                audit_setting['security_software_orgname'] = group.security_software_orgname if audit_setting['security_software_orgname'] == '' else audit_setting['security_software_orgname']
                audit_setting['security_software_address'] = group.security_software_address if audit_setting['security_software_address'] == '' else audit_setting['security_software_address']
                audit_setting['contactor'] = group.contactor if audit_setting['contactor'] == '' else audit_setting['contactor']
                audit_setting['contactor_tel'] = group.contactor_tel if audit_setting['contactor_tel'] == '' else audit_setting['contactor_tel']
                audit_setting['contactor_mail'] = group.contactor_mail if audit_setting['contactor_mail'] == '' else audit_setting['contactor_mail']

                auds = audit_setting

            for k in auds:
                if auds[k] == None:
                    auds[k] = ""
            return auds

        if kind == 'place':
            print 'xt1tx1'
            if Probe_audit_basic_status.objects.filter(mac = mac).exists():
                place_audit_corp = Probe_audit_basic_status.objects.get(mac = mac).audit_corp
            else:
                place_audit_corp = group.audit_corp
            print 'xt2xt2',place_audit_corp
            try:
                audit_setting = Probe_audit_place_status.objects.get(mac = mac)

                audit_setting.place_name = group.place_name if audit_setting.place_name == '' else audit_setting.place_name
                audit_setting.site_address = group.site_address if audit_setting.site_address == '' else audit_setting.site_address
                audit_setting.netsite_type = group.netsite_type if audit_setting.netsite_type == '' else audit_setting.netsite_type
                audit_setting.bussiness_nature = group.bussiness_nature if audit_setting.bussiness_nature == '' else audit_setting.bussiness_nature
                audit_setting.law_principal_name = group.law_principal_name if audit_setting.law_principal_name == '' else audit_setting.law_principal_name
                audit_setting.law_principal_certificate_type = group.law_principal_certificate_type if audit_setting.law_principal_certificate_type == '' else audit_setting.law_principal_certificate_type
                audit_setting.law_principal_certificate_id = group.law_principal_certificate_id if audit_setting.law_principal_certificate_id == '' else audit_setting.law_principal_certificate_id
                audit_setting.relationship_account = group.relationship_account if audit_setting.relationship_account == '' else audit_setting.relationship_account
                audit_setting.start_time = group.start_time if audit_setting.start_time == '' else audit_setting.start_time
                audit_setting.end_time = group.end_time if audit_setting.end_time == '' else audit_setting.end_time

                audit_setting.site_type = group.site_type if audit_setting.site_type == '' else audit_setting.site_type
                audit_setting.police_station_code = group.police_station_code if audit_setting.police_station_code == '' else audit_setting.police_station_code

                a_s = model_to_dict(audit_setting)
                print 'tttttt11111',a_s
                if place_audit_corp == '27':
                    a_s['policestation'] = a_s['police_station_code']
                    a_s['area'] = a_s['site_type']
                if place_audit_corp == '29':
                    a_s['policestation'] = a_s['police_station_code']
                del a_s['site_type']
                del a_s['police_station_code']

                auds = a_s
                for k in auds:
                    if auds[k] == None:
                        auds[k] = ""
                print 'tttttt22222',a_s
                return auds

            except Exception as e:
                print e,
                print 'xt3xt3'
                audit_setting = {'place_name':'','site_address':'','netsite_type':'','bussiness_nature':'','law_principal_name':'','law_principal_certificate_type':'','law_principal_certificate_id':'','relationship_account':'','start_time':'','end_time':'','mac':'','id':'','site_type':'','police_station_code':''}

                audit_setting['place_name'] = group.place_name if audit_setting['place_name'] == '' else audit_setting['place_name']
                audit_setting['site_address'] = group.site_address if audit_setting['site_address'] == '' else audit_setting['site_address']
                audit_setting['netsite_type'] = group.netsite_type if audit_setting['netsite_type'] == '' else audit_setting['netsite_type']
                audit_setting['bussiness_nature'] = group.bussiness_nature if audit_setting['bussiness_nature'] == '' else audit_setting['bussiness_nature']
                audit_setting['law_principal_name'] = group.law_principal_name if audit_setting['law_principal_name'] == '' else audit_setting['law_principal_name']
                audit_setting['law_principal_certificate_type'] = group.law_principal_certificate_type if audit_setting['law_principal_certificate_type'] == '' else audit_setting['law_principal_certificate_type']
                audit_setting['law_principal_certificate_id'] = group.law_principal_certificate_id if audit_setting['law_principal_certificate_id'] == '' else audit_setting['law_principal_certificate_id']
                audit_setting['relationship_account'] = group.relationship_account if audit_setting['relationship_account'] == '' else audit_setting['relationship_account']
                audit_setting['start_time'] = group.start_time if audit_setting['start_time'] == '' else audit_setting['start_time']
                audit_setting['end_time'] = group.end_time if audit_setting['end_time'] == '' else audit_setting['end_time']
                audit_setting['site_type'] = group.site_type if audit_setting['site_type'] == '' else audit_setting['site_type']
                audit_setting['police_station_code'] = group.police_station_code if audit_setting['police_station_code'] == '' else audit_setting['police_station_code']

                if place_audit_corp == '27':
                    audit_setting['policestation'] = audit_setting['police_station_code']
                    audit_setting['area'] = audit_setting['site_type']
                if place_audit_corp == '29':
                    audit_setting['policestation'] = audit_setting['police_station_code']
                del audit_setting['site_type']
                del audit_setting['police_station_code']

                auds = audit_setting
                for k in auds:
                    if auds[k] == None:
                        auds[k] = ""
                return auds

        if kind == 'basic':
            try:
                audit_setting = Probe_audit_basic_status.objects.get(mac = mac)

                audit_setting.audit_corp = group.audit_corp if audit_setting.audit_corp == '' else audit_setting.audit_corp

                audit_setting.ftp_name = group.ftp_name if audit_setting.ftp_name == '' else audit_setting.ftp_name
                audit_setting.ftp_passwd = group.ftp_passwd if audit_setting.ftp_passwd == '' else audit_setting.ftp_passwd
                audit_setting.ftp_port = group.ftp_port if audit_setting.ftp_port == '' else audit_setting.ftp_port

                audit_setting.audit_ip = group.audit_ip if audit_setting.audit_ip == '' else audit_setting.audit_ip
                audit_setting.audit_port = group.audit_port if audit_setting.audit_port == '' else audit_setting.audit_port
                audit_setting.location_encode = group.location_encode if audit_setting.location_encode == '' else audit_setting.location_encode
                audit_setting.device_encode = group.device_encode if audit_setting.device_encode == '' else audit_setting.device_encode
                audit_setting.longitude = group.longitude if audit_setting.longitude == '' else audit_setting.longitude
                audit_setting.latitude = group.latitude if audit_setting.latitude == '' else audit_setting.latitude
                audit_setting.ssid = json.loads(group.ssid)
                # print audit_setting.ssid,'QWEQWEQWEWQEQ'
                a_s = model_to_dict(audit_setting)
                if a_s['ssid'] == "":
                    a_s['ssid'] = []
                if dev.own_model != 'WitMAX-P550E-L':
                    del a_s['ssid']

                # renzixing,wangbo,chongqingaisi,byzoro feijing
                if audit_setting.audit_corp == '50':
                    pass
                # elif audit_setting.audit_corp == '3' or audit_setting.audit_corp == '15' or audit_setting.audit_corp == '22' or audit_setting.audit_corp == '27' or audit_setting.audit_corp == '28':
                elif audit_setting.audit_corp in Public_function.auditCropLimitList:
                    del a_s['ftp_port']
                else:
                    del a_s['ftp_port']
                    del a_s['ftp_name']
                    del a_s['ftp_passwd']

                for k in a_s:
                    if a_s[k] == None:
                        a_s[k] = ""
                print a_s
                return a_s

                # return model_to_dict(audit_setting)

            except Exception as e:
                print e
                audit_setting = {'audit_corp':'','ftp_name':'', 'ftp_passwd':'', 'ftp_port':'', 'audit_ip':'','audit_port':'','location_encode':'','device_encode':'','longitude':'','latitude':'','mac':'','id':'','ssid':''}

                audit_setting['audit_corp'] = group.audit_corp if audit_setting['audit_corp'] == '' else audit_setting['audit_corp']
                # renzixing,wangbo,chongqingaisi,byzoro feijing
                # if audit_setting['audit_corp'] == '3' or audit_setting['audit_corp'] == '15' ||  audit_setting['audit_corp'] == '22' || audit_setting['audit_corp'] == '50':
                audit_setting['ftp_name'] = group.ftp_name if audit_setting['ftp_name'] == '' else audit_setting['ftp_name']
                audit_setting['ftp_passwd'] = group.ftp_passwd if audit_setting['ftp_passwd'] == '' else audit_setting['ftp_passwd']
                audit_setting['ftp_port'] = group.ftp_port if audit_setting['ftp_port'] == '' else audit_setting['ftp_port']

                audit_setting['audit_ip'] = group.audit_ip if audit_setting['audit_ip'] == '' else audit_setting['audit_ip']
                audit_setting['audit_port'] = group.audit_port if audit_setting['audit_port'] == '' else audit_setting['audit_port']
                audit_setting['location_encode'] = group.location_encode if audit_setting['location_encode'] == '' else audit_setting['location_encode']
                audit_setting['device_encode'] = group.device_encode if audit_setting['device_encode'] == '' else audit_setting['device_encode']
                audit_setting['longitude'] = group.longitude if audit_setting['longitude'] == '' else audit_setting['longitude']
                audit_setting['latitude'] = group.latitude if audit_setting['latitude'] == '' else audit_setting['latitude']
                audit_setting['ssid'] = json.loads(group.ssid)
                # print audit_setting['ssid'],'ASDASDADS'

                if audit_setting['ssid'] == "":
                    audit_setting['ssid'] = []
                if dev.own_model != 'WitMAX-P550E-L':
                    del audit_setting['ssid']

                # renzixing,wangbo,chongqingaisi,byzoro feijing
                if audit_setting['audit_corp'] == '50':
                    pass
                # elif audit_setting['audit_corp'] == '3' or audit_setting['audit_corp'] == '15' or audit_setting['audit_corp'] == '22' or audit_setting['audit_corp'] == '27' or audit_setting['audit_corp'] == '28':
                elif audit_setting['audit_corp'] in Public_function.auditCropLimitList:
                    del audit_setting['ftp_port']
                else:
                    del audit_setting['ftp_name']
                    del audit_setting['ftp_passwd']
                    del audit_setting['ftp_port']

                auds = audit_setting
                for k in auds:
                    if auds[k] == None:
                        auds[k] = ""
                print auds
                return auds
                # return audit_setting


@login_required
def nonoperate_wlan(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)

    wlan = ''
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        wlan = Group_wlan.objects.filter(group_id__group_type = '3')
    elif request.user.administrator_permission < 4 and request.user.group_status == 1:
        wlan = Group_wlan.objects.filter(group_id__account_group_name = request.user.groupname,group_id__group_type = '3')
    for wl in wlan:
        if wl.guest_enabled == "off":
            wl.auth_type = 'off'

    plimit = get_plimit()
    return render(request,'nonoperate-wlan.html',{'admin_error_json':admin_error_json,'admin_error':admin_error,'wlan':wlan,'error':error,"error_json":error_json,'plimit':plimit})

@login_required
def choose_nonoperate_group_ajax(request):
    gp = []
    # {'group_id':'','group_account_group_name':'','group_name':'',}
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        group = Probe_group.objects.filter(group_type = "3")
    elif request.user.administrator_permission < 4 and request.user.group_status == 1:
        group = Probe_group.objects.filter(group_type = "3",account_group_name = request.user.groupname)
    for i in group:
        if i.group_wlan_set.all().count() >= 4:
            gp.append({'group_id':i.pk,'group_account_group_name':i.account_group_name,'group_name':i.group_name,'disabled':'disabled'})
        else:
            gp.append({'group_id':i.pk,'group_account_group_name':i.account_group_name,'group_name':i.group_name,'disabled':'enabled'})
    return JsonResponse(gp,safe=False)

@login_required
def add_nonoperate_wlan_ajax(request):
    errors = {'sign':'false','message':''}
    if request.method == 'POST':
        logger.info('start add wlan ->run add wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        group_id = int(request.POST.get('group_id'))
        a = Probe_group.objects.get(id = group_id)
        if not a.group_wlan_set.filter(wlan_id = '0').exists():
            wlan_id = '0'
        elif not a.group_wlan_set.filter(wlan_id = '1').exists():
            wlan_id = '1'
        elif not a.group_wlan_set.filter(wlan_id = '2').exists():
            wlan_id = '2'
        elif not a.group_wlan_set.filter(wlan_id = '3').exists():
            wlan_id = '3'
        wlan_ssid = request.POST.get('ssid')
        if request.POST.get('wlan_service') == None:
            wlan_service = "off"
        else:
            wlan_service = request.POST.get('wlan_service')
        if request.POST.get('safe') == '1':
            sec_type = 'open'
            encry_type = ''
        elif request.POST.get('safe') == '2':
            sec_type = 'psk2'
            encry_type = 'tkip'
        else:
            sec_type = 'psk2'
            encry_type = 'ccmp'
        passphrase = request.POST.get('passphrase')
        if request.POST.get('guest_enabled') == None:
            guest_enabled = "off"
        else:
            guest_enabled = request.POST.get('guest_enabled')
        if request.POST.get('vlan_enabled') == None:
            vlan_enabled = "off"
        else:
            vlan_enabled = request.POST.get('vlan_enabled')
        if request.POST.get('vlan') == '':
            vlan = '0'
        else:
            vlan = request.POST.get('vlan')
        if request.POST.get('hidden_ssid') == None:
            hidden_ssid = "off"
        else:
            hidden_ssid = request.POST.get('hidden_ssid')
        upload_speed = request.POST.get('upload_speed')
        download_speed = request.POST.get('download_speed')
        radios_enable = request.POST.get('radios_enable')
        auth_type = request.POST.get('auth_type')

        wechat_appid = request.POST.get('wechat_appid')
        wechat_appkey = request.POST.get('wechat_appkey')
        wechat_shopid = request.POST.get('wechat_shopid')
        wechat_secretkey = request.POST.get('wechat_secretkey')
        if request.POST.get('wechat_forcefollow') == None:
            wechat_forcefollow = "false"
        else:
            wechat_forcefollow = 'true'

        # auth_server_hostname = request.POST.get('group_id')
        auth_server_loginurl = request.POST.get('auth_server_loginurl')
        auth_server_portalurl = request.POST.get('auth_server_portalurl')
        a = auth_server_loginurl.split('://')
        if len(a) >= 2:
            auth_server_hostname = a[1].split('/')[0]
        else:
            auth_server_hostname = ''
        # print group_id,"###",wlan_ssid,"###",wlan_service,"###",safe,"###",passphrase,"###",guest_enabled,"###",vlan_enabled,"###",vlan,"###",hidden_ssid,"###",upload_speed,"###",download_speed,"###",radios_enable,"###",auth_type

        try:
            gw = Group_wlan()
            gw.group_id_id = group_id
            gw.wlan_id = wlan_id
            gw.wlan_ssid = wlan_ssid
            gw.wlan_service = wlan_service
            gw.sec_type = sec_type
            gw.encry_type = encry_type
            gw.passphrase = passphrase
            gw.guest_enabled = guest_enabled
            gw.vlan_enabled = vlan_enabled
            gw.vlan = vlan
            gw.hidden_ssid = hidden_ssid
            gw.upload_speed = upload_speed
            gw.download_speed = download_speed
            gw.radios_enable = radios_enable
            gw.auth_type = auth_type

            gw.wechat_appid = wechat_appid
            gw.wechat_appkey = wechat_appkey
            gw.wechat_shopid = wechat_shopid
            gw.wechat_secretkey = wechat_secretkey
            gw.wechat_forcefollow = wechat_forcefollow
            gw.auth_server_hostname = auth_server_hostname
            gw.auth_server_loginurl = auth_server_loginurl
            gw.auth_server_portalurl = auth_server_portalurl

            gw.save()
            errors['message'] = _(u'创建成功！')
            errors['sign'] = "true"
            logger.info('start add wlan ->end add wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

            task_for_add_wlan_ajax.delay(group_id)
            gp_str = portal_str(request.user.groupname)
            mac_list = []
            for i in Device.objects.filter(group_id = group_id):
                mac_list.append(i.mac)
            task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')
            # for i in Device.objects.filter(group_id = group_id):
            #     m = change_mac(i.mac)
            #     data_str = set_wlan_str(m)
            #     DataQuery.DQSetUpdateConfig('wireless',m,data_str)
            pg = Probe_group.objects.get(id = group_id)
            if pg.account_group_name == "admin" and pg.group_name == 'DefaultGroup':
                default_config_issued.delay()
            logger.info('start add wlan ->end DQ add wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        except Exception as e:
            print e
            errors['message'] = _(u'创建失败！')
            errors['sign'] = "false"


    return JsonResponse(errors,safe=False)

@login_required
def nonoperate_wlan_del(request):
    wlan_id = request.GET.get('wlan_id')
    try:
        logger.info('start del wlan ->run get wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        wl = Group_wlan.objects.get(pk = wlan_id)
        ###########################################
        ###删除组中的wlan
        #########################################
        logger.info('start del wlan ->end get wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        dev = Device.objects.filter(group_id = wl.group_id_id)
        logger.info('start del wlan ->end get dev --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        for d in dev:
            d.device_wlan_set.filter(wlan_id = wl.wlan_id).delete()
        logger.info('start del wlan ->end del device_wlan --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        logger.info('start del wlan ->print sql --- sql:{}'.format(connection.queries))
        wl.delete()
        logger.info('start del wlan ->end del wl --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        task_for_nonoperate_del_wlan.delay(wl.group_id_id)
        gp_str = portal_str(request.user.groupname)
        mac_list = []
        for i in Device.objects.filter(group_id = wl.group_id_id):
            mac_list.append(i.mac)
        task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')
        # for i in dev:
        #     m = change_mac(i.mac)
        #     data_str = set_wlan_str(m)
        #     DataQuery.DQSetUpdateConfig('wireless',m,data_str)
        pg = Probe_group.objects.get(id = wl.group_id_id)
        if pg.account_group_name == "admin" and pg.group_name == 'DefaultGroup':
            default_config_issued.delay()
        logger.info('start del wlan ->end DQconfig --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        error = {"error_type":'success',"error_msg":_(u"删除成功")}
    except Exception as e:
        print e
        error = {"error_type":'failed',"error_msg":_(u"删除失败")}

    request.session['error_dict'] = error
    return redirect('/nonoperate/nonoperate_wlan/')


@login_required
def nonoperate_wlan_info_ajax(request):
    wl_id = int(request.GET.get('id'))
    wlan = Group_wlan.objects.get(pk = wl_id)
    gp = {'gp':'','wl':''}
    gp['gp'] = {'group_id':wlan.group_id_id,'group_account_group_name':wlan.group_id.account_group_name,'group_name':wlan.group_id.group_name}
    gp['wl'] = model_to_dict(wlan)
    return JsonResponse(gp,safe=False)

@login_required
def nonoperate_modify_wlan_ajax(request):
    errors = {'sign':'false','message':''}
    if request.method == 'POST':
        wlan_ssid = request.POST.get('ssid')
        if request.POST.get('wlan_service') == None:
            wlan_service = "off"
        else:
            wlan_service = request.POST.get('wlan_service')
        print request.POST.get('safe'),123321
        if request.POST.get('safe') == '1':
            sec_type = 'open'
            encry_type = ''
        elif request.POST.get('safe') == '2':
            sec_type = 'psk2'
            encry_type = 'tkip'
        else:
            sec_type = 'psk2'
            encry_type = 'ccmp'
        passphrase = request.POST.get('passphrase')
        if request.POST.get('guest_enabled') == None:
            guest_enabled = "off"
        else:
            guest_enabled = request.POST.get('guest_enabled')
        if request.POST.get('vlan_enabled') == None:
            vlan_enabled = "off"
        else:
            vlan_enabled = request.POST.get('vlan_enabled')
        if request.POST.get('vlan') == '':
            vlan = '0'
        else:
            vlan = request.POST.get('vlan')
        if request.POST.get('hidden_ssid') == None:
            hidden_ssid = "off"
        else:
            hidden_ssid = request.POST.get('hidden_ssid')
        upload_speed = request.POST.get('upload_speed')
        download_speed = request.POST.get('download_speed')
        radios_enable = request.POST.get('radios_enable')
        auth_type = request.POST.get('auth_type')

        wechat_appid = request.POST.get('wechat_appid')
        wechat_appkey = request.POST.get('wechat_appkey')
        wechat_shopid = request.POST.get('wechat_shopid')
        wechat_secretkey = request.POST.get('wechat_secretkey')
        if request.POST.get('wechat_forcefollow') == None:
            wechat_forcefollow = "false"
        else:
            wechat_forcefollow = "true"

        # auth_server_hostname = request.POST.get('group_id')
        auth_server_loginurl = request.POST.get('auth_server_loginurl')
        auth_server_portalurl = request.POST.get('auth_server_portalurl')
        a = auth_server_loginurl.split('://')
        if len(a) >= 2:
            auth_server_hostname = a[1].split('/')[0]
        else:
            auth_server_hostname = ''

        pk = int(request.GET.get('pk'))
        try:
            gw = Group_wlan.objects.get(pk = pk)
            gw.wlan_ssid = wlan_ssid
            gw.wlan_service = wlan_service
            gw.sec_type = sec_type
            gw.encry_type = encry_type
            gw.passphrase = passphrase
            gw.guest_enabled = guest_enabled
            gw.vlan_enabled = vlan_enabled
            gw.vlan = vlan
            gw.hidden_ssid = hidden_ssid
            gw.upload_speed = upload_speed
            gw.download_speed = download_speed
            gw.radios_enable = radios_enable
            gw.auth_type = auth_type

            gw.wechat_appid = wechat_appid
            gw.wechat_appkey = wechat_appkey
            gw.wechat_shopid = wechat_shopid
            gw.wechat_secretkey = wechat_secretkey
            gw.wechat_forcefollow = wechat_forcefollow
            gw.auth_server_hostname = auth_server_hostname
            gw.auth_server_loginurl = auth_server_loginurl
            gw.auth_server_portalurl = auth_server_portalurl

            gw.save()
            errors['message'] = _(u'修改成功！')
            errors['sign'] = 'true'

            task_for_modify_wlan_ajax.delay(gw.group_id_id)
            gp_str = portal_str(request.user.groupname)
            mac_list = []
            for i in Device.objects.filter(group_id = gw.group_id_id):
                mac_list.append(i.mac)
            task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')
            pg = Probe_group.objects.get(id = gw.group_id_id)
            if pg.account_group_name == "admin" and pg.group_name == 'DefaultGroup':
                default_config_issued.delay()
            # for i in Device.objects.filter(group_id = gw.group_id_id):
            #     m = change_mac(i.mac)
            #     data_str = set_wlan_str(m)
            #     DataQuery.DQSetUpdateConfig('wireless',m,data_str)
        except Exception as e:
            print e
            errors['message'] = _(u'修改失败！')
            errors['sign'] = 'false'

    return JsonResponse(errors,safe=False)


@login_required
def nonoperate_wlan_list(request):
    ret = {'data':[]}
    ap_id = int(request.GET.get('id'))
    try:
        wlan = Probe_group.objects.get(pk = Device.objects.get(id = ap_id).group_id).group_wlan_set.all()
        for i in wlan:
            a = model_to_dict(i)
            x = ""
            try:
                apwl = Device_wlan.objects.get(device_id = ap_id,wlan_id = i.wlan_id)
                if apwl.wlan_ssid != i.wlan_ssid and apwl.wlan_ssid != "":
                    x = "SSID:" + apwl.wlan_ssid + ","
                    a['wlan_ssid'] = apwl.wlan_ssid
                if apwl.wlan_service != i.wlan_service and apwl.wlan_service != "":
                    if apwl.wlan_service == "on":
                        x = x + "Enable" + ","
                    if apwl.wlan_service == "off":
                        x = x + "Disable" + ","
                if apwl.passphrase != i.passphrase and apwl.passphrase != "":
                    x = x + "PSK2" + ","
                if apwl.vlan_enabled != i.vlan_enabled and apwl.vlan_enabled != "":
                    if apwl.vlan_enabled == "on":
                        x = x + "VLAN:Enable" + ","
                    if apwl.vlan_enabled == "off":
                        x = x + "VLAN:Disable" + ","
                if apwl.vlan != i.vlan and apwl.vlan != "":
                    x = x + "VLANID:" + apwl.vlan +","
                if apwl.upload_speed != i.upload_speed and apwl.upload_speed != "":
                    x = x + "UPLOAD SPEED:" + apwl.upload_speed + ","
                if apwl.download_speed != i.download_speed and apwl.download_speed != "":
                    x = x + "DOWNLOAD SPEED:" + apwl.download_speed + ","
                if apwl.radios_enable != i.radios_enable and apwl.radios_enable != "":
                    if apwl.radios_enable == "both":
                        x = x + "APPLY:BOTH" + ","
                    if apwl.radios_enable == "5g":
                        x = x + "APPLY:5G" + ","
                    if apwl.radios_enable == "2g":
                        x = x + "APPLY:2G" + ","
                a["change"] = x
            except Exception as e:
                print e
                a["change"] = ""
            a['ap_id'] = ap_id
            ret['data'].append(a)
    except Exception as e:
        print e

    return JsonResponse(ret,safe=False)

@login_required
def nonoperate_wlan_show_detail(request):
    wl_id = int(request.GET.get('id'))
    ap_id = int(request.GET.get('ap_id'))
    wlan = Group_wlan.objects.get(pk = wl_id)
    ret = model_to_dict(wlan)
    if Device.objects.get(pk = ap_id).device_wlan_set.filter(wlan_id = wlan.wlan_id).exists():
        dw = Device.objects.get(pk = ap_id).device_wlan_set.get(wlan_id = wlan.wlan_id)
        if dw.wlan_ssid != wlan.wlan_ssid and dw.wlan_ssid != "":
            ret['wlan_ssid'] = dw.wlan_ssid
        if dw.wlan_service != wlan.wlan_service and dw.wlan_service != "":
            ret['wlan_service'] = dw.wlan_service
        if dw.passphrase != wlan.passphrase and dw.passphrase != "":
            ret['passphrase'] = dw.passphrase
        if dw.vlan_enabled != wlan.vlan_enabled and dw.vlan_enabled != "":
            ret['vlan_enabled'] = dw.vlan_enabled
        if dw.vlan != wlan.vlan and dw.vlan != "":
            ret['vlan'] = dw.vlan
        if dw.upload_speed != wlan.upload_speed and dw.upload_speed != "":
            ret['upload_speed'] = dw.upload_speed
        if dw.download_speed != wlan.download_speed and dw.download_speed != "":
            ret['download_speed'] = dw.download_speed
        if dw.radios_enable != wlan.radios_enable and dw.radios_enable != "":
            ret['radios_enable'] = dw.radios_enable

    if wlan.encry_type == "":
        ret['pass_sign'] = "0"
    else:
        ret['pass_sign'] = "1"
    return JsonResponse(ret,safe=False)

@login_required
def nonoperate_apply_change_wlan(request):
    wl_id = int(request.GET.get('id'))
    ap_id = int(request.GET.get('ap_id'))
    errors = ""
    if request.method == 'POST':
        wlan_ssid = request.POST.get('ssid')
        if request.POST.get('wlan_service') == None:
            wlan_service = "off"
        else:
            wlan_service = request.POST.get('wlan_service')
        passphrase = request.POST.get('passphrase')
        if request.POST.get('vlan_enabled') == None:
            vlan_enabled = "off"
        else:
            vlan_enabled = request.POST.get('vlan_enabled')
        if request.POST.get('vlan') == '':
            vlan = '0'
        else:
            vlan = request.POST.get('vlan')
        upload_speed = request.POST.get('upload_speed')
        download_speed = request.POST.get('download_speed')
        radios_enable = request.POST.get('radios_enable')
        gw = Group_wlan.objects.get(pk = wl_id)

        wlan_ssid = "" if wlan_ssid == gw.wlan_ssid else wlan_ssid
        wlan_service = "" if wlan_service == gw.wlan_service else wlan_service
        passphrase = "" if passphrase == gw.passphrase else passphrase
        vlan_enabled = "" if vlan_enabled == gw.vlan_enabled else vlan_enabled
        vlan = "" if vlan == gw.vlan else vlan
        upload_speed = "" if upload_speed == gw.upload_speed else upload_speed
        download_speed = "" if download_speed == gw.download_speed else download_speed
        radios_enable = "" if radios_enable == gw.radios_enable else radios_enable

        try:
            if Device_wlan.objects.filter(device_id = ap_id,wlan_id = gw.wlan_id).exists():
                dw = Device_wlan.objects.get(device_id = ap_id,wlan_id = gw.wlan_id)
                if (wlan_ssid != "" or dw.wlan_ssid != "") or (wlan_service != "" or dw.wlan_service != "") or (passphrase != "" or dw.passphrase != "") or (vlan_enabled != "" or dw.vlan_enabled != "") or (vlan != "" or dw.vlan != "") or (wlan_ssid != "" or dw.wlan_ssid != "") or (upload_speed != "" or dw.upload_speed != "") or (download_speed != "" or dw.download_speed != "") or (radios_enable != "" or dw.radios_enable != "") :
                    dw.device_id = ap_id
                    dw.wlan_id = gw.wlan_id
                    dw.wlan_ssid = wlan_ssid
                    dw.wlan_service = wlan_service
                    dw.passphrase = passphrase
                    dw.vlan_enabled = vlan_enabled
                    dw.vlan = vlan
                    dw.upload_speed = upload_speed
                    dw.download_speed = download_speed
                    dw.radios_enable = radios_enable
                    dw.save()
            else:
                dw = Device_wlan()
                if wlan_ssid != "" or wlan_service != "" or passphrase != "" or vlan_enabled != "" or vlan != "" or wlan_ssid != "" or upload_speed != "" or download_speed != "" or radios_enable != "" :
                    dw.device_id = ap_id
                    dw.wlan_id = gw.wlan_id
                    dw.wlan_ssid = wlan_ssid
                    dw.wlan_service = wlan_service
                    dw.passphrase = passphrase
                    dw.vlan_enabled = vlan_enabled
                    dw.vlan = vlan
                    dw.upload_speed = upload_speed
                    dw.download_speed = download_speed
                    dw.radios_enable = radios_enable
                    dw.save()
                    # errors = _(u'创建成功！')
            m = Device.objects.get(pk = ap_id).mac
            create_wireless_json(Device.objects.get(mac = m))
            # data_str = set_wlan_str(m)
            # DataQuery.DQSetUpdateConfig('wireless',m,data_str)

            gp_str = portal_str(request.user.groupname)
            mac_list = []
            mac_list.append(m)
            task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')
        except Exception as e:
            print e
            errors = _(u'修改失败！')
    return JsonResponse(errors,safe=False)

@login_required
def nonoperate_reduction_change_wlan(request):
    wl_id = int(request.POST.get('id'))
    ap_id = int(request.POST.get('ap_id'))
    errors = ""
    try:
        wlan = Group_wlan.objects.get(pk = wl_id)
        if Device_wlan.objects.filter(device_id = ap_id,wlan_id = wlan.wlan_id).exists():
            apwl = Device_wlan.objects.get(device_id = ap_id,wlan_id = wlan.wlan_id)
            apwl.delete()
            m = Device.objects.get(pk = ap_id).mac
            create_wireless_json(Device.objects.get(pk = ap_id))
            # data_str = set_wlan_str(m)
            # DataQuery.DQSetUpdateConfig('wireless',m,data_str)

            gp_str = portal_str(request.user.groupname)
            mac_list = []
            mac_list.append(m)
            task_for_global_config_ajax.delay(request.user.groupname,gp_str,mac_list,'mac_list')
        else:
            pass
    except Exception as e:
        errors = _(u'还原失败！')
    return JsonResponse(errors,safe=False)

@login_required
def nonoperate_auto_accept(request):
    accept = request.POST.get('accept')
    error = {'error':''}
    if accept == "open":
        try:
            ss = Setting.objects.get(SIGN = 1)
            ss.nonoperate_auto_accept = 'open'
            ss.save()
            error['error'] = _(u'修改成功')
        except Exception as e:
            print e
            error['error'] = _(u'修改失败')
    elif accept == "get":
        try:
            ss = Setting.objects.get(SIGN = 1)
            result = ss.nonoperate_auto_accept
        except Exception as e:
            print e
            result = ""
        return JsonResponse(result,safe = False)
    else:
        try:
            ss = Setting.objects.get(SIGN = 1)
            ss.nonoperate_auto_accept = 'close'
            ss.save()
            error['error'] = _(u'修改成功')
        except Exception as e:
            print e
            error['error'] = _(u'修改失败')
    return JsonResponse(error,safe= False)


@login_required
def nonoperate_auto_update(request):
    error = {'error':''}
    if request.method == 'GET':
        get_type = request.GET.get('get_type')
        if get_type == 'option':
            result = []
            pg = Probe_group.objects.filter(account_group_name = request.user.groupname,group_type = '3')
            for i in pg:
                result.append({
                    'id': str(i.pk),
                    'text': str(i.group_name),
                })
            return JsonResponse({'results':result},safe = False)
        elif get_type == 'load_result':
            if Auto_Update_Rule.objects.filter(groupname = request.user.groupname,support_mode = '3').exists():
                auto_update_rule = Auto_Update_Rule.objects.get(groupname = request.user.groupname,support_mode = '3')
                result = auto_update_rule.rule
            else:
                result = json.dumps([])
            return JsonResponse(result,safe = False)

    elif request.method == 'POST':
        rule = request.POST.getlist('rule[]')
        if not rule:
            rule = []
        rule = json.dumps(rule)
        print rule
        if Auto_Update_Rule.objects.filter(groupname = request.user.groupname,support_mode = '3').exists():
            auto_update_rule = Auto_Update_Rule.objects.get(groupname = request.user.groupname,support_mode = '3')
        else:
            auto_update_rule = Auto_Update_Rule()
            auto_update_rule.groupname = request.user.groupname
            auto_update_rule.support_mode = '3'
        auto_update_rule.rule = rule
        try:
            auto_update_rule.save()
            error['error'] = _(u'修改成功')
        except Exception as e:
            print e
            error['error'] = _(u'修改失败')
        return JsonResponse(error,safe= False)


@login_required
def nonoperate_blacklist_add(request):
    blacklist = json.loads(request.POST.get('bl'))
    error = {'error':'','su':'false'}
    if blacklist != []:
        try:
            for i in blacklist:
                bl = ApBlackList()
                bl.mac = i
                bl.support_mode = '3'
                bl.save()
            error['error'] = _(u'执行成功')
            error['su'] = 'true'
        except Exception as e:
            print e
            error['error'] = _(u'执行失败')
    return JsonResponse(error,safe=False)

@login_required
def nonoperate_black_list_add_ajax(request):
    dev = Device.objects.filter(support_mode = "3")
    data = []
    result = {}
    for d in dev:
        if ApBlackList.objects.filter(mac = d.mac).exists():
            pass
        else:
            if d.name == "":
                name = d.model + "_" + d.mac[:6]
            else:
                name = d.name
            data.append({'name':name,'mac':d.mac})
    result = {'data':data}
    return JsonResponse(result,safe = False)

@login_required
def nonoperate_black_list_table(request):
    apb = ApBlackList.objects.filter(support_mode = '3')
    data = []
    for i in apb:
        if Device.objects.filter(mac = i.mac).exists():
            d = Device.objects.get(mac = i.mac)
        else:
            d = ''
        if d == "":
            name = i.mac[:2] +":"+ i.mac[2:4] +":"+ i.mac[4:6] +":"+ i.mac[6:8] +":"+ i.mac[8:10] +":"+ i.mac[10:]
        else:
            if d.name == "":
                name = d.model + "_" + d.mac[:6]
            else:
                name = d.name
        ap = model_to_dict(i)
        ap['name'] = name
        data.append(ap)
    return JsonResponse({'data':data},safe=False)

@login_required
def nonoperate_blacklist_remove(request):
    mac = request.POST.get('mac')
    error = {'error':''}
    try:
        ApBlackList.objects.get(mac = mac).delete()
        error['error'] = _(u'操作成功')
    except Exception as e:
        print e
        error['error'] = _(u'操作失败')
    return JsonResponse(error,safe=False)

@login_required
def nonoperate_customer_list(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)

    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,"nonoperate-customer.html",{'admin_error_json':admin_error_json,'admin_error':admin_error,'error':error,"error_json":error_json,'plimit':plimit,'oemlimit':oemlimit})

@login_required
def nonoperate_customer_table(request):
    # print 2131231,request.GET
    draw = int(request.GET.get('draw'))
    start = int(request.GET.get('start'))
    length = int(request.GET.get('length'))
    order_id = request.GET.get('order[0][column]')
    order = request.GET.get('columns['+order_id+'][data]')
    customer_type_receive = request.GET.get('customer_type')

    # radio id  0-> 2G 1-> 5G
    if customer_type_receive == '2G':
        customer_type_1 = 0
        customer_type_2 = 0
    elif customer_type_receive == '5G':
        customer_type_1 = 1
        customer_type_2 = 1
    else:
        customer_type_1 = 0
        customer_type_2 = 1

    if request.GET.get('order[0][dir]') == 'asc':
        order_type = ''
    elif request.GET.get('order[0][dir]') == 'desc':
        order_type = '-'
    search_value_old = request.GET.get('search[value]')
    search_value = search_value_old

    online_time = int(time.time()) - 900
    p_list = []
    if request.user.administrator_permission >= 4:
        p_list = [i.mac for i in Device.objects.filter(support_mode = 3)]

    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        p_list = [i.mac for i in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = 3))]
    if len(p_list) == 0:
        recordsFiltered = 0
        recordsTotal = 0
        customer_list = []
    else:
        # # 解决单元素元祖 末尾,问题
        # if len(p_list) == 1:
        #     p_list = p_list + p_list
        recordsTotal = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list)&(Q(radioid = customer_type_1)|Q(radioid = customer_type_2))).count()
        if search_value == '':
            recordsFiltered = recordsTotal
            customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list)&(Q(radioid = customer_type_1)|Q(radioid = customer_type_2))).order_by(order_type+order)[start:start+length]
        else:
            recordsFiltered = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list)&(Q(radioid = customer_type_1)|Q(radioid = customer_type_2))).filter(Q(ip__icontains=search_value)|Q(mac__icontains=search_value)|Q(ap__icontains=search_value)|Q(uphone__icontains=search_value)).count()
            customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap__in = p_list)&(Q(radioid = customer_type_1)|Q(radioid = customer_type_2))).filter(Q(ip__icontains=search_value)|Q(mac__icontains=search_value)|Q(ap__icontains=search_value)|Q(uphone__icontains=search_value))[start:start+length]


    ret = {'draw':draw,'recordsTotal':recordsTotal,'recordsFiltered':recordsFiltered,'data':[]}
    timenow = timezone.now()
    for al in customer_list:

        # al.last_heart_time = time.localtime(al.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")
        al.livetime = sec_to_hour(al.livetime)
        al.uptime = sec_to_hour(al.uptime)
        data = model_to_dict(al)
        if Customer_name.objects.filter(mac = al.mac).exists():
            cus = Customer_name.objects.get(mac = al.mac).name
        else:
            cus = ""
        if cus != "":
            name = cus
        elif al.hostname != "":
            name = al.hostname
        elif al.devtype != "":
            name = al.devtype +'_'+ al.mac
        else:
            name = al.mac[:2] +":"+ al.mac[2:4] +":"+ al.mac[4:6] +":"+ al.mac[6:8] +":"+ al.mac[8:10] +":"+ al.mac[10:]
        data['name'] = name

        if Device.objects.filter(mac = al.ap).exists():
            dev = Device.objects.get(mac = al.ap)
            if Device_wlan.objects.filter(device_id = dev.pk,wlan_id = str(al.wlanid-1)).exists():
                wlan = Device_wlan.objects.get(device_id = dev.pk,wlan_id = str(al.wlanid-1)).wlan_ssid
                if wlan == "":
                    if Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).exists():
                        wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).wlan_ssid
            elif Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).exists():
                wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).wlan_ssid
            else:
                wlan = ""
            if dev.name != "":
                ap_dev = dev.name
            else:
                ap_dev = dev.model + "_" + dev.mac

            a = int((timenow - dev.last_heart_time).total_seconds())
            dev.last_heart_time = timezone.localtime(dev.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")
            rss = dev.reboot_sign.split(",")
            if dev.last_heart_time == "1970-1-1 08:00:00":
                state = _(u"离线")
            elif a >= 86400:
                state = _(u"退服")
            elif a >= 3600:
                state = _(u"离线")
            elif a >= 900 :
                state = _(u"超时")
            elif a < 900:
                if rss[0] != '0' and rss[0] != '':
                    if rss[0] == '1':
                        state = _(u"重启")
                    elif rss[0] == '2':
                        state = _(u"升级")
                else:
                    state = _(u"在线")
            dev_id = dev.pk
        else:
            wlan = ""
            ap_dev = ""
            state = _(u"退服")
            dev_id = 0

        oemlimit = get_oemlimit()
        # print "LLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLL", oemlimit, oemlimit['oem_type']
        if oemlimit['oem_type'] == 'qingdao':
            data['wlan'] = ''
            # print "QQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQ", data['wlan']
        else:
            data['wlan'] = wlan
        data['ap_dev'] = ap_dev
        data['state'] = state
        data['dev_id'] = dev_id
        ret['data'].append(data)

    return JsonResponse(ret,safe = False)

@login_required
def nonoperate_customer_tab_table(request):
    mac = request.GET.get('mac')
    ret = {'data':[]}

    online_time = int(time.time()) - 900

    customer_list = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = 'online')&Q(ap = mac))

    print 'qwqwqwqw',customer_list
    for al in customer_list:
        # al.last_heart_time = time.localtime(al.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")
        # al.livetime = sec_to_hour(al.livetime)
        data = model_to_dict(al)
        if Customer_name.objects.filter(mac = al.mac).exists():
            cus = Customer_name.objects.get(mac = al.mac).name
        else:
            cus = ""
        if cus != "":
            name = cus
        elif al.hostname != "":
            name = al.hostname
        else:
            name = al.devtype +'_'+ al.mac
        data['name'] = name
        if Device.objects.filter(mac = al.ap).exists():
            dev = Device.objects.get(mac = al.ap)
            if Device_wlan.objects.filter(device_id = dev.pk,wlan_id = str(al.wlanid-1)).exists():
                wlan = Device_wlan.objects.get(device_id = dev.pk,wlan_id = al.wlanid-1).wlan_ssid
                if wlan == "":
                    if Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).exists():
                        wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).wlan_ssid
            elif Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).exists():
                wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(al.wlanid-1)).wlan_ssid
            else:
                wlan = ""
            if dev.name != "":
                ap_dev = dev.name
            else:
                ap_dev = dev.model + "_" + dev.mac
        else:
            wlan = ""
            ap_dev = ""

        oemlimit = get_oemlimit()
        if oemlimit['oem_type'] == 'qingdao':
            data['wlan'] = ''
        else:
            data['wlan'] = wlan
        data['ap_dev'] = ap_dev

        ret['data'].append(data)
    return JsonResponse(ret,safe = False)

@login_required
def nonoperate_customer_detail(request):
    ret = {}
    pk = request.GET.get('id')
    if Customer.objects.filter(pk = pk).exists():
        ret = model_to_dict(Customer.objects.get(pk = pk))
    else:
        ret = {}
    if ret != {}:
        ret['last_heart_time'] = time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(ret['last_heart_time']))
        ret['livetime'] = sec_to_hour(ret['livetime'])
        ret['uptime'] = sec_to_hour(ret['uptime'])
        if Customer_name.objects.filter(mac = ret['mac']).exists():
            cus = Customer_name.objects.get(mac = ret['mac']).name
        else:
            cus = ""
        if cus != "":
            name = cus
        elif ret['hostname'] != "":
            name = ret['hostname']
        elif ret['devtype'] != "":
            name = ret['devtype'] +'_'+ ret['mac']
        else:
            name = ret['mac'][:2] +":"+ ret['mac'][2:4] +":"+ret['mac'][4:6] +":"+ ret['mac'][6:8] +":"+ ret['mac'][8:10] +":"+ ret['mac'][10:]
        ret['name'] = name

        if Device.objects.filter(mac = ret['ap']).exists():
            dev = Device.objects.get(mac = ret['ap'])
            if Device_wlan.objects.filter(device_id = dev.pk,wlan_id = str(ret['wlanid']-1)).exists():
                wlan = Device_wlan.objects.get(device_id = dev.pk,wlan_id = str(ret['wlanid']-1)).wlan_ssid
                if wlan == "":
                    if Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(ret['wlanid']-1)).exists():
                        wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(ret['wlanid']-1)).wlan_ssid
            elif Group_wlan.objects.filter(group_id_id = dev.group_id,wlan_id = str(ret['wlanid']-1)).exists():
                wlan = Group_wlan.objects.get(group_id_id = dev.group_id,wlan_id = str(ret['wlanid']-1)).wlan_ssid
            else:
                wlan = ""
            if dev.name != "":
                ap_dev = dev.name
            else:
                ap_dev = dev.model + "_" + dev.mac
            timenow = timezone.now()
            a = int((timenow - dev.last_heart_time).total_seconds())
            dev.last_heart_time = timezone.localtime(dev.last_heart_time).strftime("%Y-%m-%d %H:%M:%S")
            rss = dev.reboot_sign.split(",")
            if dev.last_heart_time == "1970-1-1 08:00:00":
                state = _(u"离线")
            elif a >= 86400:
                state = _(u"退服")
            elif a >= 3600:
                state = _(u"离线")
            elif a >= 900 :
                state = _(u"超时")
            elif a < 900:
                if rss[0] != '0' and rss[0] != '':
                    if rss[0] == '1':
                        state = _(u"重启")
                    elif rss[0] == '2':
                        state = _(u"升级")
                else:
                    state = _(u"在线")
            dev_id = dev.pk

            # if hasattr(dev,'device_ap'):
            #     if ret['radioid'] == 0:
            #         channel = dev.device_ap.radios_2_channel
            #     else:
            #         channel = dev.device_ap.radios_5_channel
            # else:
            #     channel = 'auto'
        else:
            wlan = ""
            ap_dev = ""
            state = _(u"退服")
            dev_id = 0
            # channel = 'auto'

        ret['wlan'] = wlan
        ret['ap_dev'] = ap_dev
        ret['state'] = state
        ret['dev_id'] = dev_id
        # ret['channel'] = channel

    return JsonResponse(ret,safe = False)

@login_required
def nonoperate_custome_detail_table(request):
    mac = request.GET.get('mac')
    ret = {'data':[]}

    customer_list = Customer_history.objects.filter(mac = mac).order_by('-last_heart_time')
    # al_last = {}
    for al in customer_list:
        if al.online_sign == 'offline' and al.uptime >= 300:
            data = {
                'livetime':sec_to_hour(al.uptime),
                'last_heart_time':time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(al.last_heart_time)),
                'tx_bytes':al.down,
                'rx_bytes':al.up,
            }
            ret['data'].append(data)
    # al_last = {}
    # for al in customer_list:
    #     if al_last == {}:
    #         al_last = model_to_dict(al)
    #     else:
    #         if al_last['online_sign'] == 'offline':
    #             if al.online_sign == "online":
    #                 data = {
    #                     'livetime':sec_to_hour(abs(al_last['last_heart_time'] - al.last_heart_time)),
    #                     'last_heart_time':time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(al_last['last_heart_time'])),
    #                     'tx_bytes':al_last['down'],
    #                     'rx_bytes':al_last['up'],
    #                 }
    #                 ret['data'].append(data)
    #                 al_last = model_to_dict(al)
    #             else :
    #                 al_last = model_to_dict(al)
    #                 al_last['online_sign'] = "offline"
    #         else:
    #             al_last = model_to_dict(al)

    return JsonResponse(ret,safe = False)


@login_required
def nonoperate_change_customer_name(request):
    mac = request.POST.get('mac')
    name = request.POST.get('name')
    errors = {'sign':'false','mes':''}

    if Customer_name.objects.filter(mac = mac).exists():
        cus = Customer_name.objects.get(mac = mac)
    else:
        cus = Customer_name()
        cus.mac = mac
    cus.name = name
    try:
        cus.save()
        errors['sign'] = 'true'
        errors['mes'] = _(u'修改成功')
    except Exception as e:
        print e
        errors['sign'] = 'false'
        errors['mes'] = _(u'修改失败')

    return JsonResponse(errors,safe = False)


def sec_to_hour(value):
    try:
        if value != '' and value != 0:
            a = int(value)
            b = a/(24*3600)
            c = a%(24*3600)/3600
            d = a%(24*3600)%3600/60
            e = a%(24*3600)%3600%60
            if b == 0 and c == 0 and d == 0:
                return "%ds"%(e)
            elif b == 0 and c == 0:
                return "%dm %ds"%(d,e)
            elif b == 0:
                return "%dh %dm %ds"%(c,d,e)
            else:
                return "%dd %dh %dm %ds"%(b,c,d,e)
        else:
            return value
    except Exception as e:
        print e
        return value


@login_required
def nonoperate_index(request):
    admin_error = {"error_type":'',"error_msg":''}
    p_all = ''
    onldevs = 0
    ofldevs = 0
    p_list = []
    logger.info('start run sql --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    dev_online_time = timezone.now() - datetime.timedelta(minutes = 15)
    if request.user.administrator_permission >= 4 :
        p_all = Device.objects.filter(support_mode = 3)
        logger.info('end run sql and start run list --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        p_list = [i.mac for i in p_all]
        logger.info('end run list --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        ##########xuyaoxiugai##################
        # onldevs = DataQuery.DQGetOnlinedDevs(p_list)
        # ofldevs = DataQuery.DQGetOfflinedDevs(p_list)
        onldevs = Device.objects.filter(Q(support_mode = 3) & Q(last_heart_time__gte = dev_online_time)).count()
        logger.info('end run onldevs --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        ofldevs = Device.objects.filter(Q(support_mode = 3) & Q(last_heart_time__lt = dev_online_time)).count()
        logger.info('end run ofldevs --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        p_all = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3'))
        p_list = [i.mac for i in p_all]
        ##########xuyaoxiugai##################
        # onldevs = DataQuery.DQGetOnlinedDevs(p_list)
        # ofldevs = DataQuery.DQGetOfflinedDevs(p_list)
        onldevs = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = 3) & Q(last_heart_time__gte = dev_online_time)).count()
        ofldevs = Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = 3) & Q(last_heart_time__lt = dev_online_time)).count()

    online_time = int(time.time()) - 900
    logger.info('start run usercountsql --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    # pc_user = DataQuery.DQGetAPCache_nojson(request.user.groupname+"cache_pc_user")
    pc_user = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = "online")&Q(ap__in = p_list)&Q(devtype__in = ['PC','Mac PC'])).count()
    logger.info('end run pcuser --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    # phone_user = DataQuery.DQGetAPCache_nojson(request.user.groupname+"cache_phone_user")
    # phone_user = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = "online")&Q(ap__in = p_list)&Q(devtype__in = ['iphone','ipad','android'])).count()
    phone_user = Customer.objects.filter(Q(last_heart_time__gte = online_time)&Q(online_sign = "online")&Q(ap__in = p_list)).count() - pc_user
    logger.info('end run phoneuser --- time:{}'.format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    ap = {'online':onldevs,'offline':ofldevs,'pc_user':pc_user,'phone_user':phone_user}

    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'nonoperate_index.html',{'ap':ap,'admin_error_json':admin_error_json,'admin_error':admin_error,'plimit':plimit,'oemlimit':oemlimit})

def nonoperate_sort_type_change(request):
    sort_type = request.GET.get('type')
    top = []
    p_list = []
    # if request.user.administrator_permission >= 4 :
    #     p_list = [i.mac for i in Device.objects.filter(support_mode = '3')]
    # elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
    #     # p_all = Device.objects.filter(account_group_name = request.user.groupname)
    #     p_list = [i.mac for i in Device.objects.filter(Q(account_group_name = request.user.groupname)&Q(support_mode = '3'))]
    #########xuyaoxiugai##################
    if sort_type == '1':
        # top1 = {'mac':'a4e6b1300005','hour':'1h30min22s'}
        # top2 = {'mac':'fcad0f0344b1','hour':'1h30min25s'}
        # top = [top1,top2]
        # top = DataQuery.DQGetTopItem("all", "onlinetime", p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_onlinetime")
    elif sort_type == '2':
        # top2 = {'mac':'a4e6b1300005','upload':'11111mb'}
        # top1 = {'mac':'fcad0f0344b1','upload':'111121mb'}
        # top = [top1,top2]
        # top = DataQuery.DQGetTopItem("all", "upload", p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_upload")
    elif sort_type == '3':
        # top2 = {'mac':'a4e6b1300005','upload':'11111mb'}
        # top1 = {'mac':'fcad0f0344b1','upload':'111121mb'}
        # top = [top1,top2]
        # top = DataQuery.DQGetTopItem("all", "download", p_list)
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_download")
    try:
        for t in top:
            name = ''
            dev_id = ''
            state = ''
            if Device.objects.filter(mac = t['mac']).exists():
                al = Device.objects.get(mac = t['mac'])
                name = al.name
                dev_id = al.pk
                state = state_check(t['mac'])
            if name == '':
                name = Device.objects.get(mac = t['mac']).model+ '_' + t['mac'][6:12]
            t['name'] = name
            t['dev_id'] = dev_id
            t['state'] = state
    except Exception as e:
        print e

    return JsonResponse(top,safe = False)

def nonoperate_chartajax(request):
    net_type = request.GET.get('type')
    p_list = []
    data = {
        'type':'3G',
        'mac':[],
        'upload':[],
        'download':[],
        'name':[],
    }
    if net_type == '3g':
        data = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_net_type_3g")
        data['type'] = '3G'
    elif net_type == '4g':
        data = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_net_type_4g")
        data['type'] = '4G'
    elif net_type == 'car':
        data = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_net_type_car")
        data['type'] = _(u'车载')
    elif net_type == 'normal':
        data = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_net_type_normal")
        data['type'] = _(u'普通')
    try:
        for t in data['mac']:
            name = ''
            if Device.objects.filter(mac = t).exists():
                name = Device.objects.get(mac = t).name
            if name == '':
                name = Device.objects.get(mac = t).model+ '_' + t[6:12]
            data['name'].append(name)
    except Exception as e:
        print e

    return JsonResponse(data)

def nonoperate_user_chartajax(request):

    data = {
        'mac':[],
        'num':[],
        'name':[],
        'upload':[],
        'download':[],
    }
    p_list = []
    data = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_user_chart")
    try:
        for t in data['mac']:
            name = ''
            if Device.objects.filter(mac = t).exists():
                name = Device.objects.get(mac = t).name
            if name == '':
                name = Device.objects.get(mac = t).model+ '_' + t[6:12]
            data['name'].append(name)
        for uflow in data['uflow']:
            if uflow['z'] == 0:
                uflow['z'] = _(u'小于1')
            if uflow['w'] == 0:
                uflow['w'] = _(u'小于1')
    except Exception as e:
        print e
    print data,'xxxxxxxxxxxx'
    return JsonResponse(data)

def nonoperate_ssid_select(request):
        group_list = []
        for i in Probe_group.objects.filter(account_group_name = request.user.groupname, group_type = 3):
            group_list.append(i.pk)
        # print group_list, 'GGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGG'
        # for i in group_list:
        #     print i, 'GPGPGPGPGPGPGPGPGPGPGP'
        wlan_list = []
        if len(group_list) != 0:
            for i in Group_wlan.objects.filter(group_id__in = group_list):
                wlan_list.append({'group_id':i.group_id.pk,'wlan_id':i.wlan_id,'wlan_ssid':i.wlan_ssid,'group_name':i.group_id.group_name})
        print wlan_list, 'WWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW'
        return JsonResponse(wlan_list,safe = False)

def nonoperate_user_counterajax(request):
    net_type = request.GET.get('type')
    print 'NNNNNNNNNNNNN', net_type, 'NNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNN'

    gpwl = net_type.split(';')
    if len(gpwl) == 2:
        groupId = gpwl[0]
        wlanId = gpwl[1]
    else:
        # defalt group id, wlan id
        groupId = '5'
        wlanId = '1'

    data = {
        'name':[],
        'time':[],
        'num':[],
        'upload':[],
        'download':[],
        'uflow':[],
    }
    p_list = []
    # data = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_user_chart")
    # try:
    #     for t in data['mac']:
    #         name = ''
    #         if Device.objects.filter(mac = t).exists():
    #             name = Device.objects.get(mac = t).name
    #         if name == '':
    #             name = Device.objects.get(mac = t).model+ '_' + t[6:12]
    #         data['name'].append(name)
    # except Exception as e:
    #     print e



    # data['time'] = ['17:00', '18:00', '19:00', '20:00', '21:00', '22:00', '23:00']
    # if Account_Group.objects.filter(groupname = request.user.groupname).exists():
    #     account_group = Account_Group.objects.get(groupname = request.user.groupname)
    # group_list = []
    # for i in Probe_group.objects.filter(account_group_name = request.user.groupname, group_type = 3):
    #     group_list.append(i.pk)
    # # print group_list, 'GGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGGG'
    # # for i in group_list:
    # #     print i, 'GPGPGPGPGPGPGPGPGPGPGP'
    # wlan_list = []
    # if len(group_list) != 0:
    #     for i in Group_wlan.objects.filter(group_id__in = group_list):
    #         wlan_list.append({'group_id':i.group_id.pk,'wlan_id':i.wlan_id,'wlan_ssid':i.wlan_ssid,'group_name':i.group_id.group_name})

    # print wlan_list, 'WWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW'

    # groupId = '5'
    # wlanId = '3'
    timstamp = int(time.time())
    timstamp = timstamp - timstamp%3600
    uflow = []
    uflowdic = {}
    n = range(24)
    for i in reversed(n):
        timetmp = timstamp - 3600*i
        # print i, time.strftime("%H:%M", time.localtime(timetmp))
        data['time'].append(time.strftime("%H:%M", time.localtime(timetmp)))
        # data['num'].append(DataQuery.DQGetUsersByGroupID(groupId, wlanId, timetmp))
        # [
        #     {'y': 100, 'z': 1000},
        #     {'y': 1020, 'z': 1000}
        # ]
        # data['upload'].append(float(DataQuery.DQGetUsersDownLoadByGroupID(groupId, wlanId, timetmp))/1024/1024)
        # data['download'].append(float(DataQuery.DQGetUsersUpLoadByGroupID(groupId, wlanId, timetmp))/1024/1024)
        uflowdic['y'] = int(DataQuery.DQGetUsersByGroupID(groupId, wlanId, timetmp))

        uflowdic['z'] = int(float(DataQuery.DQGetUsersUpLoadByGroupID(groupId, wlanId, timetmp))/1024/1024)

        uflowdic['w'] = int(float(DataQuery.DQGetUsersDownLoadByGroupID(groupId, wlanId, timetmp))/1024/1024)
        uflow.append(uflowdic)
        uflowdic = {}

    # data['num'] = [123, 435, 123, 3457, 2456, 212, 1123]
    # data['upload'] = [456, 435, 3456, 3457, 3456, 2112, 3123]
    # data['download'] = [658, 435, 769, 3457, 3246, 2112, 1223]
    # data['name'] = ['witrusty']
    data['uflow'] = uflow
    for i in data['uflow']:
        if i['y'] != 0:
            if i['z'] == 0:
                i['z'] = _(u'小于1')
            if i['w'] == 0:
                i['w'] = _(u'小于1')
    print data,'PPPPPPPPPPPPPPPPPPPPPPPPPPPP'
    return JsonResponse(data)

def nonoperate_warningajax(request):
    warning_type = request.GET.get('type')
    top = []

    if warning_type == 'cpu':
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_cpu")
    elif warning_type == 'flash':
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_flash")
    elif warning_type == 'memory':
        top = DataQuery.DQGetAPCache_json(request.user.groupname+"nonoperate_cache_memory")

    try:
        for t in top:
            name = ''
            state = ''
            dev_id = ''
            if Device.objects.filter(mac = t['mac']).exists():
                name = Device.objects.get(mac = t['mac']).name
                dev_id = Device.objects.get(mac = t['mac']).pk
            state = state_check(t['mac'])
            if name == '':
                name = Device.objects.get(mac = t['mac']).model+ '_' + t['mac'][6:12]
            t['name'] = name
            t['state'] = state
            t['dev_id'] = dev_id
    except Exception as e:
        print e
    return JsonResponse(top,safe = False)


def ap_wlan_create_get_portal_config(request):
    mode = request.GET.get('mode')
    data = {'auth_server':''}
    if mode == 'open_form':
        groupname = Account_Group.objects.get(groupname = request.user.groupname)
        if Guest_policy.objects.filter(domainname_id = groupname.pk).exists():
            auth_server = json.loads(Guest_policy.objects.get(domainname_id = groupname.pk).auth_server)
            data['auth_server'] = auth_server
    return JsonResponse(data,safe = False)








@login_required
def nonoperate_user_policy_config(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)

    wlan = ''
    if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
        wlan = Group_wlan.objects.filter(group_id__group_type = '3')
    elif request.user.administrator_permission < 4 and request.user.group_status == 1:
        wlan = Group_wlan.objects.filter(group_id__account_group_name = request.user.groupname,group_id__group_type = '3')

    plimit = get_plimit()
    return render(request,'nonoperate-user-policy-config.html',{'admin_error_json':admin_error_json,'admin_error':admin_error,'wlan':wlan,'error':error,"error_json":error_json,'plimit':plimit})


def accessed_device2authpuppy(mac):
    try:
        db = MySQLdb.connect("localhost", settings_py.DATABASES['default']['USER'], settings_py.DATABASES['default']['PASSWORD'], "authpuppy" )
    except MySQLdb.Error, e:
        try:
            sqlError =  "Error %d:%s" % (e.args[0], e.args[1])
        except IndexError:
            print "MySQL Error:%s" % str(e)
        return False

    cursor = db.cursor()

    try:
        mac = (mac[:2] + ':' + mac[2:4] + ':' + mac[4:6] + ':' + mac[6:8] + ':' + mac[8:10] + ':' + mac[10:]).upper()
        sql = "INSERT into nodes(name,gw_id,created_at,updated_at,domainname) values ('','{}','0000-00-00 00:00:00','0000-00-00 00:00:00','admin')".format(mac)

    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.execute(sql)
    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print sql
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.close()
        db.commit()
    except Exception as e:
        print e
        cursor.close()
        db.rollback()

    db.close()


def del_device2authpuppy(mac):
    try:
        db = MySQLdb.connect("localhost", settings_py.DATABASES['default']['USER'], settings_py.DATABASES['default']['PASSWORD'], "authpuppy" )
    except MySQLdb.Error, e:
        try:
            sqlError =  "Error %d:%s" % (e.args[0], e.args[1])
        except IndexError:
            print "MySQL Error:%s" % str(e)
        return False

    cursor = db.cursor()

    try:
        mac = (mac[:2] + ':' + mac[2:4] + ':' + mac[4:6] + ':' + mac[6:8] + ':' + mac[8:10] + ':' + mac[10:]).upper()
        # sql = "INSERT into nodes(name,gw_id,created_at,updated_at,domainname) values ('','{}','0000-00-00 00:00:00','0000-00-00 00:00:00','admin')".format(mac)
        sql = "DELETE from nodes where gw_id = '{}'".format(mac)

    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.execute(sql)
    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print sql
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.close()
        db.commit()
    except Exception as e:
        print e
        cursor.close()
        db.rollback()

    db.close()

def update_device2authpuppy(mac,domainname):
    try:
        db = MySQLdb.connect("localhost", settings_py.DATABASES['default']['USER'], settings_py.DATABASES['default']['PASSWORD'], "authpuppy" )
    except MySQLdb.Error, e:
        try:
            sqlError =  "Error %d:%s" % (e.args[0], e.args[1])
        except IndexError:
            print "MySQL Error:%s" % str(e)
        return False

    cursor = db.cursor()

    try:
        mac = (mac[:2] + ':' + mac[2:4] + ':' + mac[4:6] + ':' + mac[6:8] + ':' + mac[8:10] + ':' + mac[10:]).upper()
        # sql = "INSERT into nodes(name,gw_id,created_at,updated_at,domainname) values ('','{}','0000-00-00 00:00:00','0000-00-00 00:00:00','admin')".format(mac)
        sql = "UPDATE nodes set domainname = '{1}' where gw_id = '{0}'".format(mac,domainname)

    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.execute(sql)
    except Exception as e:
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        print sql
        print e
        print "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"

    try:
        cursor.close()
        db.commit()
    except Exception as e:
        print e
        cursor.close()
        db.rollback()

    db.close()

def portal_str(groupname):
    if Account_Group.objects.filter(groupname = groupname).exists():
        ag = Account_Group.objects.get(groupname = groupname)
        if hasattr(ag,'guest_policy'):
            gp = model_to_dict(Guest_policy.objects.get(domainname = ag))
        else:
            gp = guest_policy_config
    else:
        gp = guest_policy_config

    guest_policy = gp


    white_list = json.loads(guest_policy['white_list'])
    trusted_mac_list = json.loads(guest_policy['trusted_mac_list'])
    if guest_policy['wechatallowed'] == 'true' :
        white_list = list(set(white_list)) + ['long.weixin.qq.com','short.weixin.qq.com','szlong.weixin.qq.com','szshort.weixin.qq.com','mp.weixin.qq.com','res.wx.qq.com']
    # 保序去重
    # new_list = list(set(white_list))
    # new_list.sort(key = white_list.index)
    gp_str = {"bypass": guest_policy['bypass'], "validatetime": int(guest_policy['auth_validate_timeout']), "clienttimeout": int(guest_policy['client_timeout']), "whitelist": list(set(white_list)), "trustedlist": list(set(trusted_mac_list)), "wireless":[],"personaltype":"normal"}
    gp_str = json.dumps(gp_str,sort_keys = True).replace(' ','')

    return gp_str


def user_policy_str(groupname):
    if Account_Group.objects.filter(groupname = groupname).exists():
        ag = Account_Group.objects.get(groupname = groupname)
        if hasattr(ag,'user_policy_config'):
            up = model_to_dict(User_policy_config.objects.get(domainname = ag))
        else:
            up = user_policy_config
    else:
        up = user_policy_config

    user_policy = up

    rssi_max = str(int(float(user_policy['rssi_threshold'])) + 95)
    policy = collections.OrderedDict([("access_policy",str(user_policy['access_policy'])),("band_steer",str(user_policy['band_steering'])),("dual_max_user",str(user_policy['dual_max_user'])),("l2_isolation",str(user_policy['l2_isolation'])),("load_balance",str(user_policy['load_balance'])),("reject_max",str(user_policy['reject_max'])),("roaming_assoc_rssi",str(user_policy['roaming_assoc_rssi'])),("roaming_detect",str(user_policy['roaming_policy'])),("roaming_unassoc_rssi",str(user_policy['roaming_unassoc_rssi'])),("rssi_max",rssi_max),("single_max_user",str(user_policy['single_max_user'])),("thredhold_5g",'0'),("thredhold_5g_rssi",'0')])
    mac_list = []
    if Customer_black_white_switch.objects.filter(groupname = groupname).exists():
        wb_type = Customer_black_white_switch.objects.get(groupname = groupname).switch
    else:
        wb_type = "none"
    if wb_type == "":
        wb_type = "none"
        mac_list = []
    elif wb_type == "black":
        wb_type = "blacklist"
        for i in Customer_black_list.objects.filter(groupname = groupname):
            i.mac = (i.mac[:2] + ':' + i.mac[2:4] + ':' + i.mac[4:6] + ':' + i.mac[6:8] + ':' + i.mac[8:10] + ':' + i.mac[10:]).lower()
            mac_list.append(i.mac)
    elif wb_type == "white":
        wb_type = "whitelist"
        for i in Customer_white_list.objects.filter(groupname = groupname):
            i.mac = (i.mac[:2] + ':' + i.mac[2:4] + ':' + i.mac[4:6] + ':' + i.mac[6:8] + ':' + i.mac[8:10] + ':' + i.mac[10:]).lower()
            mac_list.append(i.mac)
    wb_list = collections.OrderedDict([('maclist',mac_list),('type',wb_type)])

    if Timing_Policy.objects.filter(groupname = groupname).exists():
        time_pilocy = model_to_dict(Timing_Policy.objects.get(groupname = groupname))
    else:
        time_pilocy = {'gpon_light_switch':'off'}
    print 'xxxxx'
    if time_pilocy['gpon_light_switch'] == 'on':
        switch = 'on'
        timer = []
        timer_origin = json.loads(time_pilocy['gpon_timer'])
        print timer_origin,type(timer_origin)
        for i in timer_origin:
            timer.append(collections.OrderedDict([('end',i['end']),('start',i['start']),('weekday',','.join(i['weekday']))]))
    elif time_pilocy['gpon_light_switch'] == 'off':
        switch = 'off'
        timer = ''
    # timer = json.dumps(timer)
    ledtimer = collections.OrderedDict([('switch',switch),('timer',timer)])
    # policy_str = json.dumps({'policy':policy,'list':wb_list,'ledtimer':ledtimer},sort_keys = True).replace(' ','')
    policy_str = json.dumps(collections.OrderedDict([('ledtimer',ledtimer),('list',wb_list),('policy',policy)])).replace(' ','')
    print policy_str,321321
    return policy_str

@login_required
def nonoperate_guest_policy(request):
    admin_error = {"error_type":'',"error_msg":''}
    if request.user.administrator_permission == 6 and request.user.username == 'admin':
        admin_error = Public_function.judgment_admin_pwd(request)
    admin_error_json = json.dumps(admin_error)

    error = {"error_type":'',"error_msg":''}
    try:
        if request.session.get('error_dict') != None:
            error = request.session.get('error_dict')
        # print errors
        del request.session['error_dict']
    except Exception as e:
        pass
    error_json = json.dumps(error)

    wlan = ''
    # if request.user.administrator_permission == 0 or request.user.administrator_permission >= 4:
    #     wlan = Group_wlan.objects.all()
    # elif request.user.administrator_permission < 4 and request.user.group_status == 1:
    #     wlan = Group_wlan.objects.filter(group_id__account_group_name = request.user.groupname,group_id__group_type = '3')
    wlan = Group_wlan.objects.filter(group_id__account_group_name = request.user.groupname,group_id__group_type = '3')
    plimit = get_plimit()
    return render(request,'nonoperate-guest-policy.html',{'admin_error_json':admin_error_json,'admin_error':admin_error,'wlan':wlan,'error':error,"error_json":error_json,'plimit':plimit})




@login_required
def nonoperateByzoroXinyangList(request):
    plimit = get_plimit()
    oemlimit = get_oemlimit()
    return render(request,'nonoperate-byzoroXinyangList.html',{'plimit':plimit,'oemlimit':oemlimit})


def nonoperateByzoroXinyangTableApi(request):
    # print 2131231,request.GET
    draw = int(request.GET.get('draw'))
    start = int(request.GET.get('start'))
    length = int(request.GET.get('length'))
    order_id = request.GET.get('order[0][column]')
    order = request.GET.get('columns['+order_id+'][data]')

    # device_type_receive = request.GET.get('device_type')
    dev_online_time = timezone.now() - datetime.timedelta(minutes = 15)
    # if device_type_receive == 'online':
    #     time_condition = "last_heart_time >= '{0}'".format(dev_online_time)
    # elif device_type_receive == 'offline':
    #     time_condition = "last_heart_time < '{0}'".format(dev_online_time)
    # else:
    #     time_condition = "last_heart_time >= '{0}' or last_heart_time < '{0}'".format(dev_online_time)


    if request.GET.get('order[0][dir]') == 'asc':
        order_type = 'asc'
    elif request.GET.get('order[0][dir]') == 'desc':
        order_type = 'desc'

    # if order == 'state':
    #     order = 'last_heart_time'
    search_value_old = request.GET.get('search[value]')
    search_value = search_value_old

    condition = {
        'order':order,
        # 'device_type_receive':device_type_receive,
        'order_type':order_type,
        'search_value':search_value
    }

    if request.user.administrator_permission >= 4 or request.user.administrator_permission == 0 :
        permissionCondition = "support_mode = '{0}'".format(3)
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        permissionCondition = "account_group_name = '{0}' and support_mode = '{1}'".format(request.user.groupname,3)

    # "SELECT * from (SELECT p.place_name,b.location_encode,p.districtCode,p.policeName,CONCAT(count(d.last_heart_time >= '2018-12-19 19:24:55' or null),'/',count(d.mac)) as state from device_probe_audit_place_status p left join device_probe_audit_basic_status b on p.mac = b.mac left join device_device d on p.mac = d.mac where p.mac in (SELECT mac from device_device d where support_mode = '3') group by p.place_name) as res"

    cursor = connection.cursor()

    cursor.execute("SELECT count(*) from (SELECT p.place_name from device_probe_audit_place_status p left join device_probe_audit_basic_status b on p.mac = b.mac left join device_device d on p.mac = d.mac where p.mac in (SELECT mac from device_device d where {permissionCondition}) group by p.place_name) as res".format(permissionCondition = permissionCondition))
    rawQuery = cursor.fetchone()
    recordsTotal = rawQuery[0]
    if search_value == '':
        recordsFiltered = recordsTotal
        cursor.execute("SELECT * from (SELECT p.place_name,b.location_encode,p.districtCode,p.policeName,CONCAT(count(d.last_heart_time >= '{dev_online_time}' or null),'/',count(d.mac)) as state,if(count(d.last_heart_time >= '{dev_online_time}' or null) > 0,'在线','离线') as placeState from device_probe_audit_place_status p left join device_probe_audit_basic_status b on p.mac = b.mac left join device_device d on p.mac = d.mac where p.mac in (SELECT mac from device_device d where {permissionCondition}) group by p.place_name) as res order by {order} {order_type} limit {start},{end}".format(dev_online_time = dev_online_time,permissionCondition = permissionCondition,order = order,order_type = order_type,start = start,end = start+length))
        resultList = dictfetchall(cursor)
    else:
        search = "place_name like '%{search_value}%' or location_encode like '%{search_value}%' or policeName like '%{search_value}%' or districtCode like '%{search_value}%' or state like '%{search_value}%' or placeState like '%{search_value}%'".format(search_value = search_value)

        cursor.execute("SELECT * from (SELECT p.place_name,b.location_encode,p.districtCode,p.policeName,CONCAT(count(d.last_heart_time >= '{dev_online_time}' or null),'/',count(d.mac)) as state,if(count(d.last_heart_time >= '{dev_online_time}' or null) > 0,'在线','离线') as placeState from device_probe_audit_place_status p left join device_probe_audit_basic_status b on p.mac = b.mac left join device_device d on p.mac = d.mac where p.mac in (SELECT mac from device_device d where {permissionCondition}) group by p.place_name) as res where {search} order by {order} {order_type} limit {start},{end}".format(dev_online_time = dev_online_time,permissionCondition = permissionCondition,order = order,order_type = order_type,start = start,end = start+length,search = search))
        resultList = dictfetchall(cursor)

        cursor.execute("SELECT count(*) from (SELECT p.place_name,b.location_encode,p.districtCode,p.policeName,CONCAT(count(d.last_heart_time >= '{dev_online_time}' or null),'/',count(d.mac)) as state,if(count(d.last_heart_time >= '{dev_online_time}' or null) > 0,'在线','离线') as placeState from device_probe_audit_place_status p left join device_probe_audit_basic_status b on p.mac = b.mac left join device_device d on p.mac = d.mac where p.mac in (SELECT mac from device_device d where {permissionCondition}) group by p.place_name) as res where {search} ".format(dev_online_time = dev_online_time,permissionCondition = permissionCondition,search = search))
        countRes = cursor.fetchone()
        recordsFiltered = countRes[0]

    timenow = timezone.now()

    ret = {'draw':draw,'recordsTotal':recordsTotal,'recordsFiltered':recordsFiltered,'data':resultList,'condition':condition}
    return JsonResponse(ret,safe = False)


def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

def xinyangExportDeviceButtonApi(request):
    extype = request.GET.get('extype')
    extendsCondition = request.GET.get('extendsCondition')
    extendsCondition = json.loads(extendsCondition)
    print extendsCondition,type(extendsCondition)

    # admin
    if request.user.administrator_permission >= 4 or request.user.administrator_permission == 0 :
        permissionCondition = "support_mode = '{0}'".format(3)
    elif request.user.administrator_permission < 4 and request.user.administrator_permission != 0 and request.user.group_status == 1:
        permissionCondition = "account_group_name = '{0}' and support_mode = '{1}'".format(request.user.groupname,3)

    # device_type_receive = extendsCondition['device_type_receive']
    order = extendsCondition['order']
    order_type = extendsCondition['order_type']
    search_value = extendsCondition['search_value']
    dev_online_time = timezone.now() - datetime.timedelta(minutes = 15)
    # if device_type_receive == 'online':
    #     time_condition = "last_heart_time >= '{0}'".format(dev_online_time)
    # elif device_type_receive == 'offline':
    #     time_condition = "last_heart_time < '{0}'".format(dev_online_time)
    # else:
    #     time_condition = "last_heart_time >= '{0}' or last_heart_time < '{0}'".format(dev_online_time)

    cursor = connection.cursor()
    if search_value == '':
        cursor.execute("SELECT * from (SELECT p.place_name,b.location_encode,p.districtCode,p.policeName,CONCAT(count(d.last_heart_time >= '{dev_online_time}' or null),'/',count(d.mac)) as state,if(count(d.last_heart_time >= '{dev_online_time}' or null) > 0,'在线','离线') as placeState from device_probe_audit_place_status p left join device_probe_audit_basic_status b on p.mac = b.mac left join device_device d on p.mac = d.mac where p.mac in (SELECT mac from device_device d where {permissionCondition}) group by p.place_name) as res order by {order} {order_type}".format(dev_online_time = dev_online_time,permissionCondition = permissionCondition,order = order,order_type = order_type))
        resultList = dictfetchall(cursor)
    else:
        search = "place_name like '%{search_value}%' or location_encode like '%{search_value}%' or policeName like '%{search_value}%' or districtCode like '%{search_value}%' or state like '%{search_value}%' or placeState like '%{search_value}%'".format(search_value = search_value)
        cursor.execute("SELECT * from (SELECT p.place_name,b.location_encode,p.districtCode,p.policeName,CONCAT(count(d.last_heart_time >= '{dev_online_time}' or null),'/',count(d.mac)) as state,if(count(d.last_heart_time >= '{dev_online_time}' or null) > 0,'在线','离线') as placeState from device_probe_audit_place_status p left join device_probe_audit_basic_status b on p.mac = b.mac left join device_device d on p.mac = d.mac where p.mac in (SELECT mac from device_device d where {permissionCondition}) group by p.place_name) as res where {search} order by {order} {order_type}".format(dev_online_time = dev_online_time,permissionCondition = permissionCondition,order = order,order_type = order_type,search = search))
        resultList = dictfetchall(cursor)

    timenow = timezone.now()

    wb = Workbook()
    sheet = wb.active
    sheet.title = u"场所信息"
    sheet['A1'].value = u"ID"
    sheet['B1'].value = u"场所名称"
    sheet['C1'].value = u"场所编码"
    sheet['D1'].value = u"所属区县"
    sheet['E1'].value = u"所属派出所"
    sheet['F1'].value = u"设备状态"
    sheet['G1'].value = u"场所状态"

    excel_row = 1
    for res in resultList:
        excel_row += 1
        sheet["A%d" % (excel_row)].value = excel_row - 1
        sheet["B%d" % (excel_row)].value = res['place_name']
        sheet["C%d" % (excel_row)].value = res['location_encode']
        sheet["D%d" % (excel_row)].value = res['districtCode']
        sheet["E%d" % (excel_row)].value = res['policeName']
        sheet["F%d" % (excel_row)].value = res['state']
        sheet["G%d" % (excel_row)].value = res['placeState']

    filename = "./byzorotmp.xlsx"

    exist_file = os.path.exists("./byzorotmp.xlsx")
    if exist_file:
        os.remove(filename)
    wb.save(filename)

    print "5555555555555555555555555555555555555555"
    wrapper = FileWrapper(open(filename, 'rb'))
    response = HttpResponse(wrapper, content_type='application/vnd.ms-excel')
    response['Content-Length'] = os.path.getsize(filename)
    response['Content-Disposition'] = 'attachment; filename=PlacesInfo.xlsx'
    return response
